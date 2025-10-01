   
from flask import Flask, request, redirect, url_for, flash, render_template, jsonify, Response, session, send_file
import sqlite3
import os
from io import StringIO
#import bcrypt
import io
from functools import wraps
from datetime import datetime
import xlsxwriter
from io import BytesIO
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
import secrets
import hashlib
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from dotenv  import load_dotenv
import inspect
from flask import Flask, current_app
from urllib.parse import urlparse, urljoin
import win32print
from collections import Counter
from config import DevConfig, ProdConfig, TestConfig
from authz import require_role



TERMS=["Term 1","Term 2","Term 3"]
ALLOWED_EXTS={"csv","xlsx","xls"}
ALLOWED_EXTS_HP = ALLOWED_EXTS
SCHOOL_FEE_TYPES = ("fees","schoolfees","school_fees")
HOLIDAY_NAME = "Holiday Package"
OTHER_NAME = "Other Assessments"
CORE_CODES = ["ENG","MATH","SCI","SST"]


DB_PATH = "school.db"

# 1) Load_dotenv before anything else
load_dotenv()


 

def get_db_connection():
    """
    Single source of truth for DB connections.
    - Uses SQLITE_PATH from app config (loaded from .env via config.py)
    - Ensures parent folder exists
    - Enables foreign key constraints
    - Sets row_factory to sqlite3.Row
    """
    db_path = current_app.config.get("SQLITE_PATH", os.path.join(os.getcwd(), "school.db"))
    p = Path(db_path).expanduser()
    p.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(str(p), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    # enforce FKs across the app (same as hp_get_db used to do)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn



        

def configure_logging(app):
    if getattr(app, "_logging_configured", False):
        return
    app.logger.setLevel(app.config.get("LOG_LEVEL", "INFO"))
    console = logging.StreamHandler()
    console.setLevel(app.config.get("LOG_LEVEL", "INFO"))
    console.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)s in %(module)s: %(message)s"))
    app.logger.addHandler(console)

    log_dir = Path("logs"); log_dir.mkdir(exist_ok=True)
    file_handler = RotatingFileHandler(log_dir / "app.log", maxBytes=2_000_000, backupCount=5)
    file_handler.setLevel(app.config.get("LOG_LEVEL", "INFO"))
    file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s %(funcName)s: %(message)s"))
    app.logger.addHandler(file_handler)

    app._logging_configured = True


def bootstrap():
    """
    Initialize database schema and seed defaults.
    Tries to call your project's ensure_* and seed_* helpers if present.
    Safe to run multiple times (idempotent).
    """
    # Helper: call a function if it exists; prefer (conn) signature, fallback to no-arg
    def _call_if_exists(name, conn=None):
        fn = globals().get(name)
        if not callable(fn):
            return False, f"{name} (missing)"
        try:
            # Prefer calling with a connection if the function accepts it
            sig = inspect.signature(fn)
            if len(sig.parameters) >= 1 and conn is not None:
                fn(conn) # try fn(conn)
            else:
                fn() # try fn()
            return True, f"{name} (ok)"
        except TypeError:
            # Signature didn’t match; try alternate call
            try:
                fn() if conn is not None else fn()
                return True, f"{name} (ok: alt)"
            except Exception as e:
                return False, f"{name} (error: {e})"
        except Exception as e:
            return False, f"{name} (error: {e})"

    # Open a connection
    try:
        conn = get_db_connection()
    except NameError as e:
        # If get_db_connection isn't in this file, import it from your db helper module
        raise RuntimeError("get_db_connection() not found. Import it into this module before calling bootstrap().") from e

    ran = []
    try:
        # --- Core schema creators (call only if they exist in your project) ---
        for fname in [
            # academics
            "ensure_students_table",
            "ensure_subject_papers_schema",
            "ensure_teacher_subjects_schema",
            "ensure_streams_schema",
            "ensure_report_comments_schema",
            "ensure_expense_categories_schema",
            "ensure_grading_scale_schema",
            "ensure_subjects_table",
            "ensure_term_dates_schema",
            "ensure_results_table", # if you use 'results'
            "ensure_record_score_table", # if you use 'record_score'
            "ensure_midterms_table",
            "ensure_reports_table",
            "ensure_classes_schema",
            "ensure_archived_students_table",
            "ensure_promotions_log_schema",
            "ensure_promotion_lock",
            "ensure_transport_schema",
            "ensure_transport_as_reg",
            "ensure_holiday_package_schema",
            "ensure_promotions_log_schema",
            # finance
            "ensure_fees_table",
            "ensure_expenses_schema",
            "ensure_payroll_schema",
            "ensure_class_fees_schema",
            "ensure_join_columns",
            "ensure_bursaries_schema",
            "ensure_requirements_schema",
           # "ensure_requirements_has_year",
           # "ensure_fees_has_comment",
            #"ensure_fees_has_receipt_no",
            "ensure_assets_schema",
            "ensure_other_income_schema",
            # user/auth
            "ensure_users_table",
            "ensure_teachers_table",
            "ensure_employees_table",
            # admin/meta
            "ensure_academic_years_schema",
            "ensure_audit_trail_schema",
            # our auto-comments
            "ensure_comment_rules_schema",
            #-seeds if any
            
            
        ]:
            ok, msg = _call_if_exists(fname, conn)
            ran.append(msg)

        # --- Create comment_rules if ensure_comment_rules_schema is missing ---
        cur = conn.cursor()
        try:
            cur.execute("""
              CREATE TABLE IF NOT EXISTS comment_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role TEXT CHECK(role IN ('teacher','headteacher')) NOT NULL,
                scope TEXT CHECK(scope IN ('subject','overall')) NOT NULL,
                match_type TEXT CHECK(match_type IN ('grade','division','range')) NOT NULL,
                grade TEXT,
                division INTEGER,
                lower_limit REAL,
                upper_limit REAL,
                class_name TEXT,
                level TEXT,
                term TEXT,
                template_text TEXT NOT NULL,
                priority INTEGER DEFAULT 100,
                active INTEGER DEFAULT 1
              )
            """)
            cur.execute("""
              CREATE INDEX IF NOT EXISTS ix_comment_rules_lookup
              ON comment_rules(role, scope, active, match_type)
            """)
            conn.commit()
            ran.append("ensure_comment_rules_schema (built-in fallback ok)")
        except Exception as e:
            ran.append(f"ensure_comment_rules_schema (fallback error: {e})")

        # --- Seeders (run if present) ---
        for fname in [
            "seed_default_admin",
            "seed_default_classes",
            "seed_expense_categories",
            "seed_grading_scale",
        ]:
            ok, msg = _call_if_exists(fname, conn)
            ran.append(msg)

        # --- Optional guards / migrations if you have them ---
        for fname in [
            "apply_schema_guards", # e.g., add missing columns if tables exist
            "run_migrations", # if you wrote a simple migrations runner
        ]:
            ok, msg = _call_if_exists(fname, conn)
            ran.append(msg)

        # Commit any pending changes from ensure_* / seed_* calls
        try:
            conn.commit()
        except Exception:
            # If any ensure_* internally commits this is still fine
            pass

    finally:
        try:
            conn.close()
        except Exception:
            pass

    # Log a compact summary
    try:
        # If app logger is available (inside app context), use it
        from flask import current_app as _app
        if _app and _app.logger:
            for line in ran:
                _app.logger.info(f"[bootstrap] {line}")
    except Exception:
        # Fallback to stdout
        for line in ran:
            print(f"[bootstrap] {line}")




# 2) Create app and load config


def create_app():
    app = Flask(__name__, template_folder="templates", static_folder="static")

    env = os.getenv("FLASK_ENV", "production").lower()
    if env == "development":
        from config import DevConfig; app.config.from_object(DevConfig)
    elif env == "testing":
        from config import TestConfig; app.config.from_object(TestConfig)
    else:
        from config import ProdConfig; app.config.from_object(ProdConfig)
        
    #Set printer name before returning
        
    app.config.setdefault("ESC_POS_PRINTER_NAME", app.config["RECEIPT_PRINTER_NAME"])

    configure_logging(app)
    
    return app
    
app = create_app()
 
    


# --- normalizers (place these once near top of file, after imports) ---






def norm_class(s: str | None) -> str | None:
    if not s: return None
    s = s.strip().upper()
    order = ["BABY","MIDDLE","TOP","P1","P2","P3","P4","P5","P6","P7"]
    # allow mixed-case input, return canonical case
    key = s.capitalize() if s in ("BABY","MIDDLE","TOP") else s
    canon = {"BABY":"Baby","MIDDLE":"Middle","TOP":"Top"}
    if s in order:
        return canon.get(s, s) if s.startswith("P") else canon.get(s, s)
    if key in ("Baby","Middle","Top"): return key
    if s.startswith("P") and s[1:].isdigit():
        n = int(s[1:])
        if 1 <= n <= 7: return f"P{n}"
    return None





def norm_section(value: str | None) -> str | None:
    if not value:
        return None
    s = value.strip().lower()
    if s in ("day", "d"):
        return "Day"
    if s in ("boarding", "board", "b"):
        return "Boarding"
    return None # unknowns rejected




def class_expected_amount(student, conn):
    """
    Return the configured fee for a student based STRICTLY on
    (class_name, section). No fallback to NULL section.
    """
    if not student:
        return 0.0
    cls = (student.get("class_name") or "").strip()
    sec = (student.get("section") or "").strip()
    if not cls or not sec:
        return 0.0

    row = conn.execute(
        """
        SELECT amount
          FROM class_fees
         WHERE class_name = ?
           AND lower(section) = lower(?)
         LIMIT 1
        """,
        (cls, sec),
    ).fetchone()

    return float(row["amount"]) if row and row["amount"] is not None else 0.0

####====Students' Expected fee displayed in students finance report======

def _table_exists(conn, name: str) -> bool:
    row = conn.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (name,)).fetchone()
    return bool(row)

def _column_exists(conn, table: str, col: str) -> bool:
    try:
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        return col in cols
    except Exception:
        return False

def _norm_section(val: str) -> str:
    s = (val or "").strip().lower()
    if s in ("day", "d"): return "Day"
    if s in ("boarding", "board", "b"): return "Boarding"
    return ""




def norm_sex(v: str | None) -> str | None:
    s = (v or "").strip().upper()
    if s in ("M","MALE"): return "M"
    if s in ("F","FEMALE"): return "F"
    return None



def norm_stream(v: str | None) -> str | None:
    s = (v or "").strip().upper()
    return s or "A"

    

       

def get_active_academic_year():
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active = 1 LIMIT 1"
    ).fetchone()
    conn.close()

    if not row:
        # Fallback if no active row exists
        from datetime import datetime
        y = int(datetime.now().strftime("%Y"))
        return {"year": y, "current_term": "Term 1", "term": "Term 1"}

    ct = row["current_term"]
    return {"year": int(row["year"]), "current_term": ct, "term": ct}
    


def upsert_admin_user():
    """Create or repair an 'admin' user with a hashed password without touching employees."""
    conn = get_db_connection()
    c = conn.cursor()

    # Ensure the columns exist (defensive)
    cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    if "password_hash" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN password_hash TEXT")
    if "status" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN status TEXT DEFAULT 'active'")
    if "role" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'admin'")

    # Upsert admin (no reference to employees)
    row = c.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
    if row:
        c.execute("""
            UPDATE users
               SET password_hash = ?, role = 'admin', status = 'active'
             WHERE id = ?
        """, (generate_password_hash("admin123"), row["id"]))
    else:
        c.execute("""
            INSERT INTO users (username, password_hash, role, status)
            VALUES (?, ?, 'admin', 'active')
        """, ("admin", generate_password_hash("admin123")))

    conn.commit()
    conn.close()





def hp_resolve_student_id(conn, student_number):
    row = conn.execute("SELECT id FROM students WHERE student_number = ?", (student_number,)).fetchone()
    return row["id"] if row else None


# --- Role helpers ---
def _norm_role(val):
    """Return a canonical, lowercase role string."""
    return (str(val or "").strip().lower())

def require_login(f):
    @wraps(f)
    def _inner(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return _inner

def require_role(*roles):
    # normalize decorator inputs once
    wanted = tuple(_norm_role(r) for r in roles if r)

    def wrapper(f):
        @wraps(f)
        def inner(*args, **kwargs):
            if "user_id" not in session or "role" not in session:
                flash("Please login.", "warning")
                return redirect(url_for("login"))

            srole = _norm_role(session.get("role"))
            if wanted and srole not in wanted:
                flash("Access denied.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return inner
    return wrapper

# allowed roles
ALLOWED_ROLES = (
    "admin","bursar","teacher","headteacher",
    "director","clerk","dos","deputyheadteacher"
)

# --- normalize role every request (belt & braces) ---
@app.before_request
def _coerce_session_role():
    if "role" in session:
        session["role"] = _norm_role(session["role"])


#  FIX EXISTING FEES==============================================================

# utils/fees_fix.py
import sqlite3
from contextlib import closing
from typing import Optional, Callable

# ---- helpers ---------------------------------------------------------------

TERM_ORDER = {"Term 1": 1, "Term 2": 2, "Term 3": 3,
              "1": 1, "2": 2, "3": 3} # tolerate numeric terms in DB

def _term_rank(term: Optional[str]) -> int:
    if not term:
        return 0
    t = str(term).strip()
    return TERM_ORDER.get(t, TERM_ORDER.get(t.title(), 0))

def _fetchone(cur, sql, params=()):
    row = cur.execute(sql, params).fetchone()
    return dict(row) if row else None

def _fetchval(cur, sql, params=(), default=None):
    row = _fetchone(cur, sql, params)
    if not row:
        return default
    # return first column’s value
    return next(iter(row.values()), default)
 
# ---- main fix --------------------------------------------------------------

def fix_existing_fees(
    db_path: str = "school.db",
    get_conn: Optional[Callable[[], sqlite3.Connection]] = None,
) -> int:
    """
    Recomputes expected_amount, bursary_amount, carried_forward for all
    fees rows where payment_type='school_fees'.

    You can pass either:
      - db_path (for stand-alone use), OR
      - get_conn() -> sqlite3.Connection (for Flask integration).
    Returns the number of rows updated.
    """
    # open connection
    if get_conn is not None:
        conn = get_conn()
        close_after = False
    else:
        conn = sqlite3.connect(db_path)
        close_after = True

    conn.row_factory = sqlite3.Row
    updated = 0

    with closing(conn):
        c = conn.cursor()

        # Iterate target fee rows
        fees = c.execute("""
            SELECT * FROM fees WHERE COALESCE(payment_type,'school_fees) IN ('school_fees','fees')
        """).fetchall()

        for fee in fees:
            sid = fee["student_id"]
            term = fee["term"]
            year = int(fee["year"])
            amount_paid = float(fee["amount_paid"] or 0.0)

            # student properties we rely on
            stu = _fetchone(c, """
                SELECT class_name,
                       COALESCE(level, '') AS level,
                       COALESCE(section, '') AS section
                FROM students
                WHERE id=?
            """, (sid,))
            if not stu:
                continue

            class_name = (stu["class_name"] or "").strip()
            level = (stu["level"] or "").strip() or None
            section = (stu["section"] or "").strip() # Expect 'Day' / 'Boarding'

            # Expected fee from class_fees (enforce section!)
            expected = _fetchval(c, """
                SELECT amount
                FROM class_fees
                WHERE class_name=? AND section=?
                      AND ( (? IS NULL AND level IS NULL) OR level=? )
                LIMIT 1
            """, (class_name, section, level, level), default=0.0) or 0.0
            expected = float(expected)

            # Bursary for this term/year
            bursary_amount = _fetchval(c, """
                SELECT COALESCE(SUM(amount),0)
                FROM bursaries
                WHERE student_id=? AND term=? AND year=?
            """, (sid, term, year), default=0.0) or 0.0
            bursary_amount = float(bursary_amount)

            # Carry-forward = previous term’s (expected - bursary - paid), clamped ≥ 0
            prev = _fetchone(c, """
                SELECT expected_amount, bursary_amount, amount_paid, term, year
                FROM fees
                WHERE student_id=?
                  AND payment_type IN ('school_fees','fees')
                  AND (year < ? OR (year = ? AND ? > 0 AND
                                   CASE term
                                        WHEN 'Term 1' THEN 1
                                        WHEN 'Term 2' THEN 2
                                        WHEN 'Term 3' THEN 3
                                        WHEN '1' THEN 1
                                        WHEN '2' THEN 2
                                        WHEN '3' THEN 3
                                        ELSE 0
                                   END
                                   < ?))
                ORDER BY year DESC,
                         CASE term
                              WHEN 'Term 1' THEN 1
                              WHEN 'Term 2' THEN 2
                              WHEN 'Term 3' THEN 3
                              WHEN '1' THEN 1
                              WHEN '2' THEN 2
                              WHEN '3' THEN 3
                              ELSE 0
                         END DESC
                LIMIT 1
            """, (sid, year, year, _term_rank(term), _term_rank(term)))

            carried = 0.0
            if prev:
                prev_expected = float(prev.get("expected_amount") or 0.0)
                prev_bursary = float(prev.get("bursary_amount") or 0.0)
                prev_paid = float(prev.get("amount_paid") or 0.0)
                carried = max((prev_expected - prev_bursary) - prev_paid, 0.0)

            # Update current row
            c.execute("""
                UPDATE fees
                SET expected_amount=?,
                    bursary_amount=?,
                    carried_forward=?
                WHERE id=?
            """, (expected, bursary_amount, carried, fee["id"]))
            updated += 1

        conn.commit()

    # close if we opened
    if close_after:
        try:
            conn.close()
        except Exception:
            pass

    return updated


def generate_student_number(conn, class_name=None):
    """Generate student number as STD-YYYY-### (sequential)."""
    from datetime import datetime
    year = datetime.now().year

    # Count existing students for this year
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM students WHERE substr(student_number, 5, 4)=?", (str(year),))
    count = c.fetchone()[0] or 0

    seq = count + 1
    return f"STD-{year}-{seq:03d}"
 



def generate_fees_code(conn) -> str:
    while True:
        cand = "FC-" + secrets.token_hex(4).upper()
        if not conn.execute("SELECT 1 FROM students WHERE fees_code = ? LIMIT 1", (cand,)).fetchone():
            return cand








def log_action(user_id, action):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("INSERT INTO audit_trail (user_id, action) VALUES (?, ?)", (user_id, action))
    conn.commit()
    conn.close()


#---USB PRINTERS WITHOUT DEVICE ID---



# --- USERS schema + default admin -----------------------------------------


def ensure_students_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            Middle_name TEXT,
            last_name TEXT NOT NULL,
            sex TEXT,
            year_completed INTEGER,
            class_name TEXT CHECK(class_name IN ('Baby','Middle','Top','P1','P2','P3','P4','P5','P6','P7')) NOT NULL DEFAULT 'Baby',
            stream TEXT,
            section TEXT CHECK(section IN ('Day','Boarding')) DEFAULT 'Day',
            combination TEXT,
            fees_amount REAL,
            student_number TEXT UNIQUE,
            academic_year_id INTEGER,
            year_of_joining TEXT,
            term_joined TEXT,
            date_joined TEXT,
            cumulative_average TEXT,
            cumulative_grade TEXT,
            cumulative_comment TEXT,
            residence TEXT,
            house TEXT,
            parent_name TEXT,
            parent2_name TEXT,
            parent_contact TEXT,
            parent2_contact TEXT,
            fees_code TEXT,
            parent_email TEXT,
            archived INTEGER DEFAULT 0,
            current_class TEXT,
            status TEXT CHECK(status IN ('active','dropped','left','completed')) NOT NULL DEFAULT 'active'
        )
    """)
    # Useful lookups
    c.execute("CREATE INDEX IF NOT EXISTS ix_students_class ON students(class_name)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_students_sn ON students(student_number)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_students_order ON students(last_name, first_name, middle_name)")
    conn.commit()

def ensure_subjects_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            code TEXT
        )
    """)
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_subjects_name ON subjects(name)")
    conn.commit()

# put this ONCE (near your other ensure_* fns) and ABOVE marks_hub()
def ensure_record_score_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS record_score (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            term TEXT NOT NULL,
            initials TEXT,
            year INTEGER NOT NULL,
            bot_mark REAL,
            midterm_mark REAL,
            eot_mark REAL,
            holiday_mark REAL, -- optional; safe to keep for Holiday Package
            other_mark REAL, -- optional; for any other assessment
            ca_mark REAL, -- optional; for CA
            average_mark REAL, -- optional original source average, we will recompute anyway
            grade TEXT,
            comment TEXT,
            processed_on DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (student_id, subject_id, term, year),
            FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE,
            FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        )
    """)
    conn.commit()
# optional back-compat alias us                          ed elsewhere
ensure_record_score = ensure_record_score_table


def upgrade_record_score(conn):
    cols = {r[1].lower() for r in conn.execute("PRAGMA table_info(record_score)")}
    for col in ("holiday_mark", "other_mark", "ca_mark"):
        if col not in cols:
            conn.execute(f"ALTER TABLE record_score ADD COLUMN {col} REAL")
    conn.commit()


def ensure_results_table(conn):
    # If you also use a 'results' table (some projects keep both)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            eot REAL,
            total REAL,
            grade TEXT,
            comment TEXT,
            initials TEXT,
            term TEXT NOT NULL,
            year INTEGER NOT NULL,
            UNIQUE (student_id, subject_id, term, year),
            FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE,
            FOREIGN KEY (subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        )
    """)
    conn.commit()

def ensure_midterms_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS midterms ( 
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            term TEXT NOT NULL,
            year INTEGER NOT NULL,
            assessment TEXT NOT NULL,
            eng REAL, mat REAL, sci REAL, sst REAL,
            agg INTEGER, total INTEGER,
            UNIQUE (student_id, term, year, assessment),
            FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
        )
    """)
    conn.commit()

def ensure_reports_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            class_name TEXT NOT NULL,
            stream TEXT,
            subject_id INTEGER NOT NULL,
            term TEXT NOT NULL,
            year INTEGER NOT NULL,

            -- what we render
            average_mark REAL, -- blended average used for grading
            grade TEXT,
            comment TEXT,

            teacher_remark TEXT,
            headteacher_remark TEXT,

            teacher_id INTEGER,
            bot_mark REAL,
            midterm_mark REAL,
            eot_mark REAL,
            teacher_initial TEXT,

            processed_on DATETIME DEFAULT CURRENT_TIMESTAMP,

            UNIQUE (student_id, subject_id, year, term),
            FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE,
            FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        )
    """)
    conn.commit()



def ensure_expenses_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS expense_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT,
            amount REAL NOT NULL,
            term TEXT,
            year INTEGER,
            date_spent DATE DEFAULT CURRENT_DATE,
            category_id INTEGER,
            recorded_by TEXT,
            type TEXT,
            FOREIGN KEY(category_id) REFERENCES expense_categories(id) ON DELETE SET NULL
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS ix_expenses_date ON expenses(date_spent)")
    conn.commit()

def ensure_payroll_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS payroll (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            teacher_id INTEGER,
            term TEXT NOT NULL,
            year INTEGER NOT NULL,
            expected_salary REAL NOT NULL,
            bonus REAL DEFAULT 0,
            allowance REAL DEFAULT 0,
            total REAL NOT NULL,
            paid_amount REAL DEFAULT 0,
            status TEXT CHECK(status IN('fully_paid','partially_paid','not_paid')) NOT NULL DEFAULT 'not_paid',
            date_paid TEXT DEFAULT (DATE('now')),
            FOREIGN KEY(employee_id) REFERENCES employees(id),
            FOREIGN KEY(teacher_id) REFERENCES teachers(id)
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS ix_payroll_term_year ON payroll(term, year)")
    conn.commit()

    






def ensure_class_fees_schema(conn):
    c = conn.cursor()
    # table (same as you already have)
    c.execute("""
        CREATE TABLE IF NOT EXISTS class_fees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT NOT NULL,
            section TEXT NOT NULL, -- 'Day' or 'Boarding'
            level TEXT, -- optional (Nursery / Primary)
            amount REAL NOT NULL CHECK(amount >= 0)
        )
    """)
    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_class_fees_class_section_level
        ON class_fees (class_name, section, level)
    """)
    conn.commit()



def ensure_classes_schema(conn):
    c = conn.cursor()
    # Create if missing (minimal columns)
    c.execute("""
        CREATE TABLE IF NOT EXISTS classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT NOT NULL
        )
    """)
    # Add any missing columns safely
    cols = [r[1] for r in c.execute("PRAGMA table_info(classes)").fetchall()]
    if "level" not in cols:
        c.execute("ALTER TABLE classes ADD COLUMN level TEXT")
    if "stream" not in cols:
        c.execute("ALTER TABLE classes ADD COLUMN stream TEXT")
    # Unique index (stream may be NULL → coalesce to keep uniqueness sensible)
    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_classes_class_stream
        ON classes (class_name, COALESCE(stream,''))
    """)
    conn.commit()

    
# ---- seed: default classes ----
def seed_default_classes(conn):
    """Creates Baby–P7, stream A (primary) if missing."""
    ensure_classes_schema(conn)
    c = conn.cursor()
    rows = [
        ("Baby", "Nursery", "A"),
        ("Middle", "Nursery", "A"),
        ("Top", "Nursery", "A"),
        ("P1", "Primary", "A"),
        ("P2", "Primary", "A"),
        ("P3", "Primary", "A"),
        ("P4", "Primary", "A"),
        ("P5", "Primary", "A"),
        ("P6", "Primary", "A"),
        ("P7", "Primary", "A"),
    ]
    for class_name, level, stream in rows:
        c.execute("""
          INSERT OR IGNORE INTO classes (class_name, level, stream)
          VALUES (?, ?, ?)
        """, (class_name, level, stream))
    conn.commit()






    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_requirements_class_term_name
        ON requirements(class_name, COALESCE(term,''), name)
    """)
    # optional column on fees table to record requirement_name is already present in ensure_fees_table
    conn.commit()

def ensure_assets_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            model TEXT,
            value REAL DEFAULT 0.0,
            year_purchased TEXT,
            condition TEXT,
            qty INTEGER,
            archived_reason TEXT
        )
    """)
    conn.commit()



def ensure_teachers_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            Middle_name TEXT,
            last_name TEXT NOT NULL,
            gender TEXT,
            contact TEXT,
            email TEXT,
            residence TEXT,
            department TEXT,
            designation TEXT,
            hire_date TEXT,
            status TEXT CHECK(status IN ('active','archived')) DEFAULT 'active',
            base_salary REAL DEFAULT 0.0,
            allowance REAL DEFAULT 0.0,
            bonus REAL DEFAULT 0.0,
            pay_cycle TEXT DEFAULT 'monthly',
            bank_name TEXT,
            bank_account TEXT,
            tin TEXT,
            notes TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER UNIQUE,
            initials TEXT,
            subjects TEXT,
            class_name TEXT,
            can_reset_password INTEGER DEFAULT 0,
            status TEXT DEFAULT 'active',
            FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS teacher_subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            class_name TEXT NOT NULL,
            FOREIGN KEY(teacher_id) REFERENCES teachers(id) ON DELETE CASCADE,
            FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
        )
    """)
    conn.commit()

def ensure_employees_table(conn):
    # already created in ensure_teachers_table; keep this a no-op but present for bootstrap
    ensure_teachers_table(conn)
    
    

def add_created_at_if_missing():
    conn = get_db_connection()
    try:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
        if "created_at" not in cols:
            conn.execute("ALTER TABLE users ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            conn.commit()
            current_app.logger.info("[migrate] users.created_at added")
        else:
            current_app.logger.info("[migrate] users.created_at already present")
    finally:
        conn.close()


def ensure_academic_years_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS academic_years (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER UNIQUE NOT NULL,
            current_term TEXT DEFAULT 'Term 1',
            is_active INTEGER DEFAULT 0
        )
    """)
    # Also ensure term_dates exists
    c.execute("""
        CREATE TABLE IF NOT EXISTS term_dates (
            year INTEGER NOT NULL,
            term TEXT NOT NULL,
            next_term TEXT,
            next_term_date TEXT,
            PRIMARY KEY (year, term)
        )
    """)
    # Ensure at least current year exists as active
    from datetime import datetime
    this_year = int(datetime.now().strftime("%Y"))
    c.execute("""
        INSERT OR IGNORE INTO academic_years (year, current_term, is_active)
        VALUES (?, 'Term 1', 1)
    """, (this_year,))
    c.execute("UPDATE academic_years SET is_active = CASE WHEN year=? THEN 1 ELSE 0 END", (this_year,))
    conn.commit()




def ensure_audit_trail_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS audit_trail (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            role TEXT, outcome TEXT, severity TEXT,
            route TEXT, method TEXT, ip_address TEXT,
            target_table TEXT, target_id INTEGER,
            details_json TEXT, http_status INTEGER
        )
    """)
    conn.commit()


def _add_column_if_missing(conn, table, ddl):
    # ddl example: "created_at DATETIME DEFAULT CURRENT_TIMESTAMP"
    col = ddl.split()[0]
    if not _has_column(conn, table, col):
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")

def apply_schema_guards(app=None):
    conn = get_db_connection()
    try:
        # users: make sure columns we use exist
        _add_column_if_missing(conn, "users", "status TEXT DEFAULT 'active'")
        _add_column_if_missing(conn, "users", "role TEXT DEFAULT 'teacher'")
        _add_column_if_missing(conn, "users", "created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
        conn.commit()
    finally:
        conn.close()





def ensure_requirements_table_alias(conn):
    # kept for compatibility if some places call a different name
    ensure_requirements_schema(conn)

def ensure_expense_categories_table_alias(conn):
    ensure_expenses_schema(conn)

def ensure_other_income_schema():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS other_income (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT,
        amount REAL NOT NULL,
        term TEXT, -- 'Term 1' | 'Term 2' | 'Term 3'
        year INTEGER, -- e.g. 2025
        description TEXT,
        recorded_by TEXT, -- username or staff name
        date_received DATETIME DEFAULT CURRENT_TIMESTAMP
      )
    """)
    # helpful index for reports by term/year
    c.execute("""
      CREATE INDEX IF NOT EXISTS ix_other_income_term_year
      ON other_income(term, year)
    """)
    conn.commit()
    conn.close()


def ensure_assets_table_alias(conn):
    ensure_assets_schema(conn)
    

def ensure_fees_table(conn):
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS fees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        term TEXT,
        year INTEGER,
        amount_paid REAL DEFAULT 0,
        requirement_name TEXT, -- optional backref for requirements
        req_term TEXT, -- legacy-safe
        payment_item TEXT, -- optional notes
        bursary_amount REAL DEFAULT 0,
        carried_forward REAL DEFAULT 0,
        expected_amount REAL DEFAULT 0,
        date_paid TEXT,
        comment TEXT,
        receipt_no TEXT,
        processed_on TEXT,
        method TEXT DEFAULT 'N/A',
        payment_type TEXT DEFAULT 'school_fees', -- 'school_fees' | 'requirements' | ...
        recorded_by TEXT,
        FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
    )
    """)
    # Helpful indexes for speed
    c.execute("CREATE INDEX IF NOT EXISTS ix_fees_student ON fees(student_id)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_fees_term_year ON fees(term, year)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_fees_payment_type ON fees(payment_type)")
    conn.commit()

def ensure_fees_has_comment():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("PRAGMA table_info(fees)")
    cols = [row[1] for row in c.fetchall()]
    if "comment" not in cols:
        c.execute("ALTER TABLE fees ADD COLUMN comment TEXT")
        conn.commit()
    conn.close()




def seed_expense_categories(conn):
    conn.execute("""
      CREATE TABLE IF NOT EXISTS expense_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE
      )
    """)
    cats = [
        ('Salaries',), ('Stationery',), ('Utilities',), ('Transport',),
        ('Maintenance',), ('Service Providers',), ('Uniforms',),
        ('Examinations',), ('Meals',), ('Office supplies',), ('Medical',),
        ('Bonus',), ('Allowance',), ('Miscellaneous',)
    ]
    for (name,) in cats:
        try:
            conn.execute("INSERT OR IGNORE INTO expense_categories(name) VALUES (?)", (name,))
        except Exception:
            pass
    conn.commit()


def seed_grading_scale(conn):
    conn.execute("""
      CREATE TABLE IF NOT EXISTS grading_scale (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        grade TEXT NOT NULL,
        lower_limit INTEGER NOT NULL,
        upper_limit INTEGER NOT NULL,
        comment TEXT
      )
    """)
    # Only seed if empty
    row = conn.execute("SELECT COUNT(*) AS c FROM grading_scale").fetchone()
    if row and (row["c"] or 0) > 0:
        return

    # Example PLE-like bands (adjust to your scheme)
    bands = [
        ('D1', 90, 100, 'Excellent'),
        ('D2', 80, 89, 'Very good'),
        ('C3', 75, 79, 'Good'),
        ('C4', 70, 74, 'Good'),
        ('C5', 65, 69, 'Fair'),
        ('C6', 60, 64, 'Fair'),
        ('P7', 50, 59, 'Pass'),
        ('P8', 40, 49, 'Basic'),
        ('F9', 0, 39, 'Fail'),
    ]
    conn.executemany(
        "INSERT INTO grading_scale (grade, lower_limit, upper_limit, comment) VALUES (?, ?, ?, ?)",
        bands
    )
    conn.commit()




def run_migrations(conn):
    # place versioned migrations here if needed
    pass


def ensure_class_comments_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS class_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT,
            term TEXT,
            year INTEGER,
            comment TEXT
        )
    """)
    conn.commit()


def ensure_archived_students_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS archived_students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER,
            student_number TEXT,
            full_name TEXT,
            class_name TEXT, -- class at time of archiving (P7 for leavers)
            year_completed INTEGER NOT NULL, -- academic year of leaving/completion
            completed_stage TEXT, -- 'P7 Leaver' or 'Manual Archive' or 'Other'
            outstanding_balance REAL DEFAULT 0,
            archived_on DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE SET NULL
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS ix_archived_students_yr ON archived_students(year_completed)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_archived_students_sn ON archived_students(student_number)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_archived_students_name ON archived_students(full_name)")
    conn.commit()


def ensure_promotions_log_schema(conn=None):
    close_after = False
    if conn is None:
        conn = get_db_connection(); close_after = True
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS promotions_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        from_class TEXT NOT NULL,
        to_class TEXT NOT NULL,
        actor TEXT,
        batch_id TEXT,
        reversed INTEGER DEFAULT 0, -- 0=normal, 1=reversal (demotion)
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
      )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS ix_promolog_student ON promotions_log(student_id, created_at DESC)")
    conn.commit()
    if close_after: conn.close()


def ensure_promotion_lock(conn=None):
    """Locks the once-per-year promotion."""
    close_after = False
    if conn is None:
        conn = get_db_connection(); close_after = True
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS promotion_lock (
        year INTEGER PRIMARY KEY, -- academic year
        executed_by TEXT,
        executed_at DATETIME DEFAULT CURRENT_TIMESTAMP
      )
    """)
    conn.commit()
    if close_after: conn.close()


def promotion_already_done(year: int) -> bool:
    conn = get_db_connection(); c = conn.cursor()
    r = c.execute("SELECT 1 FROM promotion_lock WHERE year=?", (year,)).fetchone()
    conn.close()
    return bool(r)


def mark_promotion_done(year: int, actor: str):
    conn = get_db_connection(); c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO promotion_lock(year, executed_by) VALUES(?,?)", (year, actor))
    conn.commit(); conn.close()



ORDER = ["Baby","Middle","Top","P1","P2","P3","P4","P5","P6","P7"]

def next_class_name(current: str) -> str | None:
    cur = (current or "").strip()
    if cur not in ORDER: return None
    idx = ORDER.index(cur)
    return ORDER[idx+1] if idx+1 < len(ORDER) else None

def prev_class_name(current: str) -> str | None:
    cur = (current or "").strip()
    if cur not in ORDER: return None
    idx = ORDER.index(cur)
    return ORDER[idx-1] if idx-1 >= 0 else None

def write_audit(conn, *, user_id=None, role=None, action="", outcome="success",
                severity="info", route=None, method=None, ip=None,
                target_table=None, target_id=None, details=None, http_status=None):
    """Low-level insert. Never raises back to caller."""
    try:
        # Make sure table exists; safe no-op if it already does
        try:
            ensure_audit_trail_schema(conn)
        except Exception:
            pass

        dj = None
        if details is not None:
            try:
                dj = json.dumps(details, ensure_ascii=False, default=str)
            except Exception:
                dj = str(details)

        conn.execute("""
            INSERT INTO audit_trail
              (user_id, role, action, outcome, severity,
               route, method, ip_address, target_table, target_id,
               details_json, http_status, timestamp)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            user_id, role, action, outcome, severity,
            route, method, ip, target_table, target_id,
            dj, http_status,
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        ))
        conn.commit()
    except Exception:
        # swallow any audit errors silently
        pass

def audit_from_request(conn, *, action, outcome="success", severity="info",
                       target_table=None, target_id=None, details=None, http_status=None):
    """Convenience wrapper that pulls request/session info."""
    try:
        uid = session.get("user_id")
        role = session.get("role")
        route = request.path
        method = request.method
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        write_audit(conn,
            user_id=uid, role=role, action=action, outcome=outcome, severity=severity,
            route=route, method=method, ip=ip,
            target_table=target_table, target_id=target_id,
            details=details, http_status=http_status
        )
    except Exception:
        pass


def compute_outstanding_balance(student_id: int) -> float:
    # If you already have compute_student_financials(...), call it.
    # Otherwise, quick placeholder that sums fees - payments. Replace as needed.
    try:
        fin = compute_student_financials(student_id, None, None, None)
        return float(fin.get("overall_balance", 0)) if isinstance(fin, dict) else float(fin.overall_balance or 0)
    except Exception:
        return 0.0




# safe helper (put once in your codebase)

def _archive_student(conn, student_id: int, new_status: str = "completed",
                     stage: str = "Manual Archive", year_completed: int | None = None) -> int:
    """
    Soft-archive one student (students.archived=1 + allowed status) and
    ensure a row exists in archived_students for the same student/year.
    """
    conn.row_factory = sqlite3.Row
    ensure_archived_students_table(conn) # your function from earlier

    from datetime import datetime
    ay = (get_active_academic_year() or {})
    try:
        yc = int(year_completed or ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        yc = datetime.now().year

    s = conn.execute("""
        SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name,
               last_name, class_name
        FROM students WHERE id=? LIMIT 1
    """, (student_id,)).fetchone()
    if not s:
        return 0

    # 1) flip archived & use an allowed status (matches your CHECK constraint)
    cur = conn.execute("UPDATE students SET archived=1, status=? WHERE id=?",
                       (new_status, student_id))
    changed = cur.rowcount or 0

    # 2) optional best-effort outstanding (safe if compute_student_financials exists)
    outstanding = 0.0
    try:
        term = (ay.get("current_term") or ay.get("term") or "Term 1")
        fin = compute_student_financials(student_id, s["class_name"], term, yc)
        outstanding = float((fin.get("overall_balance") if isinstance(fin, dict)
                             else getattr(fin, "overall_balance", 0)) or 0)
    except Exception:
        pass

    full_name = f"{s['first_name']} {s['Middle_name']} {s['last_name']}".strip()

    # 3) write to archive table (skip duplicate for same student/year)
    conn.execute("""
        INSERT INTO archived_students (student_id, student_number, full_name, class_name,
                                       year_completed, completed_stage, outstanding_balance)
        SELECT ?,?,?,?,?,?,?
        WHERE NOT EXISTS (
          SELECT 1 FROM archived_students WHERE student_id=? AND year_completed=?
        )
    """, (s["id"], s["student_number"], full_name, s["class_name"],
          yc, stage, outstanding, s["id"], yc))

    conn.commit()
    return changed


def _unarchive_student(conn, student_id: int, *, remove_archive_rows: bool = True) -> int:
    cur = conn.execute("UPDATE students SET archived=0, status='active' WHERE id=?", (student_id,))
    changed = cur.rowcount or 0
    if remove_archive_rows: # ➤ remove snapshot(s) so it disappears from Archive Hub
        conn.execute("DELETE FROM archived_students WHERE student_id=?", (student_id,))
    conn.commit()
    return changed



def ensure_subject_papers_schema(conn):
    c = conn.cursor()
    # Base table
    c.execute("""
      CREATE TABLE IF NOT EXISTS subject_papers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        subject_id INTEGER NOT NULL,
        paper_name TEXT NOT NULL,
        paper_initial TEXT,
        FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
      )
    """)
    # Helpful indexes
    c.execute("CREATE INDEX IF NOT EXISTS ix_subject_papers_subject ON subject_papers(subject_id)")
    c.execute("""
      CREATE UNIQUE INDEX IF NOT EXISTS uq_subject_papers_unique
      ON subject_papers(subject_id, paper_name)
    """)
    conn.commit()


def ensure_teacher_subjects_schema(conn):
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS teacher_subjects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id INTEGER NOT NULL,
        subject_id INTEGER NOT NULL,
        class_name TEXT NOT NULL,
        FOREIGN KEY(teacher_id) REFERENCES teachers(id) ON DELETE CASCADE,
        FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
      )
    """)
    # Prevent duplicates per (teacher, subject, class)
    c.execute("""
      CREATE UNIQUE INDEX IF NOT EXISTS uq_teacher_subjects
      ON teacher_subjects(teacher_id, subject_id, class_name)
    """)
    # Fast lookups
    c.execute("CREATE INDEX IF NOT EXISTS ix_teacher_subjects_teacher ON teacher_subjects(teacher_id)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_teacher_subjects_subject ON teacher_subjects(subject_id)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_teacher_subjects_class ON teacher_subjects(class_name)")
    conn.commit()


def ensure_streams_schema(conn):
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS streams (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL
      )
    """)
    conn.commit()


def ensure_term_dates_schema(conn=None):
    """
    Create term_dates table if missing.
    Works with or without a supplied connection (backward compatible).
    """
    must_close = False
    if conn is None:
        from datetime import datetime # not required, but keeps style consistent
        conn = get_db_connection()
        must_close = True

    conn.execute("""
        CREATE TABLE IF NOT EXISTS term_dates (
            year INTEGER NOT NULL,
            term TEXT NOT NULL,
            next_term TEXT,
            next_term_date TEXT, -- store as 'YYYY-MM-DD' or any display string
            PRIMARY KEY (year, term)
        )
    """)
    conn.commit()
    if must_close:
        conn.close()


    





def ensure_report_comments_schema(conn):
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS report_comments (
        student_id INTEGER NOT NULL,
        term TEXT NOT NULL,
        year INTEGER NOT NULL,
        teacher_comment TEXT,
        head_comment TEXT,
        PRIMARY KEY (student_id, term, year),
        FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE
      )
    """)
    # Fast lookups by student/period
    c.execute("CREATE INDEX IF NOT EXISTS ix_report_comments_student ON report_comments(student_id)")
    c.execute("CREATE INDEX IF NOT EXISTS ix_report_comments_term_year ON report_comments(term, year)")
    conn.commit()


def ensure_expense_categories_schema(conn):
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS expense_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE
      )
    """)
    conn.commit()


def ensure_grading_scale_schema(conn):
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS grading_scale (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        grade TEXT NOT NULL,
        lower_limit INTEGER NOT NULL,
        upper_limit INTEGER NOT NULL,
        comment TEXT
      )
    """)
    # Basic sanity index
    c.execute("CREATE INDEX IF NOT EXISTS ix_grading_bounds ON grading_scale(lower_limit, upper_limit)")
    conn.commit()


def ensure_users_table(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT CHECK(role IN (
                'admin','bursar','teacher','headteacher','director','clerk','deputyheadteacher','dos'
            )) NOT NULL,
            status TEXT CHECK(status IN ('active','archived')) NOT NULL DEFAULT 'active',
            employee_id INTEGER,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE SET NULL
        )
    """)
    # Optional: bring old schemas up-to-date (idempotent)
    cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    if "employee_id" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN employee_id INTEGER")
    if "status" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN status TEXT DEFAULT 'active'")
    if "created_at" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
    conn.commit()


@app.before_request
def _sync_active_term_year():
    if "user_id" in session:
        try:
            conn = get_db_connection()
            t, y = get_active_academic_year().get("current_term"), get_active_academic_year().get("year")
            conn.close()
            if t and y:
                if session.get("current_term") != t or session.get("current_year") != y:
                    session["current_term"] = t
                    session["current_year"] = y
        except Exception:
            pass


def seed_default_admin(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT CHECK(role IN (
                'admin','bursar','teacher','headteacher','director','clerk','deputyheadteacher','dos'
            )) NOT NULL,
            status TEXT CHECK(status IN ('active','archived')) NOT NULL DEFAULT 'active',
            employee_id INTEGER,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    row = c.execute("SELECT 1 FROM users WHERE username='admin'").fetchone()
    if not row:
        c.execute("""
            INSERT INTO users (username, password_hash, role, status)
            VALUES (?, ?, 'admin', 'active')
        """, ("admin", generate_password_hash("admin123")))
    conn.commit()





def next_class_name(current: str) -> str | None:
    """Lightweight promotion map; extend as needed."""
    order = ["Baby","Middle","Top","P1","P2","P3","P4","P5","P6","P7"]
    if current not in order: return None
    idx = order.index(current)
    return order[idx+1] if idx+1 < len(order) else None
    
    

def generate_term_fees(student_row, term, year, c):
    """
    Ensures a fees row exists for this student/term/year and fills expected_amount
    using class_fees(class_name + Day/Boarding [+ optional level]).
    Returns a (upserted) fees row id.
    """
    sid = student_row["id"]
    class_name = student_row["class_name"]
    level = student_row.get("level") if hasattr(student_row, "get") else student_row["level"]
    Boarding = (
        (student_row.get("section") if hasattr(student_row, "get") else student_row["section"])
        or (student_row.get("section") if hasattr(student_row, "get") else student_row.get("section", None))
        or ""
    )

    # --- your requested lookup (strict by Day/Boarding, tolerant level) ---
    class_fee = c.execute("""
        SELECT amount
          FROM class_fees
         WHERE class_name = ?
           AND lower(section) = lower(?)
           AND (level IS NULL OR level = ?)
         LIMIT 1
    """, (class_name, Boarding, level)).fetchone()

    expected = float(class_fee["amount"]) if class_fee and class_fee["amount"] is not None else 0.0

    # bursary for this term/year
    bursary_row = c.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
          FROM bursaries
         WHERE student_id = ? AND term = ? AND year = ?
    """, (sid, term, year)).fetchone()
    bursary_amount = float(bursary_row["total"] or 0)

    # upsert fees record (payment_type = school_fees)
    c.execute("""
        INSERT INTO fees (student_id, term, year, payment_type,
                          expected_amount, bursary_amount, carried_forward, amount_paid)
        VALUES (?, ?, ?, 'school_fees', ?, ?, 0, 0)
        ON CONFLICT(student_id, term, year, payment_type) DO UPDATE SET
            expected_amount = excluded.expected_amount,
            bursary_amount = excluded.bursary_amount
    """, (sid, term, year, expected, bursary_amount))

    # return the id (optional)
    rid = c.execute("""
        SELECT id FROM fees
         WHERE student_id=? AND term=? AND year=? AND payment_type='school_fees'
    """, (sid, term, year)).fetchone()
    return rid["id"] if rid else None


# ---------------- FIX FEES (RECALC EXPECTED / BURSARY / CARRY-FORWARD) ----------------




def _recalc_all_fees(conn) -> int:
    """
    Recompute expected_amount, bursary_amount, carried_forward
    for fee rows where payment_type IN ('school_fees','fees').
    Previous-balance lookup includes ('school_fees','fees','opening_balance')
    so any imported opening balance is honored, but we do NOT modify those rows.
    """
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # rows we DO update
    rows = cur.execute("""
        SELECT id, student_id, term, year
          FROM fees
         WHERE lower(payment_type) IN ('school_fees','fees')
    """).fetchall()

    updated = 0
    for fee in rows:
        sid = fee["student_id"]
        term = (fee["term"] or "")
        year = int(fee["year"])

        # student info
        st = cur.execute("""
            SELECT class_name,
                   COALESCE(NULLIF(TRIM(section),''), NULL) AS sec
              FROM students
             WHERE id = ?
        """, (sid,)).fetchone()
        if not st:
            continue

        class_name = (st["class_name"] or "").strip()
        section = norm_section(st["sec"]) # "Day"/"Boarding"/None

        # expected from class+section
        expected = 0.0
        if class_name and section:
            class_row = cur.execute("""
                SELECT amount
                  FROM class_fees
                 WHERE class_name = ?
                   AND lower(section) = lower(?)
                 LIMIT 1
            """, (class_name, section)).fetchone()
            if class_row and class_row["amount"] is not None:
                expected = float(class_row["amount"])

        # bursary for this term/year
        bursary = cur.execute("""
            SELECT COALESCE(SUM(amount),0) AS total
              FROM bursaries
             WHERE student_id=? AND term=? AND year=?
        """, (sid, term, year)).fetchone()["total"] or 0.0

        # previous row (includes opening_balance if you have it)
        prev = cur.execute("""
            SELECT expected_amount AS exp,
                   bursary_amount AS bur,
                   amount_paid AS paid,
                   term, year
              FROM fees
             WHERE student_id = ?
               AND lower(payment_type) IN ('school_fees','fees','opening_balance')
               AND (year < ?
                    OR (year = ? AND
                        CASE lower(term)
                          WHEN 'term 1' THEN 1
                          WHEN 'term 2' THEN 2
                          WHEN 'term 3' THEN 3
                          ELSE 0
                        END
                        < ?
                   )
                 )
          ORDER BY year DESC,
                   CASE lower(term)
                     WHEN 'term 3' THEN 3
                     WHEN 'term 2' THEN 2
                     WHEN 'term 1' THEN 1
                     ELSE 0
                   END DESC
             LIMIT 1
        """, (sid, year, year, _term_rank(term))).fetchone()

        carried = 0.0
        if prev:
            carried = max(
                (float(prev["exp"] or 0) - float(prev["bur"] or 0) - float(prev["paid"] or 0)),
                0.0
            )

        cur.execute("""
            UPDATE fees
               SET expected_amount = ?,
                   bursary_amount = ?,
                   carried_forward = ?
             WHERE id = ?
        """, (expected, bursary, carried, fee["id"]))
        updated += 1

    conn.commit()
    return updated




def ensure_fee_rows_for_all(conn, term: str, year: int) -> int:
    """
    Insert a 'school_fees' row for every active student for (term, year)
    if it doesn't already exist. Returns number of rows inserted.
    """
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO fees (
            student_id, payment_type, term, year,
            expected_amount, bursary_amount, amount_paid, carried_forward, date_paid
        )
        SELECT s.id, 'school_fees', ?, ?, 0, 0, 0, 0, date('now')
        FROM students s
        WHERE COALESCE(lower(s.status),'active') = 'active'
          AND NOT EXISTS (
            SELECT 1 FROM fees f
            WHERE f.student_id = s.id
              AND f.year = ?
              AND lower(f.term) = lower(?)
              AND lower(f.payment_type) IN ('school_fees','fees')
          )
    """, (term, year, year, term))
    inserted = cur.rowcount if cur.rowcount is not None else 0
    conn.commit()
    return inserted


    
def fix_existing_fees(db_path: str = "school.db", get_conn=None) -> int:
    """
    Recomputes expected_amount, bursary_amount, carried_forward for all fees.
    Can use either db_path (standalone) or get_conn (Flask).
    """
    import sqlite3

    if get_conn is not None:
        conn = get_conn()
        close_after = True
    else:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        close_after = False

    c = conn.cursor()
    fees = c.execute("SELECT * FROM fees WHERE payment_type = 'school_fees'").fetchall()
    updated = 0

    for fee in fees:
        sid, term, year = fee['student_id'], fee['term'], fee['year']

        # student section (day/boarding enforced)
        student = c.execute(
            "SELECT class_name, level, section FROM students WHERE id = ?",
            (sid,)
        ).fetchone()
        if not student:
            continue

        class_fee = c.execute("""
            SELECT amount FROM class_fees
            WHERE class_name = ? AND (section = ? OR section IS NULL)
        """, (student['class_name'], student['section'])).fetchone()
        expected = class_fee['amount'] if class_fee else 0

        bursary = c.execute("""
            SELECT SUM(amount) as total FROM bursaries
            WHERE student_id = ? AND term = ? AND year = ?
        """, (sid, term, year)).fetchone()
        bursary_amount = bursary['total'] if bursary and bursary['total'] else 0

        prev = c.execute("""
            SELECT expected_amount, bursary_amount, amount_paid
            FROM fees
            WHERE student_id = ? AND (year < ? OR (year = ? AND term != ?))
              AND payment_type = 'school_fees'
            ORDER BY year DESC, term DESC LIMIT 1
        """, (sid, year, year, term)).fetchone()

        carried = 0
        if prev:
            carried = (prev['expected_amount'] - prev['bursary_amount']) - prev['amount_paid']
            carried = max(carried, 0)

        c.execute("""
            UPDATE fees SET expected_amount=?, bursary_amount=?, carried_forward=?
            WHERE id=?
        """, (expected, bursary_amount, carried, fee['id']))
        updated += 1

    conn.commit()
    if close_after:
        conn.close()

    return updated


  
    
# --- Helper: find where raw scores live (record_score or results) ---
def detect_scores_table(conn):
    c = conn.cursor()
    c.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='record_score'")
    if c.fetchone():
        return "record_score"
    c.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='results'")
    if c.fetchone():
        return "results"
    return None
    





# --- Helper: process snapshot for a class/term/year (idempotent) ---




def process_reports_snapshot(conn, class_name, term, year):
    """
    For each student in class:
      - Compute mid_calc = MID if present else mean(BOT, HOLIDAY, OTHER/CA).
      - blended_avg = mean(BOT, mid_calc, EOT, HOLIDAY, OTHER/CA, CA) (ignores Nones)
      - grade/comment based on blended_avg (Average).
      - Keep EOT separate so the table shows both EOT and Average.
    """
    ensure_reports_table(conn)
    ensure_record_score_table(conn)

    c = conn.cursor()

    students = c.execute("""
        SELECT id, class_name, COALESCE(stream,'') AS stream
        FROM students
        WHERE archived=0 AND class_name=?
        ORDER BY last_name, first_name
    """, (class_name,)).fetchall()

    # wipe this class/term/year snapshot
    c.execute("DELETE FROM reports WHERE class_name=? AND term=? AND year=?", (class_name, term, year))

    for s in students:
        sid = s["id"]
        rows = c.execute("""
            SELECT subject_id,
                   MAX(bot_mark) AS bot,
                   MAX(midterm_mark) AS mid,
                   MAX(eot_mark) AS eot,
                   MAX(holiday_mark) AS holiday,
                   MAX(other_mark) AS other,
                   MAX(ca_mark) AS ca,
                   MAX(initials) AS initials,
                   MAX(comment) AS comment_raw,
                   MAX(grade) AS grade_raw
            FROM record_score
            WHERE student_id=? AND term=? AND year=?
            GROUP BY subject_id
            ORDER BY subject_id
        """, (sid, term, year)).fetchall()

        for r in rows:
            mid_calc = r["mid"] if r["mid"] is not None else _mean([r["bot"], r["holiday"], r["other"], r["ca"]])
            blended_avg = _mean([r["bot"], mid_calc, r["eot"], r["holiday"], r["other"], r["ca"]])

            # grading/comment are based on Average (blended_avg)
            grd = grade_for_score(conn, blended_avg)
            cmt = comment_for_grade(conn, grd)

            initials = (r["initials"] or "") or guess_teacher_initials(conn, s["class_name"], r["subject_id"])

            c.execute("""
                INSERT OR REPLACE INTO reports
                    (student_id, class_name, stream, subject_id, term, year,
                     average_mark, grade, comment,
                     teacher_remark, headteacher_remark,
                     teacher_id, bot_mark, midterm_mark, eot_mark, teacher_initial, processed_on)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (sid, s["class_name"], s["stream"], r["subject_id"], term, year,
                  blended_avg, grd, cmt,
                  None, # teacher_remark (per-subject auto-remark optional; keep NULL)
                  r["bot"], mid_calc, r["eot"], initials))

    # headteacher remark can be filled lazily in view/print
    conn.commit()

# ---------- Single report card ----------




def ensure_simple_users_schema(db_path="school.db"):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT CHECK(role IN ('admin','bursar','teacher','headteacher','director','clerk')) NOT NULL,
        status TEXT CHECK(status IN ('active','archived')) NOT NULL DEFAULT 'active',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """)
    # Add columns if missing
    cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    if "role" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'teacher'")
    if "status" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN status TEXT DEFAULT 'active'")
    if "created_at" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP")

    # Seed an admin if none
    row = c.execute("SELECT 1 FROM users WHERE username='admin'").fetchone()
    if not row:
        c.execute("INSERT INTO users (username,password_hash,role,status,created_at) VALUES (?,?,?,?,?)",
                  ("admin", generate_password_hash("admin123"), "admin", "active",
                   datetime.now().isoformat(" ", "seconds")))
    conn.commit()
    conn.close()
    



# ==== Payroll Hub & Actions ====


def _payroll_status(total, paid):
    """Return status string based on amounts."""
    paid = paid or 0
    total = total or 0
    if paid <= 0:
        return "not_paid"
    if paid < total:
        return "partially_paid"
    return "fully_paid"

def get_or_create_expense_category(conn, name="Salaries"):
    """Return category_id for `name`, creating it if needed."""
    c = conn.cursor()
    row = c.execute("SELECT id FROM expense_categories WHERE name=?", (name,)).fetchone()
    if row:
        return row["id"] if isinstance(row, sqlite3.Row) else row[0]
    c.execute("INSERT INTO expense_categories(name) VALUES (?)", (name,))
    conn.commit()
    return c.lastrowid


# -------- Financial Hub helpers --------

def _parse_finance_filters(req, ay):
    """Return a dict of filters with sensible defaults."""
    from datetime import date, timedelta

    f = {}
    f["term"] = (req.values.get("term") or ay["current_term"]).strip()
    f["year"] = int(req.values.get("year") or ay["year"])

    # Optional date overrides (YYYY-MM-DD); if provided, they apply to date filters
    f["from_date"] = (req.values.get("from_date") or "").strip() or None
    f["to_date"] = (req.values.get("to_date") or "").strip() or None

    # If no explicit dates, we filter by term/year; if dates given, we ignore term/year date constraints
    f["use_dates"] = bool(f["from_date"] and f["to_date"])
    return f

def _date_where(fragment_date_col, f):
    """Return (where_sql, params) for date range OR term/year filter."""
    if f["use_dates"]:
        return f"({fragment_date_col} BETWEEN ? AND ?)", [f["from_date"], f["to_date"]]
    else:
        # term/year filter
        return "(term = ? AND year = ?)", [f["term"], f["year"]]

def _fetch_finance_data(f):
    """
    Pulls 5 blocks:
      fees_rows (payment_type='fees'),
      req_rows (payment_type='requirements'),
      other_rows,
      exp_rows,
      totals dict (income/expenses/net)
    """
    import sqlite3
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row

    # ----- Fees (details) -----
    w_fees, p_fees = _date_where("date_paid", f)
    fees_rows = conn.execute(f"""
        SELECT
          date_paid, student_id, term, year, method,
          amount_paid, expected_amount, bursary_amount, carried_forward,
          (SELECT student_number FROM students s WHERE s.id = fees.student_id) AS student_number,
          (SELECT first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name FROM students s WHERE s.id = fees.student_id) AS full_name
        FROM fees
        WHERE {w_fees} AND payment_type = 'fees'
        ORDER BY date_paid DESC
    """, p_fees).fetchall()

    # ----- Requirements (details) -----
    w_reqs, p_reqs = _date_where("date_paid", f)
    req_rows = conn.execute(f"""
        SELECT
          date_paid, student_id, term, year, method,
          amount_paid, requirement_name,
          (SELECT student_number FROM students s WHERE s.id = fees.student_id) AS student_number,
          (SELECT first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name FROM students s WHERE s.id = fees.student_id) AS full_name
        FROM fees
        WHERE {w_reqs} AND payment_type = 'requirements'
        ORDER BY date_paid DESC
    """, p_reqs).fetchall()

    # ----- Other income (details) -----
    w_oi, p_oi = _date_where("date_received", f)
    other_rows = conn.execute(f"""
        SELECT source, amount, recorded_by, description, date_received, term, year
        FROM other_income
        WHERE {w_oi}
        ORDER BY date_received DESC
    """, p_oi).fetchall()

    # ----- Expenses (details) -----
    w_exp, p_exp = _date_where("date_spent", f)
    exp_rows = conn.execute(f"""
        SELECT e.description, e.amount, e.recorded_by, e.type,
               e.date_spent, e.term, e.year,
               (SELECT name FROM expense_categories c WHERE c.id = e.category_id) AS category
        FROM expenses e
        WHERE {w_exp}
        ORDER BY date_spent DESC
    """, p_exp).fetchall()

    # ----- Totals for statement -----
    fees_total = sum([float(r["amount_paid"] or 0) for r in fees_rows])
    req_total = sum([float(r["amount_paid"] or 0) for r in req_rows])
    other_total = sum([float(r["amount"] or 0) for r in other_rows])
    income_total = fees_total + req_total + other_total

    exp_total = sum([float(r["amount"] or 0) for r in exp_rows])
    net_total = income_total - exp_total

    conn.close()
    return fees_rows, req_rows, other_rows, exp_rows, {
        "fees_total": fees_total,
        "requirements_total": req_total,
        "other_income_total": other_total,
        "income_total": income_total,
        "expenses_total": exp_total,
        "net_total": net_total,
    }

def _balance_sheet_snapshot(f):
    """
    Very simple snapshot:
      - Cash-in (from filters): fees+requirements (amount_paid) + other_income
      - Receivables (all-time): sum per-student of positive (expected - bursary - paid) from fees
      - Net position: cash-in - (no liabilities tracked) + (assets like 'assets' table not revalued here)
    """
    import sqlite3
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row

    # Cash-in within filter
    w_fees, p_fees = _date_where("date_paid", f)
    fees_in = conn.execute(f"""
        SELECT COALESCE(SUM(amount_paid),0) AS t
        FROM fees WHERE {w_fees} AND payment_type='fees'
    """, p_fees).fetchone()["t"]

    w_req, p_req = _date_where("date_paid", f)
    req_in = conn.execute(f"""
        SELECT COALESCE(SUM(amount_paid),0) AS t
        FROM fees WHERE {w_req} AND payment_type='requirements'
    """, p_req).fetchone()["t"]

    w_oi, p_oi = _date_where("date_received", f)
    other_in = conn.execute(f"""
        SELECT COALESCE(SUM(amount),0) AS t
        FROM other_income WHERE {w_oi}
    """, p_oi).fetchone()["t"]

    cash_in = float(fees_in or 0) + float(req_in or 0) + float(other_in or 0)

    # Receivables (all-time, not just filters) – compute per student and clamp >= 0
    rows = conn.execute("""
        SELECT student_id,
               COALESCE(SUM(expected_amount),0) AS expd,
               COALESCE(SUM(bursary_amount),0) AS bur,
               COALESCE(SUM(amount_paid),0) AS paid
        FROM fees
        GROUP BY student_id
    """).fetchall()

    receivables = 0.0
    for r in rows:
        due = float(r["expd"] or 0) - float(r["bur"] or 0) - float(r["paid"] or 0)
        if due > 0:
            receivables += due

    conn.close()
    return {
        "cash_in": cash_in,
        "receivables": receivables,
        "net_position": cash_in + receivables # simplistic view
    }
  

 

# --- Helper: compute ranking list for class/term/year (for single view) ---
def class_ranking(conn, class_name, term, year):
    """
    Rank students in a class by overall average of displayed totals:
    EOT if present else blended average.
    """
    rows = conn.execute("""
        SELECT r.student_id AS sid,
               AVG(COALESCE(r.eot_mark, r.average_mark)) AS overall
        FROM reports r
        JOIN students s ON s.id = r.student_id
        WHERE s.class_name=? AND r.term=? AND r.year=?
        GROUP BY r.student_id
        ORDER BY overall DESC
    """, (class_name, term, year)).fetchall()
    return rows # list of rows with sid, overall

# --- Helper: grading legend rows ---
def fetch_grading_scale(conn):
    return conn.execute("""
        SELECT grade, lower_limit, upper_limit, COALESCE(comment,'') AS comment
        FROM grading_scale
        ORDER BY lower_limit
    """).fetchall()


def ensure_join_columns():
    conn = get_db_connection()
    c = conn.cursor()
    for name, typ in [
        ("date_joined", "TEXT"),
        ("term_joined", "TEXT"),
        ("year_of_joining", "INTEGER"),
    ]:
        try:
            c.execute(f"ALTER TABLE students ADD COLUMN {name} {typ}")
        except Exception:
            pass
    conn.commit()
    conn.close()





TERM_ORDER = {"Term 1": 1, "Term 2": 2, "Term 3": 3}

def ensure_fees_method_column():
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("ALTER TABLE fees ADD COLUMN method TEXT DEFAULT 'N/A'")
        conn.commit()
    except sqlite3.OperationalError:
        pass
    finally:
        conn.close()

def get_active_ay():
    conn = get_db_connection()
    r = conn.execute(
        "SELECT year, current_term FROM academic_years WHERE is_active=1 LIMIT 1"
    ).fetchone()
    conn.close()
    if not r:
        return {"year": int(datetime.now().strftime("%Y")), "term": "Term 1"}
    return {"year": int(r["year"]), "term": r["current_term"]}
    
def _is_safe_url(target):
    # Only allow redirects to our own host
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target or ""))
    return (test_url.scheme in ("http", "https")) and (ref_url.netloc == test_url.netloc)

def get_student(student_number=None, last_name=None):
    """Load ONE student by student_number OR (first matching) last_name."""
    conn = get_db_connection()
    if student_number:
        row = conn.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE student_number = ? AND archived = 0
        """, (student_number,)).fetchone()
    elif last_name:
        row = conn.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE last_name LIKE ? AND archived = 0
            ORDER BY last_name, first_name
            LIMIT 1
        """, (f"%{last_name}%",)).fetchone()
    else:
        row = None
    conn.close()
    return row


def ensure_comment_rules_schema():
    conn = get_db_connection()
    c = conn.cursor()
    # main table (adds `level` if missing)
    c.execute("""
      CREATE TABLE IF NOT EXISTS comment_rules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        role TEXT CHECK(role IN ('teacher','headteacher')) NOT NULL,
        scope TEXT CHECK(scope IN ('subject','overall')) NOT NULL,
        match_type TEXT CHECK(match_type IN ('grade','division','range')) NOT NULL,
        grade TEXT, -- for match_type='grade'
        division INTEGER, -- for match_type='division'
        lower_limit REAL, -- for match_type='range'
        upper_limit REAL,
        class_name TEXT, -- optional (e.g., 'P7'), NULL = any
        level TEXT, -- optional convenience (e.g., 'Lower','Upper'), NULL = any
        term TEXT, -- optional (e.g., 'Term 1'), NULL = any
        template_text TEXT NOT NULL,
        priority INTEGER DEFAULT 100, -- lower chosen first
        active INTEGER DEFAULT 1
      )
    """)
    # add missing `level` if older DB
    try:
        c.execute("ALTER TABLE comment_rules ADD COLUMN level TEXT")
    except Exception:
        pass
    c.execute("""
      CREATE INDEX IF NOT EXISTS ix_comment_rules_lookup
      ON comment_rules(role, scope, active, match_type, class_name, term)
    """)
    conn.commit()
    conn.close()


# optional: simple hierarchy (admin can do everything)
ROLE_IMPLIES = {
    "admin": {"admin","bursar","headteacher","dos","clerk","teacher"},
    # add more if you have seniors: e.g. "headteacher": {"teacher"}
}

def _normalize_roles(value) -> set[str]:
    """
    Accepts a string like 'bursar, clerk' or a list; returns a lowercased set.
    """
    if not value:
        return set()
    if isinstance(value, (list, tuple, set)):
        items = value
    else:
        items = str(value).split(",")
    return {r.strip().lower() for r in items if str(r).strip()}

def require_role(*allowed_roles):
    allowed = _normalize_roles(allowed_roles)

    @wraps(require_role)
    def decorator(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            # must be logged in
            user_id = session.get("user_id")
            raw_role = session.get("role") # may be 'bursar, clerk'
            if not user_id or raw_role is None:
                flash("Please sign in.", "warning")
                return redirect(url_for("login", next=request.path))

            user_roles = _normalize_roles(raw_role)

            # expand with hierarchy: admin -> all, etc.
            expanded = set(user_roles)
            for r in list(user_roles):
                expanded |= ROLE_IMPLIES.get(r, set())

            if not (expanded & allowed):
                flash("You don't have permission to access this page.", "danger")
                return redirect(url_for("dashboard"))

            return fn(*args, **kwargs)
        return wrapped
    return decorator



def _asdict(row):
    """Return a plain dict for sqlite3.Row or any mapping-like object."""
    if row is None:
        return {}
    return dict(row) if isinstance(row, sqlite3.Row) else (row if isinstance(row, dict) else {})

def _mean(seq):
    vals = [float(x) for x in seq if x is not None]
    return round(sum(vals) / len(vals), 2) if vals else None

def ordinal(n):
    try:
        n = int(n)
    except Exception:
        return ""
    if 10 <= (n % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1:"st",2:"nd",3:"rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def col_exists(conn, table: str, col: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any((r[1] or "").lower() == col.lower() for r in rows)

def _code_map_from_record_score(conn, student_id: int, term: str, year: int, column: str) -> dict:
    """
    Returns { 'ENG': 80, 'MAT': 75, 'SCI': 90, 'SST': 88 } for given column in record_score.
    Missing subject codes are omitted.
    """
    rows = conn.execute(
        f"""
        SELECT UPPER(COALESCE(sub.code, SUBSTR(sub.name,1,3))) AS code,
               r.{column} AS val
        FROM record_score r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=? AND r.term=? AND r.year=? 
        """,
        (student_id, term, year)
    ).fetchall()
    out = {}
    for r in rows:
        d = _asdict(r)
        code = (d.get("code") or "").upper().strip()
        if code in BIG4_CODES:
            out[code] = d.get("val")
    return out

def _code_map_from_midterms(conn, student_id: int, term: str, year: int, assessment_name: str) -> dict:
    """
    Reads a midterms row for the given assessment and returns a code->mark map.
    """
    r = conn.execute(
        """
        SELECT eng, mat, sci, sst
        FROM midterms
        WHERE student_id=? AND term=? AND year=? AND LOWER(TRIM(assessment))=LOWER(TRIM(?))
        LIMIT 1
        """,
        (student_id, term, year, assessment_name)
    ).fetchone()
    if not r:
        return {}
    d = _asdict(r)
    return {
        "ENG": d.get("eng"),
        "MAT": d.get("mat"),
        "SCI": d.get("sci"),
        "SST": d.get("sst"),
    }


def _build_midterm_panel_dynamic(conn, student_id: int, term: str, year: int, include_eot_row: bool = True):
    """
    Builds a dynamic midterm panel for *all* subjects present in subjects table,
    but computes AGG/TOTAL only from the core four: ENG, MATH, SCI, SST.

    Returns a list of rows like:
      [
        {
          "assessment": "OTH" | "HP" | "BOT" | "MID" | "EOT",
          "per_subj": { <subject_id>: {"mark": x, "grade": "C3"} , ... },
          "agg": <sum of points for core subjects or None>,
          "total": <sum of marks for core subjects or None>,
        },
        ...
      ]
    """
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Get all subjects once (dynamic)
    subs = c.execute("SELECT id AS sid, name, code FROM subjects ORDER BY name").fetchall()
    if not subs:
        return []

    # map subject_id -> True if core
    core_codes = {"ENG", "MATH", "SCI", "SST"}
    is_core_by_sid = {s["sid"]: ((s["code"] or "").upper() in core_codes) for s in subs}

    # helper: grading
    def grade_of(v):
        return grade_for_score(conn, v) if v is not None else None

    # helper: fetch values for a given column name into {sid: mark}
    def colvals_for(colname: str):
        # Pull only subjects that exist for this student/term/year
        rows = c.execute(
            f"""
            SELECT subject_id AS sid, MAX({colname}) AS val
            FROM record_score
            WHERE student_id=? AND term=? AND year=?
            GROUP BY subject_id
            """,
            (student_id, term, year)
        ).fetchall()
        return {r["sid"]: r["val"] for r in rows}

    # Build each assessment in the desired order
    spec = [
        ("OTH", "other_mark"),
        ("HP", "holiday_mark"),
        ("BOT", "bot_mark"),
        ("MID", "midterm_mark"),
        ("EOT", "eot_mark"),
    ]

    out = []
    for label, col in spec:
        colvals = colvals_for(col)
        if not colvals:
            # nothing recorded for this assessment at all
            continue

        # Build per-subject cells and accumulate only cores
        per_subj = {}
        core_marks = []
        core_points = []

        for s in subs:
            sid = s["sid"]
            v = colvals.get(sid)
            g = grade_of(v)
            per_subj[sid] = {"mark": v, "grade": g}

            if is_core_by_sid.get(sid) and v is not None:
                core_marks.append(v)
                if g in {'D1','D2','C3','C4','C5','C6','P7','P8','F9'}:
                    _pts = {'D1':1,'D2':2,'C3':3,'C4':4,'C5':5,'C6':6,'P7':7,'P8':8,'F9':9}[g]
                    core_points.append(_pts)

        if not any(v is not None for v in colvals.values()):
            # no actual numbers anywhere — skip row
            continue

        out.append(dict(
            assessment=label,
            per_subj=per_subj,
            agg=(sum(core_points) if core_points else None),
            total=(sum(core_marks) if core_marks else None),
        ))

    if not include_eot_row:
        out = [r for r in out if r["assessment"] != "EOT"]

    return out

def _norm_section(sec: str) -> str:
        s = (sec or '').strip().lower()
        if s in ('day','d'): return 'Day'
        if s in ('boarding','board','b'): return 'Boarding'
        return (sec or '').strip()

    
def _expected_fees_for_student(conn, student_id: int, term: str, year: int) -> float:
    """
    Priority:
      1) fee_codes.amount when student has a fees_code (prefers year match if column exists)
      2) class_fees.amount for class_name + section (prefers year match if column exists)
      3) 0.0
    """
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # --- student basics
    st = c.execute(
        "SELECT class_name, section, fees_code FROM students WHERE id=?",
        (student_id,)
    ).fetchone()
    if not st:
        return 0.0

    class_name = (st["class_name"] or "").strip()
    section_raw = (st["section"] or "").strip().lower()
    if section_raw in ("day", "d"):
        section = "Day"
    elif section_raw in ("boarding", "board", "b"):
        section = "Boarding"
    else:
        section = "" # unknown → let queries ignore/handle

    # --- 1) fee_codes by student's fees_code
    fees_code = ((st["fees_code"] or "").strip() if "fees_code" in st.keys() else "")
    if fees_code and _table_exists(conn, "fee_codes"):
        # prefer year-specific if the column exists
        if _column_exists(conn, "fee_codes", "year"):
            row = c.execute(
                "SELECT amount FROM fee_codes WHERE code=? AND year=? LIMIT 1",
                (fees_code, year)
            ).fetchone()
            if row and row["amount"] is not None:
                return float(row["amount"])
        # fallback without year
        row = c.execute(
            "SELECT amount FROM fee_codes WHERE code=? LIMIT 1",
            (fees_code,)
        ).fetchone()
        if row and row["amount"] is not None:
            return float(row["amount"])

    # --- 2) class_fees by class + section (prefer year if present)
    if _table_exists(conn, "class_fees"):
        if _column_exists(conn, "class_fees", "year"):
            row = c.execute(
                """
                SELECT amount
                FROM class_fees
                WHERE class_name = ?
                  AND lower(section) = lower(?)
                  AND year = ?
                LIMIT 1
                """,
                (class_name, section, year)
            ).fetchone()
            if row and row["amount"] is not None:
                return float(row["amount"])
        # fallback without year
        row = c.execute(
            """
            SELECT amount
            FROM class_fees
            WHERE class_name = ?
              AND lower(section) = lower(?)
            LIMIT 1
            """,
            (class_name, section)
        ).fetchone()
        if row and row["amount"] is not None:
            return float(row["amount"])

    # --- 3) nothing matched
    return 0.0


def _build_mid_row(conn, title: str, by_code: dict) -> dict:
    """
    by_code: {'ENG': mark, 'MAT': mark, 'SCI': mark, 'SST': mark}
    Returns a dict with marks + *_grade + *_comment + agg + total for the row.
    """
    row = {"assessment": title}
    points_map = {'D1':1,'D2':2,'C3':3,'C4':4,'C5':5,'C6':6,'P7':7,'P8':8,'F9':9}
    agg_points = []

    for code in BIG4_CODES:
        val = by_code.get(code)
        grd = grade_for_score(conn, val) if val is not None else None
        cmt = comment_for_grade(conn, grd) if grd else None

        # pack with lowercase keys the template expects
        key = code.lower() # eng/mat/sci/sst
        row[key] = val
        row[f"{key}_grade"] = grd or ""
        row[f"{key}_comment"] = cmt or ""

        if grd in points_map:
            agg_points.append(points_map[grd])

    # aggregate (only if all 4 grades exist)
    row["agg"] = sum(agg_points) if len(agg_points) == 4 else None

    # total (sum of numeric big-4 marks present)
    total_vals = [by_code.get(c) for c in BIG4_CODES if by_code.get(c) is not None]
    row["total"] = round(sum(total_vals), 0) if total_vals else None
    return row

# ---------- main builder for the Mid-Term panel ----------
def midterm_rows_for_student(conn, student_id: int, term: str, year: int) -> list[dict]:
    """
    Returns a list of dict rows for the Mid-Term table:
      - Beginning of Term (BOT)
      - Holiday Package (if present)
      - Other Assessments (if present)
      - Mid of Term (stored midterm_mark or mean of BOT/HP/Other)
    Each row includes eng/mat/sci/sst, *_grade, *_comment, agg, total.
    """
    # source maps
    bot_map = _code_map_from_record_score(conn, student_id, term, year, "bot_mark")
    mid_map = _code_map_from_record_score(conn, student_id, term, year, "midterm_mark")
    holiday_map = _code_map_from_midterms(conn, student_id, term, year, "Holiday Package")
    other_map = _code_map_from_midterms(conn, student_id, term, year, "Other Assessments")

    rows = []

    if bot_map:
        rows.append(_build_mid_row(conn, "Beginning of Term (BOT)", bot_map))

    if holiday_map:
        rows.append(_build_mid_row(conn, "Holiday Package", holiday_map))

    if other_map:
        rows.append(_build_mid_row(conn, "Other Assessments", other_map))

    # Mid of Term (prefer stored midterm_mark; else mean of BOT/HP/Other per subject)
    final_mid = {}
    for code in BIG4_CODES:
        stored = mid_map.get(code)
        if stored is not None:
            final_mid[code] = stored
        else:
            final_mid[code] = _mean([bot_map.get(code), holiday_map.get(code), other_map.get(code)])

    # Only add the Mid row if at least one subject has a value
    if any(final_mid.get(c) is not None for c in BIG4_CODES):
        rows.append(_build_mid_row(conn, "Mid of Term Exams", final_mid))

    return rows

def _ordinal(n: int | None) -> str | None:
    if n is None:
        return None
    s = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    if 10 <= (n % 100) <= 20:
        s = "th"
    return f"{n}{s}"

def _next_term_name_and_year(cur_term: str, cur_year: int):
    order = ["Term 1", "Term 2", "Term 3"]
    try:
        i = order.index(cur_term)
    except ValueError:
        return "Term 2", cur_year
    return (order[i+1], cur_year) if i < 2 else (order[0], cur_year + 1)


def grade_for_score(conn, score: float) -> str | None:
    """Look up D1..F9 from grading_scale where min_score <= score <= max_score."""
    if score is None:
        return None
    row = conn.execute("""
        SELECT grade
        FROM grading_scale
        WHERE ? BETWEEN lower_limit AND upper_limit
        ORDER BY (upper_limit - lower_limit) ASC
        LIMIT 1
    """, (score,)).fetchone()
    return (row["grade"] if row else None)

# Used to calculate average for both single and batch
def fetch_report_rows(conn, student_id, term, year):
    conn.row_factory = sqlite3.Row
    return conn.execute("""
        SELECT sub.name AS subject, sub.code AS subject_code,
               r.eot_mark AS eot,
               COALESCE(r.average_mark, r.eot_mark) AS total_100, -- average first
               r.grade, r.comment, r.teacher_initial AS initials,
               r.teacher_remark, r.headteacher_remark
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=? AND r.term=? AND r.year=?
        ORDER BY sub.name
    """, (student_id, term, year)).fetchall()




def _bot_mid_by_sid(conn, class_name, term, year):
    """returns dict[sid][code] = {'bot':x, 'mid':y} (codes are ENG/MATH/SCI/SST)"""
    src = detect_scores_table(conn)
    if not src:
        return {}
    # join subjects to derive codes
    if not col_exists(conn, "subjects", "code"):
        code_sql = "UPPER(SUBSTR(sub.name,1,3))"
    else:
        code_sql = "UPPER(sub.code)"
    rows = conn.execute(f"""
        SELECT r.student_id, {code_sql} AS code,
               MAX(r.bot_mark) AS bot, MAX(r.midterm_mark) AS mid
        FROM {src} r
        JOIN students s ON s.id = r.student_id
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE s.archived=0 AND s.class_name=? AND r.term=? AND r.year=?
        GROUP BY r.student_id, code
    """, (class_name, term, year)).fetchall()
    out = {}
    for r in rows:
        sid = r["student_id"]; code = (r["code"] or "").upper()
        if code not in BIG4_CODES: # only big-4 here for mid-term panel
            continue
        out.setdefault(sid, {})[code] = {"bot": _to_num(r["bot"]), "mid": _to_num(r["mid"])}
    return out


def _midterm_rows(conn, student_id, term, year):
    """
    Assemble the optional midterm table from record_score columns.
    We create up to 3 logical rows if any data exists:
      - 'Beginning of Term'
      - 'Mid of Term'
      - 'Holiday Package'
    Each row shows ENG/MAT/SCI/SST marks + derived grade.
    """
    r = conn.execute("""
        SELECT sub.code AS code,
               rs.bot_mark, rs.midterm_mark, rs.eot_mark,
               rs.holiday_mark, rs.other_mark, rs.ca_mark
        FROM record_score rs
        JOIN subjects sub ON sub.id = rs.subject_id
        WHERE rs.student_id=? AND rs.term=? AND rs.year=?
    """, (student_id, term, year)).fetchall()

    # Build code->marks map
    by_code = {}
    for row in r:
        code = (row["code"] or "").upper().strip()
        if code in BIG4_CODES:
            by_code[code] = dict(bot=row["bot_mark"],
                                 mid=row["midterm_mark"],
                                 eot=row["eot_mark"],
                                 holiday=row["holiday_mark"],
                                 other=row["other_mark"],
                                 ca=row["ca_mark"])

    def _row(label, key):
        """key in {'bot','mid','holiday'} -> row dict or None if empty"""
        vals = {c: (by_code.get(c, {}).get(key)) for c in BIG4_CODES}
        if all(v is None for v in vals.values()):
            return None
        # derive grade per subject
        grades = {c: grade_for_score(conn, vals[c]) for c in BIG4_CODES}
        # aggregate (Big4 points)
        points_map = {'D1':1,'D2':2,'C3':3,'C4':4,'C5':5,'C6':6,'P7':7,'P8':8,'F9':9}
        pts = [points_map[g] for g in grades.values() if g in points_map]
        agg = sum(pts) if len(pts) == 4 else None
        total = sum([v for v in vals.values() if v is not None]) if any(vals.values()) else None
        return dict(
            assessment=label,
            eng=vals["ENG"], eng_grade=grades["ENG"],
            mat=vals["MATH"], mat_grade=grades["MATH"],
            sci=vals["SCI"], sci_grade=grades["SCI"],
            sst=vals["SST"], sst_grade=grades["SST"],
            agg=agg, total=total
        )

    rows = []
    for label, key in (("Beginning of Term", "bot"),
                       ("Mid of Term Exams", "mid"),
                       ("Holiday Package", "holiday")):
        rr = _row(label, key)
        if rr: rows.append(rr)
    return rows



# --- helpers used below (you already have these) ---
# grade_for_score(conn, score) -> 'D1'/'C3'/...
# comment_for_grade(conn, grade) -> 'Excellent'/'Good'/...
# BIG4_CODES = ['ENG','MATH','SCI','SST']

def _mid_panel_for_student(conn, student_id: int, term: str, year: int):
    """
    Build the midterm panel rows (BOT, Mid of Term, Holiday Package, Other Assessments)
    from *record_score* only. We never select *_grade columns from SQL; we compute them here.
    Returns a list of dicts with keys:
      assessment, eng, eng_grade, mat, mat_grade, sci, sci_grade, sst, sst_grade, agg, total
    Only includes assessments that actually have at least one score.
    """
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Map subject_id -> code for this student (BIG4 only)
    code_map = dict(
        (r["id"], r["code"])
        for r in c.execute("""
            SELECT id, UPPER(TRIM(code)) AS code
            FROM subjects
            WHERE UPPER(TRIM(code)) IN ('ENG','MATH','SCI','SST')
        """)
    )

    # Pull all record_score rows for this student/term/year (raw marks only)
    rows = c.execute("""
        SELECT subject_id, bot_mark, midterm_mark, holiday_mark, other_mark, eot_mark
        FROM record_score
        WHERE student_id=? AND term=? AND year=?
    """, (student_id, term, year)).fetchall()

    # Bucket marks by subject code
    marks = {code: {"BOT": None, "MID": None, "HP": None, "OTHER": None, "EOT": None}
             for code in ("ENG", "MATH", "SCI", "SST")}

    for r in rows:
        code = code_map.get(r["subject_id"])
        if not code: # ignore non-BIG4
            continue
        if r["bot_mark"] is not None: marks[code]["BOT"] = r["bot_mark"]
        if r["midterm_mark"] is not None: marks[code]["MID"] = r["midterm_mark"]
        if r["holiday_mark"] is not None: marks[code]["HP"] = r["holiday_mark"]
        if r["other_mark"] is not None: marks[code]["OTHER"] = r["other_mark"]
        if r["eot_mark"] is not None: marks[code]["EOT"] = r["eot_mark"]

    def _row_from_bucket(name, key):
        """Build one output row for an assessment key (BOT/MID/HP/OTHER) if any value exists."""
        eng = marks["ENG"][key]; mat = marks["MATH"][key]
        sci = marks["SCI"][key]; sst = marks["SST"][key]
        # skip completely empty rows
        if all(v is None for v in (eng, mat, sci, sst)):
            return None

        def g(v): # grade string for a mark (or empty)
            return grade_for_score(conn, v) if v is not None else ""

        # Optional agg/total on BIG4 for this assessment
        grade_points = {'D1':1,'D2':2,'C3':3,'C4':4,'C5':5,'C6':6,'P7':7,'P8':8,'F9':9}
        gp = [grade_points.get(g(eng)), grade_points.get(g(mat)),
              grade_points.get(g(sci)), grade_points.get(g(sst))]
        agg = sum(x for x in gp if isinstance(x, int)) if all(isinstance(x, int) for x in gp) else ""
        total = sum(v for v in (eng, mat, sci, sst) if v is not None) if any(v is not None for v in (eng, mat, sci, sst)) else ""

        return {
            "assessment": name,
            "eng": eng, "eng_grade": g(eng),
            "mat": mat, "mat_grade": g(mat),
            "sci": sci, "sci_grade": g(sci),
            "sst": sst, "sst_grade": g(sst),
            "agg": agg, "total": total,
        }

    out = []
    # Only append rows that actually exist
    for label, key in (("Beginning of Term", "BOT"),
                       ("Mid of Term Exams", "MID"),
                       ("Holiday Package", "HP"),
                       ("Other Assessments", "OTHER")):
        row = _row_from_bucket(label, key)
        if row:
            out.append(row)
    return out


def _mid_panel_from_record_score(conn, student_id, term, year):
    """
    Builds a compact mid-term panel:
      - Only shows columns among BOT / HOLIDAY / MID that actually have data.
      - No initials.
      - One concise comment per subject: prefer MID, else HOLIDAY, else BOT.
      - Grades come from grading_scale per mark.
    """
    q = """
      SELECT sub.name AS subject, sub.code AS code, r.subject_id,
             MAX(r.bot_mark) AS bot,
             MAX(r.midterm_mark) AS mid,
             MAX(r.holiday_mark) AS holiday
      FROM record_score r
      JOIN subjects sub ON sub.id = r.subject_id
      WHERE r.student_id=? AND r.term=? AND r.year=?
      GROUP BY r.subject_id
      ORDER BY sub.name
    """
    rows = conn.execute(q, (student_id, term, year)).fetchall()

    def pack(mark):
        if mark is None:
            return {"mark": None, "grade": "", "comment": ""}
        g = grade_for_score(conn, mark)
        c = comment_for_grade(conn, g) or ""
        return {"mark": mark, "grade": g, "comment": c}

    panel = []
    any_bot = any_mid = any_hol = False

    for r in rows:
        bot = pack(r["bot"])
        mid = pack(r["mid"])
        hol = pack(r["holiday"])

        if r["bot"] is not None: any_bot = True
        if r["mid"] is not None: any_mid = True
        if r["holiday"] is not None: any_hol = True

        comment_choice = mid["comment"] or hol["comment"] or bot["comment"] or ""

        panel.append({
            "subject": r["subject"],
            "code": r["code"],
            "BOT": bot,
            "HOLIDAY": hol,
            "MID": mid,
            "comment_choice": comment_choice
        })

    mid_cols = []
    if any_bot: mid_cols.append("BOT")
    if any_hol: mid_cols.append("HOLIDAY")
    if any_mid: mid_cols.append("MID")

    return panel, mid_cols


def comment_for_grade(conn, grade):
    """Use grading_scale.comment if present; else empty."""
    if not grade:
        return ""
    row = conn.execute("SELECT comment FROM grading_scale WHERE grade=? LIMIT 1", (grade,)).fetchone()
    return (row["comment"] if row and row["comment"] else "") or ""

def guess_teacher_initials(conn, class_name, subject_id):
    row = conn.execute("""
        SELECT initials FROM teachers
        WHERE class_name=? AND (subjects LIKE '%'||?||'%' OR subjects IS NULL)
        AND TRIM(COALESCE(initials,'')) <> ''
        LIMIT 1
    """, (class_name, str(subject_id))).fetchone()
    return row["initials"] if row else ""


def _is_big4(code_or_name: str) -> str | None:
    s = (code_or_name or "").upper().strip()
    # try code first
    if s in {"ENG", "ENGLISH"}: return "ENG"
    if s in {"MATH", "MATHEMATICS", "MAT"}: return "MATH"
    if s in {"SCI", "SCIENCE"}: return "SCI"
    if s in {"SST", "SOCIAL STUDIES", "SOCIAL-STUDIES"}: return "SST"
    # try name patterns
    if "ENGLISH" in s: return "ENG"
    if "MATH" in s or "MATHEMAT" in s: return "MATH"
    if "SCIENCE" in s: return "SCI"
    if "SOCIAL" in s or "SST" in s: return "SST"
    return None



# ---------- Overall average + Division (Big-4 only) ----------
def compute_overall_for_student(conn, student_id, term, year):
    """
    Overall average: mean of COALESCE(EOT, blended average) across all subjects.
    Aggregate/Division: ONLY ENG, MATH, SCI, SST; map grades -> points.
    Division bands: 4–12 => 1, 13–24 => 2, 25–29 => 3, 30–34 => 4, else U.
    """
    BIG4 = {"ENG","MATH","SCI","SST"}
    points_map = {'D1':1,'D2':2,'C3':3,'C4':4,'C5':5,'C6':6,'P7':7,'P8':8,'F9':9}

    rows = conn.execute("""
        SELECT COALESCE(sub.code, TRIM(UPPER(sub.name))) AS code,
               r.grade,
               COALESCE(r.eot_mark, r.average_mark) AS total_100
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=? AND r.term=? AND r.year=?
    """, (student_id, term, year)).fetchall()

    disp = [r["total_100"] for r in rows if r["total_100"] is not None]
    avg_overall = round(sum(disp)/len(disp), 2) if disp else None

    agg_points = []
    for r in rows:
        code = (r["code"] or "").upper().strip()
        grd = (r["grade"] or "").upper().strip()
        if code in BIG4 and grd in points_map:
            agg_points.append(points_map[grd])

    aggregate = sum(agg_points) if len(agg_points) == 4 else None

    division = None
    if aggregate is not None:
        if 4 <= aggregate <= 12: division = "1"
        elif 13 <= aggregate <= 24: division = "2"
        elif 25 <= aggregate <= 29: division = "3"
        elif 30 <= aggregate <= 34: division = "4"
        else: division = "U"

    return avg_overall, division, aggregate


def pick_comment_template(*, role: str, scope: str,
                          division=None, average=None,
                          class_name=None, term=None):
    """
    Returns the best-matching template_text (string) or None.
    Priority rules:
      1) active=1 only
      2) exact filters (class_name/term) preferred but both are optional
      3) match_type precedence by explicit priority ASC
         - 'grade' is not supported here (only overall+division/range),
           but we keep it for subject-level calls if you reuse this.
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    ensure_comment_rules_schema()

    # Gather candidates of correct role/scope and active
    q = """
      SELECT *
        FROM comment_rules
       WHERE role=? AND scope=? AND active=1
    """
    params = [role, scope]

    # optional filters: prefer exact class/term if set, otherwise allow NULL (any)
    # We'll fetch all and rank in Python so we can consider NULL as "less specific".
    rows = c.execute(q, params).fetchall()

    # Compute a list of (rank_tuple, row) and pick the smallest rank by tuple ordering
    ranked = []
    for r in rows:
        # Skip rows with missing essential fields based on match_type
        mtype = r["match_type"]
        ok = True
        if mtype == "division" and r["division"] is None:
            ok = False
        if mtype == "range" and (r["lower_limit"] is None or r["upper_limit"] is None):
            ok = False
        if not ok:
            continue

        # Apply logical match for division / range
        fits = False
        if mtype == "division" and division is not None:
            fits = (int(division) == int(r["division"]))
        elif mtype == "range" and average is not None:
            try:
                fits = (float(r["lower_limit"]) <= float(average) <= float(r["upper_limit"]))
            except Exception:
                fits = False
        elif mtype == "grade":
            # grade-based rules are typically for subject scope;
            # allow use if caller passes grade via 'average' slot as a string
            if isinstance(average, str) and r["grade"]:
                fits = (average.strip().upper() == r["grade"].strip().upper())

        if not fits:
            continue

        # Specificity: exact class/term matches outrank NULLs
        class_spec = 0 if (r["class_name"] and class_name and r["class_name"] == class_name) else (1 if r["class_name"] else 2)
        term_spec = 0 if (r["term"] and term and r["term"] == term) else (1 if r["term"] else 2)

        # priority ASC first, then specificity, then id (stable)
        rank = (int(r["priority"] or 100), class_spec, term_spec, int(r["id"]))
        ranked.append((rank, r))

    if not ranked:
        conn.close()
        return None

    ranked.sort(key=lambda x: x[0])
    best = ranked[0][1]["template_text"]
    conn.close()
    return (best or "").strip() or None
    



def autofill_head_comment(student_id, class_name, term, year):
    conn = get_db_connection()
    avgm, division = compute_overall_for_student(conn, student_id, term, year)

    text = pick_comment_template(
        role="headteacher",
        scope="overall",
        division=division,
        average=avgm,
        class_name=class_name,
        term=term
    )
    if text:
        conn.execute("""
          UPDATE reports
          SET headteacher_remark = ?
          WHERE student_id=? AND term=? AND year=? 
        """, (text, student_id, term, year))
        conn.commit()
    conn.close()



def bursary_total(student_id, term, year):
    conn = get_db_connection()
    r = conn.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM bursaries
        WHERE student_id=? AND term=? AND year=?
    """, (student_id, term, year)).fetchone()
    conn.close()
    return float(r["total"] or 0)



def paid_sum(student_id, term, year, payment_type="school_fees") -> float:
    conn = get_db_connection()
    try:
        r = conn.execute("""
            SELECT COALESCE(SUM(amount_paid), 0) AS t
            FROM fees
            WHERE student_id = ? AND term = ? AND year = ? AND payment_type = ?
        """, (student_id, term, year, payment_type)).fetchone()
        val = r["t"] if isinstance(r, sqlite3.Row) else r[0]
        return float(val or 0.0)
    finally:
        conn.close()


def carried_forward(student_id, term, year):
    """Outstanding before the active term/year."""
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT COALESCE(expected_amount,0) AS expected_amount,
               COALESCE(bursary_amount,0) AS bursary_amount,
               COALESCE(amount_paid,0) AS amount_paid,
               term, year
        FROM fees
        WHERE student_id=? AND (
              year < ?
           OR (year = ? AND
               (CASE term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END)
             < (CASE ? WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END))
        )
    """, (student_id, year, year, term)).fetchall()
    conn.close()
    outstanding = 0.0
    for r in rows:
        outstanding += (float(r["expected_amount"]) - float(r["bursary_amount"]) - float(r["amount_paid"]))
    return max(outstanding, 0.0)

def requirements_due(student):
    """Requirements configured for the student's class (term-aware if column exists)."""
    ay = get_active_ay()
    conn = get_db_connection()
    cols = [c["name"] for c in conn.execute("PRAGMA table_info(requirements)").fetchall()]
    if "term" in cols:
        rows = conn.execute("""
            SELECT id, name, qty, amount, COALESCE(term,'') AS term
            FROM requirements
            WHERE class_name = ?
              AND (term IS NULL OR term = ?)
            ORDER BY name
        """, (student["class_name"], ay["term"])).fetchall()
    else:
        rows = conn.execute("""
            SELECT id, name, qty, amount, '' AS term
            FROM requirements
            WHERE class_name = ?
            ORDER BY name
        """, (student["class_name"],)).fetchall()
    conn.close()
    return rows

def requirements_paid_sum(student_id, term, year):
    conn = get_db_connection()
    r = conn.execute("""
        SELECT COALESCE(SUM(amount_paid),0) AS total
        FROM fees
        WHERE student_id=? AND term=? AND year=? AND payment_type='requirements'
    """, (student_id, term, year)).fetchone()
    conn.close()
    return float(r["total"] or 0)


def log_action(user_id, action):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("INSERT INTO audit_trail (user_id, action) VALUES (?, ?)", (user_id, action))
    conn.commit()
    conn.close()
    

def seed_classes():
    conn = get_db_connection()
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS classes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_name TEXT NOT NULL,
        level TEXT,
        stream TEXT
    )''')

    # Clear old records
    c.execute("DELETE FROM classes")

    # Insert fresh P1–P7 Stream A
    c.executemany(
        "INSERT INTO classes (class_name, level, stream) VALUES (?, ?, ?)",
        [
            ('Baby','Nursery','A'),
            ('Middle','Nursery','A'),
            ('Top','Nursery','A'),
            ('P1','Primary','A'),
            ('P2','Primary','A'),
            ('P3','Primary','A'),
            ('P4','Primary','A'),
            ('P5','Primary','A'),
            ('P6','Primary','A'),
            ('P7','Primary','A')
        ]
    )

    # Unique index
    c.execute('''CREATE UNIQUE INDEX IF NOT EXISTS uq_classes_class_stream
                 ON classes(class_name, stream)''')

    conn.commit()
    conn.close()
    print("✅ Classes seeded successfully")
    

    


    # helpful uniqueness to avoid duplicates per class/term/item
    c.execute("""
      CREATE UNIQUE INDEX IF NOT EXISTS uq_requirements_class_term_name
      ON requirements(class_name, COALESCE(term,''), name)
    """)

    # fees table: add requirement_name column (optional, nice to keep what was paid)
    try:
        c.execute("ALTER TABLE fees ADD COLUMN requirement_name TEXT")
    except Exception:
        pass # already exists or not needed

    conn.commit()
    conn.close()




def resolve_subject_id(conn, subject_code=None, subject_name=None):
    c = conn.cursor()
    if subject_code:
        c.execute("SELECT id FROM subjects WHERE code = ?", (subject_code,))
        row = c.fetchone()
        if row: return row["id"]
    if subject_name:
        c.execute("SELECT id FROM subjects WHERE name = ?", (subject_name,))
        row = c.fetchone()
        if row: return row["id"]
    return None

def resolve_student_id(conn, student_number=None):
    if not student_number: return None
    row = conn.execute("SELECT id FROM students WHERE student_number = ?", (student_number,)).fetchone()
    return row["id"] if row else None


def get_class_requirements(class_name: str, term: str | None):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    # prefer exact term; also show term-less generic items for that class
    rows = conn.execute("""
        SELECT id, name, qty, amount, term
        FROM requirements
        WHERE class_name = ?
          AND (term = ? OR term IS NULL OR term = '')
        ORDER BY name
    """, (class_name, term)).fetchall()
    conn.close()
    return rows

def get_student_by_id(sid: int):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    row = conn.execute("""
        SELECT id, student_number, first_name, Middle_name, last_name,
               class_name, stream, section
        FROM students WHERE id = ?
    """, (sid,)).fetchone()
    conn.close()
    return row



def populate_default_expense_categories():
    default_categories = [
        'Salaries',
        'Stationery',
        'Utilities',
        'Transport',
        'Maintenance',
        'Service Providers',
        'Uniforms',
        'Examinations',
        'Meals',
        'Office supplies',
        'Medical',
        'Bonus',
        'Allowance',
        'Electricity',
        'Teachers Rent',
        'Water',
        'Religion',
        'Staf Welfare',
        'P7 Budget',
        'Outing',
        'Directors Budget',
        'Loans',
        'School Functions',
        'Donations',
        'Construction 1',
        'Construction 2',
        'Sports',
        'Computers and Printers',
        'Medical',
        'Sanitation',
        'Vans Repair',
        'Fuel',
        'Kitchen',
        'Miscellaneous'
    ]

    conn = get_db_connection()
    c = conn.cursor()

    for category in default_categories:
        try:
            c.execute("INSERT OR IGNORE INTO expense_categories (name) VALUES (?)", (category,))
        except:
            continue # Skip any insert error
    conn.commit()
    conn.close()
    print("Default expense categories inserted.")
    
def get_class_fee(conn, class_name: str, section: str | None, level: str | None = None) -> float:
    """Return amount from class_fees for (class_name, section[, level]). 0.0 if not found."""
    sec = norm_section(section) or ""
    row = conn.execute(
        """
        SELECT amount
          FROM class_fees
         WHERE class_name = ?
           AND LOWER(section) = LOWER(?)
           AND (level IS NULL OR level = ?)
         LIMIT 1
        """,
        (class_name or "", sec, level),
    ).fetchone()
    return float(row["amount"]) if row and row["amount"] is not None else 0.0
    




def ensure_bursaries_schema():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS bursaries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        sponsor_name TEXT,
        amount REAL NOT NULL,
        term TEXT, -- 'Term 1'|'Term 2'|'Term 3'
        year INTEGER NOT NULL,
        FOREIGN KEY(student_id) REFERENCES students(id)
      )
    """)
    c.execute("""
      CREATE UNIQUE INDEX IF NOT EXISTS uq_bursary_entry
      ON bursaries(student_id, year, term, sponsor_name)
    """)
    conn.commit()
    conn.close()
    
  

    




def _find_student_by_sn_or_ln(student_number, last_name):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    r = None
    if student_number:
        r = conn.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE student_number = ? AND archived = 0
        """, (student_number,)).fetchone()
    elif last_name:
        r = conn.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream, section
            FROM students
            WHERE last_name LIKE ? AND archived = 0
            ORDER BY last_name, first_name
            LIMIT 1
        """, (f"%{last_name}%",)).fetchone()
    conn.close()
    return r



def _term_order_val(t: str) -> int:
    order = {"Term 1": 1, "Term 2": 2, "Term 3": 3}
    return order.get(t, 99)





def compute_student_financials(student_id: int, class_name: str, term: str, year: int) -> dict:
    import sqlite3
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # --- student + section ---
    stu = c.execute("SELECT class_name, section FROM students WHERE id=?", (student_id,)).fetchone()
    eff_class = (stu["class_name"] if (stu and stu["class_name"]) else class_name) or ""
    try:
        sec_norm = norm_section((stu["section"] if stu else "") or "")
    except NameError:
        s = ((stu["section"] if stu else "") or "").strip().lower()
        sec_norm = "Day" if s in ("day","d") else ("Boarding" if s in ("boarding","board","b") else "")

    # --- expected fees (strict: class + section) ---
    row = c.execute("""
        SELECT amount FROM class_fees
        WHERE class_name=? AND lower(section)=lower(?) LIMIT 1
    """, (eff_class, sec_norm or "")).fetchone()
    expected_fees = float(row["amount"]) if row and row["amount"] is not None else 0.0

    # --- expected requirements ---
    row = c.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM requirements
        WHERE class_name=? AND (term=? OR term IS NULL OR term='')
    """, (eff_class, term)).fetchone()
    expected_requirements_base = float(row["total"]) if row else 0.0

    # --- transport overlay (UNCHANGED) ---
    transport_fare = 0.0
    transport_paid_term = 0.0
    try:
        tinfo = transport_subscription_info(student_id, term, year)
    except Exception:
        tinfo = None
    if tinfo and float(tinfo.get("fare_per_term") or 0) > 0:
        transport_fare = float(tinfo["fare_per_term"] or 0.0)
        try:
            transport_paid_term = transport_paid_via_requirements(conn, student_id, term, year)
        except Exception:
            transport_paid_term = 0.0
    expected_requirements = expected_requirements_base + transport_fare

    # --- bursary this term ---
    bursary_current = float(c.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM bursaries
        WHERE student_id=? AND term=? AND year=?
    """, (student_id, term, year)).fetchone()["total"])

    # --- payments this term ---
    paid_fees = float(c.execute("""
        SELECT COALESCE(SUM(amount_paid),0) AS total
        FROM fees
        WHERE student_id=? AND term=? AND year=? 
          AND lower(payment_type) IN ('school_fees','fees')
    """, (student_id, term, year)).fetchone()["total"])
    paid_requirements = float(c.execute("""
        SELECT COALESCE(SUM(amount_paid),0) AS total
        FROM fees
        WHERE student_id=? AND term=? AND year=? 
          AND lower(payment_type)='requirements'
    """, (student_id, term, year)).fetchone()["total"])

    # --- opening balance (OB) ---
    opening_balance = float(c.execute("""
        SELECT COALESCE(SUM(COALESCE(expected_amount,0) - COALESCE(amount_paid,0)), 0) AS total
        FROM fees
        WHERE student_id=?
          AND lower(payment_type) IN ('opening_balance','opening balance','ob')
          AND (comment IS NULL OR lower(comment) NOT LIKE '%voided%')
    """, (student_id,)).fetchone()["total"] or 0.0)

    # --- prior arrears (before current term) ---
    prior_rows = c.execute("""
        SELECT expected_amount, bursary_amount, amount_paid, term, year
        FROM fees
        WHERE student_id=? 
          AND lower(payment_type) IN ('school_fees','fees')
          AND (year < ?
               OR (year = ? AND
                   CASE lower(term)
                     WHEN 'term 1' THEN 1
                     WHEN 'term 2' THEN 2
                     WHEN 'term 3' THEN 3
                     ELSE 99 END < ?))
    """, (student_id, year, year, _term_order_val(term))).fetchall()
    prior_arrears = 0.0
    for r in prior_rows:
        prior_arrears += float(r["expected_amount"] or 0.0) - float(r["bursary_amount"] or 0.0) - float(r["amount_paid"] or 0.0)
    if prior_arrears < 0:
        prior_arrears = 0.0

    # --- carry forward (UNCHANGED) ---
    carry_forward = opening_balance + prior_arrears

    # --- balances (UNCHANGED math) ---
    total_due_this_term = (expected_fees + expected_requirements) - bursary_current
    balance_this_term = total_due_this_term - (paid_fees + paid_requirements)
    overall_balance = carry_forward + balance_this_term

    # ---- NEW: non-breaking helper fields for overpayments ----
    # (Do not alter existing keys; just add extra ones you can use in the UI.)
    credit_this_term = max(0.0, -balance_this_term) # >0 when overpaid this term
    overall_raw = carry_forward + balance_this_term # may be negative
    overall_credit = max(0.0, -overall_raw) # >0 when overall account is in credit
    balance_this_term_safe = max(0.0, balance_this_term) # never negative, if you prefer

    conn.close()

    return {
        # ORIGINAL FIELDS (unchanged)
        "expected_fees": expected_fees,
        "expected_requirements": expected_requirements,
        "bursary_current": bursary_current,
        "paid_fees": paid_fees,
        "paid_requirements": paid_requirements,
        "carry_forward": carry_forward,
        "total_due_this_term": total_due_this_term,
        "balance_this_term": balance_this_term,
        "overall_balance": overall_balance,
        "opening_balance_raw": opening_balance,
        "prior_arrears_raw": prior_arrears,
        "transport_due_term": transport_fare,
        "transport_paid_term": transport_paid_term,
        "transport_balance_term": max(transport_fare - transport_paid_term, 0.0),

        # NEW, OPTIONAL (safe for templates; won’t affect routes)
        "credit_this_term": credit_this_term,
        "overall_credit": overall_credit,
        "balance_this_term_safe": balance_this_term_safe,
    }


def _expand_terms_from_row(row: dict):
    """
    Accepts:
      - term = 'Term 1'
      - terms = 'Term 1, Term 2' (comma/semicolon separated)
      - apply_to = 'year' -> Term 1..3
      - apply_to = 'two' with term_one & term_two
    Returns list[str] terms.
    """
    def norm(t):
        t = (t or "").strip()
        return t if t in TERMS else None

    # multiple terms column
    terms_mult = row.get("terms") or row.get("Terms") or row.get("TERMS") or ""
    if terms_mult:
        parts = [p.strip() for p in terms_mult.replace(";", ",").split(",")]
        expanded = [p for p in (norm(p) for p in parts) if p]
        if expanded:
            return expanded

    # single term
    single = norm(row.get("term") or row.get("Term") or row.get("TERM"))
    if single:
        return [single]

    # apply_to
    apply_to = (row.get("apply_to") or row.get("Apply_To") or "").lower().strip()
    if apply_to == "year":
        return TERMS[:]
    if apply_to == "two":
        t1 = norm(row.get("term_one") or row.get("Term_One"))
        t2 = norm(row.get("term_two") or row.get("Term_Two"))
        return [t for t in (t1, t2) if t]
    return []


@app.context_processor
def inject_current_role():
    from flask import session
    return {"current_role": (session.get("role") or "").lower()}

    
@app.route('/api')
def api_home():
    return "School Management System API"

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))
    



@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # --- get active academic session (safe fallback) ---
    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1")
    try:
        year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        year = datetime.now().year

    role = session.get("role", "teacher")
    is_admin = role in ("admin", "director", "headteacher", "bursar")

    stats = {}
    if is_admin:
        conn = get_db_connection()
        c = conn.cursor()

        total_students = c.execute(
            "SELECT COALESCE(COUNT(*),0) FROM students WHERE archived = 0"
        ).fetchone()[0]

        fees_in_term = c.execute(
            "SELECT COALESCE(SUM(amount_paid),0) FROM fees WHERE term=? AND year=?",
            (term, year),
        ).fetchone()[0]

        other_in_term = c.execute(
            "SELECT COALESCE(SUM(amount),0) FROM other_income WHERE term=? AND year=?",
            (term, year),
        ).fetchone()[0]

        exp_in_term = c.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE term=? AND year=?",
            (term, year),
        ).fetchone()[0]

        term_net_income = (fees_in_term or 0) + (other_in_term or 0) - (exp_in_term or 0)

        cumulative_fees = c.execute(
            "SELECT COALESCE(SUM(amount_paid),0) FROM fees"
        ).fetchone()[0]

        cumulative_other = c.execute(
            "SELECT COALESCE(SUM(amount),0) FROM other_income"
        ).fetchone()[0]

        cumulative_income = (cumulative_fees or 0) + (cumulative_other or 0)

        total_expenses = c.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenses"
        ).fetchone()[0]

        conn.close()

        stats = {
            "total_students": total_students or 0,
            "term_net_income": term_net_income or 0,
            "cumulative_net_income": cumulative_income or 0,
            "total_expenses": total_expenses or 0,
        }

    # Pass flags + session info for template control
    return render_template(
        "dashboard.html",
        username=session.get("full_name") or session.get("username") or "User",
        user_id=session.get("user_id"),
        role=role,
        show_admin=is_admin,
        active_term=term,
        active_year=year,
        stats=stats,
    )



@app.route('/add_subject', methods=['GET', 'POST'])
@require_role("admin", "headteacher")
def add_subject():
    conn = get_db_connection()
    c = conn.cursor()

    if request.method == 'POST':
        subject_name = request.form['name']
        part_names = request.form.getlist('parts')

        # Insert subject
        c.execute("INSERT INTO subjects (name) VALUES (?)", (subject_name,))
        subject_id = c.lastrowid

        # Insert parts (if any)
        for part in part_names:
            if part.strip():
                c.execute("INSERT INTO subject_papers (subject_id, paper_name) VALUES (?, ?)", (subject_id, part.strip()))

        conn.commit()
        conn.close()
        flash("Subject and parts added successfully", "success")
        return redirect(url_for('add_subject'))

    conn.close()
    return render_template("add_subject.html")


# --------- SUBJECTS: list/search/add/edit/delete, import/export ----------
@app.route("/subjects", methods=["GET", "POST"])
@require_role("admin", "headteacher","dos") # adjust roles as you prefer
def manage_subjects():
    conn = get_db_connection()
    c = conn.cursor()

    # Add new subject
    if request.method == "POST" and request.form.get("action") == "create":
        name = (request.form.get("name") or "").strip()
        code = (request.form.get("code") or "").strip().upper() or None
        if not name:
            flash("Subject name is required.", "warning")
        else:
            try:
                c.execute("INSERT INTO subjects (name, code) VALUES (?, ?)", (name, code))
                conn.commit()
                flash("Subject added.", "success")
            except Exception as e:
                flash(f"Could not add subject: {e}", "danger")
        conn.close()
        return redirect(url_for("manage_subjects"))

    # Search/filter
    q = (request.args.get("q") or "").strip()
    params = []
    sql = "SELECT id, name, code FROM subjects"
    if q:
        sql += " WHERE name LIKE ? OR code LIKE ?"
        like = f"%{q}%"
        params.extend([like, like])
    sql += " ORDER BY name"

    rows = c.execute(sql, params).fetchall()
    conn.close()
    return render_template("manage_subjects.html", rows=rows, q=q)

@app.route("/subjects/<int:sid>/edit", methods=["GET", "POST"])
@require_role("admin", "headteacher")
def edit_subject(sid):
    conn = get_db_connection()
    c = conn.cursor()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        code = (request.form.get("code") or "").strip().upper() or None
        if not name:
            flash("Subject name is required.", "warning")
        else:
            try:
                c.execute("UPDATE subjects SET name=?, code=? WHERE id=?", (name, code, sid))
                conn.commit()
                flash("Subject updated.", "success")
                conn.close()
                return redirect(url_for("manage_subjects"))
            except Exception as e:
                flash(f"Could not update subject: {e}", "danger")

    row = c.execute("SELECT id, name, code FROM subjects WHERE id=?", (sid,)).fetchone()
    conn.close()
    if not row:
        flash("Subject not found.", "warning")
        return redirect(url_for("manage_subjects"))
    return render_template("edit_subject.html", row=row)

@app.route("/subjects/<int:sid>/delete", methods=["POST"])
@require_role("admin", "headteacher")
def delete_subject(sid):
    conn = get_db_connection()
    try:
        # if FK constraints exist (e.g., record_score.subject_id), this will error when referenced—catch it
        conn.execute("DELETE FROM subjects WHERE id=?", (sid,))
        conn.commit()
        flash("Subject deleted.", "success")
    except Exception as e:
        flash(f"Cannot delete: subject is referenced by marks/records. ({e})", "danger")
    finally:
        conn.close()
    return redirect(url_for("manage_subjects"))

# --------- Import/Export ----------
@app.route("/subjects/export")
@require_role("admin", "headteacher")
def subjects_export():
    conn = get_db_connection()
    df = pd.read_sql_query("SELECT name, code FROM subjects ORDER BY name", conn)
    conn.close()

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Subjects")
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="subjects.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/subjects/import", methods=["POST"])
@require_role("admin", "headteacher")
def subjects_import():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Choose a file to import.", "warning")
        return redirect(url_for("manage_subjects"))

    ext = "." + file.filename.rsplit(".", 1)[-1].lower()
    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(file)
        elif ext == ".csv":
            df = pd.read_csv(file)
        else:
            flash("Unsupported file type. Use .xlsx / .xls / .csv", "danger")
            return redirect(url_for("manage_subjects"))

        cols = {c.lower(): c for c in df.columns}
        need = ["name", "code"]
        missing = [k for k in need if k not in cols]
        if missing:
            flash(f"Missing columns: {', '.join(missing)}", "danger")
            return redirect(url_for("manage_subjects"))

        conn = get_db_connection()
        c = conn.cursor()
        added = updated = 0
        for _, r in df.iterrows():
            name = str(r[cols["name"]]).strip() if pd.notna(r[cols["name"]]) else None
            code = str(r[cols["code"]]).strip().upper() if pd.notna(r[cols["code"]]) else None
            if not name:
                continue
            # Upsert by name; if you prefer by code, switch WHERE clause
            existing = c.execute("SELECT id FROM subjects WHERE name=?", (name,)).fetchone()
            if existing:
                c.execute("UPDATE subjects SET code=? WHERE id=?", (code or None, existing["id"]))
                updated += 1
            else:
                c.execute("INSERT INTO subjects (name, code) VALUES (?, ?)", (name, code or None))
                added += 1
        conn.commit()
        conn.close()
        flash(f"Import done. Added: {added}, Updated: {updated}.", "success")
    except Exception as e:
        flash(f"Import failed: {e}", "danger")

    return redirect(url_for("manage_subjects"))







from datetime import datetime
from werkzeug.security import generate_password_hash
import sqlite3


def _fetch_employees_for_dropdown(conn):
    # Use correct column names: middle_name (not Middle_name)
    return conn.execute("""
        SELECT id, first_name, COALESCE(middle_name,'') AS middle_name, last_name,
               designation, status
          FROM employees
         ORDER BY CASE WHEN status='active' THEN 0 ELSE 1 END,
                  last_name, first_name
    """).fetchall()






@app.route("/users", methods=["GET","POST"])
@require_role("admin")
def manage_users():
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    employees = _fetch_employees_for_dropdown(conn)

    # If linked from Employees: ?employee_id=123 preselect in the form
    pre_emp_id = request.args.get("employee_id") or ""

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        role = (request.form.get("role") or "").strip()
        status = (request.form.get("status") or "active").strip()
        employee_id = request.form.get("employee_id") or None
        if employee_id == "": employee_id = None

        if not username or not password:
            conn.close()
            flash("Username and password are required.", "danger")
            return redirect(url_for("manage_users"))

        if role not in ALLOWED_ROLES:
            conn.close()
            flash("Invalid role.", "danger")
            return redirect(url_for("manage_users"))

        if status not in ("active","archived"):
            conn.close()
            flash("Invalid status.", "danger")
            return redirect(url_for("manage_users"))

        exists = c.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone()
        if exists:
            conn.close()
            flash("Username already exists.", "warning")
            return redirect(url_for("manage_users"))

        try:
            c.execute("""
                INSERT INTO users (username, password_hash, role, status, employee_id)
                VALUES (?, ?, ?, ?, ?)
            """, (username, generate_password_hash(password), role, status, employee_id))
            conn.commit()
            flash("User created.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Could not create user: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("manage_users"))

    # GET list + search
    q = (request.args.get("q") or "").strip()
    if q:
        users = c.execute("""
            SELECT u.*, e.first_name, e.middle_name, e.last_name, e.designation
              FROM users u
              LEFT JOIN employees e ON e.id = u.employee_id
             WHERE u.username LIKE ? OR u.role LIKE ?
             ORDER BY u.id DESC
        """, (f"%{q}%", f"%{q}%")).fetchall()
    else:
        users = c.execute("""
            SELECT u.*, e.first_name, e.middle_name, e.last_name, e.designation
              FROM users u
              LEFT JOIN employees e ON e.id = u.employee_id
             ORDER BY u.id DESC
        """).fetchall()

    conn.close()
    return render_template(
        "users.html",
        users=users,
        employees=employees,
        q=q,
        pre_emp_id=pre_emp_id,
        ALLOWED_ROLES=ALLOWED_ROLES
    )

@app.route("/users/<int:user_id>/edit", methods=["GET","POST"])
@require_role("admin")
def edit_user(user_id):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    user = c.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        flash("User not found.", "warning")
        return redirect(url_for("manage_users"))

    employees = _fetch_employees_for_dropdown(conn)

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        role = (request.form.get("role") or "").strip()
        status = (request.form.get("status") or "active").strip()
        employee_id = request.form.get("employee_id")
        new_pass = (request.form.get("new_password") or "").strip()

        if employee_id == "" or employee_id is None:
            employee_id = None
        else:
            try:
                employee_id = int(employee_id)
            except ValueError:
                employee_id = None

        if not username:
            conn.close()
            flash("Username is required.", "danger")
            return redirect(url_for("edit_user", user_id=user_id))

        if role not in ALLOWED_ROLES:
            conn.close()
            flash("Invalid role.", "danger")
            return redirect(url_for("edit_user", user_id=user_id))

        if status not in ("active","archived"):
            conn.close()
            flash("Invalid status.", "danger")
            return redirect(url_for("edit_user", user_id=user_id))

        exists = c.execute("SELECT 1 FROM users WHERE username=? AND id<>?", (username, user_id)).fetchone()
        if exists:
            conn.close()
            flash("Username already taken by another account.", "warning")
            return redirect(url_for("edit_user", user_id=user_id))

        try:
            if new_pass:
                c.execute("""
                    UPDATE users
                       SET username=?, role=?, status=?, employee_id=?, password_hash=?
                     WHERE id=?
                """, (username, role, status, employee_id,
                      generate_password_hash(new_pass), user_id))
            else:
                c.execute("""
                    UPDATE users
                       SET username=?, role=?, status=?, employee_id=?
                     WHERE id=?
                """, (username, role, status, employee_id, user_id))
            conn.commit()
            flash("User updated.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Could not update user: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("manage_users"))

    conn.close()
    return render_template("user_edit.html", user=user, employees=employees, ALLOWED_ROLES=ALLOWED_ROLES)



@app.route("/users/<int:user_id>/delete", methods=["POST"])
@require_role("admin")
def delete_user(user_id):
    if session.get("user_id") == user_id:
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("manage_users"))
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
        flash("User deleted.", "info")
    except Exception as e:
        conn.rollback()
        flash(f"Could not delete user: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("manage_users"))

@app.route("/users/<int:user_id>/toggle", methods=["POST"])
@require_role("admin")
def toggle_user(user_id):
    if session.get("user_id") == user_id:
        flash("You cannot archive/activate your own account.", "warning")
        return redirect(url_for("manage_users"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    row = c.execute("SELECT id, username, status FROM users WHERE id=?", (user_id,)).fetchone()
    if not row:
        conn.close()
        flash("User not found.", "warning")
        return redirect(url_for("manage_users"))

    new_status = "archived" if row["status"] == "active" else "active"
    try:
        c.execute("UPDATE users SET status=? WHERE id=?", (new_status, user_id))
        conn.commit()
        flash(f"User {row['username']} is now {new_status}.", "info")
    except Exception as e:
        conn.rollback()
        flash(f"Could not change status: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("manage_users"))





@app.route('/subjects', methods=['GET', 'POST'])
@require_role('admin')
def subjects():
    conn = get_db_connection()
    c = conn.cursor()
    message = ""

    # Handle adding subject
    if request.method == 'POST':
        if 'add_subject' in request.form:
            name = request.form['name'].strip()
            initial = request.form['initial'].strip().upper()
            if name and initial:
                try:
                    c.execute("INSERT INTO subjects (name, initial) VALUES (?, ?)", (name, initial))
                    conn.commit()
                    message = "Subject added successfully!"
                except sqlite3.IntegrityError:
                    message = "Subject or initial already exists!"
            else:
                message = "Please fill in all fields."

        # Handle adding paper
        elif 'add_paper' in request.form:
            subject_id = request.form['subject_id']
            paper_name = request.form['paper_name'].strip()
            paper_initial = request.form['paper_initial'].strip().upper()
            if paper_name and paper_initial:
                try:
                    c.execute('''
                        INSERT INTO subject_papers (subject_id, paper_name, paper_initial)
                        VALUES (?, ?, ?)
                    ''', (subject_id, paper_name, paper_initial))
                    conn.commit()
                    message = "Paper added successfully!"
                except sqlite3.IntegrityError:
                    message = "Paper already exists for this subject!"
            else:
                message = "Please fill in all paper fields."

    # Fetch subjects with papers
    c.execute('SELECT * FROM subjects ORDER BY name ASC')
    subjects = c.fetchall()

    # Fetch all papers grouped by subject
    papers_dict = {}
    for subj in subjects:
        c.execute('SELECT * FROM subject_papers WHERE subject_id = ?', (subj['id'],))
        papers_dict[subj['id']] = c.fetchall()

    conn.close()
    return render_template('subjects.html', subjects=subjects, papers_dict=papers_dict, message=message)
    





@app.route('/record_score', methods=['GET', 'POST'])
@require_role('admin', 'teacher')
def record_score():
    conn = get_db_connection()
    c = conn.cursor()

    students = c.execute("SELECT id, first_name || ' ' || last_name AS full_name, class FROM students WHERE status='active'").fetchall()
    subjects = c.execute("SELECT id, name FROM subjects").fetchall()
    conn.close()

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        subject_id = request.form.get('subject_id')
        subject_part_id = request.form.get('subject_part_id')
        term = request.form.get('term')
        year = request.form.get('year')
        score = request.form.get('score')

        if not all([student_id, subject_id, subject_part_id, term, year, score]):
            flash('Please fill all fields.')
            return redirect(url_for('record_score'))

        try:
            year = int(year)
            score = float(score)
        except ValueError:
            flash('Year must be integer and score must be a number.')
            return redirect(url_for('record_score'))

        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute('''
                INSERT INTO record_score (student_id, subject_part_id, term, year, score)
                VALUES (?, ?, ?, ?, ?)
            ''', (student_id, subject_part_id, term, year, score))
            conn.commit()
            flash('Score recorded successfully.')
        except Exception as e:
            flash(f'Error: {str(e)}')
        finally:
            conn.close()

        return redirect(url_for('record_score'))

    return render_template('record_score.html', students=students, subjects=subjects)
    






@app.route('/record_batch_score', methods=['GET', 'POST'])
@require_role('admin', 'headteacher')
def record_batch_score():
    conn = get_db_connection()
    c = conn.cursor()

    # Get active term and year
    academic = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    if not academic:
        flash("No active academic year found.", "warning")
        return redirect(url_for('dashboard'))

    year, term = academic['year'], academic['current_term']

    # Load filters from GET or POST
    class_name = request.args.get('class_name') or request.form.get('class_name')
    stream = request.args.get('stream') or request.form.get('stream')
    subject_id = request.args.get('subject_id') or request.form.get('subject_id')
    part_id = request.args.get('subject_part_id') or request.form.get('subject_part_id')

    subjects = c.execute("SELECT * FROM subjects").fetchall()
    parts = []
    if subject_id:
        parts = c.execute("SELECT * FROM subject_papers WHERE subject_id = ?", (subject_id,)).fetchall()

    students = []
    scores = {}

    if all([class_name, stream, subject_id]):
        students = c.execute("""
            SELECT * FROM students
            WHERE class_name = ? AND stream = ? AND status = 'active' AND archived = 0
        """, (class_name, stream)).fetchall()

        rows = c.execute("""
            SELECT * FROM record_score
            WHERE year = ? AND term = ? AND subject_id = ?
                  AND (subject_part_id = ? OR (? IS NULL AND subject_part_id IS NULL))
        """, (year, term, subject_id, part_id, part_id)).fetchall()

        for row in rows:
            scores[str(row['student_id'])] = row

    if request.method == 'POST' and students:
        for student in students:
            sid = str(student['id'])
            bot = request.form.get(f'bot_{sid}')
            mid = request.form.get(f'mid_{sid}')
            eot = request.form.get(f'eot_{sid}')

            marks = [int(m) for m in [bot, mid, eot] if m and m.isdigit()]
            avg = sum(marks) / len(marks) if marks else None

            if avg is not None:
                exists = c.execute("""
                    SELECT id FROM record_score
                    WHERE student_id = ? AND subject_id = ? AND year = ? AND term = ?
                          AND (subject_part_id = ? OR (? IS NULL AND subject_part_id IS NULL))
                """, (student['id'], subject_id, year, term, part_id, part_id)).fetchone()

                if exists:
                    c.execute("""
                        UPDATE record_score
                        SET bot_mark = ?, midterm_mark = ?, eot_mark = ?, average_mark = ?, processed_on = CURRENT_TIMESTAMP
                        WHERE id = ?
                    """, (bot or None, mid or None, eot or None, avg, exists['id']))
                else:
                    c.execute("""
                        INSERT INTO record_score (
                            student_id, subject_id, subject_part_id,
                            year, term, bot_mark, midterm_mark, eot_mark, average_mark
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        student['id'], subject_id, part_id or None, year, term,
                        bot or None, mid or None, eot or None, avg
                    ))

        conn.commit()
        flash("Scores recorded successfully.", "success")
        return redirect(url_for('record_batch_score'))

    conn.close()
    return render_template(
        'record_batch_score.html',
        students=students,
        class_name=class_name,
        stream=stream,
        term=term,
        year=year,
        subject_id=subject_id,
        subject_part_id=part_id,
        subjects=subjects,
        parts=parts,
        scores=scores
    )






@app.route('/view_scores', methods=['GET'])
@require_role('admin', 'headteacher', 'bursar')
def view_scores():
    conn = get_db_connection()
    c = conn.cursor()

    class_name = request.args.get('class_name', '')
    stream = request.args.get('stream', '')
    year = request.args.get('year', '')
    term = request.args.get('term', '')
    subject_id = request.args.get('subject_id', '')

    query = """
        SELECT rs.*, s.student_number, s.first_name, s.Middle_name, s.last_name,
               s.class_name, s.stream, subj.name AS subject, sp.paper_name
        FROM record_score rs
        JOIN students s ON rs.student_id = s.id
        JOIN subjects subj ON rs.subject_id = subj.id
        LEFT JOIN subject_papers sp ON rs.subject_part_id = sp.id
        WHERE s.status = 'active' AND s.archived = 0
    """
    params = []

    if class_name:
        query += " AND s.class_name = ?"
        params.append(class_name)
    if stream:
        query += " AND s.stream = ?"
        params.append(stream)
    if year:
        query += " AND rs.year = ?"
        params.append(year)
    if term:
        query += " AND rs.term = ?"
        params.append(term)
    if subject_id:
        query += " AND rs.subject_id = ?"
        params.append(subject_id)

    query += " ORDER BY s.last_name, sp.paper_name"
    rows = c.execute(query, params).fetchall()

    # Group scores by student and then by paper
    results_by_student = {}
    for row in rows:
        sid = row['student_id']
        if sid not in results_by_student:
            results_by_student[sid] = {
                'student_number': row['student_number'],
                'full_name': f"{row['first_name']} {row['Middle_name'] or ''} {row['last_name']}",
                'class_name': row['class_name'],
                'stream': row['stream'],
                'subject': row['subject'],
                'parts': [],
                'total_avg': 0,
                'total_parts': 0
            }

        avg = row['average_mark'] or 0
        results_by_student[sid]['parts'].append({
            'record_id': row['id'],
            'paper_name': row['paper_name'] or '-',
            'bot_mark': row['bot_mark'],
            'midterm_mark': row['midterm_mark'],
            'eot_mark': row['eot_mark'],
            'average_mark': avg,
            'grade': row['grade'],
            'comment': row['comment']
        })

        if avg > 0:
            results_by_student[sid]['total_avg'] += avg
            results_by_student[sid]['total_parts'] += 1

    for s in results_by_student.values():
        if s['total_parts']:
            s['total_avg'] = round(s['total_avg'] / s['total_parts'], 1)
        else:
            s['total_avg'] = 0

    subjects = c.execute("SELECT id, name FROM subjects").fetchall()
    available_years = [row['year'] for row in c.execute(
        "SELECT DISTINCT year FROM academic_years ORDER BY year DESC"
    ).fetchall()]

    conn.close()

    return render_template(
        "view_scores_grouped.html",
        results=results_by_student,
        class_name=class_name,
        stream=stream,
        year=year,
        term=term,
        subject_id=subject_id,
        subjects=subjects,
        available_years=available_years
    )




import csv


@app.route('/view_scores_grouped', methods=['GET', 'POST'])
@require_role('admin', 'headteacher')
def view_scores_grouped():
    conn = get_db_connection()
    c = conn.cursor()

    # Get active academic year
    academic = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    year = academic['year'] if academic else ''
    term = academic['current_term'] if academic else ''

    class_name = request.args.get('class_name', '')
    stream = request.args.get('stream', '')
    subject_id = request.args.get('subject_id', '')

    subjects = c.execute("SELECT id, name FROM subjects").fetchall()
    grouped_data = []

    if class_name and stream:
        query = '''
            SELECT rs.*, s.student_number,
                   s.first_name || ' ' || COALESCE(s.Middle_name, '') || ' ' || s.last_name AS full_name,
                   s.class_name, s.stream, subj.name AS subject, sp.paper_name
            FROM record_score rs
            JOIN students s ON s.id = rs.student_id
            JOIN subjects subj ON subj.id = rs.subject_id
            LEFT JOIN subject_papers sp ON rs.subject_part_id = sp.id
            WHERE s.class_name = ? AND s.stream = ? AND s.status = 'active' AND s.archived = 0
              AND rs.year = ? AND rs.term = ?
        '''
        params = [class_name, stream, year, term]

        if subject_id:
            query += " AND rs.subject_id = ?"
            params.append(subject_id)

        rows = c.execute(query, params).fetchall()
        grouped = {}

        for r in rows:
            key = (r['student_number'], r['subject'])

            if key not in grouped:
                grouped[key] = {
                    'student_number': r['student_number'],
                    'full_name': r['full_name'],
                    'class_name': r['class_name'],
                    'stream': r['stream'],
                    'subject': r['subject'],
                    'parts': [],
                    'overall_avg': 0,
                    'grade': '',
                    'comment': ''
                }

            marks = [r['bot_mark'], r['midterm_mark'], r['eot_mark']]
            marks = [m for m in marks if m is not None]
            avg = sum(marks) / len(marks) if marks else None

            grouped[key]['parts'].append({
                'record_id': r['id'],
                'paper': r['paper_name'] or 'N/A',
                'bot': r['bot_mark'],
                'mid': r['midterm_mark'],
                'eot': r['eot_mark'],
                'average': round(avg, 2) if avg else None
            })

        for g in grouped.values():
            part_avgs = [p['average'] for p in g['parts'] if p['average'] is not None]
            if part_avgs:
                overall = sum(part_avgs) / len(part_avgs)
                g['overall_avg'] = round(overall, 2)
                row = c.execute(
                    "SELECT grade, comment FROM grading_scale WHERE ? BETWEEN lower_limit AND upper_limit LIMIT 1",
                    (overall,)
                ).fetchone()
                g['grade'] = row['grade'] if row else ''
                g['comment'] = row['comment'] if row else ''

        grouped_data = list(grouped.values())

    conn.close()

    return render_template(
        'view_scores_grouped.html',
        class_name=class_name,
        stream=stream,
        term=term,
        year=year,
        subject_id=subject_id,
        subjects=subjects,
        grouped_scores=grouped_data
    )


@app.route('/save_scores_grouped', methods=['POST'])
@require_role('admin', 'headteacher')
def save_scores_grouped():
    conn = get_db_connection()
    c = conn.cursor()

    for key in request.form:
        if key.startswith(('bot_', 'mid_', 'eot_')):
            prefix, record_id = key.split('_')
            mark = request.form.get(key)
            if mark:
                try:
                    c.execute(
                        f"UPDATE record_score SET {prefix}_mark = ? WHERE id = ?",
                        (int(mark), int(record_id))
                    )
                except Exception as e:
                    flash(f"Error updating record {record_id}: {e}", "danger")

    conn.commit()
    conn.close()
    flash("Scores updated successfully.", "success")
    return redirect(url_for('view_scores_grouped', **request.args))


@app.route('/export_grouped_scores')
@require_role('admin', 'headteacher')
def export_grouped_scores():
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Grouped Scores")

    conn = get_db_connection()
    c = conn.cursor()

    class_name = request.args.get('class_name')
    stream = request.args.get('stream')
    subject_id = request.args.get('subject_id')

    academic = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    year = academic['year'] if academic else ''
    term = academic['current_term'] if academic else ''

    query = '''
        SELECT rs.*, s.student_number,
               s.first_name || ' ' || COALESCE(s.Middle_name, '') || ' ' || s.last_name AS full_name,
               s.class_name, s.stream, subj.name AS subject, sp.paper_name
        FROM record_score rs
        JOIN students s ON s.id = rs.student_id
        JOIN subjects subj ON subj.id = rs.subject_id
        LEFT JOIN subject_papers sp ON rs.subject_part_id = sp.id
        WHERE s.class_name = ? AND s.stream = ? AND rs.year = ? AND rs.term = ?
    '''
    params = [class_name, stream, year, term]

    if subject_id:
        query += " AND rs.subject_id = ?"
        params.append(subject_id)

    rows = c.execute(query, params).fetchall()

    worksheet.write_row(0, 0, [
        'Student No', 'Name', 'Class', 'Stream',
        'Subject', 'Paper', 'BOT', 'MID', 'EOT', 'Average'
    ])

    for row_num, r in enumerate(rows, start=1):
        marks = [r['bot_mark'], r['midterm_mark'], r['eot_mark']]
        valid_marks = [m for m in marks if m is not None]
        avg = sum(valid_marks) / len(valid_marks) if valid_marks else 0

        worksheet.write_row(row_num, 0, [
            r['student_number'], r['full_name'], r['class_name'], r['stream'],
            r['subject'], r['paper_name'] or 'N/A',
            r['bot_mark'], r['midterm_mark'], r['eot_mark'],
            round(avg, 2)
        ])

    workbook.close()
    output.seek(0)
    conn.close()

    return send_file(
        output,
        download_name="grouped_scores.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )




@app.route('/update_grouped_scores', methods=['POST'])
@require_role('admin', 'headteacher')
def update_grouped_scores():
    conn = get_db_connection()
    c = conn.cursor()

    for key, value in request.form.items():
        if key.startswith('bot_') or key.startswith('midterm_') or key.startswith('eot_'):
            field, record_id = key.split('_')
            c.execute(f"UPDATE record_score SET {field}_mark = ? WHERE id = ?", (value or None, record_id))

    conn.commit()
    conn.close()
    flash("Scores updated successfully.", "success")
    return redirect(url_for('view_scores_grouped', **request.args.to_dict()))




@app.route('/results')
def results():
    results = Result.query.options(
        joinedload(Result.student),
        joinedload(Result.subject)
    ).all()

    grouped_results = defaultdict(lambda: {"student_info": None, "subject": None, "papers": [], "total_avg": 0})

    for result in results:
        key = (result.student_id, result.subject_id)
        avg = sum(filter(None, [result.bot_mark, result.midterm_mark, result.eot_mark])) / 3 if all([result.bot_mark, result.midterm_mark, result.eot_mark]) else None

        grading = get_grade_comment(avg) if avg is not None else None
        grade = grading.grade if grading else ''
        comment = grading.comment if grading else ''

        if not grouped_results[key]["student_info"]:
            grouped_results[key]["student_info"] = result.student
            grouped_results[key]["subject"] = result.subject.name

        grouped_results[key]["papers"].append({
            "id": result.id,
            "subject_part": result.subject_part,
            "bot_mark": result.bot_mark,
            "midterm_mark": result.midterm_mark,
            "eot_mark": result.eot_mark,
            "average_mark": round(avg, 2) if avg is not None else '',
            "grade": grade,
            "comment": comment
        })

    # Add total average per group
    for group in grouped_results.values():
        valid_avgs = [float(p['average_mark']) for p in group['papers'] if p['average_mark'] != '']
        group['total_avg'] = round(sum(valid_avgs) / len(valid_avgs), 2) if valid_avgs else ''

    return render_template('your_template.html', grouped_results=grouped_results)



@app.route('/import_scores', methods=['POST'])
@require_role('admin', 'headteacher')
def import_scores():
    class_name = request.form.get('class_name')
    stream = request.form.get('stream')
    subject_id = request.form.get('subject_id')
    subject_part_id = request.form.get('subject_part_id') or None

    # Get active academic year and term
    conn = get_db_connection()
    c = conn.cursor()
    academic = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    year, term = academic['year'], academic['current_term']

    file = request.files.get('score_file')
    if not file:
        flash("No file uploaded", "danger")
        return redirect(url_for('record_batch_score'))

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == '.csv':
            df = pd.read_csv(file)
        elif ext in ['.xls', '.xlsx']:
            df = pd.read_excel(file, engine='openpyxl' if ext == '.xlsx' else 'xlrd')
        else:
            flash("Unsupported file format. Use .xlsx, .xls or .csv", "danger")
            return redirect(url_for('record_batch_score'))
    except Exception as e:
        flash(f"Failed to read file: {e}", "danger")
        return redirect(url_for('record_batch_score'))

    required_cols = ['Student Number', 'BOT', 'MIDTERM', 'EOT']
    if not all(col in df.columns for col in required_cols):
        flash("Missing required columns: Student Number, BOT, MIDTERM, EOT", "danger")
        return redirect(url_for('record_batch_score'))

    imported = 0
    for _, row in df.iterrows():
        student_number = row['Student Number']
        bot = row.get('BOT')
        midterm = row.get('MIDTERM')
        eot = row.get('EOT')

        student = c.execute("""
            SELECT id FROM students 
            WHERE student_number = ? AND class_name = ? AND stream = ? AND archived = 0
        """, (student_number, class_name, stream)).fetchone()
        if not student:
            continue

        student_id = student['id']
        marks = [m for m in [bot, midterm, eot] if pd.notnull(m)]
        avg = sum(marks) / len(marks) if marks else None

        exists = c.execute("""
            SELECT id FROM record_score
            WHERE student_id = ? AND subject_id = ? AND term = ? AND year = ? 
            AND (subject_part_id = ? OR (? IS NULL AND subject_part_id IS NULL))
        """, (student_id, subject_id, term, year, subject_part_id, subject_part_id)).fetchone()

        if exists:
            c.execute("""
                UPDATE record_score SET bot_mark = ?, midterm_mark = ?, eot_mark = ?, average_mark = ?
                WHERE id = ?
            """, (bot, midterm, eot, avg, exists['id']))
        else:
            c.execute("""
                INSERT INTO record_score (
                    student_id, subject_id, subject_part_id, year, term,
                    bot_mark, midterm_mark, eot_mark, average_mark
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (student_id, subject_id, subject_part_id, year, term, bot, midterm, eot, avg))
        imported += 1

    conn.commit()
    conn.close()

    flash(f"{imported} scores imported successfully.", "success")
    return redirect(url_for('record_batch_score',
                            class_name=class_name,
                            stream=stream,
                            subject_id=subject_id,
                            subject_part_id=subject_part_id))





@app.route('/download_score_template')
@require_role('admin', 'headteacher')
def download_score_template():
    class_name = request.args.get('class_name')
    stream = request.args.get('stream')

    conn = get_db_connection()
    c = conn.cursor()

    students = c.execute("""
        SELECT student_number, first_name, Middle_name, last_name
        FROM students
        WHERE class_name = ? AND stream = ? AND status = 'active' AND archived = 0
        ORDER BY student_number
    """, (class_name, stream)).fetchall()
    conn.close()

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    ws.append(["Student Number", "Student Name", "BOT", "MIDTERM", "EOT"])

    for s in students:
        names = [s['first_name'], s['Middle_name'], s['last_name']]
        full_name = " ".join(filter(None, names)).strip()
        ws.append([s['student_number'], full_name, "", "", ""])

    # Save to in-memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"scores_template_{class_name}_{stream}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )





@app.route('/get_subject_parts/<int:subject_id>')
def get_subject_parts(subject_id):
    conn = get_db_connection()
    c = conn.cursor()
    parts = c.execute("SELECT id, paper_name AS name FROM subject_papers WHERE subject_id = ?", (subject_id,)).fetchall()
    conn.close()
    return jsonify(parts=[dict(p) for p in parts])








@app.route('/end_of_year_process')
def end_of_year_process():
    conn = get_db_connection()
    c = conn.cursor()

    current_year = 2025 # You can make this dynamic

    # Get active students
    students = c.execute("SELECT id, name, class_name FROM students WHERE status='active'").fetchall()

    for student in students:
        student_id = student['id']
        class_name = student['class_name']

        # Get distinct subject_ids the student has results for
        subjects = c.execute('''
            SELECT sp.subject_id
            FROM student_results sr
            JOIN subject_parts sp ON sr.subject_part_id = sp.id
            WHERE sr.student_id = ? AND sr.year = ?
            GROUP BY sp.subject_id
        ''', (student_id, current_year)).fetchall()

        total_avg = 0
        subject_count = 0

        for subject in subjects:
            subject_id = subject['subject_id']

            # Average all part scores for this subject
            avg_score = c.execute('''
                SELECT AVG(score) as avg_score
                FROM student_results sr
                JOIN subject_parts sp ON sr.subject_part_id = sp.id
                WHERE sr.student_id = ? AND sp.subject_id = ? AND sr.year = ?
            ''', (student_id, subject_id, current_year)).fetchone()['avg_score']

            if avg_score is not None:
                total_avg += avg_score
                subject_count += 1

                # Archive result
                c.execute('''
                    INSERT INTO archived_results (student_id, subject_id, average_score, class_name, year)
                    VALUES (?, ?, ?, ?, ?)
                ''', (student_id, subject_id, avg_score, class_name, current_year))

        # Determine promotion
        if subject_count == 0:
            continue # No subjects found

        final_avg = total_avg / subject_count
        class_name_number = int(''.join(filter(str.isdigit, class_name)))
        class_name_prefix = ''.join(filter(str.isalpha, class_name)).upper()

        # Determine next class
        if final_avg >= 50:
            new_class_name = f"{class_name_prefix}{class_number_number + 1}"
        else:
            new_class_name = class_name # Repeat

        # If already in P7, mark graduated
        if class_name.upper() == 'P7':
            new_status = 'graduated'
        else:
            new_status = 'active'

        # Update student
        c.execute('''
            UPDATE students SET class = ?, status = ?
            WHERE id = ?
        ''', (new_class, new_status, student_id))

    conn.commit()
    conn.close()

    return 'End-of-year processing complete: results archived, students promoted or repeated.'
   
    



# ===== MARKS HUB (helpers) ================================================

import sqlite3
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import (
    request, render_template, redirect, url_for, flash, send_file, session
)

ALLOWED_EXTS = {"csv", "xlsx"}

def _safe_int(v):
    try:
        if v is None or v == "":
            return None
        # excel numeric loads may be floats; allow float->int too
        return int(float(v))
    except (TypeError, ValueError):
        return None

def _safe_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None

def _avg(*parts):
    nums = [p for p in parts if isinstance(p, (int, float))]
    return float(sum(nums) / len(nums)) if nums else None

def get_user_initials(conn, user_id):
    """
    Returns initials for the logged-in user via users -> employees -> teachers.
    Falls back to '' if not found.
    """
    if not user_id:
        return ""
    row = conn.execute("""
        SELECT COALESCE(t.initials,'') AS initials
          FROM users u
          LEFT JOIN employees e ON e.id = u.employee_id
          LEFT JOIN teachers t ON t.employee_id = e.id AND t.status='active'
         WHERE u.id = ?
         LIMIT 1
    """, (user_id,)).fetchone()
    if not row:
        return ""
    return row["initials"] if isinstance(row, sqlite3.Row) else (row[0] or "")



def resolve_student_id(conn, student_number):
    if not student_number:
        return None
    r = conn.execute("SELECT id FROM students WHERE student_number=? AND archived=0",
                     (student_number,)).fetchone()
    return r["id"] if r and isinstance(r, sqlite3.Row) else (r[0] if r else None)

def resolve_subject_id(conn, subject_id=None, subject_code=None, subject_name=None):
    # prefer explicit id
    sid = _safe_int(subject_id)
    if sid:
        exists = conn.execute("SELECT 1 FROM subjects WHERE id=?", (sid,)).fetchone()
        return sid if exists else None
    # fallbacks
    if subject_code:
        r = conn.execute("SELECT id FROM subjects WHERE code=?", (subject_code,)).fetchone()
        if r: return r["id"] if isinstance(r, sqlite3.Row) else r[0]
    if subject_name:
        r = conn.execute("SELECT id FROM subjects WHERE name=?", (subject_name,)).fetchone()
        if r: return r["id"] if isinstance(r, sqlite3.Row) else r[0]
    return None

# ===== MARKS HUB (main page) ==============================================



@app.route("/marks/hub", methods=["GET", "POST"])
@require_role("admin", "headteacher", "teacher", "dos")
def marks_hub():
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    try:
        # ---------- reference lists ----------
        classes = [r[0] for r in conn.execute(
            "SELECT DISTINCT class_name FROM classes WHERE class_name IS NOT NULL ORDER BY class_name"
        ).fetchall()]
        streams = [r[0] for r in conn.execute(
            "SELECT DISTINCT stream FROM classes WHERE stream IS NOT NULL ORDER BY stream"
        ).fetchall()] or ["A"]
        subjects = conn.execute(
            "SELECT id, name, code FROM subjects ORDER BY name"
        ).fetchall()

        # ---------- active session (FIX: define before using) ----------
        ay = get_active_academic_year()
        active_year = int(ay.get("year"))
        active_term = ay.get("current_term") or ay.get("term") or "Term 1"

        # ensure table exists (with extra columns)
        ensure_record_score_table(conn)

        # helper: average of present marks (OTH, HP, BOT, MID, EOT)
        def _avg5(oth, hp, bot, mid, eot):
            vals = [v for v in (oth, hp, bot, mid, eot) if v is not None]
            return round(sum(vals) / len(vals)) if vals else None

        # ---------- POST: save a single row ----------
        if request.method == "POST" and request.form.get("save_row") == "1":
            student_no = (request.form.get("student_number") or "").strip()
            subject_id = _safe_int(request.form.get("subject_id"))
            term = (request.form.get("term") or active_term).strip()
            year = _safe_int(request.form.get("year")) or active_year

            # new fields first (order OTH, HP, BOT, MID, EOT)
            oth = _safe_int(request.form.get("other_mark"))
            hp = _safe_int(request.form.get("holiday_mark"))
            bot = _safe_int(request.form.get("bot_mark"))
            mid = _safe_int(request.form.get("midterm_mark"))
            eot = _safe_int(request.form.get("eot_mark"))
            avg = _avg5(oth, hp, bot, mid, eot)

            if not student_no or not subject_id:
                flash("Student number and subject are required.", "warning")
                return redirect(url_for("marks_hub"))

            student_id = resolve_student_id(conn, student_no)
            if not student_id:
                flash("Student not found or archived.", "warning")
                return redirect(url_for("marks_hub"))

            initials = get_user_initials(conn, session.get("user_id"))

            conn.execute("""
                INSERT INTO record_score (
                    student_id, subject_id, term, year,
                    other_mark, holiday_mark, bot_mark, midterm_mark, eot_mark,
                    average_mark, initials
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(student_id, subject_id, term, year) DO UPDATE SET
                    other_mark = COALESCE(excluded.other_mark, record_score.other_mark),
                    holiday_mark = COALESCE(excluded.holiday_mark, record_score.holiday_mark),
                    bot_mark = COALESCE(excluded.bot_mark, record_score.bot_mark),
                    midterm_mark = COALESCE(excluded.midterm_mark, record_score.midterm_mark),
                    eot_mark = COALESCE(excluded.eot_mark, record_score.eot_mark),
                    average_mark = COALESCE(excluded.average_mark, record_score.average_mark),
                    initials = CASE
                                     WHEN excluded.initials IS NOT NULL AND excluded.initials!=''
                                     THEN excluded.initials ELSE record_score.initials
                                   END,
                    processed_on = CURRENT_TIMESTAMP
            """, (student_id, subject_id, term, year, oth, hp, bot, mid, eot, avg, initials))
            conn.commit()
            flash("Saved.", "success")

            # keep current filters
            return redirect(url_for("marks_hub", **{k: v for k, v in request.args.items()}))

        # ---------- GET: filters ----------
        filter_class = (request.args.get("class") or "").strip()
        filter_stream = (request.args.get("stream") or "").strip()
        filter_subject = _safe_int(request.args.get("subject_id"))
        filter_term = (request.args.get("term") or active_term).strip()
        filter_year = _safe_int(request.args.get("year")) or active_year

        # ---------- student list ----------
        students = []
        if filter_class:
            q = ("SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS m, "
                 "last_name, class_name, stream FROM students "
                 "WHERE class_name=? AND archived=0")
            args = [filter_class]
            if filter_stream:
                q += " AND stream=?"
                args.append(filter_stream)
            q += " ORDER BY last_name, first_name"
            students = conn.execute(q, args).fetchall()

        # ---------- existing scores for this subject/term/year ----------
        scores_by_student = {}
        if students and filter_subject:
            sids = tuple([s["id"] for s in students])
            if len(sids) == 1:
                sid_clause = f"({sids[0]})"
            else:
                sid_clause = str(sids)
            sql = f"""
                SELECT student_id,
                       other_mark, holiday_mark, bot_mark, midterm_mark, eot_mark,
                       average_mark, initials
                  FROM record_score
                 WHERE subject_id=? AND term=? AND year=? AND student_id IN {sid_clause}
            """
            for r in conn.execute(sql, (filter_subject, filter_term, filter_year)):
                scores_by_student[r["student_id"]] = r

        # initials (for display hint)
        my_initials = get_user_initials(conn, session.get("user_id"))

        return render_template(
            "marks_hub.html",
            classes=classes, streams=streams, subjects=subjects,
            active_term=active_term, active_year=active_year,
            filter_class=filter_class, filter_stream=filter_stream,
            filter_subject=filter_subject, filter_term=filter_term,
            filter_year=filter_year,
            students=students,
            scores_by_student=scores_by_student,
            my_initials=my_initials
        )
    finally:
        conn.close()

# ===== Template download (pre-fills subject/term/year) =====================




@app.route("/marks/template")
@require_role("admin", "headteacher", "teacher", "dos")
def marks_template():
    # ---- filters (pick the students you want in the file) -------------------
    class_name = (request.args.get("class_name") or "").strip()
    stream = (request.args.get("stream") or "").strip()
    subject_id = _safe_int(request.args.get("subject_id"))

    # ---- academic year context (with optional override) ---------------------
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    ay = conn.execute("SELECT year, current_term FROM academic_years WHERE is_active=1").fetchone()
    if not ay:
        conn.close()
        flash("Please activate an academic year.", "warning")
        return redirect(url_for("dashboard"))
    default_year = int(ay["year"])
    default_term = ay["current_term"]

    term = (request.args.get("term") or default_term).strip()
    try:
        year = int(request.args.get("year") or default_year)
    except ValueError:
        year = default_year

    # ---- subject info (optional, just to prefill columns) -------------------
    subj_name = subj_code = ""
    if subject_id:
        srow = conn.execute("SELECT name, code FROM subjects WHERE id=?", (subject_id,)).fetchone()
        if srow:
            subj_name, subj_code = srow["name"], srow["code"] or ""

    # ---- pull students (filtered) -------------------------------------------
    qs = """
        SELECT id, student_number,
               TRIM(first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name) AS full_name,
               class_name, stream
        FROM students
        WHERE archived = 0
    """
    params = []
    if class_name:
        qs += " AND class_name = ?"; params.append(class_name)
    if stream:
        qs += " AND stream = ?"; params.append(stream)
    qs += " ORDER BY class_name, stream, last_name, first_name"
    students = conn.execute(qs, params).fetchall()
    conn.close()

    # ---- build workbook (headers include FULL NAME) -------------------------
    wb = Workbook()
    ws = wb.active; ws.title = "Marks"

    # Order: OTH, HP, BOT, MID, EOT
    headers = [
        "student_number", "full_name", "class", "stream",
        "subject_id", "subject_name", "subject_code",
        "term", "year",
        "other_mark", "holiday_mark", "bot_mark", "midterm_mark", "eot_mark",
        "initials"
    ]
    ws.append(headers)

    if students:
        for s in students:
            ws.append([
                s["student_number"], s["full_name"], s["class_name"], s["stream"],
                subject_id or "", subj_name, subj_code,
                term, year,
                "", "", "", "", "", # OTH, HP, BOT, MID, EOT (left blank)
                ""
            ])
    else:
        # blank example row if no students matched the filter
        ws.append([
            "STD-YYYY-001", "Jane Doe", class_name or "", stream or "",
            subject_id or "", subj_name, subj_code,
            term, year,
            "", "", "", "", "",
            ""
        ])

    # column widths
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 6, 14), 40)

    out = BytesIO(); wb.save(out); out.seek(0)
    fname = f"marks_template_{class_name or 'all'}_{term}_{year}.xlsx"
    return send_file(
        out, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ===== Upload (CSV/XLSX) – auto-fills initials if missing ==================



import pandas as pd

@app.route("/marks/upload", methods=["POST"])
@require_role("admin", "headteacher", "teacher", "dos")
def marks_upload():
    inserted = skipped = 0
    conn = get_db_connection(); conn.row_factory = sqlite3.Row

    try:
        ensure_record_score_table(conn)
        auto_initials = get_user_initials(conn, session.get("user_id"))

        # ------------------ READ THE UPLOADED FILE ------------------
        file = request.files.get("file")
        if not file:
            flash("No file uploaded.", "warning")
            return redirect(url_for("marks_hub", **request.args))

        filename = file.filename.lower()

        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        elif filename.endswith((".xls", ".xlsx")):
            df = pd.read_excel(file)
        else:
            flash("Unsupported file format. Upload CSV or Excel.", "danger")
            return redirect(url_for("marks_hub", **request.args))

        # ------------------ PROCESS EACH ROW ------------------
        for _, R in df.iterrows():
            sn = str(R.get("student_number") or "").strip()
            if not sn:
                skipped += 1
                continue

            subj_id = resolve_subject_id(
                conn,
                subject_id=_safe_int(R.get("subject_id")),
                subject_code=(str(R.get("subject_code")).strip()
                              if R.get("subject_code") not in (None, float("nan")) else None),
                subject_name=(str(R.get("subject_name")).strip()
                              if R.get("subject_name") not in (None, float("nan")) else None),
            )
            if not subj_id:
                skipped += 1
                continue

            student_id = resolve_student_id(conn, sn)
            if not student_id:
                skipped += 1
                continue

            term = str(R.get("term") or default_term).strip()
            year = _safe_int(R.get("year")) or default_year

            # Marks
            oth = _safe_int(R.get("other_mark"))
            hp = _safe_int(R.get("holiday_mark"))
            bot = _safe_int(R.get("bot_mark"))
            mid = _safe_int(R.get("midterm_mark"))
            eot = _safe_int(R.get("eot_mark"))
            avg = _avg(oth, hp, bot, mid, eot)

            file_initials = R.get("initials")
            initials = (str(file_initials).strip()
                        if file_initials is not None and str(file_initials).strip() != ""
                        else auto_initials)

            conn.execute("""
                INSERT INTO record_score (
                    student_id, subject_id, term, year,
                    bot_mark, midterm_mark, holiday_mark, other_mark, eot_mark,
                    average_mark, initials
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(student_id, subject_id, term, year) DO UPDATE SET
                    bot_mark = COALESCE(excluded.bot_mark, record_score.bot_mark),
                    midterm_mark = COALESCE(excluded.midterm_mark, record_score.midterm_mark),
                    holiday_mark = COALESCE(excluded.holiday_mark, record_score.holiday_mark),
                    other_mark = COALESCE(excluded.other_mark, record_score.other_mark),
                    eot_mark = COALESCE(excluded.eot_mark, record_score.eot_mark),
                    average_mark = COALESCE(excluded.average_mark, record_score.average_mark),
                    initials = CASE WHEN excluded.initials IS NOT NULL AND excluded.initials!=''
                                    THEN excluded.initials ELSE record_score.initials END,
                    processed_on = CURRENT_TIMESTAMP
            """, (student_id, subj_id, term, year,
                  bot, mid, hp, oth, eot,
                  avg, initials))
            inserted += 1

        conn.commit()
        flash(f"Upload complete: {inserted} rows processed; skipped {skipped}.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Upload failed: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("marks_hub", **request.args))


# ===== Delete & Export =====================================================

@app.route("/marks/delete/<int:score_id>", methods=["POST"])
@require_role("admin", "headteacher", "dos")
def marks_delete(score_id):
    conn = get_db_connection()
    try:
        ensure_record_score_table(conn)
        conn.execute("DELETE FROM record_score WHERE id=?", (score_id,))
        conn.commit()
        flash("Deleted.", "info")

        # ---- AUDIT ----
        audit_from_request(
            conn,
            action="marks_delete",
            target_table="record_score",
            target_id=score_id
        )
    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
        audit_from_request(
            conn,
            action="marks_delete",
            outcome="failure",
            severity="warning",
            target_table="record_score",
            target_id=score_id,
            details={"error": str(e)}
        )
    finally:
        conn.close()
    return redirect(url_for("marks_hub", **request.args))



@app.route("/marks/export")
@require_role("admin", "headteacher", "teacher", "dos")
def marks_export():
    # ---- filters -------------------------------------------------------------
    class_name = (request.args.get("class_name") or "").strip()
    stream = (request.args.get("stream") or "").strip()
    subject_id = _safe_int(request.args.get("subject_id"))

    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    try:
        ensure_record_score_table(conn)

        # ---- academic year (with override support) ---------------------------
        ay = conn.execute("SELECT year, current_term FROM academic_years WHERE is_active=1").fetchone()
        if not ay:
            flash("Please activate an academic year.", "warning")
            return redirect(url_for("dashboard"))
        default_year = int(ay["year"])
        default_term = ay["current_term"]

        term = (request.args.get("term") or default_term).strip()
        try:
            year = int(request.args.get("year") or default_year)
        except ValueError:
            year = default_year

        # ---- query (order kept: OTH, HP, BOT, MID, EOT in export columns) ----
        q = """
        SELECT st.student_number,
               st.first_name||' '||COALESCE(st.Middle_name,'')||' '||st.last_name AS full_name,
               st.class_name, st.stream,
               sub.name AS subject_name, sub.code AS subject_code,
               rs.term, rs.year,
               rs.other_mark, rs.holiday_mark, rs.bot_mark, rs.midterm_mark, rs.eot_mark,
               rs.average_mark, rs.initials
        FROM record_score rs
        JOIN students st ON st.id = rs.student_id
        JOIN subjects sub ON sub.id = rs.subject_id
        WHERE rs.term = ? AND rs.year = ?
        """
        params = [term, year]
        if class_name:
            q += " AND st.class_name = ?"; params.append(class_name)
        if stream:
            q += " AND st.stream = ?"; params.append(stream)
        if subject_id:
            q += " AND rs.subject_id = ?"; params.append(subject_id)
        q += " ORDER BY st.class_name, st.stream, st.last_name, st.first_name, sub.name"

        rows = conn.execute(q, params).fetchall()

        # ---- workbook --------------------------------------------------------
        wb = Workbook(); ws = wb.active; ws.title = "Marks"
        headers = [
            "student_number","full_name","class","stream","subject_name","subject_code",
            "term","year",
            "other_mark","holiday_mark","bot_mark","midterm_mark","eot_mark", # OTH, HP, BOT, MID, EOT
            "average_mark","initials"
        ]
        ws.append(headers)

        for r in rows:
            ws.append([
                r["student_number"], r["full_name"], r["class_name"], r["stream"],
                r["subject_name"], r["subject_code"], r["term"], r["year"],
                r["other_mark"] if r["other_mark"] is not None else "",
                r["holiday_mark"] if r["holiday_mark"] is not None else "",
                r["bot_mark"] if r["bot_mark"] is not None else "",
                r["midterm_mark"] if r["midterm_mark"] is not None else "",
                r["eot_mark"] if r["eot_mark"] is not None else "",
                r["average_mark"] if r["average_mark"] is not None else "",
                r["initials"] or ""
            ])

        out = BytesIO(); wb.save(out); out.seek(0)
        fname = f"marks_export_{class_name or 'all'}_{term}_{year}.xlsx"
        return send_file(
            out, as_attachment=True, download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        conn.close()


# (Optional) simple summary
@app.route("/marks/summary")
@require_role("admin", "headteacher", "teacher", "dos")
def marks_summary():
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("""
            SELECT st.class_name, st.stream, sub.name AS subject, rs.term, rs.year,
                   COUNT(*) AS entries
              FROM record_score rs
              JOIN students st ON st.id = rs.student_id
              JOIN subjects sub ON sub.id = rs.subject_id
             GROUP BY st.class_name, st.stream, sub.name, rs.term, rs.year
             ORDER BY rs.year DESC,
                      CASE rs.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                      st.class_name, st.stream, sub.name
        """).fetchall()
    finally:
        conn.close()
    return render_template("marks_summary.html", rows=rows)






@app.route("/performance/summary")
@require_role("admin","headteacher","dos","bursar","teacher")
def performance_summary():
    """
    Class-by-class performance summary for the ACTIVE term/year.

    - Grades come from grading_scale (lower_limit..upper_limit).
    - Aggregate/Division use ONLY core subjects: ENG, MATH, SCI, SST.
      If any core is missing => NG (no division).
    """
    import sqlite3, math
    from collections import defaultdict

    # --- Active session ---
    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1").strip()
    year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)

    # Filters from querystring
    class_filter = (request.args.get("class_name") or "").strip()
    stream_filter = (request.args.get("stream") or "").strip()

    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # ---- Dropdown options ----
    class_options = [r[0] for r in c.execute(
        "SELECT DISTINCT class_name "
        "FROM students "
        "WHERE class_name IS NOT NULL AND TRIM(class_name)<>'' "
        "ORDER BY class_name"
    ).fetchall()]

    stream_options = [r[0] for r in c.execute(
        "SELECT DISTINCT COALESCE(stream,'') "
        "FROM students "
        "WHERE stream IS NOT NULL AND TRIM(stream)<>'' "
        "ORDER BY stream"
    ).fetchall()]

    # ---- Load grading scale (used to map score -> grade) ----
    scale_rows = c.execute(
        "SELECT grade, lower_limit, upper_limit "
        "FROM grading_scale "
        "ORDER BY lower_limit DESC"
    ).fetchall()

    def grade_for_score(score: float) -> str:
        """Map numeric score -> grade using grading_scale table."""
        s = float(score or 0.0)
        for r in scale_rows:
            lo, hi = float(r["lower_limit"]), float(r["upper_limit"])
            if lo <= s <= hi:
                return (r["grade"] or "").strip()
        return "NG"

    # Convert grade (D1..F9) to points 1..9 (fallback to band order if unknown).
    grade_point_map = {
        "D1": 1, "D2": 2, "C3": 3, "C4": 4, "C5": 5, "C6": 6, "P7": 7, "P8": 8, "F9": 9
    }
    def points_for_grade(g: str) -> int:
        g2 = (g or "").upper().replace(" ", "")
        if g2 in grade_point_map:
            return grade_point_map[g2]
        if not scale_rows:
            return 9
        ordered = sorted(scale_rows, key=lambda r: (-float(r["lower_limit"]), -float(r["upper_limit"])))
        order_map = { (r["grade"] or "").upper().replace(" ",""): i+1 for i, r in enumerate(ordered) }
        return order_map.get(g2, 9)

    # Normalize subject names to core keys
    def norm_subj(name: str) -> str:
        n = (name or "").strip().lower()
        if n.startswith("eng"): return "eng"
        if n.startswith("mat") or n.startswith("math"): return "math"
        if n.startswith("sci"): return "sci"
        if n in ("sst", "soc. studies", "social studies", "social std", "socialstudies"):
            return "sst"
        return n

    # ---- Build WHERE for marks query ----
    where = ["rs.term = ?", "rs.year = ?", "st.archived = 0"]
    params = [term, year]
    if class_filter:
        where.append("st.class_name = ?"); params.append(class_filter)
    if stream_filter:
        where.append("COALESCE(st.stream,'') = ?"); params.append(stream_filter)
    where_sql = " AND ".join(where)

    # ---- Pull marks (join subjects to get subject names) ----
    rows = c.execute(f"""
        SELECT
          st.id AS student_id,
          st.first_name,
          COALESCE(st.Middle_name,'') AS middle_name,
          st.last_name,
          st.class_name,
          COALESCE(st.stream,'') AS stream,
          sj.name AS subject_name,
          COALESCE(
            rs.average_mark,
            rs.eot_mark,
            rs.midterm_mark,
            rs.bot_mark,
            rs.other_mark,
            rs.ca_mark,
            0
          ) AS score
        FROM record_score rs
        JOIN students st ON st.id = rs.student_id
        LEFT JOIN subjects sj ON sj.id = rs.subject_id
        WHERE {where_sql}
        ORDER BY st.class_name, st.stream, st.last_name, st.first_name
    """, params).fetchall()
    conn.close()

    # ---- Crunch per student ----
    per_student = {} # sid -> dict
    for r in rows:
        sid = r["student_id"]
        stu = per_student.setdefault(sid, {
            "student_id": sid,
            "name": f"{r['first_name']} {r['middle_name']} {r['last_name']}".replace(" ", " ").strip(),
            "class_name": r["class_name"],
            "stream": r["stream"],
            "scores": [], # all numeric scores for average/total
            "by_subject": {} # norm_name -> {"score": x, "grade": g, "points": p}
        })
        score = float(r["score"] or 0.0)
        stu["scores"].append(score)

        sname = norm_subj(r["subject_name"] or "")
        g = grade_for_score(score)
        p = points_for_grade(g)
        stu["by_subject"][sname] = {"score": score, "grade": g, "points": p}

    # Division from aggregate (PLE-like)
    def division_from_aggregate(agg: int | None) -> str:
        if agg is None:
            return "NG"
        a = int(agg)
        if 4 <= a <= 12: return "Div 1"
        if 13 <= a <= 23: return "Div 2"
        if 24 <= a <= 29: return "Div 3"
        if 30 <= a <= 34: return "Div 4"
        return "U"

    classes = defaultdict(list)
    CORE = ("eng", "math", "sci", "sst")

    for s in per_student.values():
        # average/total (all subjects present for the student this term)
        n = len(s["scores"])
        total = sum(s["scores"]) if n else 0.0
        avg = (total / n) if n else 0.0

        # compute aggregate from ONLY core subjects; require all 4
        core_points = []
        has_all_core = True
        for k in CORE:
            if k in s["by_subject"]:
                core_points.append(s["by_subject"][k]["points"])
            else:
                has_all_core = False
                break

        if has_all_core:
            aggregate = sum(core_points)
            division = division_from_aggregate(aggregate)
        else:
            aggregate = None
            division = "NG"

        s["total"] = total
        s["average"] = avg
        s["aggregate"] = None if (aggregate is None) else aggregate
        s["division"] = division

        classes[(s["class_name"], s["stream"])].append(s)

    # Order students per class: graded first (by avg desc), NG last
    def _sort_key(stu):
        ng_flag = (stu["division"] == "NG")
        return (ng_flag, -stu["average"])

    for key in classes:
        classes[key].sort(key=_sort_key)

    # ---- Class stats + division counts ----
    DIV_LABELS = ["Div 1", "Div 2", "Div 3", "Div 4", "U", "NG"]
    class_stats = {}
    for key, lst in classes.items():
        graded = [x for x in lst if x["division"] != "NG"]
        avgs = [x["average"] for x in graded]
        counts = {d: 0 for d in DIV_LABELS}
        for s in lst:
            counts[s["division"]] = counts.get(s["division"], 0) + 1

        class_stats[key] = {
            "mean": (sum(avgs) / len(avgs)) if avgs else 0.0,
            "size": len(lst),
            "top_avg": max(avgs) if avgs else 0.0,
            "div_counts": counts,
        }

    # (No NaN aggregates are emitted—already set to None above.)

    return render_template(
        "performance_summary.html",
        term=term, year=year,
        classes=classes,
        class_stats=class_stats,
        class_filter=class_filter,
        stream_filter=stream_filter,
        class_options=class_options,
        stream_options=stream_options,
        BEST_N=None # core-only aggregate; BEST_N unused
    )


@app.route("/midterm/overview", methods=["GET"])
@require_role("admin","headteacher","teacher","bursar")
def midterm_overview():
    class_name = (request.args.get("class_name") or "").strip()
    term = (request.args.get("term") or "Term 1").strip()
    year = int(request.args.get("year") or datetime.now().year)

    conn = get_db_connection(); conn.row_factory = sqlite3.Row

    # roster
    students = conn.execute("""
        SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name, last_name
        FROM students
        WHERE archived=0 AND class_name=?
        ORDER BY last_name, first_name
    """, (class_name,)).fetchall()

    bm = _bot_mid_by_sid(conn, class_name, term, year)
    hol = _midterms_pick(conn, class_name, term, year, HOLIDAY_NAME)
    oth = _midterms_pick(conn, class_name, term, year, OTHER_NAME)

    table = []
    for s in students:
        sid = s["id"]
        name = f"{s['first_name']} {s['Middle_name']} {s['last_name']}".replace(" "," ").strip()
        row = {"student_number": s["student_number"], "name": name, "subjects": []}
        for code in BIG4_CODES:
            bot = bm.get(sid, {}).get(code, {}).get("bot")
            mid_stored = bm.get(sid, {}).get(code, {}).get("mid")
            hp = hol.get(sid, {}).get(code)
            other = oth.get(sid, {}).get(code)
            mid_final = mid_stored if mid_stored is not None else _mean([bot, hp, other])

            def pack(val):
                g = grade_for_score(conn, val)
                return dict(mark=val, grade=g, comment=comment_for_grade(conn, g))

            row["subjects"].append(dict(
                code=code,
                bot=pack(bot),
                holiday=pack(hp),
                other=pack(other),
                mid=pack(mid_final)
            ))
        table.append(row)

    conn.close()
    return render_template("midterm_overview.html",
                           class_name=class_name, term=term, year=year,
                           rows=table, subj_order=BIG4_CODES)



@app.route("/next_term", methods=["GET", "POST"])
@require_role('admin','headteacher')
def next_term_hub():
    # Defaults
    terms = ["Term 1", "Term 2", "Term 3"]
    try:
        # Try your active academic year helper if you have it
        ay = get_active_academic_year() or {}
        default_year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        default_year = datetime.now().year

    sel_year = request.values.get("year", type=int) or default_year
    sel_term = request.values.get("term") or terms[0]

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    ensure_term_dates_schema(conn)
    c = conn.cursor()

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "save":
            year = request.form.get("year", type=int)
            term = (request.form.get("term") or "").strip()
            next_term = (request.form.get("next_term") or "").strip() or None
            next_term_date = (request.form.get("next_term_date") or "").strip() or None

            if not (year and term):
                flash("Please select a valid Year and Term.", "warning")
            else:
                try:
                    c.execute("""
                        INSERT INTO term_dates (year, term, next_term, next_term_date)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(year, term) DO UPDATE SET
                          next_term = excluded.next_term,
                          next_term_date = excluded.next_term_date
                    """, (year, term, next_term, next_term_date))
                    conn.commit()
                    flash("Next-term info saved.", "success")
                    # Persist selection in querystring
                    return redirect(url_for("next_term_hub", year=year, term=term))
                except Exception as e:
                    conn.rollback()
                    flash(f"Could not save: {e}", "danger")

    # Load current row (if any)
    row = c.execute("""
        SELECT next_term, next_term_date
        FROM term_dates
        WHERE year=? AND term=? LIMIT 1
    """, (sel_year, sel_term)).fetchone()

    conn.close()
    return render_template(
        "next_term_hub.html",
        terms=terms,
        sel_year=sel_year,
        sel_term=sel_term,
        existing=row
    )


@app.route('/process_reports', methods=['GET','POST'])
@require_role('admin', 'headteacher','bursar','dos')
def process_reports():
    term = request.values.get('term', 'Term 1')
    year = int(request.values.get('year', datetime.now().year))
    class_id = request.values.get('class_id')

    # TODO: if you have batch compute for grades/totals, call it here.
    flash('Reports processed successfully.', 'success')
    return redirect(url_for('report_card_list', term=term, year=year, class_id=class_id))
    


# ---------------------------
# Reports Hub (Process, Print, Next-term setup)
# ---------------------------
# ---------------------------
# Reports Hub (Process, Print, Next-term setup)
# ---------------------------
@app.route("/reports", methods=["GET", "POST"])
@require_role('admin', 'headteacher','dos','bursar')
def reports_hub():
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # --- dropdown data ---
    classes = [r[0] for r in c.execute(
        "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL ORDER BY class_name"
    )]

    src = detect_scores_table(conn) # "record_score" or "results"
    if src == "record_score":
        years = [r[0] for r in c.execute("SELECT DISTINCT year FROM record_score ORDER BY 1 DESC")]
    else:
        years = [r[0] for r in c.execute("SELECT DISTINCT year FROM results ORDER BY 1 DESC")]
    terms = ["Term 1", "Term 2", "Term 3"]

    # --- selected filters (GET preserves, POST inherits) ---
    sel_class = (request.values.get("class_name") or (classes[0] if classes else "")).strip()
    sel_term = (request.values.get("term") or (terms[0] if terms else "")).strip()
    sel_year = int(request.values.get("year") or (years[0] if years else datetime.now().year))

    # --- actions ---
    action = request.form.get("action")

    # 1) Rebuild snapshot for this class/term/year
    if request.method == "POST" and action == "process":
        try:
            process_reports_snapshot(conn, sel_class, sel_term, sel_year)
            flash(f"Reports processed for {sel_class} — {sel_term} {sel_year}.", "success")
        except Exception as e:
            flash(f"Processing failed: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("reports_hub", class_name=sel_class, term=sel_term, year=sel_year))

    # 2) Save next-term details (right column form)
    if request.method == "POST" and action == "save_term_dates":
        ensure_term_dates_schema(conn)
        next_term = (request.form.get("next_term") or "").strip() or None
        next_term_date = (request.form.get("next_term_date") or "").strip() or None
        try:
            c.execute("""
                INSERT INTO term_dates (year, term, next_term, next_term_date)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(year, term) DO UPDATE SET
                    next_term = excluded.next_term,
                    next_term_date = excluded.next_term_date
            """, (sel_year, sel_term, next_term, next_term_date))
            conn.commit()
            flash("Next-term details saved.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Could not save next-term details: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("reports_hub", class_name=sel_class, term=sel_term, year=sel_year))

    # --- page data ---
    ensure_term_dates_schema(conn)
    students_ready = c.execute("""
        SELECT DISTINCT s.id, s.student_number, s.first_name, s.last_name,
                        s.class_name, r.term, r.year
        FROM reports r
        JOIN students s ON s.id = r.student_id
        WHERE r.term=? AND r.year=? AND s.class_name=?
        ORDER BY s.last_name, s.first_name
    """, (sel_term, sel_year, sel_class)).fetchall()

    existing_term = c.execute("""
        SELECT next_term, next_term_date
        FROM term_dates WHERE year=? AND term=? LIMIT 1
    """, (sel_year, sel_term)).fetchone()

    conn.close()
    return render_template(
        "reports_hub.html",
        classes=classes, terms=terms, years=years,
        sel_class=sel_class, sel_term=sel_term, sel_year=sel_year,
        students_ready=students_ready,
        existing_term=existing_term
    )



@app.route("/report_card/<int:student_id>/<term>/<int:year>")
@require_role('admin','teacher','headteacher','bursar','dos')
def report_card(student_id, term, year):
    include_mid = (request.args.get("include_mid") == "1")

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # --- Ensure aux schemas exist (safe no-ops if already there) ---
    ensure_term_dates_schema(conn)

    # --- Student ---
    student = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not student:
        conn.close()
        abort(404)

    # --- Ensure snapshot exists for this student/term/year ---
    has_reports = c.execute(
        "SELECT 1 FROM reports WHERE student_id=? AND term=? AND year=? LIMIT 1",
        (student_id, term, year)
    ).fetchone()
    if not has_reports:
        process_reports_snapshot(conn, student["class_name"], term, year)

    # --- Main subject rows (End-of-term table) ---
    rows = c.execute("""
        SELECT sub.name AS subject, sub.code AS subject_code,
               r.eot_mark AS eot,
               COALESCE(r.average_mark, r.eot_mark) AS total_100,
               r.grade, r.comment, r.teacher_initial AS initials,
               r.teacher_remark, r.headteacher_remark
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=? AND r.term=? AND r.year=?
        ORDER BY sub.name
    """, (student_id, term, year)).fetchall()

    # --- Totals / averages ---
    marks = [r["total_100"] for r in rows if r["total_100"] is not None]
    total_sum = round(sum(marks), 2) if marks else None
    avg_overall = round(sum(marks)/len(marks), 2) if marks else None

   
    # ----- Aggregate & Division: big 4 only -----
    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}
    grade_points = {"D1":1,"D2":2,"C3":3,"C4":4,"C5":5,"C6":6,"P7":7,"P8":8,"F9":9}

    pts = []
    for r in rows:
        code = (r["subject_code"] or "").upper()
        if code in CORE_CODES:
            gp = grade_points.get(r["grade"])
            if gp is not None:
                pts.append(gp)

    # only when we have all 4 core subjects
    aggregate = sum(pts) if len(pts) == 4 else None

    if aggregate is not None:
        if 4 <= aggregate <= 12:
            division = "1"
        elif 13 <= aggregate <= 24:
            division = "2"
        elif 25 <= aggregate <= 29:
            division = "3"
        elif 30 <= aggregate <= 34:
            division = "4"
        else:
            division = "U"
    else:
        division = None

    

    # --- Class position (integer; template renders ordinal) ---
    rk = class_ranking(conn, student["class_name"], term, year)
    class_size = len(rk)
    position = None
    for i, rr in enumerate(rk, start=1):
        if rr["sid"] == student_id:
            position = i
            break

    
    # ----- MIDTERM PANEL (dynamic) ---------------------------------------------
    midterms = []
    midterm_subjects = []

    if include_mid:
        CORE_CODES = ("ENG", "MATH", "SCI", "SST")

        def grade_of(v):
            return grade_for_score(conn, v) if v is not None else None

        # Pull code<->id mapping in one query (robust & fast)
        code_map_rows = c.execute("""
            SELECT UPPER(code) AS code, id
            FROM subjects
            WHERE code IS NOT NULL AND TRIM(code) <> ''
        """).fetchall()

        # Build lookup and available codes
        code_to_id = {r["code"]: r["id"] for r in code_map_rows}
        all_codes = list(code_to_id.keys())

        # Headers: Big-4 first (only if present), then others A→Z
        others = sorted([sc for sc in all_codes if sc not in CORE_CODES])
        midterm_subjects = [sc for sc in CORE_CODES if sc in all_codes] + others

        # If nothing to show, keep a minimal header so the table still renders
        if not midterm_subjects:
            midterm_subjects = []

        def fetch_marks(colname: str, *, round0: bool = False) -> dict:
            """
            Return {CODE: value_or_None} in the exact order of midterm_subjects.
            OTH/HP (round0=True) are rounded to 0 dp.
            """
            out = {}
            for sc in midterm_subjects:
                sid = code_to_id.get(sc)
                if not sid:
                    out[sc] = None
                    continue
                row = c.execute(
                    f"""SELECT MAX({colname}) AS v
                        FROM record_score
                        WHERE student_id=? AND subject_id=? AND term=? AND year=?""",
                    (student_id, sid, term, year)
                ).fetchone()
                v = row["v"] if row and row["v"] is not None else None
                if v is not None:
                    try:
                        v = int(round(float(v))) if round0 else int(v)
                    except Exception:
                        # best-effort cast
                        try:
                            v = int(float(v))
                        except Exception:
                            v = None
                out[sc] = v
            return out

        def to_grades(score_map: dict) -> dict:
            return {sc: grade_of(v) for sc, v in score_map.items()}

        panels = [
            ("OTH", fetch_marks("other_mark", round0=True)),
            ("HP", fetch_marks("holiday_mark", round0=True)),
            ("BOT", fetch_marks("bot_mark")),
            ("MID", fetch_marks("midterm_mark")),
            ("EOT", fetch_marks("eot_mark")),
        ]

        for label, score_map in panels:
            # Add a row only if at least one subject has a value
            if any(v is not None for v in score_map.values()):
                total_all = sum(int(v) for v in score_map.values() if v is not None)
                midterms.append({
                    "assessment": label,
                    "scores": score_map, # {CODE: score}
                    "grades": to_grades(score_map), # {CODE: grade}
                    "total": total_all
                })
# ---------------------------------------------------------------------------

        


    # --- Payment number (latest fee record) ---
    pay = c.execute("""
        SELECT id AS payment_number
        FROM fees
        WHERE student_id=? AND term=? AND year=?
        ORDER BY date_paid DESC, id DESC
        LIMIT 1
    """, (student_id, term, year)).fetchone()
    payment_number = pay["payment_number"] if pay else None

    # --- Auto comments ---
    head_comment = pick_comment_template(
        role="headteacher", scope="overall",
        division=(int(division) if division and str(division).isdigit() else None),
        average=avg_overall, class_name=student["class_name"], term=term
    ) or (comment_for_grade(conn, grade_for_score(conn, avg_overall)) or "")

    teacher_comment = pick_comment_template(
        role="teacher", scope="overall",
        division=(int(division) if division and str(division).isdigit() else None),
        average=avg_overall, class_name=student["class_name"], term=term
    )
    if not teacher_comment:
        per_subj = [r["teacher_remark"] for r in rows if r["teacher_remark"]]
        teacher_comment = (Counter(per_subj).most_common(1)[0][0] if per_subj else None)
    if not teacher_comment:
        teacher_comment = comment_for_grade(conn, grade_for_score(conn, avg_overall)) or ""

    # --- Next-term info (prefer current term; else fallback to computed “next”) ---
    cur = c.execute("""
        SELECT next_term, next_term_date
        FROM term_dates
        WHERE year=? AND term=? LIMIT 1
    """, (year, term)).fetchone()
    if cur and (cur["next_term"] or cur["next_term_date"]):
        next_term_info = dict(next_term=cur["next_term"], next_term_date=cur["next_term_date"])
    else:
        nt_name, nt_year = _next_term_name_and_year(term, year)
        fb = c.execute("""
            SELECT next_term, next_term_date
            FROM term_dates
            WHERE year=? AND term=? LIMIT 1
        """, (nt_year, nt_name)).fetchone()
        next_term_info = (dict(next_term=nt_name, next_term_date=fb["next_term_date"])
                          if fb and fb["next_term_date"] else None)

    # --- Grading legend + school header ---
    grading = fetch_grading_scale(conn)
    conn.close()

    school = dict(
        name="CITIZENS DAY AND BOARDING",
        tagline="PRIMARY SCHOOL – MASAJJA",
        motto="Strive for the best",
        phones="+256781757410, +256704720641, +256788529084",
        pobox="P.O Box 31882 Kampala"
    )

    return render_template(
        "report_card_citizen.html",
        school=school,
        student=student,
        term=term, year=year,
        rows=rows,
        total_sum=total_sum,
        avg_overall=avg_overall,
        aggregate=aggregate,
        division=division,
        position=position,
        class_size=class_size,
        midterms=midterms,
        midterm_subjects=midterm_subjects,
        payment_number=payment_number,
        comments={"teacher_comment": teacher_comment, "head_comment": head_comment},
        next_term_info=next_term_info, # <--- template reads this
        grading=grading
    )
# ---------------------------
# Batch printing
# ---------------------------
def build_report_payload(conn, student_id, term, year, include_mid=False):
    """
    Returns a dict with exactly the same keys your single report passes
    to report_card_citizen.html, so batch can render the same block in a loop.
    """
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    ensure_term_dates_schema(conn)

    student = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not student:
        return None

    # Ensure snapshot exists (same behavior as single)
    has_reports = c.execute(
        "SELECT 1 FROM reports WHERE student_id=? AND term=? AND year=? LIMIT 1",
        (student_id, term, year)
    ).fetchone()
    if not has_reports:
        process_reports_snapshot(conn, student["class_name"], term, year)

    # Main table rows
    rows = c.execute("""
        SELECT sub.name AS subject, sub.code AS subject_code,
               r.eot_mark AS eot,
               COALESCE(r.average_mark, r.eot_mark) AS total_100,
               r.grade, r.comment, r.teacher_initial AS initials,
               r.teacher_remark, r.headteacher_remark
        FROM reports r
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.student_id=? AND r.term=? AND r.year=?
        ORDER BY sub.name
    """, (student_id, term, year)).fetchall()

    # Totals / averages
    marks = [r["total_100"] for r in rows if r["total_100"] is not None]
    total_sum = round(sum(marks), 2) if marks else None
    avg_overall = round(sum(marks)/len(marks), 2) if marks else None

    
    # ----- Aggregate & Division: big 4 only -----
    CORE_CODES = {"ENG", "MATH", "SCI", "SST"}
    grade_points = {"D1":1,"D2":2,"C3":3,"C4":4,"C5":5,"C6":6,"P7":7,"P8":8,"F9":9}

    pts = []
    for r in rows:
        code = (r["subject_code"] or "").upper()
        if code in CORE_CODES:
            gp = grade_points.get(r["grade"])
            if gp is not None:
                pts.append(gp)

    # only when we have all 4 core subjects
    aggregate = sum(pts) if len(pts) == 4 else None

    if aggregate is not None:
        if 4 <= aggregate <= 12:
            division = "1"
        elif 13 <= aggregate <= 24:
            division = "2"
        elif 25 <= aggregate <= 29:
            division = "3"
        elif 30 <= aggregate <= 34:
            division = "4"
        else:
            division = "U"
    else:
        division = None


    # Position (as integer; template renders ordinal suffix)
    rk = class_ranking(conn, student["class_name"], term, year)
    class_size = len(rk)
    position = None
    for i, rr in enumerate(rk, start=1):
        if rr["sid"] == student_id:
            position = i
            break

    
    # ----- MIDTERM PANEL (dynamic) ---------------------------------------------
    midterms = []
    midterm_subjects = []

    if include_mid:
        CORE_CODES = ("ENG", "MATH", "SCI", "SST")

        def grade_of(v):
            return grade_for_score(conn, v) if v is not None else None

        # Pull code<->id mapping in one query (robust & fast)
        code_map_rows = c.execute("""
            SELECT UPPER(code) AS code, id
            FROM subjects
            WHERE code IS NOT NULL AND TRIM(code) <> ''
        """).fetchall()

        # Build lookup and available codes
        code_to_id = {r["code"]: r["id"] for r in code_map_rows}
        all_codes = list(code_to_id.keys())

        # Headers: Big-4 first (only if present), then others A→Z
        others = sorted([sc for sc in all_codes if sc not in CORE_CODES])
        midterm_subjects = [sc for sc in CORE_CODES if sc in all_codes] + others

        # If nothing to show, keep a minimal header so the table still renders
        if not midterm_subjects:
            midterm_subjects = []

        def fetch_marks(colname: str, *, round0: bool = False) -> dict:
            """
            Return {CODE: value_or_None} in the exact order of midterm_subjects.
            OTH/HP (round0=True) are rounded to 0 dp.
            """
            out = {}
            for sc in midterm_subjects:
                sid = code_to_id.get(sc)
                if not sid:
                    out[sc] = None
                    continue
                row = c.execute(
                    f"""SELECT MAX({colname}) AS v
                        FROM record_score
                        WHERE student_id=? AND subject_id=? AND term=? AND year=?""",
                    (student_id, sid, term, year)
                ).fetchone()
                v = row["v"] if row and row["v"] is not None else None
                if v is not None:
                    try:
                        v = int(round(float(v))) if round0 else int(v)
                    except Exception:
                        # best-effort cast
                        try:
                            v = int(float(v))
                        except Exception:
                            v = None
                out[sc] = v
            return out

        def to_grades(score_map: dict) -> dict:
            return {sc: grade_of(v) for sc, v in score_map.items()}

        panels = [
            ("OTH", fetch_marks("other_mark", round0=True)),
            ("HP", fetch_marks("holiday_mark", round0=True)),
            ("BOT", fetch_marks("bot_mark")),
            ("MID", fetch_marks("midterm_mark")),
            ("EOT", fetch_marks("eot_mark")),
        ]

        for label, score_map in panels:
            # Add a row only if at least one subject has a value
            if any(v is not None for v in score_map.values()):
                total_all = sum(int(v) for v in score_map.values() if v is not None)
                midterms.append({
                    "assessment": label,
                    "scores": score_map, # {CODE: score}
                    "grades": to_grades(score_map), # {CODE: grade}
                    "total": total_all
                })
# ---------------------------------------------------------------------------

        
        
    # Payment number
    pay = c.execute("""
        SELECT id AS payment_number
        FROM fees
        WHERE student_id=? AND term=? AND year=?
        ORDER BY date_paid DESC, id DESC
        LIMIT 1
    """, (student_id, term, year)).fetchone()
    payment_number = pay["payment_number"] if pay else None

    # Comments (same fallback chain as single)
    head_comment = pick_comment_template(
        role="headteacher", scope="overall",
        division=(int(division) if division and str(division).isdigit() else None),
        average=avg_overall, class_name=student["class_name"], term=term
     ) or (comment_for_grade(conn, grade_for_score(conn, avg_overall)) or "")

    teacher_comment = pick_comment_template(
        role="teacher", scope="overall",
        division=(int(division) if division and str(division).isdigit() else None),
        average=avg_overall, class_name=student["class_name"], term=term
    )
    if not teacher_comment:
        per_subj = [r["teacher_remark"] for r in rows if r["teacher_remark"]]
        from collections import Counter
        teacher_comment = (Counter(per_subj).most_common(1)[0][0] if per_subj else None)
    if not teacher_comment:
        teacher_comment = comment_for_grade(conn, grade_for_score(conn, avg_overall)) or ""

    # Next-term info (same logic as single)
    def _next_term_name_and_year(cur_term: str, cur_year: int):
        order = ["Term 1", "Term 2", "Term 3"]
        try:
            i = order.index(cur_term)
        except ValueError:
            return "Term 2", cur_year
        return (order[i+1], cur_year) if i < 2 else (order[0], cur_year + 1)

    cur = c.execute("""
        SELECT next_term, next_term_date
        FROM term_dates
        WHERE year=? AND term=? LIMIT 1
    """, (year, term)).fetchone()
    if cur and (cur["next_term"] or cur["next_term_date"]):
        next_term_info = dict(next_term=cur["next_term"], next_term_date=cur["next_term_date"])
    else:
        nt_name, nt_year = _next_term_name_and_year(term, year)
        fb = c.execute("""
            SELECT next_term, next_term_date
            FROM term_dates
            WHERE year=? AND term=? LIMIT 1
        """, (nt_year, nt_name)).fetchone()
        next_term_info = (dict(next_term=nt_name, next_term_date=fb["next_term_date"])
                          if fb and fb["next_term_date"] else None)

    grading = fetch_grading_scale(conn)
    school = dict(
        name="CITIZENS DAY AND BOARDING",
        tagline="PRIMARY SCHOOL – MASAJJA",
        motto="Strive for the best",
        phones="+256781757410, +256704720641, +256788529084",
        pobox="P.O Box 31882 Kampala"
    )

    return dict(
        school=school,
        student=student,
        term=term, year=year,
        rows=rows,
        total_sum=total_sum,
        avg_overall=avg_overall,
        aggregate=aggregate,
        division=division,
        position=position,
        class_size=class_size,
        midterms=midterms,
        midterm_subjects=midterm_subjects,
        payment_number=payment_number,
        comments={"teacher_comment": teacher_comment, "head_comment": head_comment},
        next_term_info=new_namedtuple_like(next_term_info), # tiny helper below
        grading=grading,
    )

def new_namedtuple_like(d):
    """Allow dot-access in Jinja for next_term_info even when None/dict."""
    if not d: 
        return None
    return type("NTI", (), d)

# ---------- batch print ----------
@app.route("/report_card/print_batch", methods=["POST"])
@require_role('admin','headteacher','bursar','dos')
def report_card_print_batch():
    class_name = (request.form.get("class_name") or "").strip()
    term = (request.form.get("term") or "").strip()
    year = request.form.get("year", type=int)
    include_mid = (request.form.get("include_midterms") == "1")
    # ids from the table
    try:
        ids = [int(x) for x in request.form.getlist("selected_ids") if str(x).strip().isdigit()]
    except Exception:
        ids = []

    if not ids:
        flash("Select at least one student to print.", "warning")
        return redirect(url_for("reports_hub", class_name=class_name, term=term, year=year))

    conn = get_db_connection()
    reports = []
    for sid in ids:
        payload = build_report_payload(conn, sid, term, year, include_mid=include_mid)
        if payload:
            reports.append(payload)
    conn.close()

    if not reports:
        flash("No printable reports found for the selected students.", "warning")
        return redirect(url_for("reports_hub", class_name=class_name, term=term, year=year))

    # Render one long page with page-breaks between reports
    return render_template("report_batch_citizen.html", reports=reports)


@app.route('/view_reports', methods=['GET'])
@require_role('admin', 'headteacher', 'bursar')
def view_reports():
    conn = get_db_connection()
    c = conn.cursor()

    # Filters
    year = request.args.get('year', '')
    term = request.args.get('term', '')
    class_name = request.args.get('class_name', '')
    stream = request.args.get('stream', '')
    subject_id = request.args.get('subject_id', '')

    query = """
        SELECT r.*, s.first_name, s.Middle_name, s.last_name, subj.name AS subject_name, sp.paper_name
        FROM reports r
        JOIN students s ON r.student_id = s.id
        JOIN subjects subj ON r.subject_id = subj.id
        LEFT JOIN subject_papers sp ON r.subject_part_id = sp.id
        WHERE 1=1
    """
    params = []

    if year:
        query += " AND r.year = ?"
        params.append(year)
    if term:
        query += " AND r.term = ?"
        params.append(term)
    if class_name:
        query += " AND r.class_name = ?"
        params.append(class_name)
    if stream:
        query += " AND r.stream = ?"
        params.append(stream)
    if subject_id:
        query += " AND r.subject_id = ?"
        params.append(subject_id)

    query += " ORDER BY r.class_name, s.last_name"

    results = c.execute(query, params).fetchall()

    years = c.execute("SELECT DISTINCT year FROM reports ORDER BY year DESC").fetchall()
    subjects = c.execute("SELECT * FROM subjects ORDER BY name").fetchall()
    conn.close()

    return render_template("view_reports.html", reports=results,
                           years=years, subjects=subjects,
                           selected_year=year, selected_term=term,
                           selected_class=class_name, selected_stream=stream,
                           selected_subject=subject_id)



@app.route('/export_reports')
@require_role('admin', 'headteacher')
def export_reports():
    year = request.args.get('year')
    term = request.args.get('term')

    if not year or not term:
        flash("Year and term are required to export reports.", "warning")
        return redirect(url_for('view_scores'))

    conn = get_db_connection()
    c = conn.cursor()

    query = """
        SELECT r.*, s.first_name, s.Middle_name, s.last_name,
               sub.name AS subject_name, sp.paper_name
        FROM reports r
        JOIN students s ON r.student_id = s.id
        JOIN subjects sub ON r.subject_id = sub.id
        LEFT JOIN subject_papers sp ON r.subject_part_id = sp.id
        WHERE r.year = ? AND r.term = ?
        ORDER BY s.last_name, sub.name
    """
    rows = c.execute(query, (year, term)).fetchall()
    headers = [desc[0] for desc in c.description]

    conn.close()

    # Generate Excel
    from io import BytesIO
    import xlsxwriter

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Reports")

    # Write header
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    # Write data rows
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, val in enumerate(row):
            worksheet.write(row_idx, col_idx, val)

    workbook.close()
    output.seek(0)

    filename = f"Reports_{term}_{year}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)




@app.route('/edit_report/<int:student_id>', methods=['GET','POST'])
@require_role('admin','dos','headteacher')
def edit_report(student_id):
    term = request.values.get('term', 'Term 1')
    year = int(request.values.get('year', datetime.now().year))
    class_id = request.args.get('class_id')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT id, first_name, Middle_name, last_name, student_number FROM students WHERE id = ?", (student_id,))
    s = cur.fetchone()
    if not s:
        conn.close(); abort(404)

    full_name = " ".join([x for x in [s["first_name"], s["Middle_name"], s["last_name"]] if x])

    cur.execute("""
        SELECT teacher_comment, head_comment
        FROM report_comments
        WHERE student_id = ? AND term = ? AND year = ?
        LIMIT 1
    """, (student_id, term, year))
    row = cur.fetchone() or {}
    teacher_comment = row.get("teacher_comment") if isinstance(row, dict) else (row["teacher_comment"] if row else "")
    head_comment = row.get("head_comment") if isinstance(row, dict) else (row["head_comment"] if row else "")

    cur.execute("""
        SELECT next_term, next_term_date
        FROM term_dates
        WHERE year = ? AND term = ?
        LIMIT 1
    """, (year, term))
    td = cur.fetchone() or {}
    next_term = td.get("next_term") if isinstance(td, dict) else (td["next_term"] if td else "Next Term")
    next_term_date = td.get("next_term_date") if isinstance(td, dict) else (td["next_term_date"] if td else None)

    if request.method == 'POST':
        teacher_comment = (request.form.get('teacher_comment') or '').strip()
        head_comment = (request.form.get('head_comment') or '').strip()
        next_term = (request.form.get('next_term') or next_term or 'Next Term').strip()
        next_term_date = request.form.get('next_term_date') or None # 'YYYY-MM-DD'

        # Upsert report_comments
        cur.execute("""
            INSERT INTO report_comments (student_id, term, year, teacher_comment, head_comment)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(student_id, term, year) DO UPDATE SET
              teacher_comment = excluded.teacher_comment,
              head_comment = excluded.head_comment
        """, (student_id, term, year, teacher_comment, head_comment))

        # Upsert term_dates
        cur.execute("""
            INSERT INTO term_dates (year, term, next_term, next_term_date)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(year, term) DO UPDATE SET
              next_term = excluded.next_term,
              next_term_date = excluded.next_term_date
        """, (year, term, next_term, next_term_date))

        conn.commit()
        conn.close()
        flash('Report updated.', 'success')
        return redirect(url_for('report_card_list', term=term, year=year, class_id=class_id))

    conn.close()
    return render_template('edit_report.html',
                           student={'id': student_id, 'name': full_name, 'number': s['student_number']},
                           term=term, year=year,
                           teacher_comment=teacher_comment, head_comment=head_comment,
                           next_term=next_term, next_term_date=next_term_date)


@app.route('/academic_years', methods=['GET', 'POST'])
@require_role('admin')
def academic_years():
    conn = get_db_connection()
    c = conn.cursor()

    if request.method == 'POST':
        if 'add' in request.form:
            year = request.form['year']
            try:
                c.execute("INSERT INTO academic_years (year) VALUES (?)", (year,))
                conn.commit()
            except:
                flash("Year already exists", "warning")

        elif 'activate' in request.form:
            selected_year = request.form['year']
            c.execute("UPDATE academic_years SET is_active = 0")
            c.execute("UPDATE academic_years SET is_active = 1 WHERE year = ?", (selected_year,))
            conn.commit()

        elif 'set_term' in request.form:
            term = request.form['term']
            c.execute("UPDATE academic_years SET current_term = ? WHERE is_active = 1", (term,))
            conn.commit()

    years = c.execute("SELECT * FROM academic_years").fetchall()
    active_year = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    conn.close()
    return render_template('academic_years.html', years=years, active_year=active_year)








 




@app.route('/fees/setup', methods=['POST'])
def setup_class_fees():
    data = request.form
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        INSERT INTO class_fees (class_name, level, section, amount)
        VALUES (?, ?, ?, ?)
    ''', (
        data['class_name'],
        data['level'],
        data['section'],
        float(data['amount'])
    ))
    conn.commit()
    conn.close()
    return f"Fees for class_name {data['class_name']} ({data['section']}) set at UGX {data['amount']}."
    





@app.route('/students/retrieve/<int:student_id>', methods=['POST'])
def retrieve_archived_student(student_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('UPDATE students SET archived = 0, status = "active" WHERE id = ?', (student_id,))
    conn.commit()
    conn.close()
    return f"Student {student_id} retrieved from archive."


@app.route('/students/drop/<int:student_id>', methods=['POST'])
def drop_student(student_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('UPDATE students SET status = "left" WHERE id = ?', (student_id,))
    conn.commit()
    conn.close()
    return f"Student {student_id} marked as left."


@app.route('/students/archive/clear', methods=['POST'])
def clear_archived_students():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('DELETE FROM students WHERE archived = 1')
    conn.commit()
    conn.close()
    return "All archived students cleared permanently."


@app.route('/expense/categories/init')
def init_expense_categories():
    categories = ['Staff Pay', 'Transport', 'Uniforms', 'Secretarial', 'Service Providers', 'Others']
    conn = get_db_connection()
    c = conn.cursor()
    for name in categories:
        c.execute('INSERT OR IGNORE INTO expense_categories (name) VALUES (?)', (name,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Default categories initialized'})



@app.route('/add_expense', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def add_expense():
    conn = get_db_connection()
    c = conn.cursor()

    current = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    if not current:
        flash("No active academic year. Please activate one first.", "warning")
        return redirect(url_for('academic_years'))

    active_year = current['year']
    current_term = current['current_term']
    categories = c.execute("SELECT id, name from expense_categories").fetchall()

    if request.method == 'POST':
        description = request.form['description']
        amount = request.form['amount']
        category_id = request.form['category_id']
        type_ = request.form.get('type', 'other')
        date_spent = request.form.get('date_spent') or datetime.now().strftime('%Y-%m-%d')
        recorded_by = session.get('username', 'System')

        try:
            amount = float(amount)
            c.execute('''
                INSERT INTO expenses (
                    description, amount, date_spent, category_id,
                    type, recorded_by, term, year
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (description, amount, date_spent, category_id, type_, recorded_by, current_term, active_year))
            conn.commit()
            flash("Expense added successfully.", "success")
            return redirect(url_for('expenditure_report'))
        except Exception as e:
            flash(f"Error: {e}", "danger")

    conn.close()
    return render_template('add_expense.html', categories=categories)




@app.route('/expenses')
def list_expenses():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        SELECT e.id, e.title, e.description, e.amount, e.date, 
               ec.name AS category, e.recorded_by
        FROM expenses e
        JOIN expense_categories ec ON e.category_id = ec.id
        ORDER BY e.date DESC
    ''')
    expenses = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(expenses)
  
  



@app.route("/export_expenses")
@require_role("admin", "headteacher", "dos")
def export_expenses():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT e.id, e.date, e.category, e.amount, e.description,
               u.username AS entered_by
        FROM expenses e
        LEFT JOIN users u ON u.id = e.user_id
        ORDER BY e.date DESC
    """).fetchall()
    conn.close()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Headers
    headers = ["ID", "Date", "Category", "Amount", "Description", "Entered By"]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Data rows
    for r in rows:
        ws.append([r["id"], r["date"], r["category"], r["amount"], r["description"], r["entered_by"]])

    # Auto column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    # Stream to browser
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="expenses.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route('/expenses/filter', methods=['GET'])
def filter_expenses_by_date():
    start_date = request.args.get('start')
    end_date = request.args.get('end')

    if not start_date or not end_date:
        return jsonify({'error': 'Start and end dates are required'}), 400

    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        SELECT e.id, e.title, e.description, e.amount, e.date,
               ec.name AS category, e.recorded_by
        FROM expenses e
        JOIN expense_categories ec ON e.category_id = ec.id
        WHERE e.date BETWEEN ? AND ?
        ORDER BY e.date DESC
    ''', (start_date, end_date))

    results = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(results)



@app.route('/pay_teacher', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def pay_teacher():
    conn = get_db_connection()
    c = conn.cursor()

    teachers = []
    searched = False
    payment_history = []
    current_year = datetime.now().year

    term_row = c.execute(
        "SELECT current_term FROM academic_years WHERE is_active = 1"
    ).fetchone()
    current_term = term_row['current_term'] if term_row else 'N/A'

    if request.method == 'POST':
        if 'search' in request.form:
            search_term = request.form['search_term']
            teachers = c.execute("""
                SELECT * FROM teachers
                WHERE id LIKE ? OR last_name LIKE ?
            """, (f"%{search_term}%", f"%{search_term}%")).fetchall()
            searched = True

        elif 'pay_teacher' in request.form:
            teacher_id = request.form['teacher_id']
            amount = float(request.form['amount_paid'])
            term = request.form['term']
            year = int(request.form['year'])
            recorded_by = session.get('full_name', 'Unknown')

            c.execute("""
                INSERT INTO expenses (
                    description, amount, term, year,
                    category_id, recorded_by, type
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                f"Payment to Teacher ID {teacher_id}",
                amount,
                term,
                year,
                1, # Assume 1 is the staff payment category
                recorded_by,
                'staff_pay'
            ))
            conn.commit()
            flash("Payment recorded successfully.", "success")
            return redirect(url_for('pay_teacher'))

    payment_history = c.execute("""
        SELECT date_spent, amount, term, year, recorded_by,
               REPLACE(description, 'Payment to ', '') AS teacher_name
        FROM expenses
        WHERE type = 'staff_pay'
        ORDER BY date_spent DESC
        LIMIT 10
    """).fetchall()

    conn.close()
    return render_template(
        'pay_teacher.html',
        teachers=teachers,
        searched=searched,
        payment_history=payment_history,
        current_term=current_term,
        current_year=current_year
    )


@app.route('/export_teacher_payments')
@require_role('admin', 'bursar')
def export_teacher_payments():
    conn = get_db_connection()
    c = conn.cursor()

    data = c.execute("""
        SELECT date_spent, amount, term, year, recorded_by,
               REPLACE(description, 'Payment to ', '') AS teacher_name
        FROM expenses
        WHERE type = 'staff_pay'
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Payments"

    headers = ['Date', 'Teacher', 'Amount', 'Term', 'Year', 'Recorded By']
    ws.append(headers)

    for row in data:
        ws.append([
            row['date_spent'],
            row['teacher_name'],
            row['amount'],
            row['term'],
            row['year'],
            row['recorded_by']
        ])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="teacher_payments.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )




# ---------- TEACHERS MODULE ----------
# Put this near your other routes (after get_db_connection, require_role, etc.)

def _teacher_classes_list():
    # Baby, Middle, Top, P1..P7
    return ["Baby","Middle","Top","P1","P2","P3","P4","P5","P6","P7"]

def _streams_list(conn):
    rows = conn.execute("SELECT name FROM streams ORDER BY name").fetchall()
    return [r["name"] for r in rows] or ["A"]

def _subjects_list(conn):
    return conn.execute("SELECT id, name, COALESCE(code,'') AS code FROM subjects ORDER BY name").fetchall()

def _employees_list(conn):
    return conn.execute("""
        SELECT id, first_name, COALESCE(Middle_name,'') AS Middle_name, last_name, designation, status
        FROM employees
        ORDER BY (status='active') DESC, last_name, first_name
    """).fetchall()


# =========================
# TEACHERS: routes (CRUD + subject/class assignments)
# =========================

# ---------- TEACHER MANAGEMENT (no login/user creation here) ----------

def _get_employees_without_teacher(conn):
    return conn.execute("""
        SELECT e.id, e.first_name, e.middle_name, e.last_name, e.designation
        FROM employees e
        LEFT JOIN teachers t ON t.employee_id = e.id
        WHERE t.id IS NULL AND e.status='active'
        ORDER BY e.last_name, e.first_name
    """).fetchall()

def _get_all_subjects(conn):
    return conn.execute("SELECT id, name, COALESCE(code,'') AS code FROM subjects ORDER BY name").fetchall()

def _get_all_classes(conn):
    return conn.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name").fetchall()

@app.route("/teachers", methods=["GET","POST"])
@require_role("admin","director","headteacher","dos")
def teachers_hub():
    conn = get_db_connection()
    c = conn.cursor()

    # Add teacher (links to an existing employee only)
    if request.method == "POST" and request.form.get("action") == "create":
        try:
            employee_id = int(request.form["employee_id"])
        except Exception:
            conn.close()
            flash("Select a valid employee.", "warning")
            return redirect(url_for("teachers_hub"))

        initials = (request.form.get("initials") or "").strip().upper()
        if not initials:
            conn.close()
            flash("Initials are required.", "warning")
            return redirect(url_for("teachers_hub"))

        # create teacher row
        c.execute("""
            INSERT INTO teachers (employee_id, initials, status)
            VALUES (?, ?, 'active')
        """, (employee_id, initials))
        teacher_id = c.lastrowid

        # multi-assign subjects/classes
        sel_subjects = request.form.getlist("subjects[]") # list of subject_id (str)
        sel_classes = request.form.getlist("classes[]") # list of class_name (str)
        pairs = []
        for sid in sel_subjects:
            for cn in sel_classes:
                pairs.append((teacher_id, int(sid), cn))
        if pairs:
            c.executemany("""
                INSERT INTO teacher_subjects (teacher_id, subject_id, class_name)
                VALUES (?,?,?)
            """, pairs)

        conn.commit()
        conn.close()
        flash("Teacher created & assignments saved.", "success")
        return redirect(url_for("teachers_hub"))

    # List teachers
    rows = c.execute("""
        SELECT t.id, t.initials, t.status, e.first_name, e.middle_name, e.last_name, e.designation
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        ORDER BY (t.status='active') DESC, e.last_name, e.first_name
    """).fetchall()

    employees_free = _get_employees_without_teacher(conn)
    subjects = _get_all_subjects(conn)
    classes = _get_all_classes(conn)
    conn.close()

    return render_template(
        "teachers.html",
        teachers=rows,
        employees_free=employees_free,
        subjects=subjects,
        classes=[r["class_name"] for r in classes]
    )

@app.route("/teachers/<int:tid>/edit", methods=["GET","POST"])
@require_role("admin","director","headteacher","dos")
def edit_teacher(tid):
    conn = get_db_connection()
    c = conn.cursor()

    teacher = c.execute("""
        SELECT t.*, e.first_name, e.middle_name, e.last_name, e.designation
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        WHERE t.id=?
    """, (tid,)).fetchone()
    if not teacher:
        conn.close()
        flash("Teacher not found.", "warning")
        return redirect(url_for("teachers_hub"))

    if request.method == "POST":
        # update initials/status
        initials = (request.form.get("initials") or "").strip().upper()
        status = request.form.get("status") or "active"
        c.execute("UPDATE teachers SET initials=?, status=? WHERE id=?", (initials, status, tid))

        # replace assignments
        c.execute("DELETE FROM teacher_subjects WHERE teacher_id=?", (tid,))
        sel_subjects = request.form.getlist("subjects[]")
        sel_classes = request.form.getlist("classes[]")
        pairs = []
        for sid in sel_subjects:
            for cn in sel_classes:
                pairs.append((tid, int(sid), cn))
        if pairs:
            c.executemany("""
                INSERT INTO teacher_subjects (teacher_id, subject_id, class_name)
                VALUES (?,?,?)
            """, pairs)

        conn.commit()
        conn.close()
        flash("Teacher updated.", "success")
        return redirect(url_for("teachers_hub"))

    # Preload current selections
    cur = c.execute("""
        SELECT subject_id, class_name
        FROM teacher_subjects
        WHERE teacher_id=?
    """, (tid,)).fetchall()
    chosen_subjects = {r["subject_id"] for r in cur}
    chosen_classes = {r["class_name"] for r in cur}

    subjects = _get_all_subjects(conn)
    classes = [r["class_name"] for r in _get_all_classes(conn)]
    conn.close()

    return render_template(
        "teacher_edit.html",
        teacher=teacher,
        subjects=subjects,
        classes=classes,
        chosen_subjects=chosen_subjects,
        chosen_classes=chosen_classes
    )

@app.route("/teachers/<int:tid>/delete", methods=["POST"])
@require_role("admin","director","headteacher","dos")
def delete_teacher(tid):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM teacher_subjects WHERE teacher_id=?", (tid,))
    c.execute("DELETE FROM teachers WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    flash("Teacher removed.", "info")
    return redirect(url_for("teachers_hub"))
    

from io import StringIO
import csv
from flask import Response

@app.route("/teachers/load", methods=["GET"])
@require_role("admin","director","headteacher","dos","bursar")
def teachers_load():
    """
    Summary view of each teacher's assignments: subjects ↔ classes.
    Optional filters: q (name), class_name, subject_id, status.
    """
    conn = get_db_connection()
    c = conn.cursor()

    q = (request.args.get("q") or "").strip()
    class_f = request.args.get("class_name") or ""
    subj_f = request.args.get("subject_id") or ""
    status_f = request.args.get("status") or "" # active/archived

    where = []
    args = []

    if q:
        where.append("(e.first_name LIKE ? OR e.middle_name LIKE ? OR e.last_name LIKE ?)")
        args += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if class_f:
        where.append("ts.class_name = ?")
        args.append(class_f)
    if subj_f:
        where.append("ts.subject_id = ?")
        args.append(subj_f)
    if status_f:
        where.append("t.status = ?")
        args.append(status_f)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    # one row per teacher, with grouped subjects/classes
    rows = c.execute(f"""
        SELECT
          t.id AS teacher_id,
          t.initials,
          t.status,
          e.first_name, e.middle_name, e.last_name, e.designation,
          -- group subjects: "Math(P5), English(P6), Science(P7)"
          GROUP_CONCAT(DISTINCT s.name || COALESCE(' ('||ts.class_name||')','') ORDER BY s.name, ts.class_name) AS load_list
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        LEFT JOIN teacher_subjects ts ON ts.teacher_id = t.id
        LEFT JOIN subjects s ON s.id = ts.subject_id
        {where_sql}
        GROUP BY t.id, t.initials, t.status, e.first_name, e.middle_name, e.last_name, e.designation
        ORDER BY (t.status='active') DESC, e.last_name, e.first_name
    """, args).fetchall()

    # filter dropdown data
    classes = c.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name").fetchall()
    subjects = c.execute("SELECT id, name, COALESCE(code,'') AS code FROM subjects ORDER BY name").fetchall()
    conn.close()

    return render_template(
        "teachers_load.html",
        rows=rows,
        classes=[r["class_name"] for r in classes],
        subjects=subjects,
        q=q, class_f=class_f, subj_f=subj_f, status_f=status_f
    )


@app.route("/teachers/load/export")
@require_role("admin","director","headteacher","dos","bursar")
def teachers_load_export():
    """
    Export the CURRENT filtered list as CSV.
    Uses same filters as /teachers/load.
    """
    conn = get_db_connection()
    c = conn.cursor()

    q = (request.args.get("q") or "").strip()
    class_f = request.args.get("class_name") or ""
    subj_f = request.args.get("subject_id") or ""
    status_f = request.args.get("status") or ""

    where = []
    args = []

    if q:
        where.append("(e.first_name LIKE ? OR e.middle_name LIKE ? OR e.last_name LIKE ?)")
        args += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if class_f:
        where.append("ts.class_name = ?")
        args.append(class_f)
    if subj_f:
        where.append("ts.subject_id = ?")
        args.append(subj_f)
    if status_f:
        where.append("t.status = ?")
        args.append(status_f)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    data = c.execute(f"""
        SELECT
          e.last_name || ', ' || e.first_name || COALESCE(' '||e.middle_name,'') AS teacher_name,
          t.initials,
          e.designation,
          t.status,
          GROUP_CONCAT(DISTINCT s.name || COALESCE(' ('||ts.class_name||')','') ORDER BY s.name, ts.class_name) AS load_list
        FROM teachers t
        JOIN employees e ON e.id = t.employee_id
        LEFT JOIN teacher_subjects ts ON ts.teacher_id = t.id
        LEFT JOIN subjects s ON s.id = ts.subject_id
        {where_sql}
        GROUP BY t.id
        ORDER BY (t.status='active') DESC, e.last_name, e.first_name
    """, args).fetchall()
    conn.close()

    # Build CSV in-memory
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(["Teacher", "Initials", "Designation", "Status", "Subjects (Class)"])
    for r in data:
        writer.writerow([
            r["teacher_name"],
            r["initials"] or "",
            r["designation"] or "",
            r["status"] or "",
            r["load_list"] or "",
        ])

    output = si.getvalue()
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=teacher_load.csv"}
    )




@app.route('/expenses/total')
def total_expense():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT SUM(amount) AS total_expense FROM expenses')
    row = c.fetchone()
    conn.close()
    return jsonify({'total_expense': row['total_expense'] or 0})
    


@app.route('/grading_scale', methods=['GET', 'POST'])
@require_role('admin', 'headteacher')
def grading_scale():
    conn = get_db_connection()
    c = conn.cursor()

    if request.method == 'POST':
        grade = request.form['grade']
        lower = int(request.form['lower_limit'])
        upper = int(request.form['upper_limit'])
        comment = request.form.get('comment', '')

        # Check for overlapping ranges
        overlapping = c.execute('''
            SELECT * FROM grading_scale
            WHERE NOT (? < lower_limit OR ? > upper_limit)
        ''', (upper, lower)).fetchall()

        if overlapping:
            flash("Grade range overlaps with existing entries.", "danger")
        else:
            c.execute('''
                INSERT INTO grading_scale (grade, lower_limit, upper_limit, comment)
                VALUES (?, ?, ?, ?)
            ''', (grade, lower, upper, comment))
            conn.commit()
            flash("Grade added successfully", "success")

    grades = c.execute('SELECT * FROM grading_scale ORDER BY lower_limit ASC').fetchall()
    conn.close()
    return render_template('grading_scale.html', grades=grades)

 
@app.route('/edit_grade/<int:grade_id>', methods=['GET', 'POST'])
@require_role('admin', 'headteacher')
def edit_grade(grade_id):
    conn = get_db_connection()
    c = conn.cursor()

    if request.method == 'POST':
        grade = request.form['grade']
        lower = int(request.form['lower_limit'])
        upper = int(request.form['upper_limit'])
        comment = request.form.get('comment', '')

        c.execute('''
            UPDATE grading_scale SET grade = ?, lower_limit = ?, upper_limit = ?, comment = ?
            WHERE id = ?
        ''', (grade, lower, upper, comment, grade_id))
        conn.commit()
        conn.close()
        flash("Grade updated successfully.", "success")
        return redirect(url_for('grading_scale'))

    grade_data = c.execute('SELECT * FROM grading_scale WHERE id = ?', (grade_id,)).fetchone()
    conn.close()
    return render_template('edit_grade.html', grade=grade_data)
    

@app.route('/delete_grade/<int:grade_id>')
@require_role('admin')
def delete_grade(grade_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('DELETE FROM grading_scale WHERE id = ?', (grade_id,))
    conn.commit()
    conn.close()
    flash("Grade deleted.", "info")
    return redirect(url_for('grading_scale'))
   


@app.route('/export_grading_scale')
@require_role('admin', 'headteacher')
def export_grading_scale():
    conn = get_db_connection()
    c = conn.cursor()
    data = c.execute("SELECT * FROM grading_scale ORDER BY lower_limit ASC").fetchall()
    conn.close()

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Grading Scale')

    headers = ['Grade', 'Lower Limit', 'Upper Limit', 'Comment']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    for row_num, row in enumerate(data, start=1):
        worksheet.write(row_num, 0, row['grade'])
        worksheet.write(row_num, 1, row['lower_limit'])
        worksheet.write(row_num, 2, row['upper_limit'])
        worksheet.write(row_num, 3, row['comment'] or '')

    workbook.close()
    output.seek(0)

    return send_file(output, download_name='grading_scale.xlsx', as_attachment=True)

   


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))



# ---------- AUDIT TRAIL (admin only) ----------

@app.route('/audit-trail')
@require_role('admin')
def audit_trail():
    import os, sqlite3
    from math import ceil
    from urllib.parse import urlencode

    q_user = request.args.get('user_id', type=int)
    q_role = (request.args.get('role') or '').strip() or None
    q_act = (request.args.get('action') or '').strip() or None
    q_out = (request.args.get('outcome')or '').strip() or None
    q_route = (request.args.get('route') or '').strip() or None
    page = max(request.args.get('page', 1, type=int), 1)
    per = min(max(request.args.get('per', 50, type=int), 10), 200)
    offset = (page - 1) * per

    # Choose display offset (default +03:00). Set AUDIT_TZ_OFFSET in .env if needed.
    tz_off = os.getenv("AUDIT_TZ_OFFSET", "+03:00")
    if not tz_off or len(tz_off) < 3:
        tz_off = "+00:00"

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    ensure_audit_trail_schema(conn)
    c = conn.cursor()

    where, params = [], []
    if q_user: where.append("a.user_id = ?"); params.append(q_user)
    if q_role: where.append("a.role = ?"); params.append(q_role)
    if q_act: where.append("a.action LIKE ?"); params.append(f"%{q_act}%")
    if q_out: where.append("a.outcome = ?"); params.append(q_out)
    if q_route: where.append("a.route LIKE ?"); params.append(f"%{q_route}%")
    where_sql = "WHERE " + " AND ".join(where) if where else ""

    total = c.execute(f"SELECT COUNT(*) AS n FROM audit_trail a {where_sql}", params).fetchone()["n"]

    # Render timestamp with explicit SQL offset so it's correct even if OS/localtime is off.
    rows = c.execute(f"""
        SELECT
            strftime('%Y-%m-%d %H:%M:%S', datetime(a.timestamp, '{tz_off}')) AS timestamp,
            a.id, a.user_id, a.role, a.action, a.outcome, a.severity,
            a.route, a.method, a.ip_address, a.target_table, a.target_id,
            a.details_json, a.http_status,
            COALESCE(
              TRIM(
                COALESCE(e.first_name,'') || ' ' ||
                COALESCE(NULLIF(e.Middle_name,''),'') ||
                CASE WHEN COALESCE(NULLIF(e.Middle_name,''),'') <> '' THEN ' ' ELSE '' END ||
                COALESCE(e.last_name,'')
              ),
              u.username, '—'
            ) AS user_name
        FROM audit_trail a
        LEFT JOIN users u ON u.id = a.user_id
        LEFT JOIN employees e ON e.id = u.employee_id
        {where_sql}
        ORDER BY a.timestamp DESC, a.id DESC
        LIMIT ? OFFSET ?
    """, (*params, per, offset)).fetchall()
    conn.close()

    pages = max(ceil(total / per), 1)
    q = {"user_id": q_user, "role": q_role, "action": q_act, "outcome": q_out, "route": q_route}
    querystring = urlencode({k: v for k, v in q.items() if v not in (None, "",)})

    return render_template(
        "audit_trail.html",
        audit_logs=[dict(r) for r in rows],
        total=total, page=page, per=per, pages=pages,
        querystring=querystring, q=q
    )



@app.route('/asset_register', methods=['GET', 'POST'])
@require_role('admin')
def asset_register():
    conn = get_db_connection()
    c = conn.cursor()

    from_date = to_date = asset_name = ''
    assets = []

    if request.method == 'POST':
        if 'add_asset' in request.form:
            name = request.form['name']
            model = request.form['model']
            year_purchased = request.form['year_purchased']
            condition = request.form['condition']
            value = request.form['value']
            qty = request.form['qty']

            c.execute('''
                INSERT INTO assets (name, model, year_purchased, condition, value, qty)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (name, model, year_purchased, condition, value, qty))
            conn.commit()

        elif 'filter' in request.form:
            from_date = request.form['from_date']
            to_date = request.form['to_date']
            asset_name = request.form['asset_name']

            query = "SELECT * FROM assets WHERE 1=1"
            params = []

            if from_date and to_date:
                query += " AND date(year_purchased) BETWEEN ? AND ?"
                params += [from_date, to_date]

            if asset_name:
                query += " AND LOWER(name) LIKE ?"
                params.append(f"%{asset_name.lower()}%")

            assets = c.execute(query, params).fetchall()

    if not assets:
        assets = c.execute("SELECT * FROM assets ORDER BY year_purchased DESC").fetchall()

    conn.close()
    return render_template('asset_register.html',
                           assets=assets,
                           from_date=from_date,
                           to_date=to_date,
                           asset_name=asset_name)



@app.route('/edit_asset/<int:asset_id>', methods=['GET', 'POST'])
@require_role('admin')
def edit_asset(asset_id):
    conn = get_db_connection()
    c = conn.cursor()

    # Fetch asset
    asset = c.execute("SELECT * FROM assets WHERE id = ?", (asset_id,)).fetchone()
    if not asset:
        flash("Asset not found.", "danger")
        return redirect(url_for('asset_register'))

    if request.method == 'POST':
        name = request.form['name']
        model = request.form['model']
        year_purchased = request.form['year_purchased']
        condition = request.form['condition']
        value = request.form['value']
        qty = request.form['qty']

        c.execute('''
            UPDATE assets
            SET name = ?, model = ?, year_purchased = ?, condition = ?, value = ?, qty = ?
            WHERE id = ?
        ''', (name, model, year_purchased, condition, value, qty, asset_id))
        conn.commit()
        conn.close()
        flash("Asset updated successfully.", "success")
        return redirect(url_for('asset_register'))

    conn.close()
    return render_template('edit_asset.html', asset=asset)



@app.route('/export_assets')
@require_role('admin')
def export_assets():
    conn = get_db_connection()
    df = pd.read_sql_query("SELECT name, model, year_purchased, condition, value, qty FROM assets", conn)
    conn.close()

    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Assets')
    writer.close()
    output.seek(0)

    return send_file(output,
                     download_name='asset_register.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/add_income', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def add_income():
    preview = None
    if request.method == 'POST':
        if 'preview' in request.form:
            # Just return form data to re-render the page with preview
            preview = {
                'source': request.form['source'],
                'amount': request.form['amount'],
                'term': request.form['term'],
                'year': request.form['year'],
                'description': request.form['description'],
                'date_received': request.form['date_received']
            }
        elif 'confirm' in request.form:
            conn = get_db_connection()
            c = conn.cursor()

            source = request.form['source']
            amount = float(request.form['amount'])
            term = request.form['term']
            year = int(request.form['year'])
            description = request.form['description']
            date_received = request.form['date_received']
            recorded_by = session.get('full_name', 'Unknown')

            c.execute('''
                INSERT INTO other_income (source, amount, term, year, description, recorded_by, date_received)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (source, amount, term, year, description, recorded_by, date_received))

            conn.commit()
            conn.close()
            flash("Income successfully recorded", "success")
            return redirect(url_for('add_income'))

    return render_template('add_income.html', preview=preview)



@app.route('/income_statement', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def income_statement():
    conn = get_db_connection()
    c = conn.cursor()

    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    query_params = ()
    date_filter = ""
    if start_date and end_date:
        date_filter = "AND date_paid BETWEEN ? AND ?"
        query_params = (start_date, end_date)

    # Income from fees
    fee_income = c.execute(f'''
        SELECT SUM(amount_paid) as total
        FROM fees
        WHERE payment_type = 'school_fees' {date_filter}
    ''', query_params).fetchone()['total'] or 0

    # Income from requirements
    req_income = c.execute(f'''
        SELECT SUM(amount_paid) as total
        FROM fees
        WHERE payment_type NOT IN ('school_fees') {date_filter}
    ''', query_params).fetchone()['total'] or 0

    # Other income
    other_income = c.execute(f'''
        SELECT SUM(amount) as total
        FROM other_income
        WHERE date_received IS NOT NULL {f"AND date_received BETWEEN ? AND ?" if date_filter else ""}
    ''', query_params).fetchone()['total'] or 0

    # Expenses
    expenses = c.execute(f'''
        SELECT SUM(amount) as total
        FROM expenses
        WHERE date_spent IS NOT NULL {f"AND date_spent BETWEEN ? AND ?" if date_filter else ""}
    ''', query_params).fetchone()['total'] or 0

    net_income = (fee_income + req_income + other_income) - expenses

    conn.close()
    return render_template(
        'income_statement.html',
        start_date=start_date,
        end_date=end_date,
        fee_income=fee_income,
        req_income=req_income,
        other_income=other_income,
        total_income=fee_income + req_income + other_income,
        expenses=expenses,
        net_income=net_income
    )



@app.route('/income_statement/export')
@require_role('admin', 'bursar')
def export_income_statement():
    import io
    import csv
    from flask import make_response, request

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    c = conn.cursor()

    date_filter = ""
    params = ()
    if start_date and end_date:
        date_filter = "AND date_paid BETWEEN ? AND ?"
        params = (start_date, end_date)

    fee_income = c.execute(f'''
        SELECT SUM(amount_paid) FROM fees
        WHERE payment_type = 'school_fees' {date_filter}
    ''', params).fetchone()[0] or 0

    req_income = c.execute(f'''
        SELECT SUM(amount_paid) FROM fees
        WHERE payment_type NOT IN ('school_fees') {date_filter}
    ''', params).fetchone()[0] or 0

    other_income = c.execute(f'''
        SELECT SUM(amount) FROM other_income
        WHERE date_received IS NOT NULL {f"AND date_received BETWEEN ? AND ?" if date_filter else ""}
    ''', params).fetchone()[0] or 0

    expenses = c.execute(f'''
        SELECT SUM(amount) FROM expenses
        WHERE date_spent IS NOT NULL {f"AND date_spent BETWEEN ? AND ?" if date_filter else ""}
    ''', params).fetchone()[0] or 0

    net = (fee_income + req_income + other_income) - expenses
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Income Statement'])
    writer.writerow(['From:', start_date or '', 'To:', end_date or ''])
    writer.writerow([])
    writer.writerow(['Category', 'Amount (UGX)'])
    writer.writerow(['School Fees', fee_income])
    writer.writerow(['Requirements', req_income])
    writer.writerow(['Other Income', other_income])
    writer.writerow(['Total Income', fee_income + req_income + other_income])
    writer.writerow(['Expenses', expenses])
    writer.writerow(['Net Income', net])

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=income_statement.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response




@app.route('/income_report/export')
@require_role('admin', 'bursar', 'headteacher')
def export_income_report():
    import io
    import csv
    from flask import make_response, request

    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    term = request.args.get('term', '').strip()

    conn = get_db_connection()
    c = conn.cursor()

    # Fees
    fees_query = '''
        SELECT student_id, amount_paid, date_paid
        FROM fees
        WHERE payment_type = 'school_fees' AND date_paid BETWEEN ? AND ?
    '''
    reqs_query = '''
        SELECT student_id, payment_type, amount_paid, date_paid
        FROM fees
        WHERE payment_type NOT IN ('school_fees') AND date_paid BETWEEN ? AND ?
    '''
    params = (from_date, to_date)

    if term:
        fees_query += ' AND term = ?'
        reqs_query += ' AND term = ?'
        params += (term,)

    fees = c.execute(fees_query, params).fetchall()
    requirements = c.execute(reqs_query, params).fetchall()

    other_income = c.execute('''
        SELECT source, amount, recorded_by, date_received
        FROM other_income
        WHERE date_received BETWEEN ? AND ?
    ''', (from_date, to_date)).fetchall()

    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Income Report'])
    writer.writerow(['From:', from_date, 'To:', to_date, 'Term:', term or 'All'])
    writer.writerow([])

    writer.writerow(['School Fees'])
    writer.writerow(['Student ID', 'Amount Paid', 'Date'])
    for f in fees:
        writer.writerow([f['student_id'], f['amount_paid'], f['date_paid']])
    writer.writerow([])

    writer.writerow(['Requirement Payments'])
    writer.writerow(['Student ID', 'Item', 'Amount Paid', 'Date'])
    for r in requirements:
        writer.writerow([r['student_id'], r['payment_type'], r['amount_paid'], r['date_paid']])
    writer.writerow([])

    writer.writerow(['Other Income'])
    writer.writerow(['Source', 'Amount', 'Recorded By', 'Date'])
    for o in other_income:
        writer.writerow([o['source'], o['amount'], o['recorded_by'], o['date_received']])

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=income_report.csv"
    response.headers["Content-type"] = "text/csv"
    return response





@app.route('/income_report', methods=['GET', 'POST'])
@require_role('admin', 'bursar', 'headteacher')
def income_report():
    conn = get_db_connection()
    c = conn.cursor()

    from_date = to_date = datetime.now().strftime('%Y-%m-%d')
    term = ''
    if request.method == 'POST':
        from_date = request.form['from_date']
        to_date = request.form['to_date']
        term = request.form.get('term', '').strip()
    else:
        term_row = c.execute("SELECT current_term FROM academic_years WHERE is_active = 1").fetchone()
        term = term_row['current_term'] if term_row else ''

    filters = {
        'from_date': from_date,
        'to_date': to_date,
        'term': term
    }

    params = [from_date, to_date]
    term_filter = ''
    if term:
        term_filter = 'AND f.term = ?'
        params.append(term)

    fees = c.execute(f'''
        SELECT f.student_id, s.student_number, s.first_name, s.last_name,
               f.amount_paid, f.date_paid, f.payment_type
        FROM fees f
        JOIN students s ON f.student_id = s.id
        WHERE f.payment_type = 'school_fees' AND f.date_paid BETWEEN ? AND ? {term_filter}
    ''', params).fetchall()

    requirements = c.execute(f'''
        SELECT f.student_id, s.student_number, s.first_name, s.last_name,
               f.amount_paid, f.date_paid, f.payment_type
        FROM fees f
        JOIN students s ON f.student_id = s.id
        WHERE f.payment_type != 'school_fees' AND f.date_paid BETWEEN ? AND ? {term_filter}
    ''', params).fetchall()

    other_income = c.execute('''
        SELECT source, amount, recorded_by, date_received
        FROM other_income
        WHERE date_received BETWEEN ? AND ?
    ''', [from_date, to_date]).fetchall()

    totals = {
        'fees_total': sum(row['amount_paid'] for row in fees),
        'requirements_total': sum(row['amount_paid'] for row in requirements),
        'other_income_total': sum(row['amount'] for row in other_income),
    }
    totals['overall'] = totals['fees_total'] + totals['requirements_total'] + totals['other_income_total']

    conn.close()
    return render_template(
        'view_income.html',
        filters=filters,
        fees=fees,
        requirements=requirements,
        other_income=other_income,
        totals=totals
    )

# ---------- Student Hub ----------
@app.route("/students/hub")
@require_role("admin", "headteacher", "dos", "bursar")
def students_hub():
    return render_template("students_hub.html")




# --- Active Academic Year guard ---------------------------------------------
def _redirect_to_existing(*endpoints):
    """Redirect to the first endpoint that exists; fall back to dashboard."""
    for ep in endpoints:
        if ep in app.view_functions:
            return redirect(url_for(ep))
    return redirect(url_for("dashboard"))

def require_active_academic_year():
    """
    Use like:
        guard = require_active_academic_year()
        if guard: return guard
    Returns None if an active AY exists; otherwise returns a redirect Response.
    """
    try:
        ay = get_active_academic_year() # your existing helper
        year = ay.get("year")
        term = ay.get("current_term") or ay.get("term")
        if year and term:
            return None
    except Exception:
        pass

    flash("Please create/activate an academic year first.", "warning")
    # Try a few likely endpoints you might have; falls back to dashboard.
    return _redirect_to_existing("manage_academic_years", "academic_years", "settings_academic_year", "dashboard")



# ---------- Register Student ----------



@app.route("/register_student", methods=["GET", "POST"])
@require_role("admin", "headteacher", "deputyheadteacher")
def register_student():
    ay = get_active_academic_year() # {"year": 2025, "current_term": "Term 1", "term": "Term 1"}
    active_year = int(ay.get("year"))
    active_term = ay.get("current_term") or ay.get("term") or "Term 1"

    if request.method == "GET":
        # Pull available streams for the dropdown
        conn = get_db_connection()
        try:
            streams = [r[0] for r in conn.execute(
                "SELECT DISTINCT stream FROM classes WHERE stream IS NOT NULL AND TRIM(stream) <> '' ORDER BY stream"
            ).fetchall()] or ["A"] # fallback
        finally:
            conn.close()

        # Render your existing template (unchanged) + streams
        return render_template(
            "register_student.html",
            active_year=active_year,
            active_term=active_term,
            streams=streams,
        )

    # ---- POST ----
    f = request.form

    # Raw values from the form
    raw_first = (f.get("first_name") or "").strip()
    raw_middle = (f.get("Middle_name") or "").strip()
    raw_last = (f.get("last_name") or "").strip()
    raw_sex = (f.get("sex") or "").strip()
    raw_class = (f.get("class_name") or "").strip()
    raw_stream = (f.get("stream") or "").strip()
    raw_section = (f.get("section") or "").strip()

    # Normalize to the exact values your DB expects
    class_name = norm_class(raw_class) # 'Baby'/'Middle'/'Top'/'P1'..'P7' or None
    stream = norm_stream(raw_stream) # 'A' (default) or uppercase value
    section = norm_section(raw_section) # 'Day'/'Boarding' or None
    sex = norm_sex(raw_sex) # 'M'/'F' or None

    # Basic validation (prevents DB CHECK errors)
    if not raw_first or not raw_last:
        flash("First and Last name are required.", "danger")
        return redirect(url_for("register_student"))
    if not class_name:
        flash("Please choose a valid Class (Baby, Middle, Top, P1–P7).", "danger")
        return redirect(url_for("register_student"))
    if not section:
        flash("Please choose Section (Day/Boarding).", "danger")
        return redirect(url_for("register_student"))
    if not sex:
        flash("Please choose Sex (M/F).", "danger")
        return redirect(url_for("register_student"))

    # Defaults from the active academic session if fields are empty
    year_of_joining = int((f.get("year_of_joining") or active_year))
    term_joined = (f.get("term_joined") or active_term)
    date_joined = (f.get("date_joined") or datetime.now().strftime("%Y-%m-%d"))

    conn = get_db_connection()
    try:
        student_number = generate_student_number(conn)
        fees_code = generate_fees_code(conn)

        # Insert using an explicit column list (safe with your schema)
        conn.execute("""
            INSERT INTO students (
                first_name, Middle_name, last_name, sex,
                class_name, stream, section,
                student_number, year_of_joining, term_joined, date_joined,
                fees_code, archived, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 'active')
        """, (
            raw_first, raw_middle, raw_last, sex,
            class_name, stream, section,
            student_number, year_of_joining, term_joined, date_joined,
            fees_code
        ))
        conn.commit()

        flash(f"Student {raw_first} {raw_last} registered (#{student_number}).", "success")
        return redirect(url_for("register_student"))

    except sqlite3.IntegrityError as e:
        conn.rollback()
        flash(f"Failed to register: {e}", "danger")
        return redirect(url_for("register_student"))
    except Exception as e:
        conn.rollback()
        flash(f"Unexpected error: {e}", "danger")
        return redirect(url_for("register_student"))
    finally:
        conn.close()




# ---------- Edit Student ----------
@app.route("/students/<int:student_id>/edit", methods=["GET","POST"])
@require_role("admin", "headteacher", "bursar")
def edit_student(student_id):
    conn = get_db_connection()
    c = conn.cursor()
    classes = [r[0] for r in c.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name").fetchall()]
    streams = [r[0] for r in c.execute("SELECT DISTINCT stream FROM classes ORDER BY stream").fetchall()]
    houses = ["Tiger","Zebra","Eagle","Lion"]

    if request.method == "POST":
        f = request.form
        first_name = (f.get("first_name") or "").strip()
        Middle_name = (f.get("Middle_name") or "").strip()
        last_name = (f.get("last_name") or "").strip()
        sex = norm_sex(f.get("sex"))
        section = norm_section(f.get("section"))
        class_name = norm_class(f.get("class_name"))
        stream = norm_stream(f.get("stream"))
        house = (f.get("house") or "").strip()
        parent_name = (f.get("parent_name") or "").strip()
        parent_contact = (f.get("parent_contact") or "").strip()
        parent2_name = (f.get("parent2_name") or "").strip()
        parent2_contact= (f.get("parent2_contact") or "").strip()
        parent_email = (f.get("parent_email") or "").strip()
        residence = (f.get("residence") or "").strip()
        student_number = (f.get("student_number") or "").strip()
        fees_code = (f.get("fees_code") or "").strip()

        # required checks
        missing = []
        for label, val in [("First name", first_name), ("Last name", last_name),
                           ("Sex", sex), ("Section", section),
                           ("Class", class_name), ("Stream", stream)]:
            if not val: missing.append(label)
        if missing:
            conn.close()
            flash("Missing/invalid fields: " + ", ".join(missing), "warning")
            return redirect(url_for("edit_student", student_id=student_id))

        c.execute("""
            UPDATE students SET
              first_name=?, Middle_name=?, last_name=?,
              sex=?, section=?, class_name=?, stream=?, house=?,
              parent_name=?, parent_contact=?, parent2_name=?, parent2_contact=?, parent_email=?,
              residence=?, student_number=?, fees_code=?
            WHERE id=?
        """, (first_name, Middle_name, last_name,
              sex, section, class_name, stream, house,
              parent_name, parent_contact, parent2_name, parent2_contact, parent_email,
              residence, student_number, fees_code, student_id))
        conn.commit()
        conn.close()
        flash("Student updated.", "success")
        return redirect(url_for("students"))

    row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    conn.close()
    if not row:
        flash("Student not found.", "warning")
        return redirect(url_for("students"))
    return render_template("edit_student.html", s=row, classes=classes, streams=streams, houses=houses)


# ---------- Import Students (CSV/XLSX) ----------
@app.route("/students/import", methods=["GET","POST"])
@require_role("admin", "headteacher","bursar")
def students_import():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Select a CSV or XLSX file.", "warning")
            return redirect(url_for("students_import"))

        filename = file.filename.lower()
        conn = get_db_connection()
        c = conn.cursor()

        # guiding list only; extras are ignored
        expected = ["first_name","Middle_name","last_name","sex","section","class_name","stream",
                    "house","parent1_name","parent1_contact","parent2_name","parent2_contact",
                    "parent_email","residence","fees_code","student_number"]

        imported = 0
        try:
            if filename.endswith(".csv"):
                content = io.StringIO(file.stream.read().decode("utf-8"))
                reader = csv.DictReader(content)
                for _row in reader:
                    def g(k): return (_row.get(k) or _row.get(k.upper()) or "").strip()

                    # normalize
                    sex = norm_sex(g("sex"))
                    section = norm_section(g("section"))
                    class_name = norm_class(g("class_name"))
                    stream = norm_stream(g("stream"))

                    # minimal required (normalized values!)
                    if not (g("first_name") and g("last_name") and sex and section and class_name and stream):
                        continue

                    sn = g("student_number") or generate_student_number(conn, class_name)
                    fc = g("fees_code") or generate_fees_code(conn)

                    c.execute("""
                        INSERT INTO students (
                          first_name, Middle_name, last_name, sex, section, class_name, stream,
                          house, parent_name, parent_contact, parent2_name, parent2_contact,
                          parent_email, residence, student_number, fees_code, archived, status
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,'active')
                    """, (g("first_name"), g("Middle_name"), g("last_name"), sex, section, class_name, stream,
                          g("house"), g("parent1_name"), g("parent1_contact"), g("parent2_name"), g("parent2_contact"),
                          g("parent_email"), g("residence"), sn, fc))
                    imported += 1

            elif filename.endswith(".xlsx"):
                from openpyxl import load_workbook
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = [ (cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1)) ]
                idx = {h:i for i,h in enumerate(headers)}

                def val(row, name):
                    i = idx.get(name.lower())
                    if i is None: return ""
                    v = row[i].value
                    return (str(v).strip() if v is not None else "")

                for row in ws.iter_rows(min_row=2):
                    first_name = val(row, "first_name")
                    Middle_name= val(row, "Middle_name") or val(row, "middle_name")
                    last_name = val(row, "last_name")
                    sex = norm_sex(val(row, "sex"))
                    section = norm_section(val(row, "section"))
                    class_name = norm_class(val(row, "class_name"))
                    stream = norm_stream(val(row, "stream"))
                    house = val(row, "house")
                    parent1_name = val(row, "parent1_name")
                    parent1_contact = val(row, "parent1_contact")
                    parent2_name = val(row, "parent2_name")
                    parent2_contact = val(row, "parent2_contact")
                    parent_email = val(row, "parent_email")
                    residence = val(row, "residence")
                    sn = val(row, "student_number") or generate_student_number(conn, class_name)
                    fc = val(row, "fees_code") or generate_fees_code(conn)

                    if not (first_name and last_name and sex and section and class_name and stream):
                        continue

                    c.execute("""
                        INSERT INTO students (
                          first_name, Middle_name, last_name, sex, section, class_name, stream,
                          house, parent_name, parent_contact, parent2_name, parent2_contact,
                          parent_email, residence, student_number, fees_code, archived, status
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,'active')
                    """, (first_name, Middle_name, last_name, sex, section, class_name, stream,
                          house, parent1_name, parent1_contact, parent2_name, parent2_contact,
                          parent_email, residence, sn, fc))
                    imported += 1
            else:
                flash("Unsupported file type. Use .csv or .xlsx", "warning")
                conn.close()
                return redirect(url_for("students_import"))

            conn.commit()
            flash(f"Imported {imported} students.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Import failed: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("students"))

    return render_template("students_import.html")





# ---------- Students list (search) ----------
# If you already have a /students route, keep its name and replace the body.
@app.route("/students")
@require_role("admin", "headteacher","dos", "bursar", "teacher")
def students():
    q_class = (request.args.get("class_name") or "").strip()
    q_num = (request.args.get("student_number") or "").strip()
    q_last = (request.args.get("last_name") or "").strip()

    conn = get_db_connection()
    classes = [r[0] for r in conn.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name").fetchall()]

    where = ["archived = 0"]
    params = []
    if q_class:
        where.append("class_name = ?"); params.append(q_class)
    if q_num:
        where.append("student_number LIKE ?"); params.append(f"%{q_num}%")
    if q_last:
        where.append("last_name LIKE ?"); params.append(f"%{q_last}%")

    sql = f"""
      SELECT id, student_number, first_name, Middle_name, last_name,
             sex, section, class_name, stream, house,
             parent_name, parent_contact, parent2_name, parent2_contact,
             parent_email, residence, fees_code, date_joined, term_joined, year_of_joining
      FROM students
      WHERE {' AND '.join(where)}
      ORDER BY class_name, last_name, first_name
    """
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return render_template("students_manage.html",
                           rows=rows, classes=classes,
                           q_class=q_class, q_num=q_num, q_last=q_last)

# ---------- Export Students (CSV) ----------
@app.route("/students/export")
@require_role("admin", "headteacher", "dos","bursar")
def students_export():
    # reuse same filters as /students
    q_class = (request.args.get("class_name") or "").strip()
    q_num = (request.args.get("student_number") or "").strip()
    q_last = (request.args.get("last_name") or "").strip()

    conn = get_db_connection()
    where = ["archived = 0"]; params = []
    if q_class: where.append("class_name = ?"); params.append(q_class)
    if q_num: where.append("student_number LIKE ?"); params.append(f"%{q_num}%")
    if q_last: where.append("last_name LIKE ?"); params.append(f"%{q_last}%")

    sql = f"""SELECT student_number, first_name, Middle_name, last_name, sex, section,
                     class_name, stream, house, parent_name, parent_contact,
                     parent2_name, parent2_contact, parent_email, residence, fees_code
              FROM students
              WHERE {' AND '.join(where)}
              ORDER BY class_name, last_name, first_name"""
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    # build CSV
    output = io.StringIO()
    w = csv.writer(output)
    header = [ "student_number","first_name","Middle_name","last_name","sex","section",
               "class_name","stream","house","parent1_name","parent1_contact",
               "parent2_name","parent2_contact","parent_email","residence","fees_code" ]
    w.writerow(header)
    for r in rows:
        w.writerow([r["student_number"], r["first_name"], r["Middle_name"], r["last_name"], r["sex"], r["section"],
                    r["class_name"], r["stream"], r["house"], r["parent_name"], r["parent_contact"],
                    r["parent2_name"], r["parent2_contact"], r["parent_email"], r["residence"], r["fees_code"]])
    csv_data = output.getvalue()
    return Response(csv_data, mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=students.csv"})

# ---------- Parents Directory + Export ----------
@app.route("/parents")
@require_role("admin", "headteacher", "bursar","dos" "teacher")
def parents_directory():
    search = (request.args.get("search") or "").strip()

    conn = get_db_connection()
    if search:
        rows = conn.execute("""
            SELECT student_number,
                   parent_name AS parent1_name, parent_contact AS parent1_contact,
                   parent2_name, parent2_contact,
                   first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name AS student_name,
                   class_name, stream
            FROM students
            WHERE archived = 0 AND (
                  parent_name LIKE ? OR parent_contact LIKE ? OR
                  parent2_name LIKE ? OR parent2_contact LIKE ? OR
                  student_number LIKE ? OR last_name LIKE ?
            )
            ORDER BY parent_name, parent2_name
        """, (f"%{search}%",)*6).fetchall()
    else:
        rows = conn.execute("""
            SELECT student_number,
                   parent_name AS parent1_name, parent_contact AS parent1_contact,
                   parent2_name, parent2_contact,
                   first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name AS student_name,
                   class_name, stream
            FROM students
            WHERE archived = 0
            ORDER BY parent_name, parent2_name
        """).fetchall()
    conn.close()
    return render_template("parents_directory.html", rows=rows, search=search)
# for Jinja templates: has_role('admin') etc.

@app.context_processor
def inject_role_utils():
    def has_role(*roles):
        r = session.get("role")
        return r in roles
    return dict(has_role=has_role)
@app.route("/parents/export")
@require_role("admin", "headteacher", "bursar")
def parents_export():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT student_number,
               parent_name AS parent1_name, parent_contact AS parent1_contact,
               parent2_name, parent2_contact, parent_email,
               first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name AS student_name,
               class_name, stream
        FROM students
        WHERE archived = 0
        ORDER BY parent_name, parent2_name
    """).fetchall()
    conn.close()

    output = io.StringIO()
    w = csv.writer(output)
    header = ["student_number","student_name","class_name","stream",
              "parent1_name","parent1_contact","parent2_name","parent2_contact","parent_email"]
    w.writerow(header)
    for r in rows:
        w.writerow([r["student_number"], r["student_name"], r["class_name"], r["stream"],
                    r["parent1_name"], r["parent1_contact"], r["parent2_name"], r["parent2_contact"], r["parent_email"]])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=parents.csv"})


    
    







@app.route('/fees/report')
@require_role('admin', 'bursar', 'headteacher')
def fees_report():
    import sqlite3
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Active academic context
    academic = c.execute("""
        SELECT year, current_term FROM academic_years WHERE is_active = 1
    """).fetchone()
    if not academic:
        flash("Please activate an academic year.", "warning")
        return redirect(url_for('dashboard'))
    year, term = int(academic['year']), academic['current_term']

    # ---- Filters ----
    class_filter = (request.args.get('class_name') or '').strip()
    stream_filter = (request.args.get('stream') or '').strip()
    status_filter = (request.args.get('status') or '').strip().lower().replace(' ', '')

    # ---- Students (filtered) ----
    sql = "SELECT * FROM students WHERE archived=0"
    params = []
    if class_filter:
        sql += " AND class_name=?"; params.append(class_filter)
    if stream_filter:
        sql += " AND stream=?"; params.append(stream_filter)
    sql += " ORDER BY class_name, stream, last_name, first_name"
    students = c.execute(sql, params).fetchall()
    conn.close()

    # ---- Build report buckets (use compute_student_financials as source of truth) ----
    report = {'full': [], 'partial': [], 'none': []}

    for s in students:
        fin = compute_student_financials(
            student_id=s['id'],
            class_name=s['class_name'],
            term=term,
            year=year
        )

        # Pull numbers strictly from compute_student_financials
        expected_fees = float(fin.get('expected_fees', 0.0))
        expected_reqs = float(fin.get('expected_requirements', 0.0))
        bursary_current = float(fin.get('bursary_current', 0.0))
        paid_fees = float(fin.get('paid_fees', 0.0))
        paid_reqs = float(fin.get('paid_requirements', 0.0))
        carried = float(fin.get('carry_forward', 0.0))
        overall_balance = float(fin.get('overall_balance', 0.0)) # may be negative (credit)

        # Totals for display
        expected_total = expected_fees + expected_reqs
        paid_total = paid_fees + paid_reqs
        final_expected = max(expected_fees - bursary_current, 0.0) + expected_reqs + carried

        # Status (fees-only logic), consistent basis:
        # base term due (fees net of bursary) vs fees paid
        base_term_due = max(expected_fees - bursary_current, 0.0)
        base_term_bal_real = (expected_fees - bursary_current) - paid_fees # allow negative
        if base_term_bal_real <= 0 and base_term_due > 0:
            status = 'full'
        elif 0 < base_term_bal_real < base_term_due:
            status = 'partial'
        else:
            status = 'none'

        if status_filter and status != status_filter:
            continue

        report[status].append({
            'student': s,
            'expected': expected_total, # fees + requirements
            'bursary': bursary_current,
            'carried': carried, # OB + prior arrears
            'final_expected': final_expected, # for table summary
            'paid': paid_total, # fees + requirements
            'balance': overall_balance # EXACTLY matches finance report
        })

    return render_template('fees_report.html', report=report, term=term, year=year)




# Download a blank template for import (CSV)
@app.route("/students/import/template")
@require_role("admin", "headteacher","dos","bursar")
def students_import_template():
    header = ["first_name","Middle_name","last_name","sex","section","class_name","stream",
            "house","parent1_name","parent1_contact","parent2_name","parent2_contact",
            "parent_email","residence","fees_code","student_number"]
    output = io.StringIO()
    csv.writer(output).writerow(header)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=students_import_template.csv"})



@app.route('/fees/report/export')
@require_role('admin', 'bursar', 'headteacher')
def export_fees_report():
    import io, csv, sqlite3
    from flask import make_response

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Active academic context
    academic = c.execute("""
        SELECT year, current_term FROM academic_years WHERE is_active = 1
    """).fetchone()
    if not academic:
        flash("Please activate an academic year.", "warning")
        return redirect(url_for('dashboard'))
    year, term = int(academic['year']), academic['current_term']

    # Filters
    class_filter = (request.args.get('class_name') or '').strip()
    stream_filter = (request.args.get('stream') or '').strip()
    status_filter = (request.args.get('status') or '').strip().lower().replace(' ', '')

    # Students with filters
    sql = "SELECT * FROM students WHERE archived=0"
    params = []
    if class_filter:
        sql += " AND class_name=?"; params.append(class_filter)
    if stream_filter:
        sql += " AND stream=?"; params.append(stream_filter)
    sql += " ORDER BY class_name, stream, last_name, first_name"
    students = c.execute(sql, params).fetchall()
    conn.close()

    # CSV header
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Student Number','Full Name','Class',
        'Expected (Term)','Bursary (Term)','Carried Forward',
        'Final Expected','Paid (Term)','Balance','Status'
    ])

    # Rows
    for s in students:
        fin = compute_student_financials(
            student_id=s['id'],
            class_name=s['class_name'],
            term=term,
            year=year
        )

        expected = float(fin.get('expected_fees', 0.0))
        bursary = float(fin.get('bursary_current', 0.0))
        paid = float(fin.get('paid_fees', 0.0))
        carried = float(fin.get('carry_forward', 0.0)) # OB + prior arrears (already clamped >= 0 in your helper)

        # Match on-screen fees_report math:
        base_term_due = max(expected - bursary, 0.0)
        base_term_bal_real = (expected - bursary) - paid # allow negative (overpayment)
        final_expected = base_term_due + carried
        balance = carried + base_term_bal_real # can be negative (credit)

        # Same status logic used in fees_report()
        if base_term_bal_real <= 0 and base_term_due > 0:
            status = 'full'
        elif 0 < base_term_bal_real < base_term_due:
            status = 'partial'
        else:
            status = 'none'

        if status_filter and status != status_filter:
            continue

        writer.writerow([
            s['student_number'],
            f"{s['first_name']} {s['last_name']}".strip(),
            f"{s['class_name']} {s['stream'] or ''}".strip(),
            round(expected, 0),
            round(bursary, 0),
            round(carried, 0),
            round(final_expected, 0),
            round(paid, 0),
            round(balance, 0),
            status
        ])

    resp = make_response(output.getvalue())
    resp.headers["Content-Disposition"] = "attachment; filename=fees_report.csv"
    resp.headers["Content-type"] = "text/csv"
    return resp




# ---------- FEES EDIT (admin only) ----------

@app.route("/fees/<int:fee_id>/edit", methods=["GET", "POST"])
@require_role("admin")
def fees_edit(fee_id: int):
    import sqlite3

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    fee = c.execute("""
        SELECT f.*, s.student_number, s.first_name, COALESCE(s.Middle_name,'') AS middle_name,
               s.last_name, s.class_name, s.stream
          FROM fees f
          JOIN students s ON s.id = f.student_id
         WHERE f.id = ?
    """, (fee_id,)).fetchone()
    if not fee:
        conn.close()
        flash("Transaction not found.", "warning")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        term = (request.form.get("term") or fee["term"] or "").strip()
        year = int(request.form.get("year") or fee["year"] or 0)
        date_paid = (request.form.get("date_paid") or fee["date_paid"] or "").strip()
        method = (request.form.get("method") or fee["method"] or "").strip()
        ptype_raw = (request.form.get("payment_type") or fee["payment_type"] or "").strip()

        ptype_lc = ptype_raw.lower().replace(" ", "_")
        if ptype_lc in ("school_fees","schoolfees","tuition"): payment_type = "school_fees"
        elif ptype_lc in ("fees",): payment_type = "fees"
        elif ptype_lc in ("requirements","requirement"): payment_type = "requirements"
        elif ptype_lc in ("opening_balance","opening-balance","ob"): payment_type = "opening_balance"
        else: payment_type = ptype_raw

        def _f(v):
            try: return float(v)
            except: return 0.0

        amount_paid = _f(request.form.get("amount_paid"))
        expected_amount = _f(request.form.get("expected_amount"))
        bursary_amount = _f(request.form.get("bursary_amount"))
        carried_forward = _f(request.form.get("carried_forward"))
        requirement_name = (request.form.get("requirement_name") or "").strip()
        comment = (request.form.get("comment") or "").strip()

        # snapshot BEFORE
        before = {
            "term": fee["term"], "year": int(fee["year"] or 0),
            "date_paid": fee["date_paid"], "method": fee["method"],
            "payment_type": fee["payment_type"],
            "amount_paid": float(fee["amount_paid"] or 0.0),
            "expected_amount": float(fee["expected_amount"] or 0.0),
            "bursary_amount": float(fee["bursary_amount"] or 0.0),
            "carried_forward": float(fee["carried_forward"] or 0.0),
            "requirement_name": fee["requirement_name"],
            "comment": fee["comment"],
        }

        make_void = request.form.get("void_entry") == "on"
        if make_void:
            comment = (comment + " | " if comment else "") + \
                      f"voided by {session.get('username','admin')} on DATE('now')"
            amount_paid = expected_amount = bursary_amount = carried_forward = 0.0

        try:
            c.execute("""
                UPDATE fees
                   SET term=?,
                       year=?,
                       date_paid=?,
                       method=?,
                       payment_type=?,
                       amount_paid=?,
                       expected_amount=?,
                       bursary_amount=?,
                       carried_forward=?,
                       requirement_name=?,
                       comment=?,
                       processed_on=CURRENT_TIMESTAMP
                 WHERE id=?
            """, (term, year, date_paid, method, payment_type,
                  amount_paid, expected_amount, bursary_amount, carried_forward,
                  requirement_name, comment, fee_id))
            conn.commit()
            flash("Transaction updated.", "success")

            # AFTER + changed-only for the audit payload
            after = {
                "term": term, "year": year, "date_paid": date_paid, "method": method,
                "payment_type": payment_type, "amount_paid": amount_paid,
                "expected_amount": expected_amount, "bursary_amount": bursary_amount,
                "carried_forward": carried_forward, "requirement_name": requirement_name,
                "comment": comment,
            }
            changed_only = {
                k: {"before": before.get(k), "after": after.get(k)}
                for k in after.keys() if before.get(k) != after.get(k)
            }

            audit_from_request(
                conn,
                action="fees_edit",
                target_table="fees",
                target_id=fee_id,
                details={
                    "student_id": fee["student_id"],
                    "before": before,
                    "after": after,
                    "changed": changed_only,
                    "voided": make_void
                },
                outcome="success",
                severity="info"
            )
        except Exception as e:
            conn.rollback()
            flash(f"Update failed: {e}", "danger")
            audit_from_request(
                conn,
                action="fees_edit",
                target_table="fees",
                target_id=fee_id,
                details={"error": str(e)},
                outcome="failure",
                severity="warning"
            )
        finally:
            conn.close()

        return redirect(url_for("student_statement_by_id", student_id=fee["student_id"]))

    conn.close()
    return render_template("fees_edit.html", fee=fee)


    


# ========= Secure Change Password (hashed) =========
# Requires: from flask import request, render_template, redirect, url_for, flash, session, current_app
# import sqlite3
from werkzeug.security import check_password_hash, generate_password_hash

def _get_db():
    conn = get_db_connection() # your existing helper
    conn.row_factory = sqlite3.Row
    return conn

def _password_col_name(conn) -> str:
    """
    Detect which column stores the hash: 'password_hash' preferred, fall back to 'password'.
    """
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if "password_hash" in cols:
        return "password_hash"
    if "password" in cols:
        return "password" # assume it already stores a hash
    # If neither exists, fail clearly:
    raise RuntimeError("No password column found in 'users' table. Expected 'password_hash' or 'password'.")

@app.route("/change_password", methods=["GET", "POST"])
def change_password():
    # You can adapt this if you let admins change others’ passwords.
    username = session.get("username")
    if not username:
        flash("You must be logged in to change your password.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        old_password = (request.form.get("old_password") or "").strip()
        new_password = (request.form.get("new_password") or "").strip()
        new_password2 = (request.form.get("new_password2") or "").strip()

        # Basic validations first
        if not old_password or not new_password:
            flash("Please fill in all password fields.", "warning")
            return redirect(url_for("change_password"))
        if new_password != new_password2:
            flash("New passwords do not match.", "danger")
            return redirect(url_for("change_password"))
        if len(new_password) < 8:
            flash("New password must be at least 8 characters.", "warning")
            return redirect(url_for("change_password"))

        conn = _get_db()
        try:
            pw_col = _password_col_name(conn)
            user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
            if not user:
                flash("User not found.", "danger")
                return redirect(url_for("change_password"))

            stored_hash = user[pw_col]
            if not stored_hash:
                flash("Password is not set for this account. Contact admin.", "danger")
                return redirect(url_for("change_password"))

            # Check old password (hash)
            if not check_password_hash(stored_hash, old_password):
                flash("Old password is incorrect.", "danger")
                return redirect(url_for("change_password"))

            # Prevent reusing the same password
            if check_password_hash(stored_hash, new_password):
                flash("New password must be different from the old one.", "warning")
                return redirect(url_for("change_password"))

            # Create and save new hash
            new_hash = generate_password_hash(new_password)
            conn.execute(f"UPDATE users SET {pw_col}=? WHERE username=?", (new_hash, username))
            conn.commit()

            flash("Password changed successfully.", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            try:
                current_app.logger.exception("[change_password] failed")
            except Exception:
                pass
            flash(f"Could not change password: {e}", "danger")
            return redirect(url_for("change_password"))
        finally:
            conn.close()

    # GET
    return render_template("change_password.html")
# ========= /Secure Change Password =========



from werkzeug.security import check_password_hash
from urllib.parse import urljoin, urlparse
from datetime import datetime

# ---------- small helpers ----------

def _is_safe_url(target: str) -> bool:
    """Prevent open redirects."""
    if not target:
        return False
    host_url = urlparse(request.host_url)
    redirect_url = urlparse(urljoin(request.host_url, target))
    return (redirect_url.scheme in ("http", "https")
            and host_url.netloc == redirect_url.netloc)

def get_user_profile(user_id: int) -> dict:
    """
    Returns {"initials": str, "full_name": str} for the given users.id.
    - Looks up users.employee_id -> employees (name) and teachers (initials).
    - Opens & closes its own connection (so it never uses a closed DB).
    """
    initials = ""
    full_name = ""

    conn = get_db_connection()
    c = conn.cursor()

    u = c.execute("SELECT employee_id FROM users WHERE id=?", (user_id,)).fetchone()
    emp_id = u["employee_id"] if u and "employee_id" in u.keys() else None

    if emp_id:
        emp = c.execute(
            "SELECT first_name, COALESCE(Middle_name,'') AS Middle_name, last_name "
            "FROM employees WHERE id=?", (emp_id,)
        ).fetchone()
        if emp:
            full_name = " ".join([
                (emp["first_name"] or "").strip(),
                (emp["Middle_name"] or "").strip(),
                (emp["last_name"] or "").strip(),
            ]).strip()

        tch = c.execute(
            "SELECT initials FROM teachers WHERE employee_id=?", (emp_id,)
        ).fetchone()
        if tch and (tch["initials"] or "").strip():
            initials = tch["initials"].strip()

    conn.close()
    return {"initials": initials, "full_name": full_name}


from functools import wraps
from flask import session, redirect, url_for, flash, render_template, request
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime

# --- Role helpers ---
def _norm_role(val):
    """Return a canonical, lowercase role string."""
    return (str(val or "").strip().lower())

def require_login(f):
    @wraps(f)
    def _inner(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return _inner

def require_role(*roles):
    # normalize decorator inputs once
    wanted = tuple(_norm_role(r) for r in roles if r)

    def wrapper(f):
        @wraps(f)
        def inner(*args, **kwargs):
            if "user_id" not in session or "role" not in session:
                flash("Please login.", "warning")
                return redirect(url_for("login"))

            srole = _norm_role(session.get("role"))
            if wanted and srole not in wanted:
                flash("Access denied.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return inner
    return wrapper

# allowed roles
ALLOWED_ROLES = (
    "admin","bursar","teacher","headteacher",
    "director","clerk","dos","deputyheadteacher"
)

# --- normalize role every request (belt & braces) ---
@app.before_request
def _coerce_session_role():
    if "role" in session:
        session["role"] = _norm_role(session["role"])

# --- Login route ---
@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        next_url = request.args.get("next") or request.form.get("next")

        conn = get_db_connection()
        c = conn.cursor()
        user = c.execute(
            "SELECT * FROM users WHERE username=? LIMIT 1", (username,)
        ).fetchone()
        conn.close()

        if not user:
            flash("Invalid credentials.", "danger")
            return render_template("login.html", current_year=datetime.now().year, next=next_url)

        if "status" in user.keys() and user["status"] != "active":
            flash("This account is archived/disabled.", "warning")
            return render_template("login.html", current_year=datetime.now().year, next=next_url)

        if not check_password_hash(user["password_hash"], password):
            flash("Invalid credentials.", "danger")
            return render_template("login.html", current_year=datetime.now().year, next=next_url)

        # success
        session.clear()
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = _norm_role(user["role"]) # normalized

        prof = get_user_profile(user["id"])
        session["initials"] = prof.get("initials", "") or ""
        session["full_name"] = (prof.get("full_name") or user["username"]).strip()

        flash("Login successful.", "success")
        if next_url and _is_safe_url(next_url):
            return redirect(next_url)
        return redirect(url_for("dashboard"))

    return render_template("login.html", current_year=datetime.now().year, next=request.args.get("next", ""))

# --- Legacy login_post (fixed) ---
def login_post():
    username = request.form["username"].strip()
    password = request.form["password"].strip()
    conn = get_db_connection(); c = conn.cursor()

    user = c.execute(
        "SELECT * FROM users WHERE username=? AND status='active' LIMIT 1",
        (username,)
    ).fetchone()
    conn.close()

    if not user or not check_password_hash(user["password_hash"], password):
        flash("Invalid credentials.", "danger")
        return redirect(url_for("login"))

    session.clear()
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = _norm_role(user["role"]) # normalized
    return redirect(url_for("dashboard"))



@app.route('/profile', methods=['GET', 'POST'])
@require_role('admin', 'teacher', 'bursar', 'headteacher', 'director')
def user_profile():
    user_id = session.get('user_id')
    if not user_id:
        flash("Please login to view your profile", "danger")
        return redirect(url_for('login'))

    conn = get_db_connection()
    c = conn.cursor()

    if request.method == 'POST':
        first_name = request.form['first_name']
        Middle_name = request.form.get('Middle_name', '')
        last_name = request.form['last_name']
        email = request.form['email']
        contact = request.form['contact']

        c.execute('''
            UPDATE users SET first_name = ?, Middle_name = ?, last_name = ?, email = ?, contact = ?
            WHERE id = ?
        ''', (first_name, Middle_name, last_name, email, contact, user_id))
        conn.commit()
        flash("Profile updated successfully.", "success")

    user = c.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return render_template('user_profile.html', user=user)




@app.route('/clearance')
@require_role('admin', 'bursar', 'headteacher')
def clearance():
    conn = get_db_connection()
    c = conn.cursor()

    # Get active term and year
    current = c.execute("SELECT year, current_term FROM academic_years WHERE is_active = 1").fetchone()
    if not current:
        flash("No active academic year found.", "warning")
        return redirect(url_for('academic_years'))

    year, term = current['year'], current['current_term']

    # Fetch students who have fully paid (school fees only) for that term and year
    cleared_students = c.execute('''
        SELECT s.id, s.first_name, s.Middle_name, s.last_name, s.class_name, SUM(f.amount_paid) as total_paid, f.expected_amount
        FROM students s
        JOIN fees f ON s.id = f.student_id
        WHERE f.term = ? AND f.year = ? AND f.payment_type = 'school_fees'
        GROUP BY s.id
        HAVING total_paid >= f.expected_amount
    ''', (term, int(year))).fetchall()

    conn.close()
    return render_template('clearance.html', students=cleared_students, term=term, year=year)

    


@app.route('/expenditure_report', methods=['GET', 'POST'])
@require_role('admin', 'bursar')
def expenditure_report():
    conn = get_db_connection()
    c = conn.cursor()

    filters = {"from_date": "", "to_date": ""}
    expenses = []
    summary = []
    type_summary = []
    total = 0

    if request.method == 'POST':
        filters['from_date'] = request.form.get('from_date')
        filters['to_date'] = request.form.get('to_date')

        expenses = c.execute('''
            SELECT e.*, ec.name as category_name
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE date_spent BETWEEN ? AND ?
            ORDER BY date_spent DESC
        ''', (filters['from_date'], filters['to_date'])).fetchall()

        total = sum([row['amount'] for row in expenses])

        summary = c.execute('''
            SELECT ec.name as category, SUM(e.amount) as total
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE date_spent BETWEEN ? AND ?
            GROUP BY ec.name
            ORDER BY total DESC
        ''', (filters['from_date'], filters['to_date'])).fetchall()

        type_summary = c.execute('''
            SELECT type, SUM(amount) as total
            FROM expenses
            WHERE date_spent BETWEEN ? AND ?
            GROUP BY type
            ORDER BY total DESC
        ''', (filters['from_date'], filters['to_date'])).fetchall()

    conn.close()
    return render_template('expenditure_report.html',
                           expenses=expenses, total=total,
                           filters=filters, summary=summary,
                           type_summary=type_summary)

    




    def _list_classes():
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT DISTINCT TRIM(class_name) AS class_name
            FROM students
            WHERE archived = 0 AND class_name IS NOT NULL AND TRIM(class_name) <> ''
            ORDER BY class_name
        """).fetchall()
        conn.close()
        return [r["class_name"] for r in rows]

    def _year_choices(base_year: int):
        # collect years seen in HP plus base±1 so the list isn't empty
        years = set()
        conn = get_db_connection()
        try:
            for r in conn.execute("SELECT DISTINCT year FROM holiday_package ORDER BY year"):
                try:
                    years.add(int(r[0]))
                except Exception:
                    pass
        finally:
            conn.close()
        if base_year:
            years.update({base_year - 1, base_year, base_year + 1})
        ys = sorted(y for y in years if y)
        # sensible fallback if db is empty
        if not ys:
            ys = [base_year] if base_year else []
        return ys

    def _list_students_for_class(class_name: str):
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream
            FROM students
            WHERE archived = 0 AND LOWER(TRIM(class_name)) = LOWER(TRIM(?))
            ORDER BY last_name, first_name
        """, (class_name,)).fetchall()
        conn.close()
        return rows

    def _fetch_hp_for_class_term_year(class_name: str, term: str, year: int):
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT hp.*
            FROM holiday_package hp
            JOIN students s ON s.id = hp.student_id
            WHERE s.archived = 0
              AND LOWER(TRIM(s.class_name)) = LOWER(TRIM(?))
              AND LOWER(TRIM(hp.term)) = LOWER(TRIM(?))
              AND hp.year = ?
        """, (class_name, term, year)).fetchall()
        conn.close()
        return {r["student_id"]: r for r in rows}

    def _upsert_hp_row(student_id: int, term: str, year: int, rowvals: dict):
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        existing = cur.execute("""
            SELECT id FROM holiday_package
            WHERE student_id=? AND LOWER(TRIM(term))=LOWER(TRIM(?)) AND year=?
        """, (student_id, term, year)).fetchone()

        cols = ["eng", "mat", "sci", "sst", "agg", "total"]
        vals = [rowvals.get(k) for k in cols]

        if existing:
            cur.execute("""
                UPDATE holiday_package
                   SET eng=?, mat=?, sci=?, sst=?, agg=?, total=?
                 WHERE id=?
            """, (*vals, existing["id"]))
        else:
            cur.execute("""
                INSERT INTO holiday_package (student_id, term, year, eng, mat, sci, sst, agg, total)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (student_id, term, year, *vals))
        conn.commit()
        conn.close()

    def _to_num(x):
        try:
            if x is None or str(x).strip() == "":
                return None
            return float(str(x).replace(",", "").strip())
        except Exception:
            return None

    # ---------- filters ----------
    class_name = (request.values.get("class_name") or "").strip()
    term = (request.values.get("term") or "").strip()
    try:
        year = int(request.values.get("year") or 0)
    except ValueError:
        year = 0

    # active year for defaults & year options
    try:
        ay = get_active_academic_year()
        active_year = int(ay.get("year") or 0)
    except Exception:
        active_year = 0

    class_options = _list_classes()
    year_options = _year_choices(active_year)

    # ---------- POST: upload ----------
    if request.method == "POST":
        f = request.files.get("file")
        if not (class_name and term and year):
            flash("Select Class, Term and Year first.", "warning")
            return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

        if not f or not f.filename:
            flash("No file selected.", "warning")
            return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

        _, ext = os.path.splitext(f.filename.lower())
        if ext not in ALLOWED_EXTS_HP:
            flash("Unsupported file type. Use .xlsx, .xls, or .csv", "warning")
            return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

        # build map for selected class
        stu_rows = _list_students_for_class(class_name)
        sn_to_id = {r["student_number"]: r["id"] for r in stu_rows}
        saved = skipped = 0

        try:
            if ext == ".csv":
                reader = csv.DictReader(TextIOWrapper(f.stream, encoding="utf-8", errors="ignore"))
                for r in reader:
                    sn = (r.get("student_number") or r.get("Student Number") or r.get("student") or "").strip()
                    sid = sn_to_id.get(sn)
                    if not sid:
                        skipped += 1
                        continue
                    _upsert_hp_row(sid, term, year, {
                        "eng": _to_num(r.get("eng")),
                        "mat": _to_num(r.get("mat")),
                        "sci": _to_num(r.get("sci")),
                        "sst": _to_num(r.get("sst")),
                        "agg": _to_num(r.get("agg")),
                        "total": _to_num(r.get("total")),
                    })
                    saved += 1
            else:
                if not load_workbook:
                    flash("openpyxl not installed; cannot read .xlsx/.xls. Use .csv or install openpyxl.", "danger")
                    return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

                wb = load_workbook(f, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
                def _idx(name): return headers.index(name) if name in headers else -1
                idx = {
                    "student_number": _idx("student_number"),
                    "eng": _idx("eng"), "mat": _idx("mat"),
                    "sci": _idx("sci"), "sst": _idx("sst"),
                    "agg": _idx("agg"), "total": _idx("total"),
                }
                if idx["student_number"] < 0:
                    flash("Template must include 'student_number' column.", "danger")
                    return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

                for row in ws.iter_rows(min_row=2):
                    sn_cell = row[idx["student_number"]].value
                    sn = (str(sn_cell).strip() if sn_cell is not None else "")
                    if not sn:
                        continue
                    sid = sn_to_id.get(sn)
                    if not sid:
                        skipped += 1
                        continue
                    _upsert_hp_row(sid, term, year, {
                        "eng": _to_num(row[idx["eng"]].value) if idx["eng"] >= 0 else None,
                        "mat": _to_num(row[idx["mat"]].value) if idx["mat"] >= 0 else None,
                        "sci": _to_num(row[idx["sci"]].value) if idx["sci"] >= 0 else None,
                        "sst": _to_num(row[idx["sst"]].value) if idx["sst"] >= 0 else None,
                        "agg": _to_num(row[idx["agg"]].value) if idx["agg"] >= 0 else None,
                        "total": _to_num(row[idx["total"]].value) if idx["total"] >= 0 else None,
                    })
                    saved += 1

            flash(f"Holiday Package saved: {saved} rows. Skipped: {skipped}.", "success")
        except Exception as e:
            current_app.logger.exception(f"[HP upload] failed: {e}")
            flash(f"Upload failed: {e}", "danger")

        return redirect(url_for("holiday_hub", class_name=class_name, term=term, year=year))

    # ---------- GET: load view ----------
    students = []
    hp_rows = {}
    if class_name and term and year:
        students = _list_students_for_class(class_name)
        hp_rows = _fetch_hp_for_class_term_year(class_name, term, year)

    # fallbacks so dropdowns aren't empty
    if not class_options:
        class_options = [class_name] if class_name else []
    if not year_options:
        year_options = [year] if year else ([active_year] if active_year else [])

    return render_template(
        "holiday_hub.html",
        class_name=class_name,
        term=term,
        year=year or (active_year or 0),
        students=students,
        hp_rows=hp_rows,
        TERMS=TERMS,
        class_options=class_options,
        year_options=year_options,
        ALLOWED_EXTS_HP={".xlsx", ".xls", ".csv"},
    )

# --- Template (download) for holiday Package
@app.route("/holiday/template")
@require_role("admin", "teacher", "headteacher")
def holiday_template():
    class_name = request.args.get("class_name")
    term = request.args.get("term")
    year = request.args.get("year", type=int, default=datetime.now().year)

    conn = get_db_connection()
    students = conn.execute("""
        SELECT student_number, first_name, last_name
        FROM students
        WHERE class_name=?
        ORDER BY last_name, first_name
    """, (class_name,)).fetchall()
    conn.close()

    df = pd.DataFrame([{
        "student_number": s["student_number"],
        "student_name": f"{s['first_name']} {s['last_name']}",
        "eng": "", "mat": "", "sci": "", "sst": "",
        "agg": "", "total": ""
    } for s in students])

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="holidayPackage")
    bio.seek(0)
    fname = f"holiday_template_{class_name}_{term}_{year}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- Export the current holiday Package view
@app.route("/holiday/export")
@require_role("admin", "teacher", "headteacher")
def holiday_export():
    class_name = request.args.get("class_name")
    term = request.args.get("term")
    year = request.args.get("year", type=int)

    conn = get_db_connection()
    df = pd.read_sql_query("""
        SELECT s.student_number, s.first_name, s.last_name, s.class_name,
               m.eng, m.mat, m.sci, m.sst, m.agg, m.total
        FROM midterms m
        JOIN students s ON s.id = m.student_id
        WHERE s.class_name=? AND m.term=? AND m.year=? AND m.assessment='holiday Package'
        ORDER BY s.last_name, s.first_name
    """, conn, params=(class_name, term, year))
    conn.close()

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="holidayPackage")
    bio.seek(0)
    fname = f"holiday_{class_name}_{term}_{year}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- Edit one holiday Package row
@app.route("/holiday/edit/<int:mid_id>", methods=["GET", "POST"])
@require_role("admin", "teacher", "headteacher")
def holiday_edit(mid_id):
    conn = get_db-connection()
    if request.method == "POST":
        eng = request.form.get("eng", type=float)
        mat = request.form.get("mat", type=float)
        sci = request.form.get("sci", type=float)
        sst = request.form.get("sst", type=float)
        agg = request.form.get("agg", type=int)
        tot = request.form.get("total", type=int)

        conn.execute("""
            UPDATE midterms
            SET eng=?, mat=?, sci=?, sst=?, agg=?, total=?
            WHERE id=?
        """, (eng, mat, sci, sst, agg, tot, mid_id))
        conn.commit()
        conn.close()
        flash("holiday Package row updated.", "success")
        return redirect(url_for("holiday_hub"))

    row = conn.execute("SELECT * FROM midterms WHERE id=?", (mid_id,)).fetchone()
    conn.close()
    if not row: abort(404)
    return render_template("holiday_edit.html", row=row)
    
    

    



# ========================= REQUIREMENTS MANAGEMENT =========================
# Assumes you already have: app, require_role, get_db_connection

import sqlite3
from flask import request, render_template, redirect, url_for, flash

TERMS = ["Term 1", "Term 2", "Term 3"] # or import your existing TERMS

def ensure_requirements_schema(conn=None):
    """
    Idempotently create requirements table + unique index.
    Can be called with or without an external connection.
    """
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True

    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_name TEXT,
            term TEXT,
            name TEXT,
            year INTEGER,
            qty INTEGER,
            amount REAL DEFAULT 0
        )
    """)
    # prevent duplicate rows per (class, term, name)

    


def ensure_requirements_has_year():
    """
    One-time migration for existing databases:
    - If 'year' column is missing on 'requirements', add it with a constant default (0)
      because SQLite forbids non-constant defaults on ALTER TABLE.
    - Then backfill current year for existing rows.
    Safe to call repeatedly.
    """
    conn = get_db_connection()
    try:
        c = conn.cursor()

        # Does the column already exist?
        c.execute("PRAGMA table_info(requirements)")
        cols = [row[1] for row in c.fetchall()]
        if "year" in cols:
            return # nothing to do

        # Step 1: add column with a constant default (allowed by SQLite)
        c.execute("ALTER TABLE requirements ADD COLUMN year INTEGER DEFAULT 0")
        conn.commit()

        # Step 2: backfill existing rows to the current year
        c.execute("UPDATE requirements SET year = CAST(strftime('%Y','now') AS INTEGER) WHERE year = 0 OR year IS NULL")
        conn.commit()
    finally:
        conn.close()





    c.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uq_requirements
        ON requirements(
            class_name,
            COALESCE(term,''),
            name
        )
    """)
    conn.commit()
    if close_after:
        conn.close()


def _classes_for_dropdown(conn):
    # prefer classes table; fall back to students if needed
    rows = conn.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name").fetchall()
    classes = [r[0] for r in rows]
    if not classes:
        rows = conn.execute("SELECT DISTINCT class_name FROM students ORDER BY class_name").fetchall()
        classes = [r[0] for r in rows]
    return classes


@app.route("/admin/requirements", methods=["GET", "POST"])
@require_role("admin","bursar","headteacher")
def admin_requirements():
    ensure_requirements_schema()

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if request.method == "POST":
        rid = (request.form.get("id") or "").strip()
        class_name = (request.form.get("class_name") or "").strip()
        name = (request.form.get("name") or "").strip()
        qty = int(request.form.get("qty") or 1)
        amount = float(request.form.get("amount") or 0)
        term = (request.form.get("term") or "").strip() or None

        if not class_name or not name:
            flash("Class and item name are required.", "warning")
            return redirect(url_for("admin_requirements"))

        try:
            if rid:
                c.execute("""
                  UPDATE requirements
                  SET class_name=?, name=?, qty=?, amount=?, term=?
                  WHERE id=?
                """, (class_name, name, qty, amount, term, rid))
            else:
                c.execute("""
                  INSERT INTO requirements (class_name, name, qty, amount, term)
                  VALUES (?,?,?,?,?)
                """, (class_name, name, qty, amount, term))
            conn.commit()
            flash("Requirement saved.", "success")
        except sqlite3.IntegrityError:
            flash("Duplicate item for this class/term.", "danger")
        except Exception as e:
            conn.rollback()
            flash(f"Failed to save: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("admin_requirements"))

    # GET list + filters
    q_class = (request.args.get("class_name") or "").strip()
    q_term = (request.args.get("term") or "").strip()

    where, params = ["1=1"], []
    if q_class:
        where.append("class_name=?"); params.append(q_class)
    if q_term:
        where.append("(term=? )"); params.append(q_term)

    rows = conn.execute(f"""
      SELECT id, class_name, name, qty, amount, term
      FROM requirements
      WHERE {' AND '.join(where)}
      ORDER BY class_name, COALESCE(term,''), name
    """, params).fetchall()
    conn.close()

    # classes for dropdown (use your source—this example derives from students/classes table)
    conn = get_db_connection()
    classes = [r[0] for r in conn.execute("SELECT DISTINCT class_name FROM classes ORDER BY 1").fetchall()]
    conn.close()

    return render_template("admin_requirements.html",
                           items=rows, classes=classes, terms=TERMS,
                           q_class=q_class, q_term=q_term)


@app.route("/admin/requirements/<int:rid>/delete", methods=["POST"])
@require_role("admin","bursar","headteacher")
def admin_requirements_delete(rid):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM requirements WHERE id=?", (rid,))
        conn.commit()
        flash("Requirement deleted.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin_requirements"))
 

# ===================== PROMOTIONS: HISTORY + UNDO + BATCH =====================







# ========================= PROMOTIONS: HELPERS =========================

from datetime import datetime
import sqlite3

CLASS_ORDER = ["Baby","Middle","Top","P1","P2","P3","P4","P5","P6","P7"]

def next_class_name(current: str) -> str | None:
    if current not in CLASS_ORDER: return None
    i = CLASS_ORDER.index(current)
    return CLASS_ORDER[i+1] if i+1 < len(CLASS_ORDER) else None

def prev_class_name(current: str) -> str | None:
    if current not in CLASS_ORDER: return None
    i = CLASS_ORDER.index(current)
    return CLASS_ORDER[i-1] if i-1 >= 0 else None

def _distinct_classes(conn) -> list[str]:
    """Collect classes from classes table + students, dedup + canonical order."""
    seen, out = set(), []
    try:
        for (cn,) in conn.execute(
            "SELECT DISTINCT class_name FROM classes WHERE class_name IS NOT NULL"):
            cn = (cn or "").strip()
            if cn and cn not in seen: seen.add(cn); out.append(cn)
    except Exception:
        pass
    for (cn,) in conn.execute(
        "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL"):
        cn = (cn or "").strip()
        if cn and cn not in seen: seen.add(cn); out.append(cn)
    # order by canonical list first, then name
    return sorted(out, key=lambda x: (CLASS_ORDER.index(x) if x in CLASS_ORDER else 999, x))




def _classes_for_dropdown(conn) -> list[str]:
    """Collect distinct classes for the class dropdown."""
    rows = conn.execute(
        "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL ORDER BY class_name"
    ).fetchall()
    return [r[0] for r in rows]

def _years_for_dropdown(conn) -> list[int]:
    """
    Collect distinct years from fees for the year dropdown (desc).
    Falls back to active academic year or current year if none.
    """
    yrs = [r[0] for r in conn.execute(
        "SELECT DISTINCT year FROM fees WHERE year IS NOT NULL ORDER BY year DESC"
    ).fetchall()]
    if not yrs:
        try:
            ay = get_active_academic_year()  # your existing helper
            yrs = [int(ay.get("year", datetime.now().year))]
        except Exception:
            yrs = [datetime.now().year]
    return yrs


# ========================= PROMOTIONS: HUB & PREVIEW =========================

@app.route("/promotions/hub", methods=["GET"])
@require_role("admin","headteacher","dos")
def promotions_hub():
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    classes = _distinct_classes(conn)
    streams = [r[0] for r in conn.execute(
        "SELECT DISTINCT stream FROM students WHERE stream IS NOT NULL AND TRIM(stream)<>'' ORDER BY stream"
    ).fetchall()]

    f_class = (request.args.get("class") or "").strip()
    f_stream = (request.args.get("stream") or "").strip()
    f_status = (request.args.get("status") or "active").strip() # default 'active'
    q = (request.args.get("q") or "").strip()

    sql = """
      SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name,
             last_name, class_name, stream, archived
      FROM students
      WHERE 1=1
    """
    params = []
    if f_class:
        sql += " AND class_name=?"; params.append(f_class)
    if f_stream:
        sql += " AND stream=?"; params.append(f_stream)
    if f_status == "active":
        sql += " AND archived=0"
    elif f_status == "archived":
        sql += " AND archived=1"
    if q:
        sql += " AND (student_number = ? OR (first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name) LIKE ?)"
        params += [q, f"%{q}%"]

    sql += " ORDER BY class_name, last_name, first_name"
    students = conn.execute(sql, params).fetchall()
    conn.close()

    return render_template("promotions_hub.html",
                           classes=classes, streams=streams,
                           preview_students=None, source_class=None, target_class=None,
                           students=students, f_class=f_class, f_stream=f_stream, f_status=f_status, q=q)
    

@app.route("/promotions/preview", methods=["POST"])
@require_role("admin","headteacher","dos")
def promotions_preview():
    source_class = (request.form.get("source_class") or "").strip()
    target_class = (request.form.get("target_class") or "").strip()
    if not source_class or not target_class:
        flash("Please choose both Source and Target classes.", "warning")
        return redirect(url_for("promotions_hub"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT id, student_number,
               first_name, COALESCE(Middle_name,'') AS middle_name, last_name,
               class_name, COALESCE(stream,'') AS stream
        FROM students
        WHERE archived=0 AND class_name=?
        ORDER BY last_name, first_name
    """, (source_class,)).fetchall()
    classes = _distinct_classes(conn)
    conn.close()

    return render_template(
        "promotions_hub.html",
        classes=classes,
        preview_students=rows,
        source_class=source_class,
        target_class=target_class,
        promoted_students=None,
        p7_archived_count=0
    )

# ========================= PROMOTIONS: COMMIT =========================


@app.route("/promotions/commit", methods=["POST"])
@require_role("admin","headteacher","dos")
def promotions_commit():
    ensure_promotions_log_schema()
    source_class = (request.form.get("source_class") or "").strip()
    target_class = (request.form.get("target_class") or "").strip()
    if not source_class or not target_class:
        flash("Missing source/target class.", "warning")
        return redirect(url_for("promotions_hub"))

    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        students = c.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS Middle_name,
                   last_name, class_name, COALESCE(stream,'') AS stream
            FROM students
            WHERE archived=0 AND class_name=?
            ORDER BY last_name, first_name
        """, (source_class,)).fetchall()
        if not students:
            flash(f"No students found in {source_class}.", "info")
            conn.close(); return redirect(url_for("promotions_hub"))

        batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        actor = session.get("username") or session.get("role") or "system"

        if source_class == "P7":
            count = 0
            for s in students:
                _archive_student(conn, s["id"], new_status="completed", stage="P7 Leaver")
                c.execute("""
                    INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                    VALUES (?, ?, ?, ?, ?, 0)
                """, (s["id"], "P7", "ARCHIVED", actor, batch_id))
                count += 1
            conn.commit()
            flash(f"Archived {count} P7 leaver(s) as completed.", "success")
            promoted_students = []
        else:
            c.execute("UPDATE students SET class_name=? WHERE archived=0 AND class_name=?",
                      (target_class, source_class))
            c.executemany("""
                INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                VALUES (?, ?, ?, ?, ?, 0)
            """, [(s["id"], source_class, target_class, actor, batch_id) for s in students])
            conn.commit()
            flash(f"Promoted {len(students)} student(s): {source_class} → {target_class}.", "success")
            promoted_students = c.execute("""
                SELECT student_number,
                       (first_name || ' ' || COALESCE(Middle_name,'') || ' ' || last_name) AS full_name,
                       class_name, COALESCE(stream,'') AS stream
                FROM students
                WHERE archived=0 AND class_name=?
                ORDER BY last_name, first_name
            """, (target_class,)).fetchall()

        classes = _distinct_classes(conn)
        conn.close()
        return render_template("promotions_hub.html",
                               classes=classes,
                               preview_students=None,
                               source_class=source_class, target_class=target_class,
                               promoted_students=promoted_students,
                               p7_archived_count=(len(students) if source_class=="P7" else 0))
    except Exception as e:
        conn.rollback(); current_app.logger.exception("[promotions_commit] failed")
        conn.close()
        flash(f"Promotion failed: {e}", "danger")
        return redirect(url_for("promotions_hub"))


@app.route("/promotions/batch_adjacent", methods=["POST"], endpoint="promotions_bulk_adjacent")
@require_role("admin","headteacher","dos")
def promotions_bulk_adjacent():
    src = (request.form.get("source_class") or "").strip()
    direction = (request.form.get("direction") or "").strip().lower()
    if not src or direction not in ("up","down"):
        flash("Invalid batch move inputs.", "warning")
        return redirect(url_for("promotions_hub"))

    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        if src == "P7" and direction == "up":
            ids = [r["id"] for r in c.execute(
                "SELECT id FROM students WHERE archived=0 AND class_name=?", (src,)
            ).fetchall()]
            for sid in ids:
                _archive_student(conn, sid, new_status="completed", stage="P7 Leaver")
            flash(f"Archived {len(ids)} P7 leaver(s).", "success")
        else:
            target = next_class_name(src) if direction=="up" else prev_class_name(src)
            if not target:
                flash(f"No adjacent target for '{src}'.", "warning")
                conn.close(); return redirect(url_for("promotions_hub"))
            c.execute("UPDATE students SET class_name=? WHERE archived=0 AND class_name=?", (target, src))
            conn.commit()
            flash(f"Batch moved students: {src} → {target}.", "success")
    except Exception as e:
        conn.rollback(); flash(f"Batch move failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("promotions_hub"))


@app.route("/promotions/schoolwide", methods=["POST"])
@require_role("admin","headteacher","dos")
def promotions_schoolwide():
    """
    Run a single, atomic school-wide promotion in this order:
    1) P7 -> ARCHIVE, then 2) P6->P7, 3) P5->P6, ... , Baby->Middle.
    Runs once per academic year (guarded by a batch marker in promotions_log).
    """
    ensure_promotions_log_schema()
    from datetime import datetime

    # academic year + batch guard
    ay = (get_active_academic_year() or {})
    try:
        year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        year = datetime.now().year

    batch_id = f"SCHOOLWIDE-{year}"
    actor = session.get("username") or session.get("role") or "system"

    ORDER = ["Baby","Middle","Top","P1","P2","P3","P4","P5","P6","P7"]

    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        # idempotency: already done?
        done = c.execute(
            "SELECT 1 FROM promotions_log WHERE batch_id=? LIMIT 1", (batch_id,)
        ).fetchone()
        if done:
            flash(f"School-wide promotion already completed for {year}.", "info")
            conn.close()
            return redirect(url_for("promotions_hub"))

        # start atomic batch
        c.execute("BEGIN IMMEDIATE")

        # 1) Archive all active P7 as leavers (frees P7)
        p7_ids = [r["id"] for r in c.execute(
            "SELECT id FROM students WHERE archived=0 AND class_name='P7'"
        ).fetchall()]
        for sid in p7_ids:
            _archive_student(conn, sid, new_status="completed", stage="P7 Leaver")
            c.execute("""
                INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                VALUES (?, 'P7', 'ARCHIVED', ?, ?, 0)
            """, (sid, actor, batch_id))

        # 2) Move classes downward into the newly vacated targets
        # Move P6->P7, P5->P6, ..., Baby->Middle (i.e., reverse ORDER excluding P7)
        ladder = ORDER[:-1][::-1]   # ['P6','P5',...,'Baby']
        for src in ladder:
            # compute target
            idx = ORDER.index(src)
            target = ORDER[idx + 1]  # safe because we dropped last from ladder

            # collect ids to log
            ids = [r["id"] for r in c.execute(
                "SELECT id FROM students WHERE archived=0 AND class_name=?", (src,)
            ).fetchall()]
            if ids:
                c.execute("UPDATE students SET class_name=? WHERE archived=0 AND class_name=?", (target, src))
                c.executemany("""
                    INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
                    VALUES (?, ?, ?, ?, ?, 0)
                """, [(sid, src, target, actor, batch_id) for sid in ids])

        conn.commit()
        moved = sum(1 for _ in p7_ids)  # rough indicator; full total is in logs
        flash(f"School-wide promotion completed for {year}. P7 archived: {len(p7_ids)}.", "success")
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[promotions_schoolwide] failed")
        flash(f"School-wide promotion failed: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("promotions_hub"))


# ========================= PROMOTIONS: HISTORY & REVERSE =========================

@app.route("/promotions/history", methods=["GET"])
@require_role("admin","headteacher","dos")
def promotions_history():
    ensure_promotions_log_schema()
    q_student = (request.args.get("q") or "").strip()
    q_from = (request.args.get("from_class") or "").strip()
    q_to = (request.args.get("to_class") or "").strip()
    q_batch = (request.args.get("batch_id") or "").strip()

    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    classes = _distinct_classes(conn)

    sql = """
      SELECT p.*, s.student_number,
             (s.first_name || ' ' || COALESCE(s.Middle_name,'') || ' ' || s.last_name) AS full_name
      FROM promotions_log p
      JOIN students s ON s.id = p.student_id
      WHERE 1=1
    """
    args = []
    if q_student:
        sql += " AND (s.student_number = ? OR full_name LIKE ?)"
        args += [q_student, f"%{q_student}%"]
    if q_from:
        sql += " AND p.from_class = ?"; args.append(q_from)
    if q_to:
        sql += " AND p.to_class = ?"; args.append(q_to)
    if q_batch:
        sql += " AND p.batch_id = ?"; args.append(q_batch)
    sql += " ORDER BY p.created_at DESC LIMIT 500"
    rows = conn.execute(sql, args).fetchall()
    conn.close()

    return render_template("promotions_history.html", rows=rows, classes=classes,
                           q_student=q_student, q_from=q_from, q_to=q_to, q_batch=q_batch)


@app.route("/promotions/demote/<int:student_id>", methods=["POST"])
@require_role("admin","headteacher","dos")
def promotions_demote(student_id: int):
    ensure_promotions_log_schema()
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        last = c.execute("""
            SELECT id, from_class, to_class
            FROM promotions_log
            WHERE student_id=?
            ORDER BY created_at DESC
            LIMIT 1
        """, (student_id,)).fetchone()
        if not last:
            flash("No promotion history for this student.", "info")
            conn.close()
            return redirect(request.referrer or url_for("promotions_history"))

        if last["to_class"] == "ARCHIVED":
            # Unarchive P7 leaver back to their from_class (typically 'P7')
            _unarchive_student(conn, student_id)
            c.execute("UPDATE students SET class_name=?, status='active' WHERE id=?",
                      (last["from_class"], student_id))
        else:
            # Plain step back a class
            c.execute("UPDATE students SET class_name=?, status='active' WHERE id=?",
                      (last["from_class"], student_id))

        actor = session.get("username") or session.get("role") or "system"
        c.execute("""
            INSERT INTO promotions_log (student_id, from_class, to_class, actor, batch_id, reversed)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (student_id, last["to_class"], last["from_class"], actor, f"UNDO-{last['id']}"))

        conn.commit()
        flash(f"Reverted. Student returned to {last['from_class']}.", "success")
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[promotions_demote] failed")
        flash(f"Demotion failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(request.referrer or url_for("promotions_history"))


ALLOWED_STATUS = {"active","dropped","archived","completed"}

def _safe_set_status(conn, student_id: int, new_status: str) -> None:
    s = (new_status or "").strip().lower()
    if s in ALLOWED_STATUS:
        conn.execute("UPDATE students SET status=? WHERE id=?", (s, student_id))
        conn.commit()


# =================== ARCHIVE / UNARCHIVE ROUTES ===================


@app.route("/students/<int:student_id>/archive", methods=["POST"])
@require_role("admin", "headteacher")
def archive_student(student_id: int):
    """Archive a single student (any class)."""
    conn = get_db_connection()
    try:
        changed = _archive_student(
            conn,
            student_id,
            new_status="completed",
            stage="Manual Archive"
        )
        flash("Student archived." if changed else "No change.", "success" if changed else "info")

        # ---- AUDIT ----
        audit_from_request(
            conn,
            action="student_archive",
            outcome="success" if changed else "warning",
            severity="info" if changed else "warning",
            target_table="students",
            target_id=student_id,
            details={"changed": bool(changed), "stage": "Manual Archive"}
        )
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[archive_student] failed")
        flash(f"Archive failed: {e}", "danger")
        audit_from_request(
            conn,
            action="student_archive",
            outcome="failure",
            severity="warning",
            target_table="students",
            target_id=student_id,
            details={"error": str(e)}
        )
    finally:
        conn.close()
    return redirect(request.referrer or url_for("archive_hub"))


@app.route("/students/<int:student_id>/unarchive", methods=["POST"])
@require_role("admin", "headteacher")
def unarchive_student(student_id: int):
    conn = get_db_connection()
    try:
        ensure_archived_students_table(conn) # safe no-op if exists
        _unarchive_student(conn, student_id, remove_archive_rows=True)
        flash("Student restored.", "success")

        # ---- AUDIT ----
        audit_from_request(
            conn,
            action="student_unarchive",
            target_table="students",
            target_id=student_id,
            details={"remove_archive_rows": True}
        )
    except Exception as e:
        conn.rollback()
        flash(f"Unarchive failed: {e}", "danger")
        audit_from_request(
            conn,
            action="student_unarchive",
            outcome="failure",
            severity="warning",
            target_table="students",
            target_id=student_id,
            details={"error": str(e)}
        )
    finally:
        conn.close()
    return redirect(request.referrer or url_for("archive_hub"))

@app.route("/students/<int:student_id>/delete", methods=["POST"])
@require_role("admin","headteacher")
def delete_student(student_id: int):
    """Treat 'delete' as soft-archive to avoid FK errors and fill archive list."""
    conn = get_db_connection()
    try:
        _archive_student(conn, student_id, new_status="completed", stage="Manual Archive")
        flash("Student archived (soft-deleted).", "success")
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("[delete_student] failed")
        flash(f"Archive failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(request.referrer or url_for("archive_hub"))




@app.route("/archive", methods=["GET"])
@require_role("admin","headteacher","dos")
def archive_hub():
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    ensure_archived_students_table(conn)

    q_class = (request.args.get("class") or "").strip()
    q_sn = (request.args.get("student_number") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_lname = (request.args.get("last_name") or "").strip()

    sql = "SELECT * FROM archived_students WHERE 1=1"
    params = []
    if q_class:
        sql += " AND class_name = ?"; params.append(q_class)
    if q_sn:
        sql += " AND student_number = ?"; params.append(q_sn)
    if q_year:
        try:
            sql += " AND year_completed = ?"; params.append(int(q_year))
        except ValueError:
            pass
    if q_lname:
        sql += " AND full_name LIKE ? COLLATE NOCASE"; params.append(f"%{q_lname}%")

    sql += " ORDER BY year_completed DESC, class_name, full_name LIMIT 1000"
    rows = conn.execute(sql, params).fetchall()

    # dropdowns
    try:
        classes = ORDER[:] # if you defined your own class order
    except NameError:
        classes = []
    if not classes:
        classes = [r[0] for r in conn.execute(
            "SELECT DISTINCT class_name FROM archived_students ORDER BY class_name"
        ).fetchall()]
    years = [r[0] for r in conn.execute(
        "SELECT DISTINCT year_completed FROM archived_students ORDER BY year_completed DESC"
    ).fetchall()]

    conn.close()
    return render_template("archive_hub.html",
                           rows=rows, classes=classes, years=years,
                           q_class=q_class, q_sn=q_sn, q_year=q_year, q_lname=q_lname)







@app.route("/adm in/class_fees", methods=["GET", "POST"])
@require_role("admin", "headteacher", "bursar", "clerk")
def admin_class_fees():
    conn = get_db_connection()
    try:
        # If you already have this, it's fine to call; it's idempotent.
        ensure_class_fees_schema(conn)

        # dropdowns
        class_options = ["Baby","Middle","Top","P1","P2","P3","P4","P5","P6","P7"]
        section_options = ["Day", "Boarding"] # IMPORTANT: capitalized to match stored values

        if request.method == "POST":
            f = request.form
            raw_class = (f.get("class_name") or "").strip()
            raw_section = (f.get("section") or "").strip()
            raw_level = (f.get("level") or "").strip() or None
            raw_amount = (f.get("amount") or "").strip()

            # normalize
            class_name = (raw_class or "").strip().title() # "P1", "Baby", etc.
            section = norm_section(raw_section) # "Day" / "Boarding" / None
            level = raw_level # optional, can be None
            try:
                amount = float(raw_amount)
            except ValueError:
                amount = None

            if not class_name or not section or amount is None:
                flash("Please provide Class, Section (Day/Boarding), and a valid Amount.", "warning")
            else:
                # allow different amounts for Day vs Boarding
                conn.execute("""
                    INSERT INTO class_fees (class_name, section, level, amount)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(class_name, section, level) DO UPDATE SET
                        amount = excluded.amount
                """, (class_name, section, level, amount))
                conn.commit()
                flash("Class fee saved.", "success")

        # list fees
        fees = conn.execute("""
            SELECT id, class_name, section, level, amount
            FROM class_fees
            ORDER BY
              CASE class_name
                WHEN 'Baby' THEN 0 WHEN 'Middle' THEN 1 WHEN 'Top' THEN 2
                WHEN 'P1' THEN 3 WHEN 'P2' THEN 4 WHEN 'P3' THEN 5
                WHEN 'P4' THEN 6 WHEN 'P5' THEN 7 WHEN 'P6' THEN 8 WHEN 'P7' THEN 9
                ELSE 99
              END,
              COALESCE(level,''),
              section
        """).fetchall()

        return render_template(
            "admin_class_fees.html",
            class_options=class_options,
            section_options=section_options,
            fees=fees
        )
    finally:
        conn.close()



@app.route("/admin/class_fees/<int:fee_id>/delete", methods=["POST"])
@require_role("admin", "headteacher", "bursar", "clerk")
def delete_class_fee(fee_id):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM class_fees WHERE id = ?", (fee_id,))
        conn.commit()
        flash("Fee deleted.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to delete: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin_class_fees"))



# ---- Bursaries: list/add/export/import/edit/delete ----
@app.route("/admin/bursaries", methods=["GET", "POST"])
@require_role("admin", "bursar", "headteacher")
def bursaries():
    ensure_bursaries_schema()
    ay = get_active_academic_year()
    default_year = ay["year"]

    # Filters for list
    q_sn = (request.args.get("student_number") or "").strip()
    q_ln = (request.args.get("last_name") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_term = (request.args.get("term") or "").strip()

    if request.method == "POST":
        f = request.form
        student_number = (f.get("student_number") or "").strip()
        last_name = (f.get("last_name") or "").strip()

        student = _find_student_by_sn_or_ln(student_number, last_name)
        if not student:
            flash("Student not found. Use Student Number or Last Name.", "warning")
            return redirect(url_for("bursaries"))

        sponsor_name = (f.get("sponsor_name") or "").strip() or None
        year_val = int(f.get("year") or default_year)
        amount_raw = (f.get("amount") or "").strip()
        apply_to = (f.get("apply_to") or "one").lower()

        try:
            amount = float(amount_raw)
        except ValueError:
            flash("Amount must be numeric.", "danger")
            return redirect(url_for("bursaries"))

        # NEW: two-term pair selection
        if apply_to == "year":
            terms_to_apply = TERMS[:] # all three
        elif apply_to == "two":
            pair_key = (f.get("term_pair") or "").strip()
            pair_map = {
                "t1_t2": ["Term 1", "Term 2"],
                "t1_t3": ["Term 1", "Term 3"],
                "t2_t3": ["Term 2", "Term 3"],
            }
            terms_to_apply = pair_map.get(pair_key, [])
            if not terms_to_apply:
                flash("Choose a valid two-term pair.", "warning")
                return redirect(url_for("bursaries"))
        else: # 'one'
            term_one = (f.get("term_one") or "").strip()
            if term_one in TERMS:
                terms_to_apply = [term_one]
            else:
                flash("Choose a valid term.", "warning")
                return redirect(url_for("bursaries"))

        conn = get_db_connection()
        c = conn.cursor()
        try:
            for t in terms_to_apply:
                c.execute("""
                    INSERT INTO bursaries (student_id, sponsor_name, amount, term, year)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(student_id, year, term, sponsor_name)
                    DO UPDATE SET amount=excluded.amount
                """, (student["id"], sponsor_name, amount, t, year_val))
            conn.commit()
            flash(f"Bursary saved for {', '.join(terms_to_apply)}.", "success")
        except Exception as e:
            conn.rollback()
            flash(f"Failed to save bursary: {e}", "danger")
        finally:
            conn.close()
        return redirect(url_for("bursaries",
                                student_number=q_sn or None,
                                last_name=q_ln or None,
                                year=q_year or None, term=q_term or None))

    # GET: list
    where, params = ["1=1"], []
    if q_year:
        where.append("b.year = ?"); params.append(int(q_year))
    if q_term:
        where.append("b.term = ?"); params.append(q_term)
    if q_sn:
        where.append("s.student_number = ?"); params.append(q_sn)
    if q_ln:
        where.append("s.last_name LIKE ?"); params.append(f"%{q_ln}%")

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    rows = conn.execute(f"""
        SELECT b.id, b.student_id, b.sponsor_name, b.amount, b.term, b.year,
               s.student_number, s.first_name, s.Middle_name, s.last_name, s.class_name, s.stream
        FROM bursaries b
        JOIN students s ON s.id = b.student_id
        WHERE {' AND '.join(where)}
        ORDER BY b.year DESC,
                 CASE b.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END,
                 s.last_name, s.first_name
    """, params).fetchall()
    conn.close()

    return render_template("admin_bursaries.html",
                           bursaries=rows,
                           terms=TERMS,
                           default_year=default_year,
                           q_sn=q_sn, q_ln=q_ln, q_year=q_year, q_term=q_term)

@app.route("/admin/bursaries/<int:bid>/delete", methods=["POST"])
@require_role("admin","bursar","headteacher")
def delete_bursary(bid):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM bursaries WHERE id=?", (bid,))
        conn.commit()
        flash("Bursary entry deleted.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Delete failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("bursaries"))

@app.route("/admin/bursaries/<int:bid>/update", methods=["POST"])
@require_role("admin","bursar","headteacher")
def update_bursary(bid):
    sponsor_name = (request.form.get("sponsor_name") or "").strip() or None
    term = (request.form.get("term") or "").strip()
    year_raw = (request.form.get("year") or "").strip()
    amount_raw = (request.form.get("amount") or "").strip()

    if term not in TERMS:
        flash("Choose a valid term.", "warning")
        return redirect(url_for("bursaries"))
    try:
        year = int(year_raw)
        amount = float(amount_raw)
    except ValueError:
        flash("Year and amount must be numeric.", "warning")
        return redirect(url_for("bursaries"))

    conn = get_db_connection()
    try:
        conn.execute("""
            UPDATE bursaries
            SET sponsor_name = ?, amount = ?, term = ?, year = ?
            WHERE id = ?
        """, (sponsor_name, amount, term, year, bid))
        conn.commit()
        flash("Bursary updated.", "success")
    except sqlite3.IntegrityError:
        flash("That change duplicates an existing bursary for the student/term/year/sponsor.", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Update failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("bursaries"))

# ---- Export / Sample / Import ----
@app.route("/admin/bursaries/export")
@require_role("admin","bursar","headteacher")
def bursaries_export():
    q_sn = (request.args.get("student_number") or "").strip()
    q_ln = (request.args.get("last_name") or "").strip()
    q_year = (request.args.get("year") or "").strip()
    q_term = (request.args.get("term") or "").strip()

    where, params = ["1=1"], []
    if q_year:
        where.append("b.year=?"); params.append(int(q_year))
    if q_term:
        where.append("b.term=?"); params.append(q_term)
    if q_sn:
        where.append("s.student_number=?"); params.append(q_sn)
    if q_ln:
        where.append("s.last_name LIKE ?"); params.append(f"%{q_ln}%")

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    rows = conn.execute(f"""
        SELECT s.student_number, s.first_name, s.Middle_name, s.last_name,
               s.class_name, s.stream,
               b.term, b.year, b.sponsor_name, b.amount
        FROM bursaries b
        JOIN students s ON s.id=b.student_id
        WHERE {' AND '.join(where)}
        ORDER BY b.year DESC,
                 CASE b.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END,
                 s.last_name, s.first_name
    """, params).fetchall()
    conn.close()

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["student_number","first_name","Middle_name","last_name",
                     "class_name","stream","term","year","sponsor_name","amount"])
    for r in rows:
        writer.writerow([
            r["student_number"], r["first_name"], r["Middle_name"], r["last_name"],
            r["class_name"], r["stream"], r["term"], r["year"], r["sponsor_name"] or "", r["amount"] or 0
        ])
    mem = io.BytesIO(out.getvalue().encode("utf-8"))
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                     download_name=f"bursaries_{ts}.csv")

@app.route("/admin/bursaries/sample")
@require_role("admin","bursar","headteacher")
def bursaries_sample():
    headers = ["student_number","last_name","year","amount","sponsor_name","terms","apply_to","term_one","term_two"]
    sample = [
        ["P7-2025-0001","", 2025, 150000, "Sponsor A", "Term 1, Term 2", "", "", ""],
        ["", "Okello", 2025, 120000, "PTA", "", "year", "", ""],
        ["P6-2025-0003","", 2025, 80000, "", "", "two", "Term 1", "Term 3"],
    ]
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    writer.writerows(sample)
    mem = io.BytesIO(out.getvalue().encode("utf-8"))
    return send_file(mem, mimetype="text/csv", as_attachment=True,
                     download_name="bursaries_sample.csv")

try:
    import pandas as pd
except Exception:
    pd = None

@app.route("/admin/bursaries/import", methods=["POST"])
@require_role("admin","bursar","headteacher")
def bursaries_import():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Choose a CSV or Excel file.", "warning")
        return redirect(url_for("bursaries"))

    filename = file.filename.lower()
    rows = []
    try:
        if filename.endswith(".csv"):
            text = file.read().decode("utf-8", errors="ignore")
            rows = list(csv.DictReader(io.StringIO(text)))
        elif (filename.endswith(".xlsx") or filename.endswith(".xls")) and pd is not None:
            df = pd.read_excel(file)
            rows = df.to_dict(orient="records")
        else:
            text = file.read().decode("utf-8", errors="ignore")
            rows = list(csv.DictReader(io.StringIO(text)))
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        return redirect(url_for("bursaries"))

    errors, processed = [], 0
    conn = get_db_connection()
    c = conn.cursor()

    def _expand_terms_from_row(row: dict):
        def norm(t):
            t = (t or "").strip()
            return t if t in TERMS else None
        terms_mult = row.get("terms") or row.get("Terms") or row.get("TERMS") or ""
        if terms_mult:
            parts = [p.strip() for p in terms_mult.replace(";", ",").split(",")]
            expanded = [p for p in (norm(p) for p in parts) if p]
            if expanded:
                return expanded
        single = norm(row.get("term") or row.get("Term") or row.get("TERM"))
        if single:
            return [single]
        apply_to = (row.get("apply_to") or row.get("Apply_To") or "").lower().strip()
        if apply_to == "year":
            return TERMS[:]
        if apply_to == "two":
            t1 = norm(row.get("term_one") or row.get("Term_One"))
            t2 = norm(row.get("term_two") or row.get("Term_Two"))
            return [t for t in (t1, t2) if t]
        return []

    for idx, row in enumerate(rows, start=2):
        sn = (row.get("student_number") or "").strip()
        ln = (row.get("last_name") or "").strip()
        sponsor = (row.get("sponsor_name") or "").strip() or None

        try:
            year = int(str(row.get("year") or "").strip())
            amount = float(str(row.get("amount") or "").strip())
        except Exception:
            errors.append(f"Row {idx}: invalid year/amount"); continue

        student = _find_student_by_sn_or_ln(sn, ln)
        if not student:
            errors.append(f"Row {idx}: student not found (SN='{sn}' LN='{ln}')")
            continue

        terms_to_apply = _expand_terms_from_row(row)
        if not terms_to_apply:
            errors.append(f"Row {idx}: no valid term(s)")
            continue

        try:
            for t in terms_to_apply:
                c.execute("""
                    INSERT INTO bursaries (student_id, sponsor_name, amount, term, year)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(student_id, year, term, sponsor_name)
                    DO UPDATE SET amount=excluded.amount
                """, (student["id"], sponsor, amount, t, year))
                processed += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")

    try:
        conn.commit()
    except Exception as e:
        conn.rollback()
        errors.append(f"Commit failed: {e}")
    finally:
        conn.close()

    msg = f"Import complete. Processed: {processed} entries."
    if errors:
        msg += " Some issues: " + "; ".join(errors[:6]) + ("..." if len(errors) > 6 else "")
        flash(msg, "warning")
    else:
        flash(msg, "success")
    return redirect(url_for("bursaries"))



def _term_order(t: str) -> int:
    t = (t or "").strip().lower()
    return 1 if t == "term 1" else 2 if t == "term 2" else 3 if t == "term 3" else 0



@app.route("/admin/fix_fees", methods=["POST"])
@require_role("admin", "headteacher", "bursar", "clerk")
def run_fix_fees():
    conn = get_db_connection()
    try:
        # get current term/year from your helper
        ay = get_active_academic_year() # must return {'term': 'Term 3', 'year': 2025} (names you use)
        term = ay.get("term") or ay.get("term_name") or "Term 1"
        year = int(ay.get("year") or ay.get("year_number"))

        # 1) create missing fee rows
        created = ensure_fee_rows_for_all(conn, term, year)

        # 2) recalc all fees (school_fees/fees)
        updated = _recalc_all_fees(conn)

        flash(f"Fee records refreshed: {updated} updated; {created} created.", "success")
    except Exception as e:
        conn.rollback()
        app.logger.exception("run_fix_fees()")
        flash(f"Fix failed: {e}", "danger")
    finally:
        conn.close()
    return redirect(url_for("admin_class_fees"))


@app.route("/finance", methods=["GET", "POST"])
@require_role("admin","bursar","headteacher","director")
def finance_hub():
    ay = get_active_academic_year()
    f = _parse_finance_filters(request, ay)

    fees_rows, req_rows, other_rows, exp_rows, totals = _fetch_finance_data(f)
    snapshot = _balance_sheet_snapshot(f)

    return render_template(
        "finance_hub.html",
        terms=TERMS, # ["Term 1","Term 2","Term 3"]
        filters=f,
        fees_rows=fees_rows,
        req_rows=req_rows,
        other_rows=other_rows,
        exp_rows=exp_rows,
        totals=totals,
        snapshot=snapshot
    )


from flask import send_file
from io import BytesIO

@app.route("/finance/export", methods=["GET"])
@require_role("admin","bursar","headteacher","director")
def finance_export():
    """
    Export a single view to XLSX.
    ?view=fees|requirements|other|expenses|income_statement|balance_sheet
    Accepts the same filters as the hub: term, year, from_date, to_date
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        # Fallback to xlsxwriter if openpyxl not available
        pass

    ay = get_active_academic_year()
    f = _parse_finance_filters(request, ay)
    view = (request.args.get("view") or "fees").strip().lower()

    # Get hub data (re-used)
    fees_rows, req_rows, other_rows, exp_rows, totals = _fetch_finance_data(f)
    snapshot = _balance_sheet_snapshot(f)

    wb = openpyxl.Workbook()
    ws = wb.active

    def autosize(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def sheet(title, headers, rows):
        nonlocal ws
        ws.title = title[:31] or "Sheet1" # Excel sheet name limit
        ws.append(headers)
        for r in rows:
            ws.append(r)
        autosize(ws)

    # Build each sheet on demand
    if view == "fees":
        headers = ["Date", "Student No.", "Full Name", "Method", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_paid"], r["student_number"], r["full_name"], r["method"], r["term"], r["year"], r["amount_paid"]]
            for r in fees_rows
        ]
        sheet("Fees", headers, data)

    elif view == "requirements":
        headers = ["Date", "Student No.", "Full Name", "Item", "Method", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_paid"], r["student_number"], r["full_name"], r["requirement_name"], r["method"], r["term"], r["year"], r["amount_paid"]]
            for r in req_rows
        ]
        sheet("Requirements", headers, data)

    elif view == "other":
        headers = ["Date", "Source", "Description", "Recorded By", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_received"], r["source"], r["description"], r["recorded_by"], r["term"], r["year"], r["amount"]]
            for r in other_rows
        ]
        sheet("Other Income", headers, data)

    elif view == "expenses":
        headers = ["Date", "Category", "Description", "Type", "Recorded By", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_spent"], r["category"], r["description"], r["type"], r["recorded_by"], r["term"], r["year"], r["amount"]]
            for r in exp_rows
        ]
        sheet("Expenses", headers, data)

    elif view == "income_statement":
        headers = ["Account", "Amount (UGX)"]
        data = [
            ["Fees Income", totals["fees_total"]],
            ["Requirements Income", totals["requirements_total"]],
            ["Other Income", totals["other_income_total"]],
            ["Total Income", totals["income_total"]],
            ["Total Expenses", totals["expenses_total"]],
            ["Net (Income - Expenses)", totals["net_total"]],
        ]
        sheet("Income Statement", headers, data)

    elif view == "balance_sheet":
        headers = ["Account", "Amount (UGX)"]
        data = [
            ["Cash & Cash Equivalents (filtered)", snapshot["cash_in"]],
            ["Accounts Receivable (all-time)", snapshot["receivables"]],
            ["Net Position (simple)", snapshot["net_position"]],
        ]
        sheet("Balance Sheet", headers, data)

    else:
        # default to fees if unknown
        headers = ["Date", "Student No.", "Full Name", "Method", "Term", "Year", "Amount (UGX)"]
        data = [
            [r["date_paid"], r["student_number"], r["full_name"], r["method"], r["term"], r["year"], r["amount_paid"]]
            for r in fees_rows
        ]
        sheet("Fees", headers, data)

    # Save to memory and send
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Compose filename reflecting filters
    label = f["use_dates"] and f'{f["from_date"]}_to_{f["to_date"]}' or f'{f["term"]}_{f["year"]}'
    filename = f"{view}_{label}.xlsx".replace(" ", "").replace(":", "-")

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



# ---- EMPLOYEES HUB ----
@app.route("/employees", methods=["GET", "POST"], endpoint="employees_hub")
@require_role("admin", "director", "headteacher", "bursar")
def employees_hub():
    conn = get_db_connection()
    c = conn.cursor()

    # Add new employee
    if request.method == "POST":
        data = {
            "first_name": (request.form.get("first_name") or "").strip(),
            "middle_name": (request.form.get("middle_name") or "").strip(),
            "last_name": (request.form.get("last_name") or "").strip(),
            "gender": request.form.get("gender"),
            "contact": request.form.get("contact"),
            "email": request.form.get("email"),
            "residence": request.form.get("residence"),
            "department": request.form.get("department"),
            "designation": request.form.get("designation"),
            "hire_date": request.form.get("hire_date"),
            "status": request.form.get("status") or "active",
            "base_salary": float(request.form.get("base_salary") or 0),
            "allowance": float(request.form.get("allowance") or 0),
            "bonus": float(request.form.get("bonus") or 0),
            "pay_cycle": request.form.get("pay_cycle") or "monthly",
            "bank_name": request.form.get("bank_name"),
            "bank_account":request.form.get("bank_account"),
            "tin": request.form.get("tin"),
            "notes": request.form.get("notes"),
        }
        if not data["first_name"] or not data["last_name"]:
            conn.close()
            flash("First and last name are required.", "warning")
            return redirect(url_for("employees_hub"))

        c.execute("""
            INSERT INTO employees (
              first_name, middle_name, last_name, gender, contact, email, residence,
              department, designation, hire_date, status, base_salary, allowance, bonus,
              pay_cycle, bank_name, bank_account, tin, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, tuple(data.values()))
        conn.commit()
        conn.close()
        flash("Employee added.", "success")
        return redirect(url_for("employees_hub"))

    # Search/list
    q = (request.args.get("q") or "").strip()
    if q:
        rows = c.execute("""
            SELECT * FROM employees
            WHERE first_name LIKE ? OR last_name LIKE ?
               OR middle_name LIKE ? OR designation LIKE ?
               OR department LIKE ?
            ORDER BY last_name, first_name
        """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
    else:
        rows = c.execute("""
            SELECT * FROM employees
            ORDER BY (status='active') DESC, last_name, first_name
        """).fetchall()

    conn.close()
    return render_template("employees_hub.html", employees=rows, q=q)


# ---- EDIT EMPLOYEE ----
@app.route("/employees/<int:eid>/edit", methods=["GET", "POST"], endpoint="employee_edit")
@require_role("admin", "director", "headteacher")
def employee_edit(eid):
    conn = get_db_connection()
    c = conn.cursor()
    emp = c.execute("SELECT * FROM employees WHERE id=?", (eid,)).fetchone()
    if not emp:
        conn.close()
        flash("Employee not found.", "warning")
        return redirect(url_for("employees_hub"))

    if request.method == "POST":
        fields = [
            "first_name","middle_name","last_name","gender","contact","email","residence",
            "department","designation","hire_date","status","base_salary","allowance","bonus",
            "pay_cycle","bank_name","bank_account","tin","notes"
        ]
        values = [request.form.get(k) for k in fields]
        # numeric cleanup
        for i,k in enumerate(fields):
            if k in ("base_salary","allowance","bonus"):
                try: values[i] = float(values[i] or 0)
                except: values[i] = 0.0

        set_clause = ", ".join([f"{k}=?" for k in fields])
        c.execute(f"UPDATE employees SET {set_clause} WHERE id=?", (*values, eid))
        conn.commit()
        conn.close()
        flash("Employee updated.", "success")
        return redirect(url_for("employees_hub"))

    conn.close()
    return render_template("employee_edit.html", emp=emp)


# ---- (Optional) DELETE ----
@app.route("/employees/<int:eid>/delete", methods=["POST"], endpoint="employee_delete")
@require_role("admin", "director")
def employee_delete(eid):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM employees WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    flash("Employee deleted.", "info")
    return redirect(url_for("employees_hub"))
    
    
    

import io
import csv
from datetime import datetime
from flask import send_file

# ---------- helpers (shared query) ----------
def _employees_query_and_params(q=None, status=None, department=None, designation=None):
    sql = ["SELECT * FROM employees"]
    params = []
    where = []

    if q:
        where.append("(first_name LIKE ? OR middle_name LIKE ? OR last_name LIKE ? OR department LIKE ? OR designation LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like, like, like]
    if status:
        where.append("status = ?")
        params.append(status)
    if department:
        where.append("department LIKE ?")
        params.append(f"%{department}%")
    if designation:
        where.append("designation LIKE ?")
        params.append(f"%{designation}%")

    if where:
        sql.append("WHERE " + " AND ".join(where))

    sql.append("ORDER BY (status='active') DESC, last_name, first_name")
    return " ".join(sql), params

# ---------- report page ----------
@app.route("/employees/report", methods=["GET"], endpoint="employees_report")
@require_role("admin", "director", "headteacher", "bursar")
def employees_report():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    department = (request.args.get("department") or "").strip()
    designation= (request.args.get("designation") or "").strip()

    conn = get_db_connection()
    c = conn.cursor()

    sql, params = _employees_query_and_params(q, status, department, designation)
    rows = c.execute(sql, params).fetchall()

    # quick summary
    summary = {
        "count_all": len(rows),
        "count_active": sum(1 for r in rows if (r["status"] or "") == "active"),
        "count_archived": sum(1 for r in rows if (r["status"] or "") == "archived"),
        "total_base_salary": sum(float(r["base_salary"] or 0) for r in rows),
        "total_allowance": sum(float(r["allowance"] or 0) for r in rows),
        "total_bonus": sum(float(r["bonus"] or 0) for r in rows),
    }
    conn.close()

    return render_template(
        "employees_report.html",
        rows=rows,
        q=q, status=status, department=department, designation=designation,
        summary=summary,
        today=datetime.now().strftime("%Y-%m-%d")
    )

# ---------- export (CSV / Excel) ----------
@app.route("/employees/export", methods=["GET"], endpoint="employees_export")
@require_role("admin", "director", "headteacher", "bursar")
def employees_export():
    fmt = (request.args.get("format") or "csv").lower() # 'csv' or 'xlsx'
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    department = (request.args.get("department") or "").strip()
    designation= (request.args.get("designation") or "").strip()

    conn = get_db_connection()
    c = conn.cursor()
    sql, params = _employees_query_and_params(q, status, department, designation)
    rows = c.execute(sql, params).fetchall()
    conn.close()

    # Normalize rows -> list of dicts
    data = []
    for r in rows:
        data.append({
            "ID": r["id"],
            "First Name": r["first_name"],
            "Middle Name": r["middle_name"],
            "Last Name": r["last_name"],
            "Gender": r["gender"],
            "Contact": r["contact"],
            "Email": r["email"],
            "Residence": r["residence"],
            "Department": r["department"],
            "Designation": r["designation"],
            "Hire Date": r["hire_date"],
            "Status": r["status"],
            "Base Salary": r["base_salary"],
            "Allowance": r["allowance"],
            "Bonus": r["bonus"],
            "Pay Cycle": r["pay_cycle"],
            "Bank": r["bank_name"],
            "Bank Account": r["bank_account"],
            "TIN": r["tin"],
            "Notes": r["notes"],
            "Created": r["created_at"],
        })

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_filename = f"staff_report_{stamp}"

    if fmt == "xlsx":
        # Try to create a real Excel file if pandas is available; otherwise fall back to CSV
        try:
            import pandas as pd
            import tempfile, os
            df = pd.DataFrame(data)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.close()
            with pd.ExcelWriter(tmp.name, engine="xlsxwriter") as writer:
                df.to_excel(writer, index=False, sheet_name="Staff")
            return send_file(
                tmp.name,
                as_attachment=True,
                download_name=f"{base_filename}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception:
            # fall through to CSV
            pass

    # CSV export (works everywhere; Excel opens it fine)
    si = io.StringIO()
    writer = csv.DictWriter(si, fieldnames=list(data[0].keys()) if data else [
        "ID","First Name","Middle Name","Last Name","Gender","Contact","Email",
        "Residence","Department","Designation","Hire Date","Status","Base Salary",
        "Allowance","Bonus","Pay Cycle","Bank","Bank Account","TIN","Notes","Created"
    ])
    writer.writeheader()
    for row in data:
        writer.writerow(row)

    mem = io.BytesIO()
    mem.write(si.getvalue().encode("utf-8-sig")) # BOM for Excel
    mem.seek(0)
    return send_file(
        mem,
        as_attachment=True,
        download_name=f"{base_filename}.csv",
        mimetype="text/csv"
    )






# ===================== PAYROLL =====================


@app.route("/payroll", methods=["GET", "POST"])
@require_role("admin", "bursar", "director", "headteacher")
def payroll_hub():
    from datetime import datetime

    conn = get_db_connection()
    c = conn.cursor()

    # Employees for dropdowns
    employees = c.execute("""
        SELECT id, first_name, Middle_name, last_name, designation, status
        FROM employees
        ORDER BY (status='active') DESC, last_name, first_name
    """).fetchall()

    # Create a payroll row
    if request.method == "POST" and request.form.get("action") == "create":
        try:
            employee_id = int(request.form["employee_id"])
            term = request.form["term"]
            year = int(request.form["year"])
            expected_salary = float(request.form.get("expected_salary") or 0)
            bonus = float(request.form.get("bonus") or 0)
            allowance = float(request.form.get("allowance") or 0)
        except Exception:
            conn.close()
            flash("Invalid input for payroll creation.", "danger")
            return redirect(url_for("payroll_hub"))

        total = expected_salary + bonus + allowance
        status = _payroll_status(total, 0)

        c.execute("""
            INSERT INTO payroll
                (employee_id, term, year, expected_salary, bonus, allowance, total, paid_amount, status, date_paid)
            VALUES (?,?,?,?,?,?,?,?,?, DATE('now'))
        """, (employee_id, term, year, expected_salary, bonus, allowance, total, 0.0, status))
        conn.commit()
        conn.close()
        flash("Payroll row created.", "success")
        return redirect(url_for("payroll_hub"))

    # Filters
    sel_term = request.args.get("term") or ""
    sel_year = request.args.get("year") or ""
    sel_emp = request.args.get("employee_id") or ""
    q_sql, q_args = [], []
    if sel_term:
        q_sql.append("p.term = ?"); q_args.append(sel_term)
    if sel_year:
        q_sql.append("p.year = ?"); q_args.append(sel_year)
    if sel_emp:
        q_sql.append("p.employee_id = ?"); q_args.append(sel_emp)
    where = ("WHERE " + " AND ".join(q_sql)) if q_sql else ""

    # Rows
    rows = c.execute(f"""
        SELECT p.*, e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        {where}
        ORDER BY p.year DESC,
                 CASE p.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                 e.last_name, e.first_name
    """, q_args).fetchall()

    # Summary
    summary = c.execute(f"""
        SELECT
          IFNULL(SUM(p.total), 0) AS total_expected,
          IFNULL(SUM(p.paid_amount), 0) AS total_paid,
          IFNULL(SUM(p.total - p.paid_amount), 0) AS total_outstanding,
          SUM(CASE WHEN p.status='fully_paid' THEN 1 ELSE 0 END) AS cnt_fully,
          SUM(CASE WHEN p.status='partially_paid' THEN 1 ELSE 0 END) AS cnt_partial,
          SUM(CASE WHEN p.status='not_paid' THEN 1 ELSE 0 END) AS cnt_none,
          COUNT(*) AS row_count
        FROM payroll p
        {where}
    """, q_args).fetchone()

    conn.close()
    default_year = datetime.now().year

    return render_template(
        "payroll.html",
        employees=employees,
        rows=rows,
        TERMS=TERMS,
        sel_term=sel_term,
        sel_year=sel_year,
        sel_emp=sel_emp,
        default_year=default_year,
        summary=summary
    )

@app.route("/payroll/pay/<int:pid>", methods=["POST"])
@require_role("admin", "bursar", "director", "headteacher")
def payroll_add_payment(pid):
    """Add a payment to payroll AND mirror it into expenses (Salaries)."""
    # Amount
    try:
        amount = float(request.form.get("amount") or 0)
        if amount <= 0:
            raise ValueError
    except Exception:
        flash("Invalid payment amount.", "warning")
        return redirect(url_for("payroll_hub"))

    conn = get_db_connection()
    c = conn.cursor()

    # Load row + employee (for description)
    row = c.execute("""
        SELECT p.id, p.employee_id, p.term, p.year, p.total, p.paid_amount,
               e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        WHERE p.id=?
    """, (pid,)).fetchone()

    if not row:
        conn.close()
        flash("Payroll row not found.", "warning")
        return redirect(url_for("payroll_hub"))

    # Update payroll amounts + status
    new_paid = (row["paid_amount"] or 0) + amount
    status = _payroll_status(row["total"], new_paid)
    c.execute("""
        UPDATE payroll
           SET paid_amount = ?, status = ?, date_paid = DATE('now')
         WHERE id = ?
    """, (new_paid, status, pid))

    # Ensure "Salaries" category exists, then insert into expenses
    cat_id = get_or_create_expense_category(conn, "Salaries")
    emp_name = f"{(row['last_name'] or '').strip()}, {(row['first_name'] or '').strip()} {(row['Middle_name'] or '' ).strip()}".strip().strip(',')
    description = f"Salary payment - {emp_name} — {row['term']} {row['year']}"
    recorded_by = session.get("username", "system")

    c.execute("""
        INSERT INTO expenses (description, amount, term, year, date_spent, category_id, recorded_by, type)
        VALUES (?, ?, ?, ?, DATE('now'), ?, ?, 'staff_pay')
    """, (description, amount, row["term"], row["year"], cat_id, recorded_by))

    conn.commit()
    conn.close()
    flash("Payment recorded and posted to expenses (Salaries).", "success")
    return redirect(url_for("payroll_hub"))

@app.route("/payroll/edit/<int:pid>", methods=["POST"])
@require_role("admin", "bursar", "director", "headteacher")
def payroll_edit(pid):
    """Edit expected/pay components; recompute total and status."""
    try:
        expected_salary = float(request.form.get("expected_salary") or 0)
        bonus = float(request.form.get("bonus") or 0)
        allowance = float(request.form.get("allowance") or 0)
    except Exception:
        flash("Invalid amounts for edit.", "danger")
        return redirect(url_for("payroll_hub"))

    conn = get_db_connection()
    c = conn.cursor()
    row = c.execute("SELECT total, paid_amount FROM payroll WHERE id=?", (pid,)).fetchone()
    if not row:
        conn.close()
        flash("Payroll row not found.", "warning")
        return redirect(url_for("payroll_hub"))

    new_total = expected_salary + bonus + allowance
    status = _payroll_status(new_total, row["paid_amount"])

    c.execute("""
        UPDATE payroll
           SET expected_salary=?, bonus=?, allowance=?, total=?, status=?
         WHERE id=?
    """, (expected_salary, bonus, allowance, new_total, status, pid))
    conn.commit()
    conn.close()
    flash("Payroll updated.", "success")
    return redirect(url_for("payroll_hub"))

@app.route("/payroll/delete/<int:pid>", methods=["POST"])
@require_role("admin", "director")
def payroll_delete(pid):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM payroll WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    flash("Payroll row deleted.", "info")
    return redirect(url_for("payroll_hub"))

@app.route("/payroll/export")
@require_role("admin", "bursar", "director", "headteacher")
def payroll_export():
    """Export current filtered view to CSV."""
    import csv, io
    sel_term = request.args.get("term") or ""
    sel_year = request.args.get("year") or ""
    sel_emp = request.args.get("employee_id") or ""

    conn = get_db_connection()
    c = conn.cursor()
    q_sql, q_args = [], []
    if sel_term:
        q_sql.append("p.term = ?"); q_args.append(sel_term)
    if sel_year:
        q_sql.append("p.year = ?"); q_args.append(sel_year)
    if sel_emp:
        q_sql.append("p.employee_id = ?"); q_args.append(sel_emp)
    where = ("WHERE " + " AND ".join(q_sql)) if q_sql else ""

    rows = c.execute(f"""
        SELECT p.*, e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        {where}
        ORDER BY p.year DESC,
                 CASE p.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                 e.last_name, e.first_name
    """, q_args).fetchall()
    conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Employee", "Designation", "Term", "Year",
                     "Expected Salary", "Bonus", "Allowance", "Total",
                     "Paid Amount", "Status", "Last Paid"])
    for r in rows:
        fullname = f"{(r['last_name'] or '')}, {(r['first_name'] or '')} {r['Middle_name'] or ''}".strip()
        writer.writerow([
            fullname, r["designation"] or "",
            r["term"], r["year"],
            r["expected_salary"] or 0, r["bonus"] or 0, r["allowance"] or 0,
            r["total"] or 0, r["paid_amount"] or 0, r["status"], r["date_paid"] or ""
        ])
    resp = Response(buf.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = "attachment; filename=payroll.csv"
    return resp
    

# ----- Payroll Report (read-only) -----
@app.route("/reports/payroll", methods=["GET"])
@require_role("admin", "bursar", "director", "headteacher")
def payroll_report():
    from datetime import datetime

    conn = get_db_connection()
    c = conn.cursor()

    # Reuse if you already define TERMS globally; else fallback:
    TERMS_LOCAL = globals().get("TERMS", ["Term 1", "Term 2", "Term 3"])

    # Employees for filter dropdown
    employees = c.execute("""
        SELECT id, first_name, Middle_name, last_name, designation, status
        FROM employees
        ORDER BY (status='active') DESC, last_name, first_name
    """).fetchall()

    # Filters
    sel_term = request.args.get("term") or ""
    sel_year = request.args.get("year") or ""
    sel_emp = request.args.get("employee_id") or ""

    q_sql, q_args = [], []
    if sel_term:
        q_sql.append("p.term = ?"); q_args.append(sel_term)
    if sel_year:
        q_sql.append("p.year = ?"); q_args.append(sel_year)
    if sel_emp:
        q_sql.append("p.employee_id = ?"); q_args.append(sel_emp)
    where = ("WHERE " + " AND ".join(q_sql)) if q_sql else ""

    rows = c.execute(f"""
        SELECT p.*, e.first_name, e.Middle_name, e.last_name, e.designation
        FROM payroll p
        LEFT JOIN employees e ON e.id = p.employee_id
        {where}
        ORDER BY p.year DESC,
                 CASE p.term WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 9 END,
                 e.last_name, e.first_name
    """, q_args).fetchall()

    summary = c.execute(f"""
        SELECT
          IFNULL(SUM(p.total), 0) AS total_expected,
          IFNULL(SUM(p.paid_amount), 0) AS total_paid,
          IFNULL(SUM(p.total - p.paid_amount), 0) AS total_outstanding,
          SUM(CASE WHEN p.status='fully_paid' THEN 1 ELSE 0 END) AS cnt_fully,
          SUM(CASE WHEN p.status='partially_paid' THEN 1 ELSE 0 END) AS cnt_partial,
          SUM(CASE WHEN p.status='not_paid' THEN 1 ELSE 0 END) AS cnt_none,
          COUNT(*) AS row_count
        FROM payroll p
        {where}
    """, q_args).fetchone()

    conn.close()

    return render_template(
        "payroll_report.html",
        employees=employees,
        rows=rows,
        TERMS=TERMS_LOCAL,
        sel_term=sel_term,
        sel_year=sel_year,
        sel_emp=sel_emp,
        summary=summary,
        default_year=datetime.now().year
    )
   

@app.route("/comment_rules", methods=["GET","POST"])
@require_role("admin")
def comment_rules():
    ensure_comment_rules_schema()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Dropdown helpers
    classes = [r[0] for r in c.execute(
        "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL ORDER BY class_name"
    )]
    terms = ["Term 1","Term 2","Term 3"]
    # Pull grades from grading_scale for convenience
    grades = [r[0] for r in c.execute("SELECT DISTINCT grade FROM grading_scale ORDER BY 1")]
    divisions = [1,2,3,4,5,6,7,8,9]

    if request.method == "POST":
        data = request.form
        try:
            c.execute("""
              INSERT INTO comment_rules
                (role, scope, match_type, grade, division, lower_limit, upper_limit,
                 class_name, level, term, template_text, priority, active)
              VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                data.get("role"),
                data.get("scope"),
                data.get("match_type"),
                (data.get("grade") or None),
                (int(data.get("division")) if data.get("division") else None),
                (float(data.get("lower_limit")) if data.get("lower_limit") else None),
                (float(data.get("upper_limit")) if data.get("upper_limit") else None),
                (data.get("class_name") or None),
                (data.get("level") or None),
                (data.get("term") or None),
                (data.get("template_text") or "").strip(),
                int(data.get("priority") or 100),
                int(data.get("active") or 1),
            ))
            conn.commit()
            flash("Rule added.", "success")
        except Exception as e:
            app.logger.error(f"Add comment rule failed: {e}", exc_info=True)
            conn.rollback()
            flash("Failed to add rule.", "danger")

    rows = c.execute(
        "SELECT * FROM comment_rules ORDER BY role, scope, priority, id"
    ).fetchall()
    conn.close()
    return render_template(
        "comment_rules.html",
        rows=rows,
        classes=classes,
        terms=terms,
        grades=grades,
        divisions=divisions
    )

    

@app.route("/comment_rules/<int:rid>/delete", methods=["POST"])
@require_role("admin")
def delete_comment_rule(rid):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM comment_rules WHERE id=?", (rid,))
        conn.commit()
        flash("Rule deleted.", "success")
    except Exception as e:
        app.logger.error(f"Delete comment rule failed: {e}", exc_info=True)
        flash("Failed to delete rule.", "danger")
    finally:
        conn.close()
    return redirect(url_for("comment_rules"))
    


@app.route("/students/template")
@require_role("admin", "headteacher") # adjust roles
def download_students_template():
    from io import BytesIO
    import pandas as pd

    # Define expected columns
    cols = [
        "first_name","Middle_name","last_name",
        "sex","section","class_name","stream",
        "parent_name","parent_contact",
        "parent2_name","parent2_contact",
        "student_number","fees_code"
    ]
    df = pd.DataFrame(columns=cols)

    # Write empty sheet with headers
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Students")
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="students_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.route("/admin/check_ensures")
def admin_check_ensures():
    names = [
        "ensure_expenses_schema",
        "ensure_other_income_schema",
    ]
    visible = {n: callable(globals().get(n)) for n in names}
    return jsonify(visible)
    



# ================= ESC/POS RAW PRINT HELPERS (Windows) =================
import os

# Try Pillow for logo support (gracefully skip if missing)
try:
    from PIL import Image
except Exception:
    Image = None

# --- ESC/POS command bytes ---
ESC_INIT = b"\x1b\x40" # Initialize
TXT_BOLD_ON = b"\x1b\x45\x01"
TXT_BOLD_OFF = b"\x1b\x45\x00"
ALIGN_LEFT = b"\x1b\x61\x00" # Left
ALIGN_CTR = b"\x1b\x61\x01" # Center
ALIGN_RGT = b"\x1b\x61\x02" # Right
CUT_FULL = b"\x1d\x56\x00" # Full cut (some models prefer \x1d\x56\x42\x00)
FEED_6 = b"\n" * 6

def _img_to_escpos_raster(img: "Image.Image", max_width_dots: int = 576) -> bytes:
    """
    Convert a PIL image to ESC/POS Raster format (GS v 0).
    max_width_dots: 58mm ≈ 384; 80mm ≈ 576 (some printers use 640).
    """
    # Resize keeping aspect ratio
    w, h = img.size
    if w > max_width_dots:
        new_h = int(h * (max_width_dots / float(w)))
        img = img.resize((max_width_dots, max(1, new_h)), Image.LANCZOS)

    # 1-bit dithered image works best for thermal
    img = img.convert("L")
    img = img.point(lambda x: 0 if x < 160 else 255, "1") # threshold
    w, h = img.size

    row_bytes = (w + 7) // 8
    pixels = img.load()
    data = bytearray(row_bytes * h)
    idx = 0

    for y in range(h):
        byte = 0
        bit_count = 0
        for x in range(w):
            # In mode "1": black=0, white=255. For ESC/POS, 1 = black dot.
            bit = 1 if pixels[x, y] == 0 else 0
            byte = (byte << 1) | bit
            bit_count += 1
            if bit_count == 8:
                data[idx] = byte
                idx += 1
                byte = 0
                bit_count = 0
        if bit_count: # pad final byte
            byte <<= (8 - bit_count)
            data[idx] = byte
            idx += 1

    # GS v 0 m=0 xL xH yL yH data...
    xL = row_bytes & 0xFF
    xH = (row_bytes >> 8) & 0xFF
    yL = h & 0xFF
    yH = (h >> 8) & 0xFF
    header = b"\x1d\x76\x30\x00" + bytes([xL, xH, yL, yH])
    return header + bytes(data)

def _logo_payload(logo_path: str, max_width_dots: int) -> bytes:
    """Safely open JPG/PNG and return ESC/POS raster bytes. Empty bytes if not available."""
    if not (logo_path and Image):
        return b""
    try:
        path = logo_path if os.path.isabs(logo_path) else os.path.join(os.getcwd(), logo_path)
        if not os.path.exists(path):
            return b""
        with Image.open(path) as im:
            return _img_to_escpos_raster(im, max_width_dots)
    except Exception:
        return b""

def _send_raw_to_printer(payload: bytes, printer_name: str) -> bool:
    """Send RAW bytes to a Windows printer."""
    try:
        import win32print
        h = win32print.OpenPrinter(printer_name)
        try:
            win32print.StartDocPrinter(h, 1, ("ESC/POS Receipt", None, "RAW"))
            win32print.StartPagePrinter(h)
            win32print.WritePrinter(h, payload)
            win32print.EndPagePrinter(h)
            win32print.EndDocPrinter(h)
            return True
        finally:
            try:
                win32print.ClosePrinter(h)
            except Exception:
                pass
    except Exception as e:
        try:
            from flask import current_app
            current_app.logger.exception(f"[PRINT] RAW send failed: {e}")
        except Exception:
            print("[PRINT ERROR]", e)
        return False

# --------- Public helpers your code calls ----------

def print_receipt_windows_raw(text: str, printer_name: str) -> bool:
    """
    Print plain text as ESC/POS RAW (no logo).
    """
    body = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    payload = bytearray()
    payload += ESC_INIT
    payload += ALIGN_LEFT
    payload += body.encode("utf-8") + b"\n"
    payload += FEED_6 + CUT_FULL
    return _send_raw_to_printer(bytes(payload), printer_name)








def print_receipt_with_logo_windows_raw(
    fee, stu,
    printer_name: str,
    logo_path: str = "",
    school_name: str = "",
    paper_width_dots: int = 576
) -> bool:
    """
    Same logic as before; ensures the logo prints once at the top.
    Adds bold for: school name, receipt title, receipt no, student name, amounts.
    """
    # Build the body
    try:
        body_text = build_receipt_text_clean(
            fee, stu,
            school_name=current_app.config.get("SCHOOL_NAME", school_name),
            school_address=current_app.config.get("SCHOOL_ADDRESS_LINE1", ""),
            school_tagline=current_app.config.get("SCHOOL_TAGLINE", ""),
            width=int(current_app.config.get("RECEIPT_CHARS", 42)),
        )
    except TypeError:
        body_text = build_receipt_text_clean(fee, stu)

    body_text = (body_text or "").replace("\r\n", "\n").replace("\r", "\n").strip("\n")

    # Resolve logo path
    logo_cfg = (logo_path or current_app.config.get("RECEIPT_LOGO_PATH", "")).strip()
    if logo_cfg and not os.path.isabs(logo_cfg):
        logo_cfg = os.path.join(current_app.root_path, logo_cfg)

    # Cap the logo width
    max_logo = int(current_app.config.get("RECEIPT_LOGO_MAX_DOTS", 200))
    max_dots = min(max_logo, int(paper_width_dots or 576))

    # Build payload
    payload = bytearray()
    payload += ESC_INIT

    # --- Logo ---
    try:
        logo_bytes = _logo_payload(logo_cfg, max_dots) if logo_cfg else b""
    except Exception:
        logo_bytes = b""

    if logo_bytes:
        payload += ALIGN_CTR
        payload += logo_bytes + b"\n"

    # --- Body with selective bolding ---
    payload += ALIGN_LEFT
    for line in body_text.split("\n"):
        if school_name and school_name.upper() in line.upper():
            payload += ALIGN_CTR + TXT_BOLD_ON + line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n"
        elif "PAYMENT RECEIPT" in line.upper():
            payload += ALIGN_CTR + TXT_BOLD_ON + line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n"
        elif line.strip().startswith("Receipt No:"):
            payload += TXT_BOLD_ON + line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n"
        elif line.strip().startswith("Name :"):
            payload += TXT_BOLD_ON + line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n"
        elif "This Payment" in line or "Amount Due" in line:
            payload += TXT_BOLD_ON + line.encode("utf-8", "ignore") + TXT_BOLD_OFF + b"\n"
        else:
            payload += line.encode("utf-8", "ignore") + b"\n"

    # Feed and cut
    payload += FEED_6 + CUT_FULL

    return _send_raw_to_printer(bytes(payload), printer_name)




def handle_payment_and_print(fee_id: int) -> bool:
    """
    Print a receipt for fee_id.
    - Uses the receipt number stored in fees.receipt_no.
    - If fees.receipt_no is NULL/blank, generate one, SAVE it, then print.
    - Never overwrites an existing receipt_no.
    """
    from datetime import datetime
    import os
    from flask import current_app

    try:
        # Load fee + student
        fee, stu = load_payment_with_student(fee_id)
        if not (fee and stu):
            current_app.logger.warning(f"[PRINT] No fee/student for id={fee_id}")
            return False

        # --- ALWAYS come from DB ---
        rec_no = None
        try:
            rec_no = fee["receipt_no"]
        except Exception:
            rec_no = getattr(fee, "receipt_no", None)

        # If missing, generate once and persist to DB, then use that value
        if not rec_no or str(rec_no).strip() == "":
            try:
                # Use your helper if present
                rec_no = generate_receipt_no(None, fee_id)
            except Exception:
                today = datetime.now().strftime("%Y%m%d")
                rec_no = f"RCPT-{today}-{int(fee_id):06d}"

            conn = get_db_connection()
            try:
                conn.execute("UPDATE fees SET receipt_no=? WHERE id=?", (rec_no, fee_id))
                conn.commit()
            finally:
                conn.close()

            # reflect into row we pass to the builder
            fee_for_print = dict(fee)
            fee_for_print["receipt_no"] = rec_no
        else:
            fee_for_print = fee # already has table value

        # --- Print config ---
        cfg = current_app.config
        printer_name = cfg.get("RECEIPT_PRINTER_NAME", r"GP-80220(Cut) Series")
        logo_rel = cfg.get("RECEIPT_LOGO_PATH", "") or ""
        school_name = cfg.get("SCHOOL_NAME", "") or ""
        paper_dots = int(cfg.get("RECEIPT_PAPER_DOTS", 576))

        logo_abs = ""
        if logo_rel:
            logo_abs = logo_rel if os.path.isabs(logo_rel) else os.path.join(current_app.root_path, logo_rel)
            if not os.path.exists(logo_abs):
                current_app.logger.warning(f"[PRINT] Logo file not found: {logo_abs}")
                logo_abs = "" # pass empty to skip

        current_app.logger.info(
            f"[PRINT] send payment_id={fee_id}; printer='{printer_name}', rec_no='{rec_no}', logo='{logo_abs or logo_rel}'"
        )

        # --- Send to printer (header/title handled in helper) ---
        ok = print_receipt_with_logo_windows_raw(
            fee_for_print,
            stu,
            printer_name=printer_name,
            logo_path=logo_abs,
            school_name=school_name,
            paper_width_dots=paper_dots,
        )

        current_app.logger.info(f"[PRINT] done payment_id={fee_id}, success={ok}")
        return bool(ok)

    except Exception as e:
        try:
            current_app.logger.exception(f"[PRINT] send failed for payment_id={fee_id}: {e}")
        except Exception:
            print(f"[PRINT ERROR] payment_id={fee_id}: {e}")
        return False



# ===================== RECEIPT NUMBER UTILITIES (drop-in) =====================
# Requirements:
# - You already have: get_db_connection()
# - Table: fees (with columns id, date_paid, receipt_no TEXT, ...)
# If receipt_no column / unique index might be missing, call
# ensure_fees_has_receipt_no_column_and_index() once at startup.

import sqlite3
from datetime import datetime

def ensure_fees_has_receipt_no_column_and_index():
    """
    Make sure fees.receipt_no exists and is (uniquely) indexed.
    Safe to call on every startup.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cols = [r[1] for r in cur.execute("PRAGMA table_info(fees)").fetchall()]
    if "receipt_no" not in cols:
        cur.execute("ALTER TABLE fees ADD COLUMN receipt_no TEXT")
        conn.commit()
    # Unique for non-NULL values; NULLs allowed to repeat
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_fees_receipt_no ON fees(receipt_no)")
    conn.commit()
    conn.close()


def generate_receipt_no(conn, fee_id: int) -> str:
    """
    Generate or fetch a receipt number from the fees table.
    Always ensure it's consistent with what is stored.
    """
    c = conn.cursor()
    row = c.execute("SELECT receipt_no FROM fees WHERE id=?", (fee_id,)).fetchone()
    if row and row["receipt_no"]:
        return row["receipt_no"]

    # If no receipt_no yet, create one based on date + fee_id
    from datetime import datetime
    today = datetime.now().strftime("%Y%m%d")
    new_no = f"RCPT-{today}-{fee_id:06d}"

    # Save back to DB
    c.execute("UPDATE fees SET receipt_no=? WHERE id=?", (new_no, fee_id))
    conn.commit()
    return new_no


def ensure_fee_has_receipt_no(conn: sqlite3.Connection, fee_id: int) -> None:
    """
    If fees.receipt_no is NULL/blank for the given id, set it.
    Leaves existing receipt numbers untouched.
    """
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    row = cur.execute(
        "SELECT id, receipt_no, date_paid FROM fees WHERE id = ?",
        (fee_id,)
    ).fetchone()
    if not row:
        return
    current = (row["receipt_no"] or "").strip()
    if current:
        return # already set
    new_no = generate_receipt_no_for_row(row["id"], row["date_paid"])
    cur.execute("UPDATE fees SET receipt_no = ? WHERE id = ?", (new_no, row["id"]))
    conn.commit()

def backfill_missing_receipt_numbers() -> int:
    """
    Assign receipt numbers to all existing fees rows that are missing one.
    Returns number of rows updated.
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    rows = cur.execute("""
        SELECT id, date_paid
        FROM fees
        WHERE receipt_no IS NULL OR TRIM(receipt_no) = ''
        ORDER BY id
    """).fetchall()

    updated = 0
    for r in rows:
        rec = generate_receipt_no_for_row(r["id"], r["date_paid"])
        cur.execute("UPDATE fees SET receipt_no = ? WHERE id = ?", (rec, r["id"]))
        updated += 1

    conn.commit()
    conn.close()
    return updated

# ---------------------- OPTIONAL HOOKS / USAGE EXAMPLES ----------------------
# 1) Call once at startup (after your schema migrations):
# ensure_fees_has_receipt_no_column_and_index()
# backfill_missing_receipt_numbers()

# 2) Right after inserting a new payment (so printing sees the number immediately):
# conn = get_db_connection()
# cur = conn.cursor()
# cur.execute("INSERT INTO fees (...) VALUES (...)", (...,))
# fee_id = cur.lastrowid
# ensure_fee_has_receipt_no(conn, fee_id)
# conn.commit()
# conn.close()

# 3) Extra safety inside your print flow (before loading fee+student):
# def handle_payment_and_print(fee_id: int) -> bool:
# try:
# conn = get_db_connection()
# ensure_fee_has_receipt_no(conn, fee_id) # guarantees receipt_no exists
# conn.close()
# fee, stu = load_payment_with_student(fee_id)
# # ... continue with your existing printing logic ...
# except Exception:
# ...
# ===========================================================================




def _logo_bytes(path, max_width):
    if not Image: return b""
    try:
        if not os.path.isabs(path):
            path = os.path.join(os.getcwd(), path)
        if os.path.exists(path):
            with Image.open(path) as im:
                return _img_to_escpos_raster_bytes(im, max_width=max_width)
    except Exception:
        pass
    return b""


# --- JPG/PNG logo → ESC/POS + print helper -----------------------------------





def finalize_new_payment(fee_id: int):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Set receipt_no if missing
    fee = c.execute("SELECT id, receipt_no FROM fees WHERE id=?", (fee_id,)).fetchone()
    if fee and not (fee["receipt_no"] or "").strip():
        rcpt = generate_receipt_no(conn, fee_id)
        c.execute("UPDATE fees SET receipt_no=? WHERE id=?", (rcpt, fee_id))

    # Set recorded_by from the logged-in operator (fallback System)
    op_name = get_current_operator_name() # see helper below
    c.execute("UPDATE fees SET recorded_by=COALESCE(?, recorded_by) WHERE id=?", (op_name, fee_id))

    conn.commit()
    conn.close()


def get_current_operator_name() -> str:
    # adapt keys to your login/session shape
    u = session.get("user") or {}
    name = (u.get("full_name") or u.get("username") or u.get("email") or "").strip()
    return name or "System"



# ================= Pretty Receipt Builder (text for ESC/POS) =================
from datetime import datetime

# School constants for the header lines (already printed by the print helper too)
SCHOOL_TITLE = "CITIZENS DAY AND BOARDING PRIMARY SCHOOL"
SCHOOL_POBOX = "P.O. Box 31882 Kampala"

# --- tiny helpers ---
def _nz(v, d=0):
    try:
        return d if v is None else v
    except Exception:
        return d

def _get(row, key, default=None):
    try:
        # sqlite3.Row supports dict-like access
        return row[key] if row is not None else default
    except Exception:
        return default

def _fmt_money(x):
    try:
        return f"UGX {float(x):,.0f}"
    except Exception:
        return f"UGX {x}"

def _term_short(t):
    # e.g. "Term 3" -> "III", else keep as-is
    mapping = {"Term 1": "I", "Term 2": "II", "Term 3": "III"}
    return mapping.get((t or "").strip(), t or "")






def build_receipt_text_clean(
    fee,
    stu,
    *,
    school_name: str = "",
    school_address: str = "",
    school_tagline: str = "",
    width: int = 42,
) -> str:
    """Mono receipt body; uses fees.receipt_no and recorded_by. (no duplicate school name)"""
    W = max(30, int(width))

    # helpers
    def line(ch: str = "-") -> str: return ch * W
    def center(s: str) -> str: return (s or "").center(W)
    def money(x) -> str:
        try: return f"UGX {float(x or 0):,.0f}"
        except: return f"UGX {x}"
    def cols(left: str, right: str, mid: int = 24) -> str:
        mid = min(max(10, mid), W - 6)
        L = (left or "")[:mid]
        Rw = max(0, W - mid - 1)
        R = (right or "")[:Rw].rjust(Rw)
        return f"{L} {R}"
    def g(row, key, default=None):
        try:
            if isinstance(row, dict): return row.get(key, default)
            return row[key]
        except Exception:
            return default

    from datetime import datetime

    # ---- fee + student fields ----
    rec_no = g(fee, "receipt_no")
    rec_id = rec_no if rec_no else f"ID:{g(fee, 'id', '')}"
    try:
        amount = float(g(fee, "amount_paid", 0.0) or 0.0)
    except Exception:
        amount = 0.0
    term = g(fee, "term", "") or "" # keep “Term 1/2/3” as-is
    year = g(fee, "year", "") or ""
    paid_dt = g(fee, "date_paid") or datetime.now().strftime("%Y-%m-%d")
    method = g(fee, "method") or "N/A"
    cashier = g(fee, "recorded_by") or "System"

    stu_no = g(stu, "student_number", "") or ""
    first = g(stu, "first_name", "") or ""
    middle = g(stu, "Middle_name", "") or ""
    last = g(stu, "last_name", "") or ""
    klass = g(stu, "class_name", "") or ""
    stream = g(stu, "stream", "") or ""
    sid = g(stu, "id", 0)

    # Amount Due (overall balance)
    try:
        fin = compute_student_financials(sid, klass, term, int(year or datetime.now().year)) or {}
    except Exception:
        fin = {}
    try:
        overall = float(fin.get("overall_balance", 0) or 0)
    except Exception:
        overall = 0.0

    full_name = " ".join(p for p in [first, middle, last] if p).strip()
    cls_line = f"{klass}{(' ' + stream) if stream else ''}".strip()

    # ---- assemble ----
    L = []
    # Header block: (logo is added by the print function), then school name/address/tagline
    # Then the single title line and a divider. No duplicate school name afterwards.
    if school_name:
        L += [center(school_name)]
        if school_address: L.append(center(school_address))
        if school_tagline: L.append(center(school_tagline))
    # Title (the print function will render this line bold/bigger)
    L += [center("*** PAYMENT RECEIPT ***"), line("-")]

    # Body
    L += [
        cols("Receipt No:", str(rec_id)),
        "",
        cols("Name :", full_name),
        cols("Number :", stu_no),
        cols("Class :", cls_line),
        cols("Term :", f"{term}, {year}"),
        cols("Date :", paid_dt),
        cols("Method :", method),
        cols("Cashier :", cashier),
        line("-"),
        " # Payment Details".ljust(W),
        line("-"),
        cols("This Payment", money(amount)), # ONLY what is paid now
        line("-"),
        cols("Amount Due :", money(overall)),
        line("-"),
        "Thank you.".ljust(W),
    ]
    return "\n".join(L) + "\n"


 



# EXAMPLE usage inside your payment route (you already have fee_id, load helpers):
#
# fee, stu = load_payment_with_student(fee_id)
# if fee and stu:
# ok = print_receipt_with_logo_windows_raw(
# fee, stu,
# printer_name=r"GP-80220(Cut) Series", # EXACT Windows name
# logo_path="static/logo.jpg", # jpg or png
# school_name=current_app.config.get("SCHOOL_NAME", "My School"),
# paper_width_dots=576 # 80mm = 576; 58mm = 384
# )
# if ok:
# flash("Payment saved and sent to printer.", "success")
# else:
# flash("Payment saved. Printer not confirmed — open receipt and Reprint.", "warning")


def load_payment_with_student(payment_id: int):
    """Read one fees row + its student record."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    fee = c.execute("""
        SELECT id, student_id, term, year, amount_paid, method, payment_type,
               expected_amount, carried_forward, date_paid
        FROM fees
        WHERE id = ?
    """, (payment_id,)).fetchone()
    stu = None
    if fee:
        stu = c.execute("""
            SELECT id, student_number, first_name, Middle_name, last_name,
                   class_name, stream
            FROM students
            WHERE id = ?
        """, (fee["student_id"],)).fetchone()
    conn.close()
    return fee, stu

   








def handle_payment_and_print(fee_id: int) -> bool:
    """
    Print a receipt for fee_id.
    - Uses the receipt number stored in fees.receipt_no.
    - If fees.receipt_no is NULL/blank, generate one, SAVE it, then print.
    - Never overwrites an existing receipt_no.
    """
    from datetime import datetime
    import os
    from flask import current_app

    try:
        # Load fee + student
        fee, stu = load_payment_with_student(fee_id)
        if not (fee and stu):
            current_app.logger.warning(f"[PRINT] No fee/student for id={fee_id}")
            return False

        # --- ALWAYS come from DB ---
        rec_no = None
        try:
            rec_no = fee["receipt_no"]
        except Exception:
            rec_no = getattr(fee, "receipt_no", None)

        # If missing, generate once and persist to DB, then use that value
        if not rec_no or str(rec_no).strip() == "":
            try:
                # Use your helper if present
                rec_no = generate_receipt_no(None, fee_id)
            except Exception:
                today = datetime.now().strftime("%Y%m%d")
                rec_no = f"RCPT-{today}-{int(fee_id):06d}"

            conn = get_db_connection()
            try:
                conn.execute("UPDATE fees SET receipt_no=? WHERE id=?", (rec_no, fee_id))
                conn.commit()
            finally:
                conn.close()

            # reflect into row we pass to the builder
            fee_for_print = dict(fee)
            fee_for_print["receipt_no"] = rec_no
        else:
            fee_for_print = fee # already has table value

        # --- Print config ---
        cfg = current_app.config
        printer_name = cfg.get("RECEIPT_PRINTER_NAME", r"GP-80220(Cut) Series")
        logo_rel = cfg.get("RECEIPT_LOGO_PATH", "") or ""
        school_name = cfg.get("SCHOOL_NAME", "") or ""
        paper_dots = int(cfg.get("RECEIPT_PAPER_DOTS", 576))

        logo_abs = ""
        if logo_rel:
            logo_abs = logo_rel if os.path.isabs(logo_rel) else os.path.join(current_app.root_path, logo_rel)
            if not os.path.exists(logo_abs):
                current_app.logger.warning(f"[PRINT] Logo file not found: {logo_abs}")
                logo_abs = "" # pass empty to skip

        current_app.logger.info(
            f"[PRINT] send payment_id={fee_id}; printer='{printer_name}', rec_no='{rec_no}', logo='{logo_abs or logo_rel}'"
        )

        # --- Send to printer (header/title handled in helper) ---
        ok = print_receipt_with_logo_windows_raw(
            fee_for_print,
            stu,
            printer_name=printer_name,
            logo_path=logo_abs,
            school_name=school_name,
            paper_width_dots=paper_dots,
        )

        current_app.logger.info(f"[PRINT] done payment_id={fee_id}, success={ok}")
        return bool(ok)

    except Exception as e:
        try:
            current_app.logger.exception(f"[PRINT] send failed for payment_id={fee_id}: {e}")
        except Exception:
            print(f"[PRINT ERROR] payment_id={fee_id}: {e}")
        return False


# =============================== ROUTE UPDATES ==========================================



# =================== TRANSPORT as REQUIREMENT (Single Start Payment) ===================
# Minimal additions: new tables + helpers + 2 routes + a small change in start_payment GET

import sqlite3
from flask import request, redirect, url_for, flash, current_app, session

# ---- 1) Schemas (idempotent) ---------------------------------------------------------

def get_student_requirements(class_name:str, term:str, student_id:int, year:int):
    """
    Combine class requirements (your existing 'requirements' table) with any
    active student extras for the term/year. Return as a list of rows shaped
    like your template expects: id, name, amount, qty(optional).
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # Existing class requirements (include generic/no-term items, preserving your logic)
    class_rows = c.execute("""
        SELECT id, name, amount, qty
        FROM requirements
        WHERE class_name = ?
          AND (term = ? OR term IS NULL OR term = '')
    """, (class_name, term)).fetchall()

    # Active extras for this student/term/year (Transport lives here)
    extra_rows = c.execute("""
        SELECT id, item_name AS name, amount, NULL AS qty
        FROM student_extra_requirements
        WHERE student_id=? AND term=? AND year=? AND active=1
    """, (student_id, term, year)).fetchall()

    conn.close()

    # Return a single list (template treats them equally)
    # To let you identify extras in the template if desired, add a flag in memory:
    result = [dict(r) | {"_extra": False} for r in class_rows] + [dict(r) | {"_extra": True} for r in extra_rows]
    return result



# ---- 5) (Optional) Tiny admin utility to add routes -------------------------------
@app.route("/transport/routes/save", methods=["POST"])
@require_role("admin","bursar","headteacher")
def transport_route_save():
    ensure_transport_as_requirement_schema()
    name = (request.form.get("route_name") or "").strip()
    fare = float(request.form.get("fare_per_term") or 0)
    if not name:
        flash("Route name required.", "warning")
        return redirect(request.referrer or url_for("start_payment"))
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT OR REPLACE INTO transport_routes (id, name, fare_per_term) "
            "VALUES ((SELECT id FROM transport_routes WHERE name=?), ?, ?)",
            (name, name, fare)
        )
        conn.commit(); conn.close()
        flash("Route saved.", "success")
    except Exception as e:
        flash(f"Failed to save route: {e}", "danger")
    return redirect(request.referrer or url_for("start_payment"))
    
    
# ======================= TRANSPORT (as Other Income) =========================
# Uses your existing `other_income` table exactly.

import sqlite3
from datetime import datetime
from flask import (
    request, redirect, url_for, flash, render_template, make_response,
    session, current_app, jsonify
)

TRANSPORT_DESC = "Transport" # stored in other_income.description

def ensure_transport_as_req(term: str, year: int):
    # intentionally do nothing now
    return

# ---------- 1) Schema (unchanged from your version) ----------

def ensure_transport_schema(conn=None):
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True

    c = conn.cursor()
    # enable FK checks (safe even if already on)
    c.execute("PRAGMA foreign_keys = ON")

    c.execute("""
      CREATE TABLE IF NOT EXISTS transport_routes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        fare_per_term REAL NOT NULL DEFAULT 0
      )
    """)

    c.execute("""
      CREATE TABLE IF NOT EXISTS transport_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        route_id INTEGER NOT NULL,
        start_term TEXT NOT NULL,
        start_year INTEGER NOT NULL,
        active INTEGER NOT NULL DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(student_id, route_id, start_term, start_year),
        FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE,
        FOREIGN KEY(route_id) REFERENCES transport_routes(id) ON DELETE CASCADE
      )
    """)

    conn.commit()
    if close_after:
        conn.close()


def transport_get_routes(conn=None):
    """
    Return all transport routes ordered by name.
    Ensures the transport tables exist before querying.
    Accepts an optional sqlite3 connection; if omitted, opens/closes its own.
    """
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True

    # make sure tables exist
    ensure_transport_schema(conn)

    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id, name, fare_per_term FROM transport_routes ORDER BY name"
    ).fetchall()

    if close_after:
        conn.close()
    return rows





def get_class_requirements_without_transport(class_name: str, term: str):
    """
    Your original requirements (no transport). We simply filter names that look like 'Transport (...)'.
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT id, class_name, term, name, qty, amount
        FROM requirements
        WHERE class_name=? AND term=?
          AND (name NOT LIKE 'Transport (%' AND name NOT LIKE 'Transport% - %')
        ORDER BY name
    """, (class_name, term)).fetchall()
    conn.close()
    return rows


def build_virtual_transport_requirement(student_id:int, term:str, year:int):
    """
    If the student is subscribed, return a synthetic requirement row for the UI:
    {id: 't-<route_id>', name: 'Transport - <route>', qty:1, amount: fare}
    else return None.
    """
    info = transport_subscription_info(student_id, term, year)
    if not info["has_sub"]:
        return None
    return {
        "id": f"t-{info['route_id']}",
        "class_name": "", "term": term,
        "name": f"Transport - {info['route_name']}",
        "qty": 1,
        "amount": info["fare_per_term"] or 0.0,
    }
    
def transport_active_for_student(conn, student_id:int, term:str, year:int):
    """
    Returns None if not subscribed for this term.
    Else returns row with route_name, fare_per_term, route_id.
    """
    conn.row_factory = sqlite3.Row
    row = conn.execute("""
      SELECT tr.id as route_id, tr.name AS route_name, tr.fare_per_term
      FROM transport_subscriptions ts
      JOIN transport_routes tr ON tr.id = ts.route_id
      WHERE ts.active=1
        AND ts.student_id=?
        AND (ts.start_year < ? OR (ts.start_year = ? AND ts.start_term <= ?))
      ORDER BY ts.created_at DESC
      LIMIT 1
    """, (student_id, year, year, term)).fetchone()
    return row



    


TRANSPORT_PREFIX = "Transport — " # use the same text you have in requirements names

def _get_term_index(term: str) -> int:
    return {"Term 1": 1, "Term 2": 2, "Term 3": 3}.get(term, 1)



    
def compute_transport_term_status(student_id:int, class_name:str, term:str, year:int):
    """
    Compute Transport due/paid/balance for balances card.
    Paid is pulled from your existing FEES table where payment_type='requirements'
    and requirement_name matches Transport (...).
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row

    sub = transport_active_for_student(conn, student_id, term, year)
    if not sub:
        conn.close()
        return {
            "has_sub": False,
            "route_name": None,
            "fare_per_term": 0.0,
            "paid": 0.0,
            "balance": 0.0,
            "route_id": None,
        }

    req_name = f"Transport ({sub['route_name']})"

    # Due (from requirements table)
    due_row = conn.execute("""
      SELECT amount FROM requirements
      WHERE class_name=? AND term=? AND year=? AND name=?
      LIMIT 1
    """, (class_name, term, year, req_name)).fetchone()
    due = float(due_row["amount"]) if due_row else float(sub["fare_per_term"] or 0)

    # Paid (from fees entries recorded as requirements)
    paid_row = conn.execute("""
      SELECT COALESCE(SUM(amount_paid), 0) AS tot
      FROM fees
      WHERE student_id=? AND term=? AND year=?
        AND payment_type='requirements' AND requirement_name=?
    """, (student_id, term, year, req_name)).fetchone()
    paid = float(paid_row["tot"] if paid_row else 0.0)

    conn.close()
    return {
        "has_sub": True,
        "route_name": sub["route_name"],
        "fare_per_term": due,
        "paid": paid,
        "balance": max(due - paid, 0.0),
        "route_id": sub["route_id"],
    }




def transport_subscribe(student_id:int, route_id:int, term:str, year:int):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
      INSERT OR IGNORE INTO transport_subscriptions
      (student_id, route_id, start_term, start_year, active)
      VALUES (?, ?, ?, ?, 1)
    """, (student_id, route_id, term, year))
    conn.commit()
    conn.close()

def transport_unsubscribe(student_id:int, route_id:int):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
      UPDATE transport_subscriptions
         SET active=0
       WHERE student_id=? AND route_id=? AND active=1
    """, (student_id, route_id))
    conn.commit()
    conn.close()


def transport_subscription_info(student_id: int, term: str, year: int) -> dict:
    """
    Returns {'has_sub': bool, 'route_name': str, 'fare_per_term': float}
    based on active subscription that started on/before the selected term/year.
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    row = conn.execute("""
        SELECT tr.name AS route_name, tr.fare_per_term
        FROM transport_subscriptions ts
        JOIN transport_routes tr ON tr.id = ts.route_id
        WHERE ts.student_id=? AND ts.active=1
          AND (ts.start_year < ? OR (ts.start_year = ? AND ts.start_term <= ?))
        ORDER BY ts.start_year DESC,
                 CASE ts.start_term
                   WHEN 'Term 3' THEN 3
                   WHEN 'Term 2' THEN 2
                   WHEN 'Term 1' THEN 1
                   ELSE 0
                 END DESC
        LIMIT 1
    """, (student_id, year, year, term)).fetchone()
    conn.close()
    if not row:
        return {"has_sub": False, "route_name": "", "fare_per_term": 0.0}
    return {"has_sub": True,
            "route_name": row["route_name"],
            "fare_per_term": float(row["fare_per_term"] or 0.0)}

def transport_paid_via_requirements(conn, student_id: int, term: str, year: int) -> float:
    """
    Sums payments captured in 'fees' table as payment_type='requirements'
    whose requirement_name starts with 'Transport (' for the given term/year.
    """
    row = conn.execute("""
        SELECT COALESCE(SUM(amount_paid),0) AS total
          FROM fees
         WHERE student_id=? AND term=? AND year=? AND lower(payment_type)='requirements'
           AND requirement_name LIKE 'Transport (%'
    """, (student_id, term, year)).fetchone()
    return float(row["total"] if row else 0.0)


def transport_get_active_subscribers(conn, term: str, year: int, route_id: int|None=None) -> list[sqlite3.Row]:
    conn.row_factory = sqlite3.Row
    base = """
      SELECT ts.*, s.student_number, s.first_name, s.Middle_name, s.last_name, s.class_name, s.stream,
             tr.name AS route_name, tr.fare_per_term
      FROM transport_subscriptions ts
      JOIN students s ON s.id = ts.student_id
      JOIN transport_routes tr ON tr.id = ts.route_id
      WHERE ts.active=1
        AND (ts.start_year < ? OR (ts.start_year = ? AND ts.start_term <= ?))
    """
    params = [year, year, term]
    if route_id:
        base += " AND ts.route_id=?"
        params.append(route_id)
    base += " ORDER BY tr.name, s.class_name, s.last_name"
    return conn.execute(base, params).fetchall()

def transport_is_already_subscribed(conn, student_id:int, route_id:int) -> bool:
    r = conn.execute(
        "SELECT 1 FROM transport_subscriptions WHERE active=1 AND student_id=? AND route_id=?",
        (student_id, route_id)
    ).fetchone()
    return bool(r)

def transport_has_active_subscription(conn, student_id:int, term:str, year:int, route_id:int|None=None) -> bool:
    """
    True iff student has an active subscription that applies to (term, year).
    If route_id is provided, it must match; otherwise any active route qualifies.
    """
    base = """
      SELECT 1
      FROM transport_subscriptions
      WHERE active=1
        AND student_id=?
        AND (start_year < ? OR (start_year = ? AND start_term <= ?))
    """
    params = [student_id, year, year, term]
    if route_id:
        base += " AND route_id=?"
        params.append(route_id)
    row = get_db_connection().execute(base + " LIMIT 1", params).fetchone()
    return bool(row)






def _transport_source_string(sn: str, route_name: str) -> str:
    return f"Transport (SN: {sn}) - {route_name}"

def transport_record_payment_for_student_number(student_number:str, route_name:str, term:str, year:int,
                                                amount:float, method:str, recorded_by:str):
    conn = get_db_connection()
    conn.execute("""
      INSERT INTO other_income (source, amount, term, year, description, recorded_by, date_received)
      VALUES (?, ?, ?, ?, ?, ?, DATE('now'))
    """, (_transport_source_string(student_number, route_name), amount, term, year, TRANSPORT_DESC, recorded_by))
    conn.commit()
    conn.close()

def transport_paid_total_for_sn(conn, student_number:str, term:str, year:int) -> float:
    like_token = f"%SN: {student_number}%"
    row = conn.execute("""
      SELECT COALESCE(SUM(amount),0) AS total
      FROM other_income
      WHERE description=? AND term=? AND year=? AND source LIKE ?
    """, (TRANSPORT_DESC, term, year, like_token)).fetchone()
    return float(row[0] if row else 0.0)

# ---------- 3) Student lookup API (for auto-fill + guard hint) ----------

@app.route("/transport/subscribe", methods=["POST"])
@require_role("admin","bursar")
def transport_simple_subscribe():
    try:
        student_number = (request.form.get("student_number") or "").strip()
        route_id = int(request.form.get("route_id") or 0)
        term = (request.form.get("term") or "").strip()
        year = int(request.form.get("year") or 0)

        if not student_number or not route_id or not term or not year:
            flash("Missing inputs for subscribe.", "warning")
            return redirect(url_for("start_payment", student_number=student_number))

        conn = get_db_connection(); conn.row_factory = sqlite3.Row
        stu = conn.execute("SELECT id FROM students WHERE student_number=? AND archived=0", (student_number,)).fetchone()
        conn.close()
        if not stu:
            flash("Student not found or archived.", "warning")
            return redirect(url_for("start_payment", student_number=student_number))

        transport_subscribe(stu["id"], route_id, term, year)
        flash("Subscribed: transport requirement will now appear in Requirements.", "success")
        return redirect(url_for("start_payment", student_number=student_number, term=term))
    except Exception as e:
        flash(f"Subscribe failed: {e}", "danger")
        return redirect(url_for("start_payment"))

@app.route("/transport/unsubscribe", methods=["POST"])
@require_role("admin","bursar")
def transport_simple_unsubscribe():
    try:
        student_number = (request.form.get("student_number") or "").strip()
        route_id = int(request.form.get("route_id") or 0)
        term = (request.form.get("term") or "").strip()

        if not student_number or not route_id:
            flash("Missing inputs for unsubscribe.", "warning")
            return redirect(url_for("start_payment", student_number=student_number, term=term))

        conn = get_db_connection(); conn.row_factory = sqlite3.Row
        stu = conn.execute("SELECT id FROM students WHERE student_number=?", (student_number,)).fetchone()
        conn.close()
        if not stu:
            flash("Student not found.", "warning")
            return redirect(url_for("start_payment", student_number=student_number, term=term))

        transport_unsubscribe(stu["id"], route_id)
        flash("Unsubscribed: transport requirement removed for this term.", "info")
        return redirect(url_for("start_payment", student_number=student_number, term=term))
    except Exception as e:
        flash(f"Unsubscribe failed: {e}", "danger")
        return redirect(url_for("start_payment"))

# ---------- 4) CSV template (unchanged) ----------
@app.route("/transport/template")
@require_role("admin","bursar","headteacher")
def transport_template_download():
    csv_text = (
        "student_number,route_name,action,amount,method\n"
        "STU-0001,City Route,SUBSCRIBE,,\n"
        "STU-0002,City Route,PAY,50000,Cash\n"
        "STU-0003,City Route,UNSUBSCRIBE,,\n"
    )
    resp = make_response(csv_text)
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = "attachment; filename=transport_template.csv"
    return resp

# ---------- 5) Hub (kept, with pay-guard + lookup auto-fill) ----------
@app.route("/transport/hub", methods=["GET","POST"])
@require_role("admin","bursar","headteacher")
def transport_hub():
    ensure_transport_schema()
    ay = get_active_academic_year()
    active_term = ay.get("current_term") or ay.get("term") or "Term 1"
    active_year = int(ay.get("year"))

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # --- Add/Update route ---
        if action == "add_route":
            name = (request.form.get("route_name") or "").strip()
            fare = float(request.form.get("fare_per_term") or 0)
            if not name:
                flash("Route name is required.", "warning")
            else:
                try:
                    conn = get_db_connection()
                    conn.execute(
                        "INSERT OR REPLACE INTO transport_routes (id, name, fare_per_term) "
                        "VALUES ((SELECT id FROM transport_routes WHERE name=?), ?, ?)",
                        (name, name, fare)
                    )
                    conn.commit(); conn.close()
                    flash("Route saved.", "success")
                except Exception as e:
                    flash(f"Failed to save route: {e}", "danger")
            return redirect(url_for("transport_hub"))

        # --- Delete route ---
        if action == "delete_route":
            rid = request.form.get("route_id")
            try:
                conn = get_db_connection()
                conn.execute("DELETE FROM transport_routes WHERE id=?", (rid,))
                conn.commit(); conn.close()
                flash("Route deleted.", "info")
            except Exception as e:
                flash(f"Delete failed: {e}", "danger")
            return redirect(url_for("transport_hub"))

        # --- Subscribe ---
        if action == "subscribe":
            student_number = (request.form.get("student_number") or "").strip()
            route_id = int(request.form.get("route_id") or 0)
            term = (request.form.get("term") or active_term)
            year = int(request.form.get("year") or active_year)

            if not student_number or not route_id:
                flash("Student number and route are required.", "warning")
                return redirect(url_for("transport_hub"))

            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            stu = conn.execute("SELECT id FROM students WHERE archived=0 AND student_number=?", (student_number,)).fetchone()
            conn.close()
            if not stu:
                flash("Student not found or archived.", "warning")
                return redirect(url_for("transport_hub"))

            conn = get_db_connection()
            if transport_is_already_subscribed(conn, stu["id"], route_id):
                conn.close()
                flash("Already subscribed to this route.", "info")
                return redirect(url_for("transport_hub"))
            conn.close()

            transport_subscribe(stu["id"], route_id, term, year)
            flash("Subscribed.", "success")
            return redirect(url_for("transport_hub"))

        # --- Unsubscribe ---
        if action == "unsubscribe":
            student_number = (request.form.get("student_number") or "").strip()
            route_id = int(request.form.get("route_id") or 0)

            if not student_number or not route_id:
                flash("Student number and route are required.", "warning")
                return redirect(url_for("transport_hub"))

            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            stu = conn.execute("SELECT id FROM students WHERE student_number=?", (student_number,)).fetchone()
            conn.close()
            if not stu:
                flash("Student not found.", "warning")
                return redirect(url_for("transport_hub"))

            transport_unsubscribe(stu["id"], route_id)
            flash("Unsubscribed.", "success")
            return redirect(url_for("transport_hub"))

        # --- Pay (into other_income) with GUARD ---
        if action == "pay":
            student_number = (request.form.get("student_number") or "").strip()
            route_id = int(request.form.get("route_id") or 0)
            amount = float(request.form.get("amount") or 0)
            method = (request.form.get("method") or "Cash").strip()
            term = (request.form.get("term") or active_term)
            year = int(request.form.get("year") or active_year)

            if not student_number or amount <= 0:
                flash("Student number and a positive amount are required.", "warning")
                return redirect(url_for("transport_hub"))

            # Resolve student + subscription check
            conn = get_db_connection()
            conn.row_factory = sqlite3.Row
            stu = conn.execute("SELECT id FROM students WHERE archived=0 AND student_number=?", (student_number,)).fetchone()
            conn.close()
            if not stu:
                flash("Student not found or archived.", "warning")
                return redirect(url_for("transport_hub"))

            if not transport_has_active_subscription(get_db_connection(), stu["id"], term, year, (route_id or None)):
                flash("Payment rejected: student is NOT an active transport subscriber for this term/route.", "danger")
                return redirect(url_for("transport_hub"))

            # Pretty route name (optional)
            route_name = ""
            if route_id:
                conn = get_db_connection()
                row = conn.execute("SELECT name FROM transport_routes WHERE id=?", (route_id,)).fetchone()
                conn.close()
                route_name = row["name"] if row else ""

            recorded_by = session.get("full_name") or session.get("username") or session.get("role") or "system"
            try:
                transport_record_payment_for_student_number(student_number, route_name, term, year, amount, method, recorded_by)
                flash("Transport payment recorded (Other Income).", "success")
            except Exception as e:
                current_app.logger.exception("[transport pay] insert failed")
                flash(f"Payment failed: {e}", "danger")

            return redirect(url_for("transport_hub"))

        # --- Bulk CSV (unchanged logic) ---
        if action == "upload_csv":
            file = request.files.get("file")
            if not file:
                flash("No file provided.", "warning")
                return redirect(url_for("transport_hub"))

            import pandas as pd
            df = pd.read_csv(file) if file.filename.lower().endswith(".csv") else pd.read_excel(file)
            ok, skip = 0, 0

            for _, r in df.iterrows():
                try:
                    sn = str(r.get("student_number") or "").strip()
                    rname = str(r.get("route_name") or "").strip()
                    act = str(r.get("action") or "").strip().upper()

                    # resolve route id (if provided)
                    route_id = None
                    if rname:
                        conn = get_db_connection()
                        row = conn.execute("SELECT id FROM transport_routes WHERE name=?", (rname,)).fetchone()
                        conn.close()
                        route_id = row["id"] if row else None

                    if act == "SUBSCRIBE":
                        if not route_id: skip += 1; continue
                        conn = get_db_connection(); conn.row_factory = sqlite3.Row
                        stu = conn.execute("SELECT id FROM students WHERE archived=0 AND student_number=?", (sn,)).fetchone()
                        conn.close()
                        if not stu: skip += 1; continue
                        transport_subscribe(stu["id"], route_id, active_term, active_year)
                        ok += 1

                    elif act == "UNSUBSCRIBE":
                        if not route_id: skip += 1; continue
                        conn = get_db_connection(); conn.row_factory = sqlite3.Row
                        stu = conn.execute("SELECT id FROM students WHERE student_number=?", (sn,)).fetchone()
                        conn.close()
                        if not stu: skip += 1; continue
                        transport_unsubscribe(stu["id"], route_id); ok += 1

                    elif act == "PAY":
                        amt = float(r.get("amount") or 0)
                        if amt <= 0: skip += 1; continue
                        method = str(r.get("method") or "Cash")

                        # optional guard here too: require active sub (any route)
                        conn = get_db_connection(); conn.row_factory = sqlite3.Row
                        stu = conn.execute("SELECT id FROM students WHERE archived=0 AND student_number=?", (sn,)).fetchone()
                        conn.close()
                        if not stu or not transport_has_active_subscription(get_db_connection(), stu["id"], active_term, active_year, route_id):
                            skip += 1; continue

                        transport_record_payment_for_student_number(sn, rname or "", active_term, active_year, amt, method,
                            session.get("full_name") or session.get("username") or "system")
                        ok += 1
                    else:
                        skip += 1
                except Exception:
                    skip += 1

            flash(f"Processed: {ok}, Skipped: {skip}.", "info")
            return redirect(url_for("transport_hub"))

    # GET: show lists/balances
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    routes = transport_get_routes(conn)

    route_filter = request.args.get("route_id")
    try:
        route_filter_id = int(route_filter) if route_filter else None
    except Exception:
        route_filter_id = None

    subs = transport_get_active_subscribers(conn, active_term, active_year, route_filter_id)

    balances = []
    for row in subs:
        due = float(row["fare_per_term"] or 0.0)
        paid = transport_paid_total_for_sn(conn, row["student_number"], active_term, active_year)
        balances.append({
            "subscription_id": row["id"],
            "student_number": row["student_number"],
            "full_name": f"{row['first_name']} {row['Middle_name'] or ''} {row['last_name']}".replace(" "," ").strip(),
            "class_name": row["class_name"],
            "stream": row["stream"],
            "route_name": row["route_name"],
            "fare_per_term": due,
            "paid": paid,
            "balance": max(due - paid, 0.0),
        })

    conn.close()

    return render_template(
        "transport_hub.html",
        routes=routes,
        balances=balances,
        active_term=active_term,
        active_year=active_year,
        TERMS=TERMS,
        route_filter=(route_filter_id or "")
    )
    
    
    
    
from flask import jsonify

@app.route("/api/student/by_number")
@require_role("admin","bursar","headteacher")
def api_student_by_number():
    """
    Returns basic student info (active only) and current transport route (if subscribed).
    Query: ?sn=STU-0001
    Response:
      { success: true,
        data: { id, student_number, full_name, class_name, stream,
                route_id, route_name } }
      or { success: false, message: "..." }
    """
    sn = (request.args.get("sn") or "").strip()
    if not sn:
        return jsonify({"success": False, "message": "Missing student number."}), 400

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    stu = conn.execute("""
        SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS mid, last_name,
               class_name, stream
        FROM students
        WHERE archived=0 AND student_number=?
        LIMIT 1
    """, (sn,)).fetchone()

    if not stu:
        conn.close()
        return jsonify({"success": False, "message": "Student not found or archived."}), 404

    # try find an active transport subscription (any route)
    sub = conn.execute("""
        SELECT ts.route_id, tr.name AS route_name
        FROM transport_subscriptions ts
        JOIN transport_routes tr ON tr.id=ts.route_id
        WHERE ts.active=1 AND ts.student_id=?
        ORDER BY ts.created_at DESC
        LIMIT 1
    """, (stu["id"],)).fetchone()
    conn.close()

    full_name = f"{stu['first_name']} {stu['mid']} {stu['last_name']}".replace(" "," ").strip()
    data = {
        "id": stu["id"],
        "student_number": stu["student_number"],
        "full_name": full_name,
        "class_name": stu["class_name"],
        "stream": stu["stream"],
        "route_id": (sub["route_id"] if sub else None),
        "route_name": (sub["route_name"] if sub else None),
    }
    return jsonify({"success": True, "data": data})
# --------------------------------------------------------------------

# ===================== /TRANSPORT (as Other Income) =====================
# ================= /TRANSPORT as REQUIREMENT (Single Start Payment) ===================



# 2) /fees/pay (unchanged UX; after commit we print via the wrapper)

@app.route("/fees/pay", methods=["POST"])
@require_role("admin", "bursar", "headteacher", "clerk")
def pay_fees():
    f = request.form
    student_id = int(f.get("student_id") or 0)
    amount_paid = float(f.get("amount_paid") or 0)
    method = (f.get("method") or "Cash").strip()
    payment_type= (f.get("payment_type") or "fees").strip()

    if not student_id or amount_paid <= 0:
        flash("Student and a positive amount are required.", "warning")
        return redirect(request.referrer or url_for("dashboard"))

    ay = get_active_academic_year()
    term = f.get("term") or ay.get("current_term") or ay.get("term") or "Term 1"
    year = int(f.get("year") or ay.get("year"))

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    try:
        c.execute("""
            INSERT INTO fees (
                student_id, term, year, amount_paid, method, payment_type, date_paid, recorded_by
            ) VALUES (?, ?, ?, ?, ?, ?, DATE('now'), ?)
        """, (student_id, term, year, amount_paid, method, payment_type, (session.get("username") or "system")))
        fee_id = c.lastrowid
        conn.commit()

        ok = handle_payment_and_print(fee_id)
        if ok:
            flash("Payment saved and sent to printer.", "success")
        else:
            flash("Payment saved. Printer not confirmed — open the receipt and try Reprint.", "warning")

        return redirect(url_for("receipt_view", payment_id=fee_id))

    except Exception as e:
        conn.rollback()
        current_app.logger.exception(f"[fees/pay] insert failed: {e}")
        flash(f"Failed to save payment: {e}", "danger")
        return redirect(request.referrer or url_for("dashboard"))
    finally:
        conn.close()




# ========================= START PAYMENT (with Transport-as-Requirement) =========================
# Assumes you pasted earlier helpers:
# ensure_transport_as_requirement_schema, get_student_requirements
# If not, paste those first (from my previous message).


# --- Put these near your other helpers (once) -------------------------------
TRANSPORT_PREFIX = "Transport — " # must match the names you inject in requirements

def _term_index(t: str) -> int:
    return 1 if t == "Term 1" else 2 if t == "Term 2" else 3


def transport_subscription_info(student_id: int, term: str, year: int):
    """
    Return the active subscription for this student at/before (term,year).
    {has_sub: bool, route_id: int|None, route_name: str, fare_per_term: float}
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    t_idx = _term_index(term)

    row = conn.execute(
        """
        SELECT ts.route_id, tr.name AS route_name, tr.fare_per_term
        FROM transport_subscriptions ts
        JOIN transport_routes tr ON tr.id = ts.route_id
        WHERE ts.student_id = ? AND ts.active=1
          AND (ts.start_year < ?
               OR (ts.start_year = ? AND
                   CASE ts.start_term
                     WHEN 'Term 1' THEN 1
                     WHEN 'Term 2' THEN 2
                     WHEN 'Term 3' THEN 3
                     ELSE 1
                   END <= ?)
              )
        ORDER BY ts.created_at DESC
        LIMIT 1
        """,
        (student_id, year, year, t_idx)
    ).fetchone()
    conn.close()

    if not row:
        return {"has_sub": False, "route_id": None, "route_name": "", "fare_per_term": 0.0}
    return {
        "has_sub": True,
        "route_id": int(row["route_id"]),
        "route_name": row["route_name"],
        "fare_per_term": float(row["fare_per_term"] or 0.0),
    }




def transport_subscribe(student_id:int, route_id:int, term:str, year:int):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
      INSERT OR IGNORE INTO transport_subscriptions (student_id, route_id, start_term, start_year, active)
      VALUES (?, ?, ?, ?, 1)
    """, (student_id, route_id, term, year))
    # If there was an inactive sub for the same route earlier, make it active:
    c.execute("""
      UPDATE transport_subscriptions
         SET active=1
       WHERE student_id=? AND route_id=? AND start_term=? AND start_year=? AND active=0
    """, (student_id, route_id, term, year))
    conn.commit()
    conn.close()



@app.route("/fees/<int:fee_id>/print")
@require_role("admin","clerk","headteacher","bursar")
def print_fee_receipt(fee_id):
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    c = conn.cursor()
    fee = c.execute("SELECT * FROM fees WHERE id=?", (fee_id,)).fetchone()
    if not fee:
        conn.close()
        flash("Payment not found.", "danger")
        return redirect(url_for("start_payment"))

    stu = c.execute("SELECT * FROM students WHERE id=?", (fee["student_id"],)).fetchone()
    conn.close()

    ok = send_receipt_to_printer(
        fee, stu,
        printer_name=current_app.config["RECEIPT_PRINTER_NAME"],
        school_name=current_app.config.get("SCHOOL_NAME","School"),
        logo_path=current_app.config.get("RECEIPT_LOGO_PATH"),
        paper_width_dots=current_app.config.get("RECEIPT_PAPER_DOTS",576)
    )
    flash("Receipt sent to printer." if ok else "Could not print receipt.", "success" if ok else "warning")
    return redirect(request.referrer or url_for("start_payment"))


@app.route("/receipt/reprint/<int:fee_id>")
@require_role("admin", "bursar", "headteacher")
def reprint_receipt(fee_id):
    """Reprint an existing receipt by fee_id."""
    try:
        ok = handle_payment_and_print(fee_id)
        if ok:
            flash(f"Receipt {fee_id} sent to printer.", "success")
        else:
            flash(f"Receipt {fee_id} could not be printed.", "warning")
    except Exception as e:
        current_app.logger.exception(f"[REPRINT] failed fee_id={fee_id}: {e}")
        flash("Unexpected error while reprinting.", "danger")
    return redirect(request.referrer or url_for("student_statement"))





@app.route("/start_payment", methods=["GET", "POST"])
@require_role("admin", "bursar")
def start_payment():
    # Academic context
    ay = get_active_academic_year()
    current_term = ay.get("current_term") or ay.get("term") or "Term 1"
    current_year = int(ay.get("year"))
    sel_term = (request.values.get("term") or current_term).strip()

    # Dropdown selection
    q_student_id = request.values.get("student_id", type=int)

    # --- helpers -------------------------------------------------------------
    def _get_student_by_id(stid: int):
        if not stid:
            return None
        conn = get_db_connection(); conn.row_factory = sqlite3.Row
        row = conn.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS middle_name,
                   last_name, class_name, stream, section
              FROM students
             WHERE id=? AND archived=0
        """, (stid,)).fetchone()
        conn.close()
        return row

    def _all_students_for_dropdown():
        conn = get_db_connection(); conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT id, student_number, first_name, COALESCE(Middle_name,'') AS m,
                   last_name, class_name, COALESCE(stream,'') AS stream
              FROM students
             WHERE archived=0
             ORDER BY last_name, first_name
        """).fetchall()
        conn.close()
        return rows

    # -------------------- POST: record payment (kept logic) ------------------
    if request.method == "POST":
        sid = request.form.get("sid", type=int)
        if not sid:
            flash("Select a student before recording payment.", "warning")
            return redirect(url_for("start_payment", student_id=q_student_id, term=sel_term))

        payment_type = (request.form.get("payment_type") or "school_fees").strip()
        method = (request.form.get("method") or "cash").strip()
        term = (request.form.get("term") or current_term).strip()
        year = int(request.form.get("year") or current_year)
        comment = (request.form.get("comment") or request.form.get("payment_item") or "").strip()
        recorded_by = session.get("full_name") or session.get("username") or session.get("role") or "system"

        conn = get_db_connection()
        c = conn.cursor()
        try:
            if payment_type == "requirements":
                ids = request.form.getlist("req_id[]")
                names = request.form.getlist("req_name[]")
                amts = request.form.getlist("req_amount[]")
                total = 0.0
                for rid, rname, ramt in zip(ids, names, amts):
                    try: amt = float(ramt)
                    except ValueError: amt = 0.0
                    if amt <= 0: continue
                    total += amt
                    c.execute("""
                        INSERT INTO fees (
                          student_id, term, year, amount_paid,
                          expected_amount, bursary_amount, carried_forward,
                          date_paid, method, payment_type, requirement_name,
                          comment, recorded_by
                        ) VALUES (?, ?, ?, ?, ?, 0, 0, DATE('now'), ?, 'requirements', ?, ?, ?)
                    """, (sid, term, year, amt, amt, method, rname, comment, recorded_by))
                conn.commit()
                flash(f"Requirements payment recorded (UGX {total:,.0f}).", "success")
            else:
                amount_paid_raw = (request.form.get("amount_paid") or "0").strip()
                try: amount_paid = float(amount_paid_raw)
                except ValueError: amount_paid = 0.0
                if amount_paid <= 0:
                    flash("Amount must be greater than zero.", "warning")
                    conn.close()
                    return redirect(url_for("start_payment", student_id=q_student_id, term=term))

                c.execute("""
                    INSERT INTO fees (
                      student_id, term, year, amount_paid,
                      expected_amount, bursary_amount, carried_forward,
                      date_paid, method, payment_type, comment, recorded_by
                    ) VALUES (?, ?, ?, ?, ?, 0, 0, DATE('now'), ?, 'fees', ?, ?)
                """, (sid, term, year, amount_paid, amount_paid, method, comment, recorded_by))
                conn.commit()
                flash("Fees payment recorded.", "success")

                fee_id = c.lastrowid
                try:
                    current_app.logger.info(f"[PRINT] start_payment -> printing id={fee_id}")
                    handle_payment_and_print(fee_id)
                except Exception as e:
                    current_app.logger.exception(f"[PRINT] failed: {e}")
        except Exception as e:
            conn.rollback()
            flash(f"Payment failed: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("start_payment", student_id=sid, term=term))

    # ---------------------- GET: student + reqs + balances -------------------
    student = _get_student_by_id(q_student_id)
    students_list = _all_students_for_dropdown()
    reqs, fin, transport = [], None, None

    # transport routes for quick subscribe
    try:
        routes = transport_get_routes()
    except Exception:
        routes = []

    if student:
        reqs = get_class_requirements(student["class_name"], sel_term)
        fin = compute_student_financials(student["id"], student["class_name"], sel_term, current_year)

        try:
            tinfo = transport_subscription_info(student["id"], sel_term, current_year)
        except Exception:
            tinfo = None
        if tinfo and float(tinfo.get("fare_per_term") or 0) > 0:
            conn = get_db_connection()
            try:
                tp_paid = transport_paid_via_requirements(conn, student["id"], sel_term, current_year)
            finally:
                conn.close()
            transport = {
                "route_name": tinfo["route_name"],
                "fare_per_term": float(tinfo["fare_per_term"] or 0.0),
                "paid": tp_paid,
                "balance": max(float(tinfo["fare_per_term"] or 0.0) - tp_paid, 0.0),
            }
            reqs.append({
                "id": "transport",
                "name": f"Transport ({tinfo['route_name']})",
                "qty": 1,
                "amount": float(tinfo["fare_per_term"] or 0.0),
            })

    return render_template(
        "start_payment.html",
        terms=TERMS,
        current_term=sel_term or current_term,
        current_year=current_year,
        student=student,
        students_list=students_list, # <-- for dropdown
        reqs=reqs,
        fin=fin,
        transport=transport,
        routes=routes,
    )




@app.route("/payments/confirm", methods=["POST"])
@require_role("admin", "bursar", "clerk", "headteacher")
def payments_confirm():
    f = request.form
    ay = get_active_academic_year()
    term = (f.get("term") or ay.get("current_term") or ay.get("term") or "Term 1").strip()
    year = int(f.get("year") or ay.get("year") or datetime.now().year)

    # Resolve student
    student_id = f.get("student_id")
    student_number = (f.get("student_number") or "").strip()
    if not student_id and student_number:
        conn = get_db_connection()
        try:
            r = conn.execute(
                "SELECT id FROM students WHERE student_number=? AND archived=0",
                (student_number,)
            ).fetchone()
            if not r:
                flash("Student not found.", "warning")
                return redirect(url_for("dashboard"))
            student_id = r["id"]
        finally:
            conn.close()

    try:
        student_id = int(student_id)
    except (TypeError, ValueError):
        flash("Invalid or missing student.", "danger")
        return redirect(url_for("dashboard"))

    # Amount + basics
    try:
        amount_paid = float(f.get("amount_paid") or 0)
        if amount_paid <= 0:
            raise ValueError
    except Exception:
        flash("Enter a valid positive amount.", "danger")
        return redirect(url_for("dashboard"))

    method = (f.get("method") or "N/A").strip()
    payment_type = (f.get("payment_type") or "fees").strip().lower()
    payment_item = (f.get("payment_item") or "").strip()
    recorded_by = session.get("username") or session.get("role") or "System"

    # Save
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO fees (
              student_id, term, year, amount_paid,
              payment_item, bursary_amount, carried_forward, expected_amount,
              date_paid, method, payment_type, recorded_by
            ) VALUES (?, ?, ?, ?, ?, 0, 0, 0, DATE('now'), ?, ?, ?)
        """, (student_id, term, year, amount_paid, payment_item, method, payment_type, recorded_by))
        fee_id = cur.lastrowid
        conn.commit()
    except Exception as e:
        conn.rollback()
        current_app.logger.exception("Failed saving payment")
        flash(f"Failed to save payment: {e}", "danger")
        return redirect(url_for("dashboard"))
    finally:
        conn.close()

    # Auto print
    ok = handle_payment_and_print(fee_id)
    if ok:
        flash("Payment saved and sent to printer.", "success")
    else:
        flash("Payment saved. Printer not confirmed — open the receipt and use Reprint.", "warning")

    return redirect(url_for("receipt_view", payment_id=fee_id))
    

@app.route("/receipt/reprint/<int:payment_id>", methods=["POST"])
@require_role("admin", "bursar", "headteacher")
def receipt_reprint(payment_id: int):
    ok = handle_payment_and_print(payment_id)
    if ok:
        flash("Receipt sent to printer.", "success")
    else:
        flash("Printer not confirmed. Check connection and try again.", "warning")
    return redirect(url_for("receipt_view", payment_id=payment_id))
# ========================================================================================



  

# ---------- Opening Balance: BULK UPLOAD ----------
from werkzeug.utils import secure_filename
import csv
import io
import pandas as pd

def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")

def _read_opening_rows(file_storage, filename: str):
    """
    Returns a list of dict rows with keys:
      student_number (str), amount (float), asof_year (int|None)
    Accepts .csv or .xlsx
    """
    name = (filename or "").lower()
    raw = file_storage.read()

    rows = []
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = [ _normalize_header(h) for h in reader.fieldnames or [] ]
        for r in reader:
            row = { _normalize_header(k): (v or "").strip() for k, v in r.items() }
            rows.append(row)

    elif name.endswith(".xlsx"):
        bio = io.BytesIO(raw)
        df = pd.read_excel(bio, dtype=str) # read everything as text first
        df.columns = [_normalize_header(c) for c in df.columns]
        for _, r in df.fillna("").iterrows():
            rows.append({k: str(v).strip() for k, v in r.to_dict().items()})

    else:
        raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")

    # Map & coerce
    normalized = []
    for i, r in enumerate(rows, start=2): # start=2: header is row 1
        sn = (r.get("student_number") or r.get("studentno") or "").strip()
        amt_raw = (r.get("amount") or r.get("balance") or "").strip()
        asof_raw = (r.get("asof_year") or r.get("as_of_year") or r.get("year") or "").strip()

        if not sn or not amt_raw:
            normalized.append({
                "_row": i, "student_number": sn, "amount": None, "asof_year": None,
                "_error": "Missing student_number or amount"
            })
            continue

        try:
            amt = float(amt_raw)
        except Exception:
            normalized.append({
                "_row": i, "student_number": sn, "amount": None, "asof_year": None,
                "_error": f"Amount '{amt_raw}' is not a number"
            })
            continue

        asof_year = None
        if asof_raw:
            try:
                asof_year = int(asof_raw)
            except Exception:
                normalized.append({
                    "_row": i, "student_number": sn, "amount": amt, "asof_year": None,
                    "_error": f"asof_year '{asof_raw}' is not a valid year"
                })
                continue

        normalized.append({
            "_row": i, "student_number": sn, "amount": amt, "asof_year": asof_year
        })
    return normalized
    






def set_opening_balance(conn, student_id: int, amount: float, year: int, note: str = ""):
    """
    Insert an opening balance record into the fees table.
    Uses payment_type='opening_balance' so it's distinguishable.
    """
    conn.execute("""
        INSERT INTO fees (
            student_id, term, year,
            expected_amount, bursary_amount, amount_paid,
            date_paid, method, payment_type, recorded_by, payment_item
        )
        VALUES (?, 'Term 3', ?, ?, 0, 0, DATE('now'), 'N/A', 'opening_balance', ?, ?)
    """, (
        student_id, int(year), float(amount),
        session.get("username") or "system",
        note
    ))
    
    
def carried_forward(student_id, term, year):
    """Outstanding before the active term/year + any opening_balance rows (always included)."""
    conn = get_db_connection()
    prev_rows = conn.execute("""
        SELECT expected_amount, bursary_amount, amount_paid, term, year
        FROM fees
        WHERE student_id=? AND (
              year < ?
           OR (year = ? AND
               (CASE term
                 WHEN 'Term 1' THEN 1
                 WHEN 'Term 2' THEN 2
                 WHEN 'Term 3' THEN 3
                 ELSE 99
               END)
               < (CASE ? WHEN 'Term 1' THEN 1 WHEN 'Term 2' THEN 2 WHEN 'Term 3' THEN 3 ELSE 99 END))
        ) AND (payment_type IS NULL OR payment_type!='requirements')
    """, (student_id, year, year, term)).fetchall()

    out_prev = 0.0
    for r in prev_rows:
        out_prev += (float(r["expected_amount"] or 0) -
                     float(r["bursary_amount"] or 0) -
                     float(r["amount_paid"] or 0))

    # Always include all opening_balance rows (no date/term filter)
    ob_row = conn.execute("""
        SELECT COALESCE(SUM(expected_amount - bursary_amount - amount_paid), 0) AS t
        FROM fees
        WHERE student_id=? AND payment_type='opening_balance'
    """, (student_id,)).fetchone()
    conn.close()

    total = max(out_prev, 0.0) + float(ob_row["t"] or 0.0)
    return max(total, 0.0)
    


@app.route("/balances/opening-balance/bulk", methods=["GET", "POST"])
@require_role("admin", "bursar", "clerk")
def opening_balance_bulk():
    """
    Upload a CSV/XLSX with columns:
      - student_number (required)
      - amount (required)
      - asof_year (optional; falls back to active year)

    Saves one row per student into fees table using payment_type='opening_balance'.
    """
    ay = get_active_academic_year()
    active_year = int(ay.get("year"))

    # GET: show page
    if request.method == "GET":
        sample = [
            {"row": 1, "student_number": "STD-2025-001", "amount": 150000, "asof_year": active_year},
            {"row": 2, "student_number": "STD-2025-002", "amount": 90000, "asof_year": active_year},
        ]
        return render_template("opening_balance_bulk.html",
                               active_year=active_year, results=[],
                               sample_rows=sample)

    # POST: process upload
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose a CSV or XLSX file.", "warning")
        return redirect(url_for("opening_balance_bulk"))

    # --- Load rows (CSV/XLSX) ---
    try:
        fname = file.filename.lower()
        if fname.endswith(".csv"):
            df = pd.read_csv(file)
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            df = pd.read_excel(file)
        else:
            flash("Only .csv, .xlsx are supported.", "danger")
            return redirect(url_for("opening_balance_bulk"))
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        return redirect(url_for("opening_balance_bulk"))

    # Normalize column names
    cols = {c.strip().lower(): c for c in df.columns}
    need = {"student_number", "amount"}
    if not need.issubset(cols.keys()):
        flash("Required columns: student_number, amount (optional: asof_year).", "danger")
        return redirect(url_for("opening_balance_bulk"))

    # Prepare iteration
    results = []
    ok = bad = 0

    # Open one DB connection for all rows
    conn = get_db_connection()

    # Iterate rows
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        sn = str(row.get(cols.get("student_number"), "")).strip()
        amt = row.get(cols.get("amount"))
        asof = row.get(cols.get("asof_year"), active_year)

        # Validate basics
        if not sn or pd.isna(amt):
            bad += 1
            results.append({
                "row": i, "student_number": sn, "amount": amt,
                "asof_year": asof, "status": "ERROR",
                "message": "Missing student_number or amount"
            })
            continue

        # Resolve student_id from student_number
        sid = resolve_student_id(conn, student_number=sn)
        if not sid:
            bad += 1
            results.append({
                "row": i, "student_number": sn, "amount": amt,
                "asof_year": asof, "status": "ERROR",
                "message": "Student not found"
            })
            continue

        # Coerce types (any failure -> error row)
        try:
            amount = float(amt)
            asof_year = int(asof) if str(asof).strip() else active_year
        except Exception:
            bad += 1
            results.append({
                "row": i, "student_number": sn, "amount": amt,
                "asof_year": asof, "status": "ERROR",
                "message": "Amount or asof_year is invalid"
            })
            continue

        # ✅ Save opening balance (no surrounding try/except; let hard errors surface)
        set_opening_balance(conn, sid, amount, asof_year,
                            note=f"opening_balance import row {i}")
        ok += 1
        results.append({
            "row": i, "student_number": sn, "amount": amount,
            "asof_year": asof_year, "status": "OK", "message": "saved"
        })

    # Commit once for the batch
    conn.commit()
    conn.close()

    if ok and not bad:
        flash(f"{ok} opening balances saved.", "success")
    elif ok and bad:
        flash(f"{ok} saved; {bad} failed. See details below.", "warning")
    else:
        flash("No rows saved. Check the errors below.", "danger")

    return render_template("opening_balance_bulk.html",
                           active_year=active_year,
                           results=results,
                           sample_rows=[])




@app.route("/balances/opening/template.csv")
@require_role("admin", "bursar", "headteacher")
def opening_balance_template_csv():
    """
    Download a CSV template: student_number,amount,asof_year
    """
    csv_text = "student_number,amount,asof_year\nSTD-2025-001,250000,2024\n"
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=opening_balance_template.csv"}
    )


@app.route("/balances/opening/template.xlsx")
@require_role("admin", "bursar", "headteacher")
def opening_balance_template_xlsx():
    """
    Download an XLSX template with the same columns.
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "opening_balance"
    ws.append(["student_number", "amount", "asof_year"])
    ws.append(["STD-2025-001", 250000, 2024])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="opening_balance_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )




# ========================== HOLIDAY PACKAGE MODULE (updated) ==========================
import sqlite3, io, csv
import pandas as pd
from flask import request, redirect, url_for, flash, render_template, session, Response
from datetime import datetime

# 1) Schema (idempotent)
def ensure_holiday_package_schema(conn=None):
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True
    c = conn.cursor()
    c.execute("""
      CREATE TABLE IF NOT EXISTS holiday_package_scores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        subject_id INTEGER NOT NULL,
        assessment_type TEXT,
        score REAL NOT NULL,
        max_score REAL NOT NULL DEFAULT 100,
        weight REAL,
        term TEXT NOT NULL,
        year INTEGER NOT NULL,
        date_recorded DATETIME DEFAULT CURRENT_TIMESTAMP
      )
    """)
    conn.commit()
    if close_after:
        conn.close()

# small internal guard so every route has schema available
def _hp_ensure():
    try:
        ensure_holiday_package_schema()
    except Exception:
        pass

# 2) Subject helper
def get_or_create_subject_by_name(name: str) -> int:
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    row = c.execute("SELECT id FROM subjects WHERE lower(name)=lower(?)", (name.strip(),)).fetchone()
    if row:
        sid = row["id"]
    else:
        c.execute("INSERT INTO subjects(name) VALUES (?)", (name.strip(),))
        sid = c.lastrowid
        conn.commit()
    conn.close()
    return sid

# 3) Add a score
def hp_add_score(student_id:int, subject_id:int, assessment_type:str,
                 score:float, term:str, year:int, max_score:float=100.0, weight=None):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
      INSERT INTO holiday_package_scores
      (student_id, subject_id, assessment_type, score, max_score, weight, term, year)
      VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (student_id, subject_id, (assessment_type or "").strip(), float(score),
          float(max_score or 100), (None if weight in ("", None) else float(weight)),
          term, int(year)))
    conn.commit()
    conn.close()

# 4) Aggregation
def hp_aggregate_student_subject(conn, student_id:int, subject_id:int, term:str, year:int) -> float|None:
    rows = conn.execute("""
      SELECT score, max_score, weight
      FROM holiday_package_scores
      WHERE student_id=? AND subject_id=? AND term=? AND year=?
    """, (student_id, subject_id, term, year)).fetchall()
    if not rows:
        return None
    any_weight = any(r["weight"] is not None for r in rows)
    if any_weight:
        total, total_w = 0.0, 0.0
        for r in rows:
            w = float(r["weight"] or 0.0)
            if w <= 0: 
                continue
            pct = (float(r["score"] or 0.0) / float(r["max_score"] or 100.0)) * 100.0
            total += pct * w
            total_w += w
        return (total / total_w) if total_w > 0 else None
    else:
        vals = [(float(r["score"]) / float(r["max_score"] or 100.0)) * 100.0 for r in rows if r["max_score"]]
        return (sum(vals)/len(vals)) if vals else None

# 5) Sync into record_score
def hp_sync_into_record_score(class_name:str, term:str, year:int, initials:str="HP"):
    ensure_record_score_table(get_db_connection())
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    students = c.execute(
        "SELECT id FROM students WHERE archived=0 AND class_name=?",
        (class_name,)
    ).fetchall()
    subject_rows = c.execute("""
      SELECT DISTINCT h.subject_id
      FROM holiday_package_scores h
      JOIN students s ON s.id=h.student_id
      WHERE s.class_name=? AND h.term=? AND h.year=?
    """, (class_name, term, year)).fetchall()
    subject_ids = [r["subject_id"] for r in subject_rows]
    if not students or not subject_ids:
        conn.close(); 
        return 0

    grading = c.execute("SELECT grade, lower_limit, upper_limit FROM grading_scale").fetchall()
    def _grade_from_mark(mark):
        if mark is None: return None
        for g in grading:
            if float(g["lower_limit"]) <= mark <= float(g["upper_limit"]):
                return g["grade"]
        return None

    n = 0
    for s in students:
        for subj_id in subject_ids:
            avg = hp_aggregate_student_subject(conn, s["id"], subj_id, term, year)
            if avg is None: 
                continue
            grd = _grade_from_mark(avg)
            c.execute("""
              INSERT INTO record_score
              (student_id, subject_id, term, year, average_mark, grade, comment, initials, processed_on)
              VALUES (?, ?, ?, ?, ?, ?, 'Holiday Package', ?, CURRENT_TIMESTAMP)
              ON CONFLICT(student_id, subject_id, term, year) DO UPDATE SET
                average_mark=excluded.average_mark,
                grade=excluded.grade,
                comment=excluded.comment,
                initials=excluded.initials,
                processed_on=CURRENT_TIMESTAMP
            """, (s["id"], subj_id, term, year, avg, grd, initials))
            n += 1
    conn.commit()
    conn.close()
    return n

# ---------- (NEW) CSV template download ----------
@app.route("/holiday/template.csv")
@require_role("admin","teacher","headteacher","dos")
def holiday_template_csv():
    _hp_ensure()
    ay = get_active_academic_year()
    term = ay.get("current_term") or ay.get("term") or "Term 1"
    year = int(ay.get("year"))

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["student_number","subject_name","assessment_type","score","max_score","weight","term","year"])
    w.writerow(["STD-2025-001","Holiday Package","Test","85","100","","%s"%term,"%d"%year])
    data = output.getvalue()
    output.close()
    return Response(
        data,
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="holiday_package_template.csv"'}
    )

# ---------- (NEW) subjects that actually have HP marks in class/period ----------
def _hp_subjects_for_class_term_year(conn, class_name: str, term: str, year: int):
    rows = conn.execute("""
      SELECT DISTINCT h.subject_id,
             (SELECT name FROM subjects s WHERE s.id=h.subject_id) AS name
      FROM holiday_package_scores h
      JOIN students st ON st.id = h.student_id
      WHERE st.class_name = ? AND h.term = ? AND h.year = ?
      ORDER BY name
    """, (class_name, term, year)).fetchall()
    return [{"id": r["subject_id"], "name": r["name"]} for r in rows]

# 6) Route (original logic kept; preview & template link added)
@app.route("/holiday/hub", methods=["GET","POST"])
@require_role("admin","teacher","headteacher","dos")
def holiday_hub():
    _hp_ensure()
    ay = get_active_academic_year()
    active_term = ay.get("current_term") or ay.get("term") or "Term 1"
    active_year = int(ay.get("year"))

    # -------- POST: ORIGINAL ACTIONS (unchanged) --------
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "add_one":
            sn = (request.form.get("student_number") or "").strip()
            sid = request.form.get("student_id")
            if not sid and sn:
                conn = get_db_connection(); conn.row_factory = sqlite3.Row
                r = conn.execute("SELECT id FROM students WHERE student_number=? AND archived=0",(sn,)).fetchone()
                conn.close()
                if not r:
                    flash("Student not found.","warning")
                    return redirect(url_for("holiday_hub"))
                sid = r["id"]

            subject_id = request.form.get("subject_id")
            subj_name = (request.form.get("subject_name") or "").strip()
            if not subject_id and subj_name:
                subject_id = get_or_create_subject_by_name(subj_name)

            try:
                hp_add_score(
                    int(sid),
                    int(subject_id),
                    request.form.get("assessment_type") or "Test",
                    float(request.form.get("score") or 0),
                    request.form.get("term") or active_term,
                    int(request.form.get("year") or active_year),
                    float(request.form.get("max_score") or 100),
                    request.form.get("weight")
                )
                flash("Holiday score recorded.","success")
            except Exception as e:
                flash(f"Failed: {e}","danger")
            return redirect(url_for("holiday_hub"))

        if action == "upload_csv":
            file = request.files.get("file")
            if not file:
                flash("No file","warning")
                return redirect(url_for("holiday_hub"))

            if file.filename.lower().endswith(".csv"):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)

            ok, skip = 0, 0
            conn = get_db_connection(); conn.row_factory = sqlite3.Row
            for _, r in df.iterrows():
                try:
                    rs = conn.execute(
                        "SELECT id FROM students WHERE student_number=?",
                        (str(r["student_number"]),)
                    ).fetchone()
                    if not rs:
                        skip += 1; 
                        continue
                    sid = rs["id"]
                    subj = get_or_create_subject_by_name(str(r["subject_name"]))
                    hp_add_score(
                        sid, subj,
                        str(r.get("assessment_type","Test")),
                        float(r["score"]),
                        str(r.get("term", active_term)),
                        int(r.get("year", active_year)),
                        float(r.get("max_score", 100)),
                        r.get("weight", None)
                    )
                    ok += 1
                except Exception:
                    skip += 1
            conn.close()
            flash(f"Uploaded {ok}, skipped {skip}.","info")
            return redirect(url_for("holiday_hub"))

        if action == "sync":
            class_name = (request.form.get("class_name") or "").strip()
            try:
                n = hp_sync_into_record_score(class_name, active_term, active_year, session.get("initials") or "HP")
                flash(f"Synced {n} scores","success")
            except Exception as e:
                flash(f"Sync failed: {e}","danger")
            return redirect(url_for("holiday_hub"))

    # -------- GET: options + (NEW) preview --------
    conn = get_db_connection(); conn.row_factory = sqlite3.Row
    classes = [r[0] for r in conn.execute("SELECT DISTINCT class_name FROM classes ORDER BY class_name").fetchall()]
    subjects = conn.execute("SELECT id,name FROM subjects ORDER BY name").fetchall()

    # Preview params (optional)
    preview_class = (request.args.get("class_name") or "").strip()
    preview_term = (request.args.get("term") or active_term).strip()
    preview_year = int(request.args.get("year") or active_year)
    do_preview = request.args.get("preview") == "1"

    preview = None
    if do_preview and preview_class:
        subj_list = _hp_subjects_for_class_term_year(conn, preview_class, preview_term, preview_year)
        students = conn.execute("""
          SELECT id, student_number, first_name, Middle_name, last_name
          FROM students
          WHERE archived=0 AND class_name=?
          ORDER BY last_name, first_name
        """, (preview_class,)).fetchall()

        rows = []
        for s in students:
            row = {
                "student_id": s["id"],
                "student_number": s["student_number"],
                "name": f'{s["first_name"]} {s["Middle_name"] or ""} {s["last_name"]}'.replace(" "," ").strip(),
                "subjects": {}
            }
            for sj in subj_list:
                avg = hp_aggregate_student_subject(conn, s["id"], sj["id"], preview_term, preview_year)
                row["subjects"][sj["id"]] = avg









# ======================= STUDENTS FINANCE REPORT ============================
import sqlite3
from io import BytesIO
import pandas as pd
from flask import request, render_template, send_file, flash, url_for
from datetime import datetime

# ---- If not already defined elsewhere in your code ----
def _term_order_val(t: str) -> int:
    t = (t or "").strip()
    return 1 if t == "Term 1" else 2 if t == "Term 2" else 3 if t == "Term 3" else 99

def _is_subscribed_this_term(conn, student_id: int, term: str, year: int):
    """
    Returns (is_subscribed: bool, route_name: str or '', fare: float).
    Safe if transport tables don't exist (returns False, '', 0.0).
    """
    try:
        conn.row_factory = sqlite3.Row
        row = conn.execute("""
            SELECT tr.name AS route_name, tr.fare_per_term AS fare, ts.start_term, ts.start_year
            FROM transport_subscriptions ts
            JOIN transport_routes tr ON tr.id = ts.route_id
            WHERE ts.student_id=? AND ts.active=1
        """, (student_id,)).fetchone()
        if not row:
            return False, "", 0.0
        # active if start <= current (term/year)
        start_rank = (_term_order_val(row["start_term"]), int(row["start_year"]))
        now_rank = (_term_order_val(term), int(year))
        if start_rank <= now_rank:
            return True, row["route_name"] or "", float(row["fare"] or 0.0)
        return False, "", 0.0
    except Exception:
        # transport tables may not exist yet
        return False, "", 0.0

def _transport_paid_total(conn, student_number: str, term: str, year: int) -> float:
    like_token = f"%SN: {student_number}%"
    row = conn.execute("""
        SELECT COALESCE(SUM(amount),0) AS total
        FROM other_income
        WHERE description='Transport' AND term=? AND year=? AND source LIKE ?
    """, (term, int(year), like_token)).fetchone()
    return float(row["total"] if row else 0.0)

def _expected_fees_for_class(conn, class_name: str) -> float:
    row = conn.execute("""
        SELECT amount
        FROM class_fees
        WHERE class_name = ?
        ORDER BY id DESC LIMIT 1
    """, (class_name,)).fetchone()
    return float(row["amount"]) if row and row["amount"] is not None else 0.0

def _expected_requirements_for_class(conn, class_name: str, term: str) -> float:
    row = conn.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM requirements
        WHERE class_name = ?
          AND (term = ? OR term IS NULL OR term = '')
    """, (class_name, term)).fetchone()
    return float(row["total"] if row else 0.0)

def _bursary_for_term(conn, student_id: int, term: str, year: int) -> float:
    row = conn.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM bursaries
        WHERE student_id=? AND term=? AND year=?
    """, (student_id, term, int(year))).fetchone()
    return float(row["total"] if row else 0.0)

def _fees_paid_for_term(conn, student_id: int, term: str, year: int) -> float:
    row = conn.execute("""
        SELECT COALESCE(SUM(amount_paid), 0) AS total
        FROM fees
        WHERE student_id=? AND term=? AND year=? AND payment_type='fees'
    """, (student_id, term, int(year))).fetchone()
    return float(row["total"] if row else 0.0)

def _requirements_paid_for_term(conn, student_id: int, term: str, year: int) -> float:
    row = conn.execute("""
        SELECT COALESCE(SUM(amount_paid), 0) AS total
        FROM fees
        WHERE student_id=? AND term=? AND year=? AND payment_type='requirements'
    """, (student_id, term, int(year))).fetchone()
    return float(row["total"] if row else 0.0)

def _filters_from_request():
    ay = get_active_academic_year()
    default_term = ay.get("current_term") or ay.get("term") or "Term 1"
    default_year = int(ay.get("year") or datetime.now().year)

    return {
        "student_number": (request.args.get("student_number") or "").strip(),
        "last_name": (request.args.get("last_name") or "").strip(),
        "class_name": (request.args.get("class_name") or "").strip(),
        "term": (request.args.get("term") or default_term).strip(),
        "year": int(request.args.get("year") or default_year),
        "export": (request.args.get("export") == "1"),
    }



@app.route("/reports/students_finance", methods=["GET"])
@require_role("admin","bursar","headteacher")
def students_finance_report():
    """
    Students finance report with Carried Forward baked into Overall Outstanding.
    Overall Outstanding can be negative (credit) to match the new netting logic.
    """
    f = _filters_from_request()
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # --- helpers ---
    def _term_order_val(term: str) -> int:
        t = (term or "").strip().lower()
        return 1 if t == "term 1" else 2 if t == "term 2" else 3 if t == "term 3" else 99

    def _opening_balance_total(conn, sid: int) -> float:
        row = conn.execute("""
            SELECT COALESCE(SUM(expected_amount - amount_paid), 0) AS total
              FROM fees
             WHERE student_id = ?
               AND lower(payment_type) = 'opening_balance'
        """, (sid,)).fetchone()
        return float(row["total"] or 0.0)

    def _carry_forward_prior_terms(conn, sid: int, term: str, year: int) -> float:
        rows = conn.execute("""
            SELECT expected_amount, bursary_amount, amount_paid, term AS t, year AS y
              FROM fees
             WHERE student_id = ?
               AND lower(payment_type) IN ('school_fees','fees')
               AND (y < ?
                 OR (y = ? AND
                     CASE lower(t)
                       WHEN 'term 1' THEN 1
                       WHEN 'term 2' THEN 2
                       WHEN 'term 3' THEN 3
                       ELSE 99
                     END < ?))
        """, (sid, year, year, _term_order_val(term))).fetchall()
        cf = 0.0
        for r in rows:
            exp = float(r["expected_amount"] or 0.0)
            bur = float(r["bursary_amount"] or 0.0)
            paid = float(r["amount_paid"] or 0.0)
            cf += (exp - bur - paid)
        # keep prior-terms carry-forward behavior as-is
        return max(cf, 0.0)

    # --- students (active) ---
    sql = "SELECT * FROM students WHERE archived=0"
    params = []
    if f["student_number"]:
        sql += " AND student_number = ?"
        params.append(f["student_number"])
    if f["last_name"]:
        sql += " AND last_name LIKE ?"
        params.append(f"%{f['last_name']}%")
    if f["class_name"]:
        sql += " AND class_name = ?"
        params.append(f["class_name"])
    sql += " ORDER BY class_name, stream, last_name, first_name"
    students = conn.execute(sql, params).fetchall()

    # --- build rows ---
    rows = []
    totals = {
        "fees_expected": 0.0, "fees_paid": 0.0, "fees_balance": 0.0,
        "req_expected": 0.0, "req_paid": 0.0, "req_balance": 0.0,
        "tr_due": 0.0, "tr_paid": 0.0, "tr_balance": 0.0,
        "carried_forward": 0.0,
        "overall_expected": 0.0, "overall_paid": 0.0,
        "overall_outstanding": 0.0 # may be negative (credit)
    }

    for s in students:
        sid = int(s["id"])
        sn = s["student_number"]
        cls = s["class_name"]
        strm = s["stream"]

        fees_expected = _expected_fees_for_student(conn, sid, f["term"], f["year"])
        req_expected = _expected_requirements_for_class(conn, cls, f["term"])
        bursary = _bursary_for_term(conn, sid, f["term"], f["year"])

        fees_paid = _fees_paid_for_term(conn, sid, f["term"], f["year"])
        req_paid = _requirements_paid_for_term(conn, sid, f["term"], f["year"])

        # Transport (due if subscribed)
        is_sub, route_name, fare = _is_subscribed_this_term(conn, sid, f["term"], f["year"])
        tr_due = fare if is_sub else 0.0
        tr_paid = _transport_paid_total(conn, sn, f["term"], f["year"]) if is_sub else 0.0

        # Balances by head (leave per-head balances clamped as you had)
        fees_balance = max((fees_expected - bursary) - fees_paid, 0.0)
        req_balance = max(req_expected - req_paid, 0.0)
        tr_balance = max(tr_due - tr_paid, 0.0)

        # Totals for this term
        overall_expected = max(fees_expected - bursary, 0.0) + req_expected + tr_due
        overall_paid = fees_paid + req_paid + tr_paid

        # Carried forward (OB + prior tuition shortfall)
        opening_balance = _opening_balance_total(conn, sid)
        prior_shortfall = _carry_forward_prior_terms(conn, sid, f["term"], f["year"])
        carried_forward = opening_balance + prior_shortfall

        # ---- KEY CHANGE: allow negative (credit) ----
        # Remove the max(..., 0.0) clamp so credits show as negative.
        overall_outstanding = (overall_expected - overall_paid) + carried_forward

        row = {
            "student_number": sn,
            "full_name": f"{s['first_name']} {(s['Middle_name'] or '')} {s['last_name']}".replace(" "," ").strip(),
            "class_name": cls,
            "stream": strm,
            "fees_expected": fees_expected,
            "fees_paid": fees_paid,
            "fees_balance": fees_balance,
            "req_expected": req_expected,
            "req_paid": req_paid,
            "req_balance": req_balance,
            "transport_route": route_name if is_sub else "",
            "tr_due": tr_due,
            "tr_paid": tr_paid,
            "tr_balance": tr_balance,
            "carried_forward": carried_forward, # visible column
            "overall_expected": overall_expected,
            "overall_paid": overall_paid,
            "overall_outstanding": overall_outstanding, # can be negative (credit)
        }
        rows.append(row)

        # totals
        for k in totals.keys():
            totals[k] += row.get(k, 0.0)

    # dropdown helpers
    class_options = [r[0] for r in conn.execute(
        "SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL ORDER BY class_name"
    ).fetchall()]
    conn.close()

    # Export
    if f["export"]:
        import pandas as pd
        from io import BytesIO
        df = pd.DataFrame(rows)
        cols = [
            "student_number","full_name","class_name","stream",
            "fees_expected","fees_paid","fees_balance",
            "req_expected","req_paid","req_balance",
            "transport_route","tr_due","tr_paid","tr_balance",
            "carried_forward",
            "overall_expected","overall_paid","overall_outstanding" # may be negative
        ]
        df = df[cols]
        totals_row = {
            "student_number":"", "full_name":"TOTALS", "class_name":"", "stream":"",
            **{k: round(v,2) for k,v in totals.items()}
        }
        df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Students Finance")
        bio.seek(0)
        filename = f"students_finance_{f['term']}_{f['year']}.xlsx"
        return send_file(
            bio, as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template(
        "students_finance_report.html",
        rows=rows, totals=totals,
        terms=TERMS,
        filters=f,
        class_options=class_options,
    )

# ===================== /STUDENTS FINANCE REPORT ============================



# ---------- STUDENT STATEMENT ----------
from collections import defaultdict

# Use this name (or keep your renamed 'student_statement') consistently in url_for(...)
@app.route("/student-statement", methods=["GET"])
@require_role("admin","director","bursar","headteacher")
def student_statement():
    """
    If no query is provided, render a simple search form.
    If student_number or last_name is provided, find the student and redirect to their statement.
    """
    sn = (request.args.get("student_number") or "").strip()
    ln = (request.args.get("last_name") or "").strip()

    # No input → show search page (do NOT redirect)
    if not sn and not ln:
        # Optional: show the info message on the page instead of flashing
        return render_template("student_statement_search.html")

    # With input → do lookup
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    try:
        if sn:
            stu = conn.execute("""
                SELECT id, student_number, first_name,
                       COALESCE(Middle_name,'') AS middle_name, last_name,
                       class_name, stream, section, sex, parent_name, parent_contact
                FROM students
                WHERE student_number=? AND archived=0
                LIMIT 1
            """, (sn,)).fetchone()
        else:
            stu = conn.execute("""
                SELECT id, student_number, first_name,
                       COALESCE(Middle_name,'') AS middle_name, last_name,
                       class_name, stream, section, sex, parent_name, parent_contact
                FROM students
                WHERE last_name LIKE ? AND archived=0
                ORDER BY last_name, first_name
                LIMIT 1
            """, (f"%{ln}%",)).fetchone()
    finally:
        conn.close()

    if not stu:
        flash("Student not found or archived.", "warning")
        return redirect(url_for("student_statement"))

    return redirect(url_for("student_statement_by_id", student_id=stu["id"]))


@app.route("/student-statement/<int:student_id>", methods=["GET"])
@require_role("admin","bursar","headteacher","director")
def student_statement_by_id(student_id: int):
    """
    Printable Student Statement (transactions + summary).
    Uses your existing compute_student_financials for the active term/year.
    Also ensures every fees row has a receipt_no; generates one if missing.
    """
    # ensure schema has the receipt_no column
    try:
        ensure_fees_has_receipt_no()
    except Exception:
        # don't break view if migration helper errors; continue gracefully
        pass

    # active term/year
    ay = get_active_academic_year() or {}
    term = (ay.get("current_term") or ay.get("term") or "Term 1")
    try:
        year = int(ay.get("year") or ay.get("active_year") or datetime.now().year)
    except Exception:
        year = datetime.now().year

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    try:
        # student
        stu = conn.execute("""
            SELECT id, student_number, first_name,
                   COALESCE(Middle_name,'') AS middle_name, last_name,
                   class_name, stream, section, sex, parent_name, parent_contact, residence
            FROM students
            WHERE id=? LIMIT 1
        """, (student_id,)).fetchone()
        if not stu:
            flash("Student not found.", "warning")
            return redirect(url_for("register_student"))

        # transactions (full history). We now include: comment & receipt_no
        tx = conn.execute("""
            SELECT id, term, year, date_paid, method, payment_type,
                   amount_paid, expected_amount, bursary_amount, carried_forward,
                   COALESCE(requirement_name, '') AS requirement_name,
                   COALESCE(comment, '') AS comment,
                   receipt_no
            FROM fees
            WHERE student_id = ?
            ORDER BY year,
                     CASE LOWER(term)
                        WHEN 'term 1' THEN 1
                        WHEN 'term 2' THEN 2
                        WHEN 'term 3' THEN 3
                        ELSE 99 END,
                     id
        """, (student_id,)).fetchall()

        # backfill missing receipt numbers in-place
        missing = [row["id"] for row in tx if not row["receipt_no"]]
        if missing:
            for fee_id in missing:
                rcpt = generate_receipt_no(conn, fee_id)
                # If a very rare collision happens (e.g., same day and manual reuse),
                # append a short suffix based on ROWID/time to keep it unique.
                try:
                    conn.execute("UPDATE fees SET receipt_no=? WHERE id=?", (rcpt, fee_id))
                except sqlite3.IntegrityError:
                    rcpt2 = f"{rcpt}-{fee_id%1000:03d}"
                    conn.execute("UPDATE fees SET receipt_no=? WHERE id=?", (rcpt2, fee_id))
            conn.commit()

            # re-fetch with the newly assigned receipt_no values
            tx = conn.execute("""
                SELECT id, term, year, date_paid, method, payment_type,
                       amount_paid, expected_amount, bursary_amount, carried_forward,
                       COALESCE(requirement_name, '') AS requirement_name,
                       COALESCE(comment, '') AS comment,
                       receipt_no
                FROM fees
                WHERE student_id = ?
                ORDER BY year,
                         CASE LOWER(term)
                            WHEN 'term 1' THEN 1
                            WHEN 'term 2' THEN 2
                            WHEN 'term 3' THEN 3
                            ELSE 99 END,
                         id
            """, (student_id,)).fetchall()
    finally:
        conn.close()

    # financial summary for the ACTIVE term/year
    fin = compute_student_financials(student_id, stu["class_name"], term, year)

    # Group transactions by year->term for friendly display
    grouped = defaultdict(lambda: defaultdict(list))
    for r in tx:
        grouped[r["year"]][r["term"]].append(r)

    # school header data
    school = {
        "name": current_app.config.get("SCHOOL_NAME", "CITIZEN DAY AND BOARDING PRI SCHOOL"),
        "address": current_app.config.get("SCHOOL_ADDRESS", "P.O. Box 31882, Kampala"),
        "phone": current_app.config.get("SCHOOL_PHONE", "(+256) 781757410/788529084/704720641"),
        "logo_url": url_for("static", filename="logo.jpg"),
    }

    return render_template(
        "student_statement.html",
        school=school,
        student=stu,
        grouped=grouped,
        fin=fin,
        active_term=term,
        active_year=year,
        today=datetime.now().strftime("%d %b %Y %H:%M")
    )
# ---------- /STUDENT STATEMENT ----------






with app.app_context():
    bootstrap()


if __name__ == "__main__":
    configure_logging(app)
    with app.app_context():
        bootstrap()
        #populate_default_expense_categories()
        #apply_schema_guards(app)
        #upsert_admin_user()
        #add_created_at_if_missing()
    app.run(host="0.0.0.0", port=1400, debug=True)


















